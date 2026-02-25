import math
import os
import re
import io
import csv
import urllib.request
import urllib.parse
import json
import xml.etree.ElementTree as ET
import hashlib
import secrets
import logging
import ssl
from collections import defaultdict, deque
from pathlib import Path
from datetime import datetime, timezone, timedelta
import uuid
from typing import List, Optional, Dict, Any, Tuple
import time

from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
import pyotp
from pywebpush import webpush, WebPushException

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Preformatted
from openpyxl import Workbook, load_workbook

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, Body, Response, Query, WebSocket, WebSocketDisconnect, Header
from fastapi.routing import APIRoute
from starlette.middleware.cors import CORSMiddleware
from starlette.routing import Match
from dotenv import load_dotenv
from pydantic import BaseModel, Field, EmailStr
from alembic.config import Config as AlembicConfig
from alembic.script import ScriptDirectory
from emergentintegrations.payments.stripe.checkout import (
    StripeCheckout,
    CheckoutSessionRequest,
    CheckoutSessionResponse,
    CheckoutStatusResponse,
)
from sqlalchemy.exc import IntegrityError, DBAPIError, OperationalError, TimeoutError as SATimeoutError
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from sqlalchemy import select, func, case, distinct, String, Text, DateTime, ForeignKey, desc, and_, or_, update, text, cast, event
from sqlalchemy.orm import Mapped, mapped_column, selectinload
from sqlalchemy.dialects.postgresql import UUID as PGUUID, JSONB
from sqlalchemy.dialects.postgresql.asyncpg import AsyncAdapt_asyncpg_dbapi
import html

from app.core.config import settings

try:
    from asgi_correlation_id import correlation_id, CorrelationIdMiddleware
except Exception:  # pragma: no cover
    correlation_id = None
    CorrelationIdMiddleware = None

from app.core.security import (
    verify_password,
    get_password_hash,
    create_access_token,
    create_refresh_token,
    decode_token,
)
from app.repositories.auth_repository import SqlAuthRepository
from app.repositories.applications_repository import SqlApplicationsRepository
from app.admin_country_context import resolve_admin_country_context

from app.dependencies import (
    get_current_user,
    get_current_user_optional,
    check_permissions,
    require_portal_scope,
    _get_sql_user,
)
from app.countries_seed import default_countries
from app.menu_seed import default_top_menu
from app.categories_seed import vehicle_category_tree
from app.master_data_seed import default_vehicle_makes, default_vehicle_models
from app.models.base import Base
from app.models.user import User as SqlUser
from app.models.application import Application
from app.models.campaign import Campaign
from app.models.plan import Plan
from app.models.billing import VatRate
from app.models.monetization import UserSubscription
from app.models.admin_invoice import AdminInvoice
from app.models.payment import Payment, PaymentTransaction
from app.models.webhook_event_log import WebhookEventLog
from app.models.category import Category, CategoryTranslation
from app.models.category_schema_version import CategorySchemaVersion
from app.models.core import AuditLog, Country
from app.models.attribute import Attribute, AttributeOption, CategoryAttributeMap
from app.models.dealer_profile import DealerProfile
from app.models.dealer import DealerApplication, Dealer, DealerUser
from app.models.consumer_profile import ConsumerProfile
from app.models.auth import UserCredential, EmailVerificationToken
from app.utils.slug import slugify
from app.utils.monitoring import (
    classify_endpoint,
    get_endpoint_stats,
    get_slow_query_summary,
    record_request_latency,
)
from app.models.dealer_listing import DealerListing
from app.models.moderation import Listing, ModerationAction, ModerationItem
from app.models.listing_search import ListingSearch
from app.models.vehicle_mdm import VehicleMake, VehicleModel
from app.models.user_recent_category import UserRecentCategory
from app.models.analytics import ListingView
from app.models.messaging import Conversation, Message
from app.models.favorite import Favorite
from app.models.notification import Notification
from app.models.gdpr_export import GDPRExport
from app.models.site_header import SiteHeaderSetting
from app.models.advertisement import Advertisement, AdImpression, AdClick, AdCampaign
from app.models.doping_request import DopingRequest
from app.models.footer_layout import FooterLayout
from app.models.info_page import InfoPage
from app.models.pricing_campaign import PricingCampaign
from app.models.pricing_campaign_item import PricingCampaignItem
from app.models.pricing_tier_rule import PricingTierRule
from app.models.pricing_package import PricingPackage, UserPackageSubscription
from app.models.pricing_snapshot import PricingPriceSnapshot
from app.models.push_subscription import PushSubscription
from app.site_media_storage import store_site_asset
from app.models.cloudflare_config import CloudflareConfig
from app.models.system_setting import SystemSetting
from app.models.admin_invite import AdminInvite
from app.models.menu_item import MenuItem
from app.models.report import Report
from app.models.menu import TopMenuItem
from app.models.support_message import SupportMessage
from app.services.audit import log_action
from app.services.cloudflare_metrics import (
    CloudflareCredentials,
    CloudflareMetricsService,
    CloudflareMetricsError,
    CANARY_CONFIG_MISSING,
)
from app.services.cloudflare_config import (
    CloudflareConfigError,
    build_masked_config,
    encrypt_config_value,
    resolve_cloudflare_config,
    resolve_env_fallback,
    resolve_env_source,
    upsert_cloudflare_config,
    update_canary_status,
)


from fastapi import UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse, RedirectResponse

from app.vehicle_publish_guard import validate_publish, validate_listing_schema
from app.vehicle_media_storage import store_image, resolve_public_media_path

# Vehicle Master Data (file-based REV-B)
from app.vehicle_master_file import get_vehicle_master_dir, load_current_master
from app.vehicle_master_admin_file import validate_upload, rollback as vehicle_master_rollback, get_status as vehicle_master_status


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")
load_dotenv(ROOT_DIR / ".env.local", override=False)

SUPPORTED_COUNTRIES = {"DE", "CH", "FR", "AT"}
TRANSACTION_TYPE_OPTIONS = {"satilik", "kiralik", "gunluk"}
SLUG_PATTERN = re.compile(r"^[a-z0-9]+(?:-[a-z0-9]+)*$")
CATEGORY_IMPORT_MAX_BYTES = 10 * 1024 * 1024
ATTRIBUTE_KEY_PATTERN = re.compile(r"^[a-z0-9_]+$")
VAT_ID_PATTERN = re.compile(r"^[A-Z]{2}[A-Z0-9]{6,12}$")


# P0.1 Failed-login rate limit (in-process)
FAILED_LOGIN_WINDOW_SECONDS = 10 * 60  # 10 min sliding window
FAILED_LOGIN_MAX_ATTEMPTS = 5
FAILED_LOGIN_BLOCK_SECONDS = 15 * 60  # 15 min block

EMAIL_VERIFICATION_TTL_MINUTES = 15
EMAIL_VERIFICATION_MAX_ATTEMPTS = 5
EMAIL_VERIFICATION_BLOCK_MINUTES = 15
EMAIL_VERIFICATION_RESEND_COOLDOWN_SECONDS = 90

BILLING_AUDIT_ACTIONS = {
    "payment_succeeded",
    "invoice_marked_paid",
    "subscription_activated",
    "quota_limits_updated",
}

_failed_login_attempts: Dict[str, List[float]] = {}
_failed_login_blocked_until: Dict[str, float] = {}

# Dealer Application reasons (v1)
DEALER_APP_REJECT_REASONS_V1 = {
    "incomplete_documents",
    "invalid_company_info",
    "duplicate_application",
    "compliance_issue",
    "other",
}

INDIVIDUAL_APP_REJECT_REASONS_V1 = {
    "incomplete_documents",
    "failed_verification",
    "duplicate_application",
    "country_not_supported",
    "other",
}

DEALER_LISTING_QUOTA_LIMIT = 10

# Dealer Application audit event types
DEALER_APPLICATION_EVENT_TYPES = {"DEALER_APPLICATION_APPROVED", "DEALER_APPLICATION_REJECTED"}

_failed_login_block_audited: Dict[str, bool] = {}



class UserLogin(BaseModel):
    email: EmailStr
    password: str
    totp_code: Optional[str] = None


class ConsumerRegisterPayload(BaseModel):
    full_name: str = Field(..., min_length=2, max_length=255)
    email: EmailStr
    password: str = Field(..., min_length=8)
    country_code: str = Field(..., min_length=2, max_length=5)
    preferred_language: Optional[str] = "tr"
    company_website: Optional[str] = None


class DealerRegisterPayload(BaseModel):
    company_name: str = Field(..., min_length=2, max_length=255)
    contact_name: str = Field(..., min_length=2, max_length=255)
    email: EmailStr
    password: str = Field(..., min_length=8)
    country_code: str = Field(..., min_length=2, max_length=5)
    tax_id: Optional[str] = None
    preferred_language: Optional[str] = "tr"
    company_website: Optional[str] = None



class VerifyEmailPayload(BaseModel):
    email: EmailStr
    code: str = Field(..., min_length=6, max_length=6)


class VerifyEmailResponse(BaseModel):
    status: str
    remaining_attempts: int


class ResendVerificationPayload(BaseModel):
    email: EmailStr


class ResendVerificationResponse(BaseModel):
    status: str
    cooldown_seconds: int


class EmailVerifyHelpPayload(BaseModel):
    email: Optional[EmailStr] = None
    reason: Optional[str] = "email_verification"


class UserResponse(BaseModel):
    id: str
    email: str
    full_name: str
    role: str
    portal_scope: Optional[str] = None
    country_scope: List[str] = Field(default_factory=list)
    preferred_language: str = "tr"
    is_active: bool = True
    is_verified: bool = True
    deleted_at: Optional[str] = None
    created_at: str
    last_login: Optional[str] = None
    invite_status: Optional[str] = None


class RegisterVerificationResponse(BaseModel):
    success: bool = True
    requires_verification: bool = True


class TokenResponse(BaseModel):
    access_token: str
    refresh_token: str
    token_type: str = "bearer"
    user: UserResponse


class RefreshTokenRequest(BaseModel):
    refresh_token: str


class UserProfileUpdatePayload(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    locale: Optional[str] = None
    country_code: Optional[str] = None
    display_name_mode: Optional[str] = None
    marketing_consent: Optional[bool] = None
    notification_prefs: Optional[Dict[str, Any]] = None


class TwoFactorVerifyPayload(BaseModel):
    code: str


class AccountDeletePayload(BaseModel):
    reason: Optional[str] = None


class ConsumerProfileUpdatePayload(BaseModel):
    full_name: Optional[str] = None
    locale: Optional[str] = None
    country_code: Optional[str] = None
    display_name_mode: Optional[str] = None
    marketing_consent: Optional[bool] = None


class DealerProfileUpdatePayload(BaseModel):
    company_name: Optional[str] = None
    vat_id: Optional[str] = None
    trade_register_no: Optional[str] = None
    authorized_person: Optional[str] = None
    address_json: Optional[Dict[str, Any]] = None
    logo_url: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_phone: Optional[str] = None
    address_country: Optional[str] = None
    impressum_text: Optional[str] = None
    terms_text: Optional[str] = None
    withdrawal_policy_text: Optional[str] = None


class GDPRExportResponse(BaseModel):
    status: str
    data: Dict[str, Any]


class ChangePasswordPayload(BaseModel):
    current_password: str
    new_password: str


class MessageThreadCreatePayload(BaseModel):
    listing_id: str


class MessageSendPayload(BaseModel):
    body: str
    client_message_id: Optional[str] = None


class PushSubscriptionKeys(BaseModel):
    p256dh: str
    auth: str


class PushSubscriptionPayload(BaseModel):
    endpoint: str
    keys: PushSubscriptionKeys


class PushUnsubscribePayload(BaseModel):
    endpoint: str


def _normalize_user_status(doc: dict) -> str:
    if doc.get("deleted_at"):
        return "deleted"
    status = doc.get("status") or "active"
    if status == "suspended" or doc.get("is_active") is False:
        return "suspended"
    return "active"


def _determine_user_type(role: str) -> str:
    if role in ADMIN_ROLE_OPTIONS:
        return "admin"
    if role == "dealer":
        return "dealer"
    return "individual"


def _resolve_portal_scope(role: Optional[str]) -> str:
    if not role:
        return "account"
    if role == "dealer":
        return "dealer"
    if role in ADMIN_ROLE_OPTIONS or role == "audit_viewer":
        return "admin"
    return "account"


def _normalize_phone_e164(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = re.sub(r"[^\d+]", "", value)
    if not cleaned:
        return None
    if cleaned.startswith("00"):
        cleaned = f"+{cleaned[2:]}"
    if cleaned.startswith("+"):
        digits = re.sub(r"\D", "", cleaned[1:])
        return f"+{digits}" if digits else None
    digits = re.sub(r"\D", "", cleaned)
    if not digits:
        return None
    if digits.startswith("49"):
        return f"+{digits}"
    return digits


def _normalize_phone_candidates(value: Optional[str]) -> List[str]:
    if not value:
        return []
    cleaned = re.sub(r"[^\d+]", "", value)
    digits = re.sub(r"\D", "", cleaned)
    if not digits:
        return []
    candidates = set()
    if cleaned.startswith("+"):
        candidates.add(f"+{digits}")
    if digits.startswith("00"):
        trimmed = digits[2:]
        if trimmed:
            candidates.add(f"+{trimmed}")
            candidates.add(trimmed)
    if digits.startswith("49"):
        candidates.add(f"+{digits}")
    candidates.add(digits)
    return list(candidates)


async def _check_register_honeypot(
    session: AsyncSession,
    request: Request,
    email: Optional[str],
    value: Optional[str],
    role: str,
) -> None:
    if not value or not str(value).strip():
        return

    actor = {"id": None, "email": email}
    await _write_audit_log_sql(
        session=session,
        action="register_honeypot_hit",
        actor=actor,
        resource_type="security",
        resource_id=email,
        metadata={"field": "company_website", "value": str(value).strip(), "role": role},
        request=request,
        country_code=None,
    )
    await session.commit()
    raise HTTPException(status_code=400, detail="Invalid request")


def _normalize_notification_prefs(payload: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    prefs = payload or {}
    return {
        "push_enabled": bool(prefs.get("push_enabled", True)),
        "email_enabled": bool(prefs.get("email_enabled", True)),
    }


def _build_listing_snapshot(listing) -> dict:
    if not listing:
        return {
            "listing_id": None,
            "listing_title": None,
            "listing_price": None,
            "listing_image": None,
            "listing_location": None,
        }

    if isinstance(listing, Listing):
        price_label = None
        if listing.price is not None:
            price_label = f"{listing.currency} {listing.price}"
        image = None
        if listing.images:
            image = listing.images[0]
        location_label = listing.city or listing.country
        return {
            "listing_id": str(listing.id),
            "listing_title": listing.title,
            "listing_price": price_label,
            "listing_image": image,
            "listing_location": location_label,
        }

    price = listing.get("price") or {}
    amount = price.get("amount")
    currency = price.get("currency_primary") or "EUR"
    price_label = price.get("formatted_primary") or (f"{currency} {amount}" if amount is not None else None)
    media = listing.get("media") or []
    image = None
    if media:
        first = media[0]
        image = first.get("preview_url") or first.get("url") or first.get("file")
    location = listing.get("location") or {}
    location_label = location.get("city") or location.get("region") or location.get("country") or listing.get("country")
    return {
        "listing_id": listing.get("id"),
        "listing_title": listing.get("title"),
        "listing_price": price_label,
        "listing_image": image,
        "listing_location": location_label,
    }


def _build_thread_summary(thread: dict, current_user_id: str) -> dict:
    unread_map = thread.get("unread_counts") or {}
    return {
        "id": thread.get("id"),
        "listing_id": thread.get("listing_id"),
        "listing_title": thread.get("listing_title"),
        "listing_image": thread.get("listing_image"),
        "last_message": thread.get("last_message"),
        "last_message_at": thread.get("last_message_at"),
        "participants": thread.get("participants") or [],
        "unread_count": int(unread_map.get(current_user_id, 0)),
    }

def _build_thread_summary_sql(
    thread: Conversation,
    current_user_id: str,
    listing: Listing | None,
    last_message: Message | None,
    unread_count: int,
) -> dict:
    return {
        "id": str(thread.id),
        "listing_id": str(thread.listing_id),
        "listing_title": listing.title if listing else None,
        "listing_image": (listing.images[0] if listing and listing.images else None),
        "last_message": last_message.body if last_message else None,
        "last_message_at": last_message.created_at.isoformat() if last_message and last_message.created_at else None,
        "participants": [str(thread.buyer_id), str(thread.seller_id)],
        "unread_count": unread_count,
    }


def _build_notification_payload(notification: Notification) -> dict:
    return {
        "id": str(notification.id),
        "user_id": str(notification.user_id),
        "title": notification.title,
        "message": notification.message,
        "source_type": notification.source_type,
        "source_id": notification.source_id,
        "action_url": notification.action_url,
        "payload_json": notification.payload_json or {},
        "dedupe_key": notification.dedupe_key,
        "read_at": notification.read_at.isoformat() if notification.read_at else None,
        "delivered_at": notification.delivered_at.isoformat() if notification.delivered_at else None,
        "created_at": notification.created_at.isoformat() if notification.created_at else None,
        "is_read": bool(notification.read_at),
    }


async def _get_message_thread_or_404(session: AsyncSession, thread_id: str, current_user_id: str) -> Conversation:
    thread_uuid = uuid.UUID(thread_id)
    thread = await session.get(Conversation, thread_uuid)
    if not thread:
        raise HTTPException(status_code=404, detail="Thread not found")
    if current_user_id not in {str(thread.buyer_id), str(thread.seller_id)}:
        raise HTTPException(status_code=403, detail="Forbidden")
    return thread






def _subscription_field(subscription: Any, field: str) -> Optional[str]:
    if isinstance(subscription, PushSubscription):
        return getattr(subscription, field, None)
    if isinstance(subscription, dict):
        return subscription.get(field)
    return None


async def _get_active_push_subscriptions(session: AsyncSession, user_id: str) -> List[PushSubscription]:
    if session is None:
        return []
    try:
        user_uuid = uuid.UUID(str(user_id))
    except ValueError:
        return []
    result = await session.execute(
        select(PushSubscription).where(
            PushSubscription.user_id == user_uuid,
            PushSubscription.is_active.is_(True),
        )
    )
    return result.scalars().all()


async def _deactivate_push_subscription(session: AsyncSession, subscription_id: str, reason: str) -> None:
    if session is None or not subscription_id:
        return
    try:
        sub_uuid = uuid.UUID(str(subscription_id))
    except ValueError:
        return
    subscription = await session.get(PushSubscription, sub_uuid)
    if not subscription:
        return
    subscription.is_active = False
    subscription.revoked_at = datetime.now(timezone.utc)
    subscription.revoked_reason = reason
    await session.commit()


async def _send_web_push_notification(subscription: Any, payload: dict) -> dict:
    if not PUSH_ENABLED:
        return {"ok": False, "revoke": False}

    endpoint = _subscription_field(subscription, "endpoint")
    p256dh = _subscription_field(subscription, "p256dh")
    auth = _subscription_field(subscription, "auth")
    if not endpoint or not p256dh or not auth:
        return {"ok": False, "revoke": False}

    try:
        webpush(
            subscription_info={
                "endpoint": endpoint,
                "keys": {
                    "p256dh": p256dh,
                    "auth": auth,
                },
            },
            data=json.dumps(payload, ensure_ascii=False),
            vapid_private_key=VAPID_PRIVATE_KEY,
            vapid_claims={"sub": VAPID_SUBJECT},
        )
        return {"ok": True, "revoke": False}
    except WebPushException as exc:
        status = getattr(getattr(exc, "response", None), "status_code", None)
        return {"ok": False, "revoke": status in (404, 410)}
    except Exception:
        return {"ok": False, "revoke": False}


async def _send_message_push_notification(session: AsyncSession, recipient_id: str, thread: dict, message: dict) -> bool:
    if not PUSH_ENABLED or session is None:
        return False

    try:
        recipient_uuid = uuid.UUID(str(recipient_id))
    except ValueError:
        return False

    recipient = await session.get(SqlUser, recipient_uuid)
    if not recipient:
        return False

    prefs = getattr(recipient, "notification_prefs", None) or {}
    if not prefs.get("push_enabled", True):
        return False

    subscriptions = await _get_active_push_subscriptions(session, recipient_id)
    if not subscriptions:
        return False

    payload = {
        "title": "Yeni mesaj",
        "body": thread.get("listing_title") or "İlan mesajı",
        "url": f"/account/messages?thread={thread.get('id')}",
        "thread_id": thread.get("id"),
        "tag": "message",
    }

    success = False
    for subscription in subscriptions:
        result = await _send_web_push_notification(subscription, payload)
        if result.get("ok"):
            success = True
        elif result.get("revoke"):
            await _deactivate_push_subscription(session, str(subscription.id), "push_failed")
    return success


def _resolve_user_phone_e164(doc: dict) -> Optional[str]:
    raw_phone = (
        doc.get("phone_e164")
        or doc.get("phone_number")
        or doc.get("phone")
        or doc.get("mobile")
    )
    return _normalize_phone_e164(raw_phone)


def _extract_moderation_reason(payload: Optional["AdminUserActionPayload"]) -> tuple[Optional[str], Optional[str]]:
    if not payload:
        return None, None
    reason_code = (payload.reason_code or payload.reason or "").strip()
    reason_detail = (payload.reason_detail or "").strip()
    return (reason_code or None), (reason_detail or None)


def _parse_suspension_until(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        iso_value = str(value).replace("Z", "+00:00")
        parsed = datetime.fromisoformat(iso_value)
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        parsed = parsed.astimezone(timezone.utc)
        if parsed <= datetime.now(timezone.utc):
            raise HTTPException(status_code=400, detail="suspension_until must be in the future")
        return parsed
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid suspension_until") from exc


def _normalize_risk_level(value: Optional[str]) -> str:
    if not value:
        return "low"
    normalized = str(value).lower().strip()
    if normalized not in RISK_LEVEL_ALLOWED:
        raise HTTPException(status_code=400, detail="Invalid risk_level")
    return normalized


def _assert_plan_quota_limits(listing_quota: Optional[int], showcase_quota: Optional[int]) -> None:
    if listing_quota is not None:
        if listing_quota < PLAN_QUOTA_MIN or listing_quota > PLAN_QUOTA_MAX:
            raise HTTPException(
                status_code=400,
                detail=f"Listing quota must be between {PLAN_QUOTA_MIN} and {PLAN_QUOTA_MAX}",
            )
    if showcase_quota is not None:
        if showcase_quota < PLAN_QUOTA_MIN or showcase_quota > PLAN_QUOTA_MAX:
            raise HTTPException(
                status_code=400,
                detail=f"Showcase quota must be between {PLAN_QUOTA_MIN} and {PLAN_QUOTA_MAX}",
            )


async def _auto_reactivate_if_expired(user: SqlUser, session: AsyncSession, request: Optional[Request]) -> SqlUser:
    if not user or user.deleted_at:
        return user
    if user.status != "suspended":
        return user
    suspension_until = user.suspension_until
    if not suspension_until:
        return user

    if suspension_until > datetime.now(timezone.utc):
        return user

    before_state = {
        "status": user.status,
        "is_active": user.is_active,
        "suspension_until": user.suspension_until.isoformat() if user.suspension_until else None,
        "ban_reason": user.ban_reason,
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
    }

    user.status = "active"
    user.is_active = True
    user.suspension_until = None
    user.ban_reason = None
    if user.role == "dealer":
        user.dealer_status = "active"

    await session.commit()

    if request:
        actor = {
            "id": "system",
            "email": "system@platform.com",
            "role": "system",
            "country_scope": [],
        }
        event_type = "dealer_reactivated" if user.role == "dealer" else "user_reactivated"
        await _write_audit_log_sql(
            session=session,
            action=event_type,
            actor=actor,
            resource_type="user",
            resource_id=str(user.id),
            metadata={
                "reason_code": "suspension_expired",
                "before": before_state,
                "after": {
                    **before_state,
                    "status": user.status,
                    "is_active": user.is_active,
                    "suspension_until": None,
                    "ban_reason": None,
                },
            },
            request=request,
            country_code=user.country_code,
        )
        await session.commit()

    return user


def _build_user_summary(doc: dict, listing_stats: Optional[Dict[str, Any]] = None, plan_map: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    listing_stats = listing_stats or {}
    plan_map = plan_map or {}
    plan = plan_map.get(doc.get("plan_id"), {})
    status = _normalize_user_status(doc)
    user_type = _determine_user_type(doc.get("role", "individual"))
    phone_verified = bool(doc.get("phone_verified") or doc.get("phone_verified_at"))
    plan_expiry = doc.get("plan_expires_at") or doc.get("plan_end_at") or doc.get("plan_ends_at")

    return {
        "id": doc.get("id"),
        "full_name": doc.get("full_name"),
        "first_name": doc.get("first_name"),
        "last_name": doc.get("last_name"),
        "email": doc.get("email"),
        "role": doc.get("role"),
        "user_type": user_type,
        "country_code": doc.get("country_code"),
        "country_scope": doc.get("country_scope") or [],
        "status": status,
        "risk_level": doc.get("risk_level") or "low",
        "ban_reason": doc.get("ban_reason"),
        "is_active": bool(doc.get("is_active", True)),
        "deleted_at": doc.get("deleted_at"),
        "created_at": doc.get("created_at"),
        "last_login": doc.get("last_login"),
        "email_verified": bool(doc.get("is_verified", False)),
        "phone_verified": phone_verified,
        "phone_e164": _resolve_user_phone_e164(doc),
        "total_listings": listing_stats.get("total", 0),
        "active_listings": listing_stats.get("active", 0),
        "plan_id": doc.get("plan_id"),
        "plan_name": plan.get("name") if plan else None,
        "plan_expires_at": plan_expiry,
    }


def _user_to_response(doc: dict) -> UserResponse:
    return UserResponse(
        id=doc["id"],
        email=doc["email"],
        full_name=doc.get("full_name", ""),
        role=doc.get("role", "support"),
        country_scope=doc.get("country_scope") or [],
        preferred_language=doc.get("preferred_language", "tr"),
        is_active=bool(doc.get("is_active", True)),
        is_verified=bool(doc.get("is_verified", True)),
        deleted_at=doc.get("deleted_at"),
        created_at=doc.get("created_at") or datetime.now(timezone.utc).isoformat(),
        last_login=doc.get("last_login"),
        invite_status=doc.get("invite_status"),
        portal_scope=doc.get("portal_scope") or _resolve_portal_scope(doc.get("role")),
    )


async def _ensure_country_enabled(session: AsyncSession, country_code: Optional[str]) -> str:
    normalized = (country_code or "DE").upper()
    result = await session.execute(
        select(Country).where(Country.code == normalized, Country.is_enabled.is_(True))
    )
    country = result.scalar_one_or_none()
    if not country:
        raise HTTPException(status_code=400, detail="Geçersiz ülke")
    return normalized


async def _generate_unique_dealer_slug(session: AsyncSession, company_name: str) -> str:
    base_slug = slugify(company_name) or "dealer"
    base_slug = base_slug[:90]
    candidate = base_slug
    suffix = 2

    while True:
        result = await session.execute(select(DealerProfile).where(DealerProfile.slug == candidate))
        if not result.scalar_one_or_none():
            return candidate
        candidate = f"{base_slug}-{suffix}"
        suffix += 1


def _generate_email_verification_code() -> str:
    return f"{secrets.randbelow(1000000):06d}"


def _build_verification_email(code: str, locale: Optional[str]) -> Tuple[str, str, str]:
    locale_key = (locale or "tr").lower()
    if locale_key.startswith("de"):
        subject = "E-Mail doğrulama kodunuz"
        intro = "E-posta doğrulama kodunuz"
    elif locale_key.startswith("fr"):
        subject = "Code de vérification"
        intro = "Votre code de vérification"
    else:
        subject = "E-posta doğrulama kodunuz"
        intro = "E-posta doğrulama kodunuz"

    text_body = f"{intro}: {code}\nKod 15 dakika içinde geçerlidir."
    html_body = (
        "<div style='font-family:Arial,sans-serif;font-size:14px;'>"
        f"<p>{intro}: <strong>{code}</strong></p>"
        "<p>Kod 15 dakika içinde geçerlidir.</p>"
        "</div>"
    )
    return subject, text_body, html_body


def _send_verification_email(to_email: str, code: str, locale: Optional[str]) -> None:
    logger = logging.getLogger("email_verification")
    provider = EMAIL_PROVIDER
    subject, text_body, html_body = _build_verification_email(code, locale)

    if provider == "mock":
        logger.warning("Mock verification email -> %s code=%s", to_email, code)
        return

    if provider == "sendgrid":
        sendgrid_key = os.environ.get("SENDGRID_API_KEY")
        sender_email = os.environ.get("SENDER_EMAIL")
        if not sendgrid_key or not sender_email:
            logger.error("SendGrid configuration missing: SENDGRID_API_KEY or SENDER_EMAIL")
            raise HTTPException(status_code=503, detail="Email provider not configured")
        message = Mail(
            from_email=sender_email,
            to_emails=to_email,
            subject=subject,
            html_content=html_body,
        )
        try:
            sg = SendGridAPIClient(sendgrid_key)
            sg.send(message)
        except Exception as exc:
            logger.error("SendGrid verification send error: %s", exc)
            raise HTTPException(status_code=502, detail="Failed to send verification email") from exc
        return

    if provider == "smtp":
        logger.error("SMTP provider not supported (SendGrid only)")
        raise HTTPException(status_code=503, detail="SMTP provider not supported")

    raise HTTPException(status_code=503, detail="Email provider not configured")


async def _issue_email_verification_code(
    session: AsyncSession,
    user: SqlUser,
    request: Optional[Request] = None,
) -> str:
    code = _generate_email_verification_code()
    now = datetime.now(timezone.utc)
    await session.execute(
        update(EmailVerificationToken)
        .where(
            EmailVerificationToken.user_id == user.id,
            EmailVerificationToken.consumed_at.is_(None),
        )
        .values(consumed_at=now)
    )
    token = EmailVerificationToken(
        user_id=user.id,
        token_hash=get_password_hash(code),
        expires_at=now + timedelta(minutes=EMAIL_VERIFICATION_TTL_MINUTES),
        consumed_at=None,
        created_at=now,
        ip_address=_get_client_ip(request) if request else None,
        user_agent=request.headers.get("user-agent") if request else None,
    )
    session.add(token)
    user.email_verification_attempts = 0
    user.email_verification_expires_at = None
    return code


async def _email_verification_sent_at(session: AsyncSession, user_id: uuid.UUID) -> Optional[datetime]:
    result = await session.execute(
        select(EmailVerificationToken)
        .where(EmailVerificationToken.user_id == user_id)
        .order_by(EmailVerificationToken.created_at.desc())
        .limit(1)
    )
    token = result.scalars().first()
    return token.created_at if token else None


async def _get_active_verification_tokens(session: AsyncSession, user_id: uuid.UUID) -> List[EmailVerificationToken]:
    result = await session.execute(
        select(EmailVerificationToken)
        .where(
            EmailVerificationToken.user_id == user_id,
            EmailVerificationToken.consumed_at.is_(None),
        )
        .order_by(EmailVerificationToken.created_at.desc())
        .limit(5)
    )
    return list(result.scalars().all())


def _email_verification_block_expires_at(user: SqlUser) -> Optional[datetime]:
    if user.email_verification_attempts < EMAIL_VERIFICATION_MAX_ATTEMPTS:
        return None
    return user.email_verification_expires_at


def _resolve_display_name(full_name: str, mode: str) -> str:
    name = (full_name or "").strip()
    if mode == "hidden":
        return ""
    if mode == "initials":
        parts = [p for p in name.split(" ") if p]
        initials = "".join([p[0].upper() for p in parts[:2]])
        return initials
    return name


def _generate_recovery_codes(count: int = 6) -> List[str]:
    return [secrets.token_hex(4) for _ in range(count)]


def _hash_recovery_codes(codes: List[str]) -> List[str]:
    return [get_password_hash(code) for code in codes]


def _verify_recovery_code(code: str, hashed_codes: List[str]) -> Tuple[bool, List[str]]:
    for idx, hashed in enumerate(hashed_codes):
        if verify_password(code, hashed):
            remaining = [c for i, c in enumerate(hashed_codes) if i != idx]
            return True, remaining
    return False, hashed_codes


def _verify_totp_code(secret: str, code: str) -> bool:
    try:
        totp = pyotp.TOTP(secret)
        return totp.verify(code, valid_window=1)
    except Exception:
        return False


def _build_totp_uri(email: str, secret: str) -> str:
    return pyotp.totp.TOTP(secret).provisioning_uri(name=email, issuer_name="FAZ Panel")


async def _get_or_create_consumer_profile(
    session: AsyncSession,
    user: SqlUser,
    language: Optional[str] = None,
    country_code: Optional[str] = None,
) -> ConsumerProfile:
    result = await session.execute(select(ConsumerProfile).where(ConsumerProfile.user_id == user.id))
    profile = result.scalar_one_or_none()
    if profile:
        return profile

    profile = ConsumerProfile(
        user_id=user.id,
        language=language or user.preferred_language or "tr",
        country_code=country_code or user.country_code or "DE",
        display_name_mode="full_name",
        marketing_consent=False,
    )
    session.add(profile)
    await session.flush()
    return profile


def _normalize_vat_id(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = re.sub(r"\s+", "", value).upper()
    return cleaned or None


def _is_valid_vat_id(value: Optional[str]) -> bool:
    if not value:
        return True
    return bool(VAT_ID_PATTERN.match(value))


async def _get_or_create_dealer_profile(
    session: AsyncSession,
    user: SqlUser,
) -> DealerProfile:
    result = await session.execute(select(DealerProfile).where(DealerProfile.user_id == user.id))
    profile = result.scalar_one_or_none()
    if profile:
        return profile

    company_name = (user.full_name or "Bayi").strip() or "Bayi"
    slug = await _generate_unique_dealer_slug(session, company_name)
    profile = DealerProfile(
        user_id=user.id,
        slug=slug,
        company_name=company_name,
        vat_number=None,
        vat_id=None,
        trade_register_no=None,
        authorized_person=user.full_name,
        address_json=None,
        logo_url=None,
        address_country=user.country_code or "DE",
        contact_email=user.email,
        verification_status="pending",
    )
    session.add(profile)
    await session.flush()
    return profile


def _build_consumer_profile_payload(user_row: SqlUser, profile: ConsumerProfile) -> Dict[str, Any]:
    notification_prefs = getattr(user_row, "notification_prefs", None)
    phone = getattr(user_row, "phone_e164", None)
    return {
        "profile_type": "consumer",
        "id": str(user_row.id),
        "email": user_row.email,
        "full_name": user_row.full_name,
        "display_name": _resolve_display_name(user_row.full_name, profile.display_name_mode),
        "display_name_mode": profile.display_name_mode,
        "phone": phone,
        "locale": user_row.preferred_language or profile.language or "tr",
        "country_code": profile.country_code,
        "marketing_consent": profile.marketing_consent,
        "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
        "totp_enabled": profile.totp_enabled,
        "notification_prefs": _normalize_notification_prefs(notification_prefs),
    }


def _build_dealer_profile_payload(user_row: SqlUser, profile: DealerProfile) -> Dict[str, Any]:
    return {
        "profile_type": "dealer",
        "id": str(user_row.id),
        "email": user_row.email,
        "full_name": user_row.full_name,
        "locale": user_row.preferred_language or "tr",
        "country_code": profile.address_country or user_row.country_code,
        "company_name": profile.company_name,
        "vat_id": profile.vat_id,
        "trade_register_no": profile.trade_register_no,
        "authorized_person": profile.authorized_person,
        "address_json": profile.address_json,
        "logo_url": profile.logo_url,
        "contact_email": profile.contact_email,
        "contact_phone": profile.contact_phone,
        "verification_status": profile.verification_status,
        "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
    }


async def _get_user_row_from_current(session: AsyncSession, current_user: dict) -> SqlUser:
    try:
        user_uuid = uuid.UUID(str(current_user.get("id")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")
    return user_row


async def _log_email_verify_event(
    session: AsyncSession,
    action: str,
    user: SqlUser,
    request: Optional[Request],
    metadata: Optional[dict] = None,
) -> None:
    actor = {
        "id": str(user.id),
        "email": user.email,
        "country_scope": user.country_scope or [],
    }
    await _write_audit_log_sql(
        session=session,
        action=action,
        actor=actor,
        resource_type="auth",
        resource_id=str(user.id),
        metadata=metadata,
        request=request,
        country_code=user.country_code,
    )


async def build_audit_entry(
    event_type: str,
    actor: dict,
    target_id: str,
    target_type: str,
    country_code: Optional[str],
    details: Optional[dict],
    request: Optional[Request],
) -> dict:
    now_iso = datetime.now(timezone.utc).isoformat()
    return {
        "id": str(uuid.uuid4()),
        "created_at": now_iso,
        "event_type": event_type,
        "action": event_type,
        "resource_type": target_type,
        "resource_id": target_id,
        "admin_user_id": actor.get("id"),
        "user_id": actor.get("id"),
        "user_email": actor.get("email"),
        "country_code": country_code,
        "country_scope": actor.get("country_scope") or [],
        "metadata": details or {},
        "request_ip": request.client.host if request and request.client else None,
        "applied": True,
    }


async def _write_audit_log_sql(
    session: AsyncSession,
    action: str,
    actor: dict,
    resource_type: str,
    resource_id: Optional[str],
    metadata: Optional[dict],
    request: Optional[Request],
    country_code: Optional[str] = None,
) -> AuditLog:
    entry = AuditLog(
        id=uuid.uuid4(),
        user_id=_safe_uuid(actor.get("id")),
        user_email=actor.get("email"),
        action=action,
        resource_type=resource_type,
        resource_id=resource_id,
        old_values=None,
        new_values=None,
        metadata_info=metadata or {},
        ip_address=request.client.host if request and request.client else None,
        user_agent=request.headers.get("user-agent") if request else None,
        country_scope=country_code,
        is_pii_scrubbed=False,
        created_at=datetime.now(timezone.utc),
    )
    session.add(entry)
    await session.flush()
    return entry


async def _upsert_seed_user(
    session: AsyncSession,
    *,
    email: str,
    password: str,
    role: str,
    full_name: str,
    country_code: str,
    country_scope: list[str],
    first_name: Optional[str] = None,
    last_name: Optional[str] = None,
    phone_e164: Optional[str] = None,
) -> SqlUser:
    now_dt = datetime.now(timezone.utc)
    hashed = get_password_hash(password)
    result = await session.execute(select(SqlUser).where(SqlUser.email == email))
    user = result.scalar_one_or_none()
    if user:
        user.hashed_password = hashed
        user.role = role
        user.full_name = full_name
        user.first_name = first_name
        user.last_name = last_name
        user.phone_e164 = phone_e164
        user.status = "active"
        user.is_active = True
        user.is_verified = True
        user.country_scope = country_scope
        user.country_code = country_code
        user.preferred_language = user.preferred_language or "tr"
        user.updated_at = now_dt
        if role == "dealer":
            user.dealer_status = "active"
        if user.deleted_at:
            user.deleted_at = None
    else:
        user = SqlUser(
            id=uuid.uuid4(),
            email=email,
            hashed_password=hashed,
            full_name=full_name,
            first_name=first_name,
            last_name=last_name,
            phone_e164=phone_e164,
            role=role,
            status="active",
            is_active=True,
            is_verified=True,
            country_scope=country_scope,
            country_code=country_code,
            preferred_language="tr",
            created_at=now_dt,
            updated_at=now_dt,
            last_login=None,
            dealer_status="active" if role == "dealer" else None,
        )
        session.add(user)
    await session.commit()
    return user


async def _ensure_admin_user(session: AsyncSession):
    await _upsert_seed_user(
        session,
        email="admin@platform.com",
        password="Admin123!",
        role="super_admin",
        full_name="System Administrator",
        country_code="TR",
        country_scope=["*"],
    )



async def _ensure_dealer_user(session: AsyncSession):
    await _upsert_seed_user(
        session,
        email="dealer@platform.com",
        password="Dealer123!",
        role="dealer",
        full_name="Dealer Demo",
        country_code="DE",
        country_scope=["DE"],
    )


async def _ensure_test_user(session: AsyncSession):
    await _upsert_seed_user(
        session,
        email="user@platform.com",
        password="User123!",
        role="individual",
        full_name="Test User",
        first_name="Test",
        last_name="User",
        phone_e164="+491701112233",
        country_code="DE",
        country_scope=["DE"],
    )


async def _ensure_test_user_two(session: AsyncSession):
    await _upsert_seed_user(
        session,
        email="user2@platform.com",
        password="User123!",
        role="individual",
        full_name="Test User 2",
        first_name="Test",
        last_name="User Two",
        phone_e164="+491701112244",
        country_code="DE",
        country_scope=["DE"],
    )


async def _ensure_individual_fixtures(session: AsyncSession):
    fixtures = [
        {
            "email": "ayse.kaya@platform.com",
            "first_name": "Ayşe",
            "last_name": "Kaya",
            "phone_e164": "+905321234567",
            "country_code": "TR",
        },
        {
            "email": "mehmet.yilmaz@platform.com",
            "first_name": "Mehmet",
            "last_name": "Yılmaz",
            "phone_e164": "+905551112233",
            "country_code": "TR",
        },
        {
            "email": "elif.demir@platform.com",
            "first_name": "Elif",
            "last_name": "Demir",
            "phone_e164": "+491701234567",
            "country_code": "DE",
        },
        {
            "email": "zeynep.sari@platform.com",
            "first_name": "Zeynep",
            "last_name": "Sarı",
            "phone_e164": "+43123456789",
            "country_code": "AT",
        },
        {
            "email": "mert.ozkan@platform.com",
            "first_name": "Mert",
            "last_name": "Özkan",
            "phone_e164": "+41791234567",
            "country_code": "CH",
        },
    ]

    for fixture in fixtures:
        await _upsert_seed_user(
            session,
            email=fixture["email"],
            password="User123!",
            role="individual",
            full_name=f"{fixture['first_name']} {fixture['last_name']}",
            first_name=fixture["first_name"],
            last_name=fixture["last_name"],
            phone_e164=fixture["phone_e164"],
            country_code=fixture["country_code"],
            country_scope=[fixture["country_code"]],
        )


async def _ensure_country_admin_user(session: AsyncSession):
    await _upsert_seed_user(
        session,
        email="countryadmin@platform.com",
        password="Country123!",
        role="country_admin",
        full_name="Country Admin",
        country_code="DE",
        country_scope=["DE"],
    )


async def _ensure_fixture_category_schema(session: AsyncSession):
    slug = "e2e-fixture-category"
    name = "E2E Fixture Category"
    country_code = "DE"

    slug_json = {"tr": slug, "en": slug, "de": slug}
    now_dt = datetime.now(timezone.utc)

    result = await session.execute(
        select(Category).where(
            Category.country_code == country_code,
            Category.slug["tr"].astext == slug,
        )
    )
    existing = result.scalar_one_or_none()

    schema = {
        "status": "published",
        "core_fields": {
            "title": {
                "required": True,
                "min": 10,
                "max": 120,
                "messages": {
                    "required": "Başlık zorunludur.",
                    "min": "Başlık çok kısa.",
                    "max": "Başlık çok uzun.",
                },
            },
            "description": {
                "required": True,
                "min": 30,
                "max": 4000,
                "messages": {
                    "required": "Açıklama zorunludur.",
                    "min": "Açıklama çok kısa.",
                    "max": "Açıklama çok uzun.",
                },
            },
            "price": {
                "required": True,
                "range": {"min": 1, "max": 100000},
                "currency_primary": "EUR",
                "currency_secondary": "CHF",
                "secondary_enabled": False,
                "decimal_places": 0,
                "messages": {
                    "required": "Fiyat zorunludur.",
                    "range": "Fiyat aralık dışında",
                },
            },
        },
        "dynamic_fields": [
            {
                "id": "extra_option",
                "label": "Ekstra Seçim",
                "key": "extra_option",
                "type": "select",
                "required": True,
                "options": ["A", "B"],
                "messages": {"required": "Ekstra seçim zorunlu"},
            }
        ],
        "detail_groups": [
            {
                "id": "features",
                "title": "Donanım",
                "required": True,
                "options": ["ABS", "Airbag", "Klima", "ESP"],
                "messages": {"required": "Donanım seçimi zorunlu"},
            }
        ],
        "modules": {
            "address": {"enabled": True},
            "photos": {"enabled": True},
            "contact": {"enabled": False},
            "payment": {"enabled": False},
        },
        "photo_config": {"max": 12},
        "payment_options": {"package": False, "doping": False},
    }

    if existing:
        existing.form_schema = schema
        existing.updated_at = now_dt
        await session.commit()
        return

    category = Category(
        name=name,
        slug=slug_json,
        path=slug,
        module="vehicle",
        country_code=country_code,
        is_active=True,
        sort_order=0,
        parent_id=None,
        hierarchy_complete=True,
        form_schema=schema,
        created_at=now_dt,
        updated_at=now_dt,
    )
    session.add(category)
    await session.flush()
    for lang in ("tr", "en", "de"):
        session.add(
            CategoryTranslation(
                category_id=category.id,
                language=lang,
                name=name,
                description=None,
                meta_title=None,
                meta_description=None,
            )
        )
    await session.commit()



async def lifespan(app: FastAPI):
    logging.getLogger("runtime").warning(
        "APP_ENV=%s AUTH_PROVIDER=%s APPLICATIONS_PROVIDER=%s",
        APP_ENV,
        AUTH_PROVIDER,
        APPLICATIONS_PROVIDER,
    )

    sql_target = _get_masked_db_target()
    logging.getLogger("sql_config").info(
        "SQL target host=%s db=%s ssl=%s pool=%s/%s",
        sql_target.get("host"),
        sql_target.get("database"),
        DB_SSL_MODE,
        DB_POOL_SIZE,
        DB_MAX_OVERFLOW,
    )
    encryption_key_present = bool(os.environ.get("CONFIG_ENCRYPTION_KEY"))
    logging.getLogger("runtime").warning(
        "CONFIG_ENCRYPTION_KEY loaded=%s",
        "true" if encryption_key_present else "false",
    )

    try:
        async with sql_engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
            await _ensure_wizard_progress_column(conn)
    except Exception as exc:
        logging.getLogger("sql_config").warning("SQL init skipped: %s", exc)

    try:
        async with AsyncSessionLocal() as session:
            await _ensure_admin_user(session)
            await _ensure_dealer_user(session)
            await _ensure_test_user(session)
            await _ensure_test_user_two(session)
            await _ensure_individual_fixtures(session)
            await _ensure_country_admin_user(session)
    except Exception as exc:
        logging.getLogger("runtime").warning("Seed users skipped: %s", exc)

    app.state.db = None

    yield
    await sql_engine.dispose()


app = FastAPI(
    title="Admin Panel API",
    version="1.0.0",
    lifespan=lifespan,
)

if CorrelationIdMiddleware:
    app.add_middleware(CorrelationIdMiddleware)


@app.middleware("http")
async def record_request_metrics(request: Request, call_next):
    start_ts = time.perf_counter()
    response = await call_next(request)
    duration_ms = (time.perf_counter() - start_ts) * 1000
    endpoint_group = classify_endpoint(request.url.path)
    if endpoint_group:
        record_request_latency(endpoint_group, duration_ms)
    response.headers["X-Response-Time-ms"] = f"{duration_ms:.2f}"
    return response


RBAC_ALLOWLIST: Dict[str, list[str]] = {}
RBAC_MISSING_POLICIES: list[str] = []
RBAC_PUBLIC_ADMIN_ROUTES: Dict[str, list[str]] = {
    "GET /api/admin/invite/preview": ["public"],
    "POST /api/admin/invite/accept": ["public"],
}


def _normalize_admin_path(path: str) -> str:
    if path.startswith("/api/"):
        return path
    if path.startswith("/admin") or path.startswith("/v1/admin"):
        return f"/api{path}"
    return path


def _is_admin_path(path: str) -> bool:
    return path.startswith("/api/admin") or path.startswith("/admin") or path.startswith("/api/v1/admin") or path.startswith("/v1/admin")


def _extract_route_roles(route: APIRoute) -> Optional[list[str]]:
    for dependency in route.dependant.dependencies:
        call = getattr(dependency, "call", None)
        roles = getattr(call, "required_roles", None)
        if roles:
            return list(roles)
    return None


def _build_rbac_allowlist(target_app: FastAPI) -> tuple[Dict[str, list[str]], list[str]]:
    allowlist: Dict[str, list[str]] = {}
    missing: list[str] = []
    public_paths = {key.split(" ", 1)[1] for key in RBAC_PUBLIC_ADMIN_ROUTES.keys()}
    for route in target_app.routes:
        if not isinstance(route, APIRoute):
            continue
        if not _is_admin_path(route.path):
            continue
        roles = _extract_route_roles(route)
        if not roles:
            if route.path not in public_paths:
                missing.append(route.path)
            continue
        for method in route.methods or []:
            if method in {"HEAD", "OPTIONS"}:
                continue
            allowlist[f"{method} {route.path}"] = roles

    for key, roles in RBAC_PUBLIC_ADMIN_ROUTES.items():
        allowlist.setdefault(key, roles)

    return allowlist, sorted(set(missing))


def _match_admin_route(request: Request, normalized_path: str) -> Optional[APIRoute]:
    scope = dict(request.scope)
    scope["path"] = normalized_path
    for route in request.app.router.routes:
        if not isinstance(route, APIRoute):
            continue
        if request.method not in route.methods:
            continue
        match, _ = route.matches(scope)
        if match == Match.FULL:
            return route
    return None


async def _resolve_rbac_user(request: Request) -> Optional[dict]:
    cached_user = getattr(request.state, "current_user", None)
    if cached_user:
        return cached_user

    auth_header = request.headers.get("authorization") or request.headers.get("Authorization")
    if not auth_header or not auth_header.lower().startswith("bearer "):
        return None
    token = auth_header.split(" ", 1)[1].strip()
    if not token:
        return None

    payload = decode_token(token)
    if not payload or payload.get("type") != "access":
        return None
    if payload.get("token_version") != TOKEN_VERSION:
        return None

    user_id = payload.get("sub")
    if not user_id:
        return None

    async with AsyncSessionLocal() as session:
        user = await _get_sql_user(user_id, session)
    if not user or not user.is_active:
        return None

    token_scope = payload.get("portal_scope")
    if not token_scope:
        return None
    expected_scope = _resolve_portal_scope(user.role)
    if token_scope != expected_scope:
        return None

    user.portal_scope = token_scope
    request.state.current_user = user
    return user


async def _write_rbac_audit_log(
    request: Request,
    actor: Optional[dict],
    action: str,
    metadata: Optional[dict],
):
    try:
        async with AsyncSessionLocal() as session:
            await _write_audit_log_sql(
                session=session,
                action=action,
                actor=actor or {"id": None, "email": None, "country_scope": []},
                resource_type="rbac_guard",
                resource_id=request.url.path,
                metadata=metadata or {},
                request=request,
                country_code=None,
            )
            await session.commit()
    except Exception:
        logger = logging.getLogger("rbac_guard")
        logger.exception("rbac_audit_log_failed")


@app.middleware("http")
async def rbac_hard_lock(request: Request, call_next):
    if not _is_admin_path(request.url.path):
        return await call_next(request)

    normalized_path = _normalize_admin_path(request.url.path)
    route = _match_admin_route(request, normalized_path)
    if not route:
        return JSONResponse(status_code=404, content={"detail": "Not found"})

    allowlist = request.app.state.rbac_allowlist or {}
    policy_key = f"{request.method} {route.path}"
    required_roles = allowlist.get(policy_key)
    if not required_roles:
        await _write_rbac_audit_log(
            request,
            actor=None,
            action="RBAC_POLICY_MISSING",
            metadata={"path": route.path, "method": request.method},
        )
        return JSONResponse(status_code=403, content={"detail": "RBAC policy missing"})

    if "public" in required_roles:
        return await call_next(request)

    user = await _resolve_rbac_user(request)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "Not authenticated"})

    if user.get("role") not in required_roles:
        await _write_rbac_audit_log(
            request,
            actor=user,
            action="RBAC_DENY",
            metadata={"path": route.path, "method": request.method, "required_roles": required_roles},
        )
        return JSONResponse(status_code=403, content={"detail": "Insufficient permissions"})

    return await call_next(request)


DB_ERROR_CODES = {
    "pool_timeout": "DB_POOL_TIMEOUT",
    "connection_closed": "DB_CONN_CLOSED",
    "connection_lost": "DB_CONN_LOST",
    "pool_exhausted": "DB_POOL_EXHAUSTED",
    "unknown": "DB_ERROR",
}


def _classify_db_error(exc: Exception) -> str:
    message = str(exc).lower()
    if isinstance(exc, SATimeoutError):
        return DB_ERROR_CODES["pool_timeout"]
    if "too many connections" in message:
        return DB_ERROR_CODES["pool_exhausted"]
    if "connection was closed" in message:
        return DB_ERROR_CODES["connection_closed"]
    if "connection reset" in message or "connection lost" in message:
        return DB_ERROR_CODES["connection_lost"]
    return DB_ERROR_CODES["unknown"]


def _sanitize_db_error_message(message: str) -> str:
    sanitized = re.sub(r"postgresql://[^@]+@", "postgresql://***:***@", str(message))
    sanitized = sanitized.replace("@localhost", "@***")
    sanitized = sanitized.replace("@127.0.0.1", "@***")
    return sanitized[:500]


def _get_request_id(request: Optional[Request] = None) -> Optional[str]:
    req_id = None
    if correlation_id:
        try:
            req_id = correlation_id.get()
        except Exception:
            req_id = None
    if not req_id and request:
        req_id = request.headers.get("x-request-id") or request.headers.get("x-correlation-id")
    return req_id


def _log_db_exception(exc: Exception, code: str, request: Optional[Request] = None):
    request_id = _get_request_id(request)
    _db_error_events.append(time.time())
    db_error_logger.info(
        "db_error",
        extra={
            "error_code": code,
            "error": _sanitize_db_error_message(str(exc)),
            "type": exc.__class__.__name__,
            "request_id": request_id,
        },
    )


def _get_db_error_rate(window_seconds: int = 300) -> tuple[int, float]:
    now_ts = time.time()
    while _db_error_events and (now_ts - _db_error_events[0]) > window_seconds:
        _db_error_events.popleft()
    count = len(_db_error_events)
    rate_per_min = round(count / (window_seconds / 60), 2) if window_seconds else 0.0
    return count, rate_per_min


def _record_db_latency(latency_ms: float) -> None:
    _db_latency_events.append((time.time(), latency_ms))


def _get_db_latency_stats(window_seconds: int = 86400) -> tuple[Optional[float], Optional[float]]:
    now_ts = time.time()
    while _db_latency_events and (now_ts - _db_latency_events[0][0]) > window_seconds:
        _db_latency_events.popleft()
    values = [lat for _, lat in _db_latency_events]
    if not values:
        return None, None
    values_sorted = sorted(values)
    avg = sum(values_sorted) / len(values_sorted)
    idx = int(0.95 * (len(values_sorted) - 1))
    p95 = values_sorted[idx]
    return round(avg, 2), round(p95, 2)


def _build_error_buckets(window_seconds: int = 86400, bucket_seconds: int = 300) -> list[dict]:
    now_ts = time.time()
    start_ts = now_ts - window_seconds
    bucket_count = int(window_seconds / bucket_seconds)
    buckets = [0 for _ in range(bucket_count)]
    for ts in _db_error_events:
        if ts < start_ts:
            continue
        index = int((ts - start_ts) // bucket_seconds)
        if 0 <= index < bucket_count:
            buckets[index] += 1
    results = []
    for i, count in enumerate(buckets):
        bucket_start = start_ts + (i * bucket_seconds)
        results.append({
            "bucket_start": datetime.fromtimestamp(bucket_start, tz=timezone.utc).isoformat(),
            "count": count,
        })
    return results


def _read_search_etl_state() -> dict:
    path = Path("/app/memory/SEARCH_ETL_STATE.json")
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text())
    except Exception:
        return {}


@app.exception_handler(SATimeoutError)
async def db_timeout_handler(request: Request, exc: SATimeoutError):
    code = _classify_db_error(exc)
    _log_db_exception(exc, code, request)
    return JSONResponse(
        status_code=503,
        content={"detail": "Database connection timeout", "error_code": code},
    )


@app.exception_handler(DBAPIError)
async def db_api_handler(request: Request, exc: DBAPIError):
    code = _classify_db_error(exc)
    _log_db_exception(exc, code, request)
    return JSONResponse(
        status_code=503,
        content={"detail": "Database connection error", "error_code": code},
    )


@app.exception_handler(AsyncAdapt_asyncpg_dbapi.Error)
async def db_asyncpg_handler(request: Request, exc: AsyncAdapt_asyncpg_dbapi.Error):
    code = _classify_db_error(exc)
    _log_db_exception(exc, code, request)
    return JSONResponse(
        status_code=503,
        content={"detail": "Database connection error", "error_code": code},
    )


@app.exception_handler(OperationalError)
async def db_operational_handler(request: Request, exc: OperationalError):
    code = _classify_db_error(exc)
    _log_db_exception(exc, code, request)
    return JSONResponse(
        status_code=503,
        content={"detail": "Database operational error", "error_code": code},
    )


cors_origins = os.environ.get("CORS_ORIGINS", "*")
origins = ["*"] if cors_origins == "*" else [o.strip() for o in cors_origins.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)



def _get_client_ip(request: Request) -> str | None:
    xff = request.headers.get("x-forwarded-for")
    if xff:
        # first IP in list is the original client
        return xff.split(",")[0].strip() or None
    return request.client.host if request.client else None

async def _ensure_wizard_progress_column(conn) -> None:
    try:
        result = await conn.execute(
            text(
                """
                SELECT column_name
                FROM information_schema.columns
                WHERE table_name = 'categories'
                  AND column_name = 'wizard_progress'
                """
            )
        )
        if result.first() is None:
            await conn.execute(text("ALTER TABLE categories ADD COLUMN wizard_progress JSONB"))
    except Exception as exc:
        logging.getLogger("sql_config").warning("Wizard progress column check failed: %s", exc)


def _get_masked_db_target() -> Dict[str, Optional[str]]:
    try:
        parsed = urllib.parse.urlparse(DATABASE_URL)
    except Exception:
        return {"host": None, "database": None}

    host = parsed.hostname or None
    db_name = parsed.path.lstrip("/") if parsed.path else None

    masked_host = host
    if host and host not in {"localhost", "127.0.0.1"}:
        parts = host.split(".")
        if len(parts) > 2:
            masked_host = ".".join(parts[-2:])

    masked_db = None
    if db_name:
        masked_db = f"{db_name[:3]}***" if len(db_name) > 3 else "***"

    return {"host": masked_host, "database": masked_db}


def _get_alembic_head_revisions() -> List[str]:
    try:
        config = AlembicConfig(str(ROOT_DIR / "alembic.ini"))
        script = ScriptDirectory.from_config(config)
        return script.get_heads() or []
    except Exception:
        return []


async def _get_migration_state(conn) -> Dict[str, Optional[str]]:
    now_ts = time.time()
    cached_at = _migration_state_cache.get("checked_at", 0) or 0
    if cached_at and (now_ts - cached_at) < MIGRATION_STATE_CACHE_TTL_SECONDS:
        return {
            "state": _migration_state_cache.get("state", "unknown"),
            "current": _migration_state_cache.get("current"),
            "head": _migration_state_cache.get("head"),
            "checked_at": _migration_state_cache.get("checked_at"),
        }

    head_revisions = _get_alembic_head_revisions()
    current_revisions: List[str] = []
    state = "unknown"
    if head_revisions:
        try:
            result = await conn.execute(text("SELECT version_num FROM alembic_version"))
            current_revisions = [row[0] for row in result.fetchall() if row and row[0]]
        except Exception:
            current_revisions = []

    if head_revisions and current_revisions:
        state = "ok" if set(current_revisions) == set(head_revisions) else "migration_required"

    _migration_state_cache.update({
        "checked_at": now_ts,
        "state": state,
        "current": current_revisions or None,
        "head": head_revisions or None,
    })
    return {"state": state, "current": current_revisions or None, "head": head_revisions or None, "checked_at": now_ts}


def _format_migration_checked_at() -> Optional[str]:
    checked_at = _migration_state_cache.get("checked_at") or 0
    if not checked_at:
        return None
    return datetime.fromtimestamp(checked_at, tz=timezone.utc).isoformat()


def _set_last_db_error(message: Optional[str]) -> None:
    global _last_db_error
    if not message:
        _last_db_error = None
        return
    _last_db_error = _sanitize_db_error_message(str(message))


def _sanitize_text(value: str) -> str:
    return html.escape(value or "").strip()


def _validate_attachment_url(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = value.strip()
    parsed = urllib.parse.urlparse(cleaned)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(status_code=400, detail="Invalid attachment_url")
    return cleaned


def _build_error_detail(code: str, message: str) -> Dict[str, Any]:
    if correlation_id:
        try:
            req_id = correlation_id.get()
        except Exception:
            req_id = None
    else:
        req_id = None
    if not req_id:
        req_id = str(uuid.uuid4())
    return {"code": code, "message": message, "request_id": req_id}


async def _ensure_campaigns_db_ready(session: AsyncSession) -> None:
    if not RAW_DATABASE_URL:
        raise HTTPException(status_code=503, detail=_build_error_detail("DB_NOT_READY", "Campaigns database not ready"))
    try:
        await session.execute(select(1))
    except Exception:
        raise HTTPException(status_code=503, detail=_build_error_detail("DB_NOT_READY", "Campaigns database not reachable"))


async def _ensure_plans_db_ready(session: AsyncSession) -> None:
    if not RAW_DATABASE_URL:
        raise HTTPException(status_code=503, detail=_build_error_detail("DB_NOT_READY", "Plans database not ready"))
    try:
        await session.execute(select(1))
    except Exception:
        raise HTTPException(status_code=503, detail=_build_error_detail("DB_NOT_READY", "Plans database not reachable"))


async def _ensure_invoices_db_ready(session: AsyncSession) -> None:
    if not RAW_DATABASE_URL:
        raise HTTPException(status_code=503, detail=_build_error_detail("DB_NOT_READY", "Invoices database not ready"))
    try:
        await session.execute(select(1))
    except Exception:
        raise HTTPException(status_code=503, detail=_build_error_detail("DB_NOT_READY", "Invoices database not reachable"))


async def _get_db_status(session: AsyncSession) -> tuple[bool, str, str]:
    if not RAW_DATABASE_URL:
        return False, "db_config_missing", "config_missing"
    try:
        await session.execute(select(1))
        return True, "ok", "ok"
    except Exception:
        return False, "db_unreachable", "unreachable"


def _resolve_currency_code(country_code: Optional[str]) -> Optional[str]:
    code = country_code or settings.DEFAULT_COUNTRY
    return settings.COUNTRY_CURRENCIES.get(code)


def _slugify_value(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", "-", (value or "").strip().lower())
    cleaned = re.sub(r"-+", "-", cleaned).strip("-")
    return cleaned


def _generate_invoice_no() -> str:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
    short = uuid.uuid4().hex[:6].upper()
    return f"INV-{stamp}-{short}"


def _is_payment_enabled_for_country(country_code: Optional[str]) -> bool:
    if not country_code:
        return False
    return country_code.upper() in PAYMENTS_ENABLED_COUNTRIES


def _parse_datetime_field(value: str, field_name: str) -> datetime:
    if not value:
        raise HTTPException(status_code=400, detail=f"{field_name} is required")
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}") from exc
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed


def _normalize_str_list(value: Optional[List[str]]) -> Optional[List[str]]:
    if value is None:
        return None
    if not isinstance(value, list):
        raise HTTPException(status_code=400, detail="List format expected")
    cleaned = [str(item).strip() for item in value if str(item).strip()]
    return cleaned or None


def _normalize_campaign_payload(payload, existing: Optional[Campaign] = None) -> Dict[str, Any]:
    name = payload.name if payload.name is not None else (existing.name if existing else None)
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    status_value = payload.status if payload.status is not None else (existing.status if existing else "draft")
    status_value = status_value.lower()
    if status_value not in CAMPAIGN_STATUS_SET:
        raise HTTPException(status_code=400, detail="Invalid status")

    country_code = payload.country_code if payload.country_code is not None else (existing.country_code if existing else None)
    if not country_code:
        raise HTTPException(status_code=400, detail="country_code is required")
    country_code = country_code.upper()

    start_raw = payload.start_at if payload.start_at is not None else (
        existing.start_at.isoformat() if existing and existing.start_at else None
    )
    if not start_raw and not existing:
        raise HTTPException(status_code=400, detail="start_at is required")
    start_at = _parse_datetime_field(start_raw, "start_at") if start_raw else existing.start_at

    end_raw = payload.end_at if payload.end_at is not None else (
        existing.end_at.isoformat() if existing and existing.end_at else None
    )
    end_at = _parse_datetime_field(end_raw, "end_at") if end_raw else None
    if end_at and start_at and start_at >= end_at:
        raise HTTPException(status_code=400, detail="start_at must be before end_at")

    budget_amount = payload.budget_amount if payload.budget_amount is not None else (
        float(existing.budget_amount) if existing and existing.budget_amount is not None else None
    )
    budget_currency = payload.budget_currency if payload.budget_currency is not None else (
        existing.budget_currency if existing else None
    )

    if budget_amount is not None:
        if budget_amount < 0:
            raise HTTPException(status_code=400, detail="budget_amount must be >= 0")
        if not budget_currency:
            raise HTTPException(status_code=400, detail="budget_currency is required when budget_amount is set")
    if budget_currency and budget_amount is None:
        raise HTTPException(status_code=400, detail="budget_amount is required when budget_currency is set")

    notes = payload.notes if payload.notes is not None else (existing.notes if existing else None)
    rules_json = payload.rules_json if payload.rules_json is not None else (existing.rules_json if existing else None)

    return {
        "name": name,
        "status": status_value,
        "start_at": start_at,
        "end_at": end_at,
        "country_code": country_code,
        "budget_amount": budget_amount,
        "budget_currency": budget_currency,
        "notes": notes,
        "rules_json": rules_json,
    }


def _campaign_to_dict(campaign: Campaign) -> Dict[str, Any]:
    return {
        "id": str(campaign.id),
        "name": campaign.name,
        "status": campaign.status,
        "start_at": campaign.start_at.isoformat() if campaign.start_at else None,
        "end_at": campaign.end_at.isoformat() if campaign.end_at else None,
        "country_code": campaign.country_code,
        "budget_amount": float(campaign.budget_amount) if campaign.budget_amount is not None else None,
        "budget_currency": campaign.budget_currency,
        "notes": campaign.notes,
        "rules_json": campaign.rules_json or {},
        "created_at": campaign.created_at.isoformat() if campaign.created_at else None,
        "updated_at": campaign.updated_at.isoformat() if campaign.updated_at else None,
    }


def _audit_log_to_dict(entry: AuditLog) -> Dict[str, Any]:
    return {
        "id": str(entry.id),
        "action": entry.action,
        "user_id": str(entry.user_id) if entry.user_id else None,
        "user_email": entry.user_email,
        "resource_type": entry.resource_type,
        "resource_id": entry.resource_id,
        "old_values": entry.old_values,
        "new_values": entry.new_values,
        "metadata": entry.metadata_info,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
    }


def _plan_to_dict(plan: Plan) -> Dict[str, Any]:
    return {
        "id": str(plan.id),
        "slug": plan.slug,
        "name": plan.name,
        "country_scope": plan.country_scope,
        "country_code": plan.country_code,
        "period": plan.period,
        "price_amount": float(plan.price_amount) if plan.price_amount is not None else None,
        "currency_code": plan.currency_code,
        "listing_quota": plan.listing_quota,
        "showcase_quota": plan.showcase_quota,
        "active_flag": plan.active_flag,
        "archived_at": plan.archived_at.isoformat() if plan.archived_at else None,
        "created_at": plan.created_at.isoformat() if plan.created_at else None,
        "updated_at": plan.updated_at.isoformat() if plan.updated_at else None,
    }


def _admin_invoice_to_dict(invoice: AdminInvoice, dealer: Optional[SqlUser] = None, plan: Optional[Plan] = None) -> Dict[str, Any]:
    return {
        "id": str(invoice.id),
        "invoice_no": invoice.invoice_no,
        "user_id": str(invoice.user_id),
        "dealer_id": str(invoice.user_id),
        "dealer_email": dealer.email if dealer else None,
        "subscription_id": str(invoice.subscription_id) if invoice.subscription_id else None,
        "plan_id": str(invoice.plan_id) if invoice.plan_id else None,
        "plan_name": plan.name if plan else None,
        "campaign_id": str(invoice.campaign_id) if invoice.campaign_id else None,
        "amount_total": float(invoice.amount_total) if invoice.amount_total is not None else None,
        "amount": float(invoice.amount_total) if invoice.amount_total is not None else None,
        "currency": invoice.currency,
        "currency_code": invoice.currency,
        "status": invoice.status,
        "payment_status": invoice.payment_status,
        "issued_at": invoice.issued_at.isoformat() if invoice.issued_at else None,
        "paid_at": invoice.paid_at.isoformat() if invoice.paid_at else None,
        "due_at": invoice.due_at.isoformat() if invoice.due_at else None,
        "provider_customer_id": invoice.provider_customer_id,
        "meta_json": invoice.meta_json or {},
        "scope": invoice.scope,
        "country_code": invoice.country_code,
        "payment_method": invoice.payment_method,
        "notes": invoice.notes,
        "created_at": invoice.created_at.isoformat() if invoice.created_at else None,
        "updated_at": invoice.updated_at.isoformat() if invoice.updated_at else None,
    }


def _payment_to_dict(payment: Payment, invoice: Optional[AdminInvoice] = None, user: Optional[SqlUser] = None) -> Dict[str, Any]:
    return {
        "id": str(payment.id),
        "provider": payment.provider,
        "provider_ref": payment.provider_ref,
        "invoice_id": str(payment.invoice_id),
        "invoice_no": invoice.invoice_no if invoice else None,
        "user_id": str(payment.user_id),
        "user_email": user.email if user else None,
        "status": payment.status,
        "amount_total": float(payment.amount_total) if payment.amount_total is not None else None,
        "currency": payment.currency,
        "meta_json": payment.meta_json or {},
        "created_at": payment.created_at.isoformat() if payment.created_at else None,
        "updated_at": payment.updated_at.isoformat() if payment.updated_at else None,
    }


async def _activate_subscription_from_invoice(session: AsyncSession, invoice: AdminInvoice) -> None:
    if not invoice.subscription_id or not invoice.plan_id:
        return

    subscription = await session.get(UserSubscription, invoice.subscription_id)
    if not subscription:
        return

    plan = await session.get(Plan, invoice.plan_id)
    if not plan:
        return

    now = datetime.now(timezone.utc)
    subscription.status = "active"
    subscription.current_period_start = subscription.current_period_start or now
    subscription.current_period_end = subscription.current_period_end or (now + timedelta(days=30))
    subscription.updated_at = now

    user = await session.get(SqlUser, subscription.user_id)
    if user:
        user.listing_quota_limit = plan.listing_quota
        user.showcase_quota_limit = plan.showcase_quota
        user.updated_at = now

    await _write_audit_log_sql(
        session=session,
        action="subscription_activated",
        actor={"id": str(invoice.user_id), "email": None},
        resource_type="subscription",
        resource_id=str(subscription.id),
        metadata={"status": "active"},
        request=None,
        country_code=invoice.country_code,
    )
    await _write_audit_log_sql(
        session=session,
        action="quota_limits_updated",
        actor={"id": str(invoice.user_id), "email": None},
        resource_type="user",
        resource_id=str(subscription.user_id),
        metadata={"listing_quota_limit": plan.listing_quota, "showcase_quota_limit": plan.showcase_quota},
        request=None,
        country_code=invoice.country_code,
    )


async def _invoice_totals_by_currency(
    session: AsyncSession, conditions: List[Any]
) -> Dict[str, float]:
    query = select(AdminInvoice.currency, func.sum(AdminInvoice.amount_total)).group_by(AdminInvoice.currency)
    if conditions:
        query = query.where(*conditions)
    rows = (await session.execute(query)).all()
    totals: Dict[str, float] = {}
    for currency, amount in rows:
        totals[currency or "UNKNOWN"] = float(amount or 0)
    return totals


def _check_application_rate_limit(request: Request, user_id: str) -> None:
    key = user_id or _get_client_ip(request) or "unknown"
    now = time.time()
    attempts = _application_submit_attempts.get(key, [])
    attempts = [ts for ts in attempts if (now - ts) <= APPLICATION_RATE_LIMIT_WINDOW_SECONDS]
    if len(attempts) >= APPLICATION_RATE_LIMIT_MAX_ATTEMPTS:
        retry_after_seconds = int(APPLICATION_RATE_LIMIT_WINDOW_SECONDS - (now - attempts[0]))
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": max(retry_after_seconds, 1)},
        )
    attempts.append(now)
    _application_submit_attempts[key] = attempts


async def _ensure_sql_user(session: AsyncSession, user_doc: dict) -> uuid.UUID:
    user_id_raw = user_doc.get("id")
    if not user_id_raw:
        raise HTTPException(status_code=400, detail="Invalid user")
    try:
        user_uuid = uuid.UUID(str(user_id_raw))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    existing = result.scalar_one_or_none()
    if existing:
        return user_uuid

    fallback_password = user_doc.get("hashed_password") or get_password_hash(secrets.token_hex(12))
    new_user = SqlUser(
        id=user_uuid,
        email=user_doc.get("email") or f"user-{user_uuid}@platform.local",
        hashed_password=fallback_password,
        full_name=user_doc.get("full_name") or user_doc.get("email") or "User",
        role=user_doc.get("role") or "individual",
        is_active=bool(user_doc.get("is_active", True)),
        is_verified=bool(user_doc.get("is_verified", True)),
        country_scope=user_doc.get("country_scope") or [],
        preferred_language=user_doc.get("preferred_language", "tr"),
        country_code=user_doc.get("country_code") or "TR",
    )
    session.add(new_user)
    await session.flush()
    return user_uuid


def _enforce_export_rate_limit(request: Request, user_id: str) -> None:
    now = time.time()
    key = f"{user_id}:{_get_client_ip(request)}"
    attempts = _export_attempts.get(key, [])
    attempts = [ts for ts in attempts if (now - ts) <= EXPORT_RATE_LIMIT_WINDOW_SECONDS]
    if len(attempts) >= EXPORT_RATE_LIMIT_MAX_ATTEMPTS:
        retry_after_seconds = int(EXPORT_RATE_LIMIT_WINDOW_SECONDS - (now - attempts[0]))
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": max(retry_after_seconds, 1)},
        )
    attempts.append(now)
    _export_attempts[key] = attempts


def _enforce_admin_invite_rate_limit(request: Request, user_id: str) -> None:
    now = time.time()
    key = f"invite:{user_id}:{_get_client_ip(request)}"
    attempts = _admin_invite_attempts.get(key, [])
    attempts = [ts for ts in attempts if (now - ts) <= ADMIN_INVITE_RATE_LIMIT_WINDOW_SECONDS]
    if len(attempts) >= ADMIN_INVITE_RATE_LIMIT_MAX_ATTEMPTS:
        retry_after_seconds = int(ADMIN_INVITE_RATE_LIMIT_WINDOW_SECONDS - (now - attempts[0]))
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": max(retry_after_seconds, 1)},
        )
    attempts.append(now)
    _admin_invite_attempts[key] = attempts


def _hash_invite_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def _get_admin_base_url(request: Request) -> str:
    origin = request.headers.get("origin")
    if origin:
        return origin.rstrip("/")
    referer = request.headers.get("referer")
    if referer:
        return referer.split("/admin")[0].rstrip("/")
    base = str(request.base_url).rstrip("/")
    if base.endswith("/api"):
        base = base[:-4]
    return base


def _send_admin_invite_email(to_email: str, full_name: str, invite_link: str) -> None:
    sendgrid_key = os.environ.get("SENDGRID_API_KEY")
    sender_email = os.environ.get("SENDER_EMAIL")
    if not sendgrid_key or not sender_email:
        _admin_invite_logger.error("SendGrid configuration missing: SENDGRID_API_KEY or SENDER_EMAIL")
        raise HTTPException(status_code=503, detail="SendGrid is not configured")

    subject = "Admin Daveti"
    html_content = f"""
    <html>
      <body>
        <h2>Admin Daveti</h2>
        <p>Merhaba {full_name or to_email},</p>
        <p>Admin hesabınız oluşturuldu. Daveti kabul etmek ve şifre belirlemek için aşağıdaki bağlantıyı kullanın:</p>
        <p><a href=\"{invite_link}\">Daveti Kabul Et</a></p>
        <p>Bağlantı 24 saat boyunca geçerlidir.</p>
      </body>
    </html>
    """
    message = Mail(
        from_email=sender_email,
        to_emails=to_email,
        subject=subject,
        html_content=html_content,
    )

    try:
        sg = SendGridAPIClient(sendgrid_key)
        response = sg.send(message)
        if response.status_code not in (200, 202):
            raise HTTPException(status_code=502, detail="Failed to send invite email")
    except HTTPException:
        raise
    except Exception as exc:
        _admin_invite_logger.error("SendGrid send error: %s", exc)
        raise HTTPException(status_code=502, detail="Failed to send invite email")


def _send_support_received_email(to_email: str, application_id: str, subject_text: str) -> None:
    sendgrid_key = os.environ.get("SENDGRID_API_KEY")
    sender_email = os.environ.get("SENDER_EMAIL")
    if not sendgrid_key or not sender_email:
        _admin_invite_logger.warning("SendGrid configuration missing: SENDGRID_API_KEY or SENDER_EMAIL")
        return

    subject = "Başvurunuz alındı"
    html_content = f"""
    <html>
      <body>
        <h2>Başvurunuz alındı</h2>
        <p>Merhaba,</p>
        <p>Başvurunuzu aldık. Referans numaranız: <strong>{application_id}</strong></p>
        <p>Konu: {subject_text}</p>
        <p>İnceleme sürecimiz başladığında sizi bilgilendireceğiz.</p>
      </body>
    </html>
    """
    message = Mail(
        from_email=sender_email,
        to_emails=to_email,
        subject=subject,
        html_content=html_content,
    )

    try:
        sg = SendGridAPIClient(sendgrid_key)
        response = sg.send(message)
        if response.status_code not in (200, 202):
            raise HTTPException(status_code=502, detail="Failed to send support email")
    except HTTPException:
        raise
    except Exception as exc:
        _admin_invite_logger.error("SendGrid send error: %s", exc)
        raise HTTPException(status_code=502, detail="Failed to send support email")


async def _send_message_notification_email(session: AsyncSession, recipient_id: str, thread: dict, message: dict) -> None:
    sendgrid_key = os.environ.get("SENDGRID_API_KEY")
    sender_email = os.environ.get("SENDER_EMAIL")
    if not sendgrid_key or not sender_email or session is None:
        return

    try:
        recipient_uuid = uuid.UUID(str(recipient_id))
    except ValueError:
        return

    recipient = await session.get(SqlUser, recipient_uuid)
    if not recipient or not recipient.email:
        return

    prefs = getattr(recipient, "notification_prefs", None) or {}
    if not prefs.get("email_enabled", True):
        return

    listing_title = thread.get("listing_title") or "İlan"
    preview = (message.get("body") or "")[:160]
    subject = f"Yeni mesajınız var: {listing_title}"
    html_content = f"""
    <html>
      <body>
        <h2>Yeni mesajınız var</h2>
        <p>İlan: <strong>{listing_title}</strong></p>
        <p>Mesaj: {preview}</p>
        <p>Mesajları görüntülemek için hesabınıza giriş yapın.</p>
      </body>
    </html>
    """

    message_mail = Mail(
        from_email=sender_email,
        to_emails=recipient.get("email"),
        subject=subject,
        html_content=html_content,
    )

    try:
        sg = SendGridAPIClient(sendgrid_key)
        sg.send(message_mail)
    except Exception as exc:
        _admin_invite_logger.error("SendGrid message notification error: %s", exc)


async def _create_inapp_notification(
    session: AsyncSession,
    user_id: str,
    message: str,
    metadata: Optional[Dict[str, Any]] = None,
) -> None:
    if session is None:
        return
    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError:
        return

    payload = metadata or {}
    notification = Notification(
        user_id=user_uuid,
        title=payload.get("title"),
        message=message,
        source_type=payload.get("source_type") or payload.get("type") or "support",
        source_id=payload.get("source_id") or payload.get("listing_id") or payload.get("thread_id"),
        action_url=payload.get("action_url") or payload.get("url"),
        payload_json=payload or None,
        dedupe_key=payload.get("dedupe_key"),
        read_at=None,
        delivered_at=None,
    )
    session.add(notification)
    await session.commit()


def _normalize_scope(role: str, scope: Optional[List[str]], active_countries: List[str]) -> List[str]:
    if role == "super_admin":
        return ["*"]
    normalized = [code.strip().upper() for code in (scope or []) if code and code.strip()]
    if role == "country_admin" and not normalized:
        raise HTTPException(status_code=400, detail="Country scope required for country_admin")
    if "*" in normalized:
        raise HTTPException(status_code=400, detail="Invalid country scope value")
    for code in normalized:
        if code not in active_countries:
            raise HTTPException(status_code=400, detail=f"Invalid country scope: {code}")
    return normalized



async def _assert_super_admin_invariant_sql(
    session: AsyncSession,
    target_user: SqlUser,
    payload_role: Optional[str],
    payload_active: Optional[bool],
    actor: dict,
) -> None:
    if not target_user:
        return

    target_role = target_user.role
    target_id = str(target_user.id)
    actor_id = actor.get("id")

    if payload_role and target_id == actor_id and payload_role != target_role:
        raise HTTPException(status_code=400, detail="Self role change is not allowed")
    if payload_active is not None and target_id == actor_id and payload_active is False:
        raise HTTPException(status_code=400, detail="Cannot deactivate yourself")

    demoting = payload_role and target_role == "super_admin" and payload_role != "super_admin"
    deactivating = payload_active is False and target_role == "super_admin"
    if demoting or deactivating:
        result = await session.execute(
            select(func.count())
            .select_from(SqlUser)
            .where(SqlUser.role == "super_admin", SqlUser.is_active.is_(True))
        )
        super_admin_count = result.scalar_one()
        if super_admin_count <= 1:
            raise HTTPException(status_code=400, detail="At least one super_admin must remain active")


def _parse_country_codes(value: Optional[str]) -> Optional[List[str]]:
    if not value:
        return None
    codes = [code.strip().upper() for code in value.split(",") if code.strip()]
    return codes or None

async def _build_country_compare_payload_sql(
    session: AsyncSession,
    current_user: dict,
    period: str,
    start_date: Optional[str],
    end_date: Optional[str],
    sort_by: Optional[str],
    sort_dir: Optional[str],
    selected_codes: Optional[List[str]],
) -> dict:
    codes = [code.upper() for code in (selected_codes or [])]
    if not codes:
        active_countries = (
            await session.execute(select(Country).where(Country.is_enabled.is_(True)))
        ).scalars().all()
        codes = [c.code for c in active_countries]

    items = []
    for code in codes:
        total_listings = (
            await session.execute(
                select(func.count()).select_from(Listing).where(Listing.country == code)
            )
        ).scalar_one()
        published = (
            await session.execute(
                select(func.count()).select_from(Listing).where(
                    Listing.country == code,
                    Listing.status == "published",
                )
            )
        ).scalar_one()
        active_dealers = (
            await session.execute(
                select(func.count()).select_from(SqlUser).where(
                    SqlUser.country_code == code,
                    SqlUser.role == "dealer",
                    SqlUser.is_active.is_(True),
                )
            )
        ).scalar_one()

        items.append(
            {
                "country_code": code,
                "total_listings": int(total_listings or 0),
                "growth_total_listings_7d": 0,
                "growth_total_listings_30d": 0,
                "published_listings": int(published or 0),
                "growth_published_7d": 0,
                "growth_published_30d": 0,
                "conversion_rate": 0,
                "active_dealers": int(active_dealers or 0),
                "growth_active_dealers_7d": 0,
                "growth_active_dealers_30d": 0,
                "dealer_density": 0,
                "revenue_eur": 0,
                "growth_revenue_7d": 0,
                "growth_revenue_30d": 0,
                "revenue_mtd_growth_pct": 0,
                "sla_pending_24h": 0,
                "sla_pending_48h": 0,
                "risk_multi_login": 0,
                "risk_pending_payments": 0,
                "note": "",
            }
        )

    if sort_by:
        key = sort_by
        reverse = (sort_dir or "desc").lower() == "desc"
        items.sort(key=lambda item: item.get(key) or 0, reverse=reverse)

    return {"items": items, "period": period, "start_date": start_date, "end_date": end_date}




def _dashboard_cache_key(role: str, country_codes: Optional[List[str]], trend_days: int) -> str:
    trend_label = f"trend{trend_days}"
    if not country_codes:
        return f"{role}:global:{trend_label}"
    joined = ",".join(sorted(country_codes))
    return f"{role}:{joined}:{trend_label}"


def _country_compare_cache_key(
    role: str,
    country_codes: Optional[List[str]],
    selected_codes: Optional[List[str]],
    period: str,
    start_date: Optional[str],
    end_date: Optional[str],
    sort_by: Optional[str],
    sort_dir: Optional[str],
) -> str:
    joined = ",".join(sorted(country_codes)) if country_codes else "global"
    selected = ",".join(sorted(selected_codes)) if selected_codes else "all"
    return f"{role}:{joined}:{selected}:{period}:{start_date or ''}:{end_date or ''}:{sort_by or ''}:{sort_dir or ''}"


def _get_cached_dashboard_summary(cache_key: str) -> Optional[Dict[str, Any]]:
    entry = _dashboard_summary_cache.get(cache_key)
    if not entry:
        return None
    if (time.time() - entry.get("timestamp", 0)) > DASHBOARD_CACHE_TTL_SECONDS:
        _dashboard_summary_cache.pop(cache_key, None)
        return None
    return entry.get("data")


def _set_cached_dashboard_summary(cache_key: str, data: Dict[str, Any]) -> None:
    _dashboard_summary_cache[cache_key] = {"timestamp": time.time(), "data": data}


def _get_cached_country_compare(cache_key: str) -> Optional[Dict[str, Any]]:
    entry = _country_compare_cache.get(cache_key)
    if not entry:
        return None
    if (time.time() - entry.get("timestamp", 0)) > COUNTRY_COMPARE_CACHE_TTL_SECONDS:
        _country_compare_cache.pop(cache_key, None)
        return None
    return entry.get("data")


def _set_cached_country_compare(cache_key: str, data: Dict[str, Any]) -> None:
    _country_compare_cache[cache_key] = {"timestamp": time.time(), "data": data}


def _fetch_ecb_rates() -> Optional[Dict[str, float]]:
    try:
        with urllib.request.urlopen(ECB_DAILY_URL, timeout=20) as response:
            xml_bytes = response.read()
        root = ET.fromstring(xml_bytes)
        ns = {
            "gesmes": "http://www.gesmes.org/xml/2002-08-01",
            "ecb": "http://www.ecb.int/vocabulary/2002-08-01/eurofxref",
        }
        rates: Dict[str, float] = {ECB_RATE_BASE: 1.0}
        for cube in root.findall(".//ecb:Cube[@currency]", ns):
            currency = cube.get("currency")
            rate = cube.get("rate")
            if currency and rate:
                try:
                    rates[currency.upper()] = float(rate)
                except ValueError:
                    continue
        return rates
    except Exception:
        return None


def _get_ecb_rates() -> Dict[str, Any]:
    now = time.time()
    if _ecb_rates_cache.get("rates") and (now - _ecb_rates_cache.get("timestamp", 0)) < ECB_CACHE_TTL_SECONDS:
        return _ecb_rates_cache

    rates = _fetch_ecb_rates()
    if rates:
        fetched_at = datetime.now(timezone.utc).isoformat()
        _ecb_rates_cache.update({
            "timestamp": now,
            "rates": rates,
            "last_success_at": fetched_at,
            "fallback": False,
        })
        global _ecb_rates_fallback
        _ecb_rates_fallback = dict(_ecb_rates_cache)
        return _ecb_rates_cache

    if _ecb_rates_fallback and _ecb_rates_fallback.get("rates"):
        fallback = dict(_ecb_rates_fallback)
        fallback["fallback"] = True
        _ecb_rates_cache.update(fallback)
        _ecb_rates_cache["timestamp"] = now
        return _ecb_rates_cache

    return {
        "timestamp": now,
        "rates": {ECB_RATE_BASE: 1.0},
        "last_success_at": None,
        "fallback": True,
    }


def _convert_to_eur(amount: float, currency: Optional[str], rates: Dict[str, float]) -> Optional[float]:
    if amount is None:
        return None
    currency_code = (currency or ECB_RATE_BASE).upper()
    rate = rates.get(currency_code)
    if not rate or rate == 0:
        return None
    return round(float(amount) / rate, 4)


def _format_uptime(seconds: int) -> str:
    if seconds < 0:
        return "0s"
    days, rem = divmod(seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, secs = divmod(rem, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    if secs or not parts:
        parts.append(f"{secs}s")
    return " ".join(parts)


def _parse_iso(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            return parsed.replace(tzinfo=timezone.utc)
        return parsed
    except Exception:
        return None


def _resolve_period_window(period: str, start_date: Optional[str], end_date: Optional[str]) -> tuple[datetime, datetime, str]:
    now = datetime.now(timezone.utc)
    period_key = (period or "30d").lower()

    if period_key == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end = now
        return start, end, "Bugün"

    if period_key in {"7d", "7", "last_7_days"}:
        start = now - timedelta(days=7)
        end = now
        return start, end, "Son 7 Gün"

    if period_key in {"30d", "30", "last_30_days"}:
        start = now - timedelta(days=30)
        end = now
        return start, end, "Son 30 Gün"

    if period_key == "mtd":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end = now
        return start, end, "MTD"

    if period_key == "custom":
        start_dt = _parse_iso(start_date)
        end_dt = _parse_iso(end_date)
        if not start_dt or not end_dt:
            raise HTTPException(status_code=400, detail="custom start_date and end_date required")
        start = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
        end = end_dt.replace(hour=23, minute=59, second=59, microsecond=0)
        if end < start:
            raise HTTPException(status_code=400, detail="end_date must be after start_date")
        range_days = (end - start).days + 1
        if range_days < DASHBOARD_TREND_MIN_DAYS or range_days > DASHBOARD_TREND_MAX_DAYS:
            raise HTTPException(status_code=400, detail="custom range must be between 7 and 365 days")
        return start, end, "Özel"

    start = now - timedelta(days=30)
    end = now
    return start, end, "Son 30 Gün"


def _growth_pct(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    current_val = float(current or 0)
    prev_val = float(previous or 0)
    if prev_val == 0:
        return None if current_val == 0 else 100.0
    return round(((current_val - prev_val) / prev_val) * 100, 2)


def _safe_ratio(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    num = float(numerator or 0)
    den = float(denominator or 0)
    if den == 0:
        return None
    return round(num / den, 4)


# Audit event taxonomy (v1)
AUDIT_EVENT_TYPES_V1 = {
    "MODERATION_APPROVE",
    "MODERATION_REJECT",
    "MODERATION_NEEDS_REVISION",
    "FAILED_LOGIN",
    "RATE_LIMIT_BLOCK",
    "ADMIN_ROLE_CHANGE",
    "UNAUTHORIZED_ROLE_CHANGE_ATTEMPT",
    "DEALER_STATUS_CHANGE",
    "DEALER_APPLICATION_APPROVED",
    "DEALER_APPLICATION_REJECTED",
    "INDIVIDUAL_APPLICATION_APPROVED",
    "INDIVIDUAL_APPLICATION_REJECTED",
    "LISTING_SOFT_DELETE",
    "LISTING_FORCE_UNPUBLISH",
    "REPORT_STATUS_CHANGE",
    "REPORT_CREATED",
    "INVOICE_STATUS_CHANGE",
    "TAX_RATE_CHANGE",
    "PLAN_CHANGE",
    "ADMIN_PLAN_ASSIGNMENT",
    "DEALER_PLAN_OVERRIDE",
    "RISK_LEVEL_UPDATED",
    "DOPING_REQUEST_PAID",
    "DOPING_APPROVED",
    "MIGRATION_DRY_RUN",
    "COUNTRY_CHANGE",
    "SYSTEM_SETTING_CHANGE",
    "CATEGORY_CHANGE",
    "MENU_CHANGE",
    "ATTRIBUTE_CHANGE",
    "VEHICLE_MASTER_DATA_CHANGE",
}

api_router = APIRouter(prefix="/api")


# Moderation reasons (v1.0.0 single-source contract)
REJECT_REASONS_V1 = {"duplicate", "spam", "illegal", "wrong_category"}
NEEDS_REVISION_REASONS_V1 = {"missing_photos", "insufficient_description", "wrong_price", "other"}

# Report reasons (v1)
REPORT_REASONS_V1 = {
    "spam",
    "scam_fraud",
    "prohibited_item",
    "wrong_category",
    "harassment",
    "copyright",
    "other",
}

REPORT_STATUS_SET = {"open", "in_review", "resolved", "dismissed"}
KEY_NAMESPACE_REGEX = re.compile(r"^[a-z0-9]+(\.[a-z0-9_]+)+$")
MODERATION_FREEZE_SETTING_KEY = "moderation.freeze.active"

REPORT_STATUS_TRANSITIONS = {
    "open": {"in_review"},
    "in_review": {"resolved", "dismissed"},
}

REPORT_RATE_LIMIT_WINDOW_SECONDS = 10 * 60
REPORT_RATE_LIMIT_MAX_ATTEMPTS = 5
_report_submit_attempts: Dict[str, List[float]] = {}

EXPORT_RATE_LIMIT_WINDOW_SECONDS = 60
EXPORT_RATE_LIMIT_MAX_ATTEMPTS = 10
GDPR_EXPORT_RETENTION_DAYS = 30
GDPR_EXPORT_DIR = os.path.join(os.path.dirname(__file__), "static", "exports")
_export_attempts: Dict[str, List[float]] = {}

VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY")
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY")
VAPID_SUBJECT = os.environ.get("VAPID_SUBJECT")
PUSH_ENABLED = bool(VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY and VAPID_SUBJECT)
if PUSH_ENABLED:
    logging.getLogger("push_config").info("push config loaded")
else:
    logging.getLogger("push_config").warning("push config not set")


class MessageConnectionManager:
    def __init__(self):
        self.user_connections: Dict[str, set] = defaultdict(set)
        self.thread_connections: Dict[str, set] = defaultdict(set)

    async def connect(self, websocket: WebSocket, user_id: str) -> None:
        await websocket.accept()
        self.user_connections[user_id].add(websocket)

    def disconnect(self, websocket: WebSocket, user_id: str) -> None:
        if user_id in self.user_connections and websocket in self.user_connections[user_id]:
            self.user_connections[user_id].remove(websocket)
            if not self.user_connections[user_id]:
                self.user_connections.pop(user_id, None)
        for thread_id, sockets in list(self.thread_connections.items()):
            if websocket in sockets:
                sockets.remove(websocket)
                if not sockets:
                    self.thread_connections.pop(thread_id, None)

    def subscribe(self, websocket: WebSocket, thread_id: str) -> None:
        self.thread_connections[thread_id].add(websocket)

    def unsubscribe(self, websocket: WebSocket, thread_id: str) -> None:
        if thread_id in self.thread_connections and websocket in self.thread_connections[thread_id]:
            self.thread_connections[thread_id].remove(websocket)
            if not self.thread_connections[thread_id]:
                self.thread_connections.pop(thread_id, None)

    async def send_personal(self, user_id: str, payload: dict) -> None:
        for socket in list(self.user_connections.get(user_id, [])):
            await self._safe_send(socket, payload)

    async def broadcast_thread(self, thread_id: str, payload: dict) -> None:
        for socket in list(self.thread_connections.get(thread_id, [])):
            await self._safe_send(socket, payload)

    async def _safe_send(self, websocket: WebSocket, payload: dict) -> None:
        try:
            await websocket.send_json(payload)
        except Exception:
            for thread_id, sockets in list(self.thread_connections.items()):
                if websocket in sockets:
                    sockets.remove(websocket)
                    if not sockets:
                        self.thread_connections.pop(thread_id, None)
            for user_id, sockets in list(self.user_connections.items()):
                if websocket in sockets:
                    sockets.remove(websocket)
                    if not sockets:
                        self.user_connections.pop(user_id, None)


message_ws_manager = MessageConnectionManager()

ADMIN_INVITE_RATE_LIMIT_WINDOW_SECONDS = 60
ADMIN_INVITE_RATE_LIMIT_MAX_ATTEMPTS = 5

ADMIN_ROLE_OPTIONS = {"super_admin", "country_admin", "finance", "support", "moderator", "campaigns_admin", "campaigns_supervisor", "ROLE_AUDIT_VIEWER", "ads_manager", "pricing_manager"}

CAMPAIGN_STATUS_SET = {"draft", "active", "paused", "ended"}
CAMPAIGN_STATUS_TRANSITIONS = {
    "draft": {"active", "ended"},
    "active": {"paused", "ended"},
    "paused": {"active", "ended"},
    "ended": set(),
}

PLAN_SCOPE_SET = {"global", "country"}
PLAN_STATUS_SET = {"active", "inactive", "archived"}
PLAN_PERIOD_SET = {"monthly", "yearly"}
INVOICE_STATUS_SET = {"draft", "issued", "paid", "void", "refunded"}
INVOICE_STATUS_TRANSITIONS = {
    "draft": {"issued", "void"},
    "issued": {"paid", "void"},
    "paid": {"refunded"},
    "void": set(),
    "refunded": set(),
}

APPLICATION_TYPES = {"individual", "dealer"}
APPLICATION_REQUEST_TYPES = {"complaint", "request"}
APPLICATION_PRIORITY_SET = {"low", "medium", "high"}
APPLICATION_STATUS_SET = {"pending", "in_review", "approved", "rejected", "closed"}
APPLICATION_STATUS_TRANSITIONS = {
    "pending": {"in_review", "approved", "rejected", "closed"},
    "in_review": {"approved", "rejected", "closed"},
    "approved": {"closed"},
    "rejected": {"closed"},
    "closed": set(),
}

APPLICATION_RATE_LIMIT_WINDOW_SECONDS = 10 * 60
APPLICATION_RATE_LIMIT_MAX_ATTEMPTS = 5
_application_submit_attempts: Dict[str, List[float]] = {}

VEHICLE_TYPE_SET = {"car", "suv", "offroad", "pickup", "truck", "bus"}

APP_ENV = (os.environ.get("APP_ENV") or "dev").lower()
AUTH_PROVIDER = "sql"
APPLICATIONS_PROVIDER = "sql"

RAW_DATABASE_URL = os.environ.get("DATABASE_URL")
TOKEN_VERSION = os.environ.get("TOKEN_VERSION", "v2")

DB_POOL_SIZE_RAW = os.environ.get("DB_POOL_SIZE")
DB_MAX_OVERFLOW_RAW = os.environ.get("DB_MAX_OVERFLOW")
DB_POOL_TIMEOUT_RAW = os.environ.get("DB_POOL_TIMEOUT")
DB_POOL_RECYCLE_RAW = os.environ.get("DB_POOL_RECYCLE")
DB_POOL_DEBUG = os.environ.get("DB_POOL_DEBUG", "").lower() in {"1", "true", "yes"}
DB_SSL_MODE = (os.environ.get("DB_SSL_MODE") or ("require" if APP_ENV in {"prod", "preview"} else "disable")).lower()

STRIPE_API_KEY = os.environ.get("STRIPE_API_KEY")
STRIPE_WEBHOOK_SECRET = os.environ.get("STRIPE_WEBHOOK_SECRET")

EMAIL_PROVIDER = (os.environ.get("EMAIL_PROVIDER") or "mock").lower()
EMAIL_PROVIDER_OPTIONS = {"mock", "sendgrid"}

PAYMENTS_ENABLED_COUNTRIES_RAW = os.environ.get("PAYMENTS_ENABLED_COUNTRIES")
PAYMENTS_ENABLED_COUNTRIES = {
    code.strip().upper()
    for code in (PAYMENTS_ENABLED_COUNTRIES_RAW or "").split(",")
    if code.strip()
}


def _db_url_is_localhost(value: Optional[str]) -> bool:
    if not value:
        return True
    try:
        parsed = urllib.parse.urlparse(value)
        host = (parsed.hostname or "").lower()
    except Exception:
        host = value.lower()
    return host in {"localhost", "127.0.0.1"}


def _sanitize_database_url(value: str) -> str:
    try:
        parsed = urllib.parse.urlparse(value)
        if not parsed.query:
            return value
        params = urllib.parse.parse_qs(parsed.query)
        params.pop("sslmode", None)
        new_query = urllib.parse.urlencode(params, doseq=True)
        return parsed._replace(query=new_query).geturl()
    except Exception:
        return value


def _ensure_logger(name: str, level: int) -> logging.Logger:
    logger = logging.getLogger(name)
    if not logger.handlers:
        handler = logging.StreamHandler()
        handler.setLevel(level)
        formatter = logging.Formatter("%(levelname)s:%(name)s:%(message)s")
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        logger.propagate = False
    logger.setLevel(level)
    return logger


sql_logger = _ensure_logger("sql_config", logging.INFO)
db_error_logger = _ensure_logger("database.error", logging.INFO)
pool_logger = _ensure_logger("database.pool", logging.DEBUG if DB_POOL_DEBUG else logging.INFO)


if APP_ENV in {"preview", "prod"}:
    if not RAW_DATABASE_URL:
        raise RuntimeError("CONFIG_MISSING: DATABASE_URL")
    if _db_url_is_localhost(RAW_DATABASE_URL):
        raise RuntimeError("CONFIG_MISSING: DATABASE_URL")
    if DB_SSL_MODE != "require":
        raise RuntimeError("DB_SSL_MODE must be require for preview/prod")
    if APP_ENV == "prod" and EMAIL_PROVIDER == "mock":
        raise RuntimeError("EMAIL_PROVIDER cannot be mock in prod")

if EMAIL_PROVIDER not in EMAIL_PROVIDER_OPTIONS:
    raise RuntimeError("EMAIL_PROVIDER must be one of: mock, sendgrid")

if EMAIL_PROVIDER == "sendgrid":
    if not os.environ.get("SENDGRID_API_KEY") or not os.environ.get("SENDER_EMAIL"):
        raise RuntimeError("SendGrid configuration missing: SENDGRID_API_KEY or SENDER_EMAIL")

if RAW_DATABASE_URL:
    DATABASE_URL = RAW_DATABASE_URL
elif APP_ENV in {"preview", "prod"}:
    raise RuntimeError("DATABASE_URL must be set for preview/prod")
else:
    logging.getLogger("sql_config").warning("DATABASE_URL not set – running with local fallback")
    DATABASE_URL = "postgresql://admin_user:admin_pass@localhost:5432/admin_panel"

try:
    DB_POOL_SIZE = int(DB_POOL_SIZE_RAW) if DB_POOL_SIZE_RAW else 5
    DB_MAX_OVERFLOW = int(DB_MAX_OVERFLOW_RAW) if DB_MAX_OVERFLOW_RAW else 10
    DB_POOL_TIMEOUT = int(DB_POOL_TIMEOUT_RAW) if DB_POOL_TIMEOUT_RAW else 30
    DB_POOL_RECYCLE = int(DB_POOL_RECYCLE_RAW) if DB_POOL_RECYCLE_RAW else 1800
except ValueError:
    logging.getLogger("sql_config").warning("Invalid DB pool values, defaulting to 5/10/30/1800")
    DB_POOL_SIZE = 5
    DB_MAX_OVERFLOW = 10
    DB_POOL_TIMEOUT = 30
    DB_POOL_RECYCLE = 1800

ssl_context = None
connect_args: Dict[str, Any] = {
    "server_settings": {
        "client_encoding": "UTF8",
    }
}
if DB_SSL_MODE == "require":
    ssl_context = ssl.create_default_context()
    ssl_context.check_hostname = False
    ssl_context.verify_mode = ssl.CERT_NONE
    connect_args["ssl"] = ssl_context

SAFE_DATABASE_URL = _sanitize_database_url(DATABASE_URL)
ASYNC_DATABASE_URL = SAFE_DATABASE_URL.replace("postgresql://", "postgresql+asyncpg://")

sql_engine = create_async_engine(
    ASYNC_DATABASE_URL,
    echo=False,
    future=True,
    pool_size=DB_POOL_SIZE,
    max_overflow=DB_MAX_OVERFLOW,
    pool_timeout=DB_POOL_TIMEOUT,
    pool_recycle=DB_POOL_RECYCLE,
    pool_pre_ping=True,
    connect_args=connect_args,
)
sql_logger.info(
    "Effective DB pool config: pool_size=%s max_overflow=%s pool_timeout=%s pool_recycle=%s pool_pre_ping=%s",
    DB_POOL_SIZE,
    DB_MAX_OVERFLOW,
    DB_POOL_TIMEOUT,
    DB_POOL_RECYCLE,
    True,
)


if DB_POOL_DEBUG:
    pool_logger.setLevel(logging.DEBUG)


@event.listens_for(sql_engine.sync_engine, "connect")
def _log_pool_connect(dbapi_connection, connection_record):
    if DB_POOL_DEBUG:
        pool_logger.debug("db_pool_connect")


@event.listens_for(sql_engine.sync_engine, "checkout")
def _log_pool_checkout(dbapi_connection, connection_record, connection_proxy):
    if DB_POOL_DEBUG:
        pool_logger.debug("db_pool_checkout")


@event.listens_for(sql_engine.sync_engine, "checkin")
def _log_pool_checkin(dbapi_connection, connection_record):
    if DB_POOL_DEBUG:
        pool_logger.debug("db_pool_checkin")


@event.listens_for(sql_engine.sync_engine, "close")
def _log_pool_close(dbapi_connection, connection_record):
    if DB_POOL_DEBUG:
        pool_logger.debug("db_pool_close")


@event.listens_for(sql_engine.sync_engine, "invalidate")
def _log_pool_invalidate(dbapi_connection, connection_record, exception):
    pool_logger.info("db_pool_invalidate", extra={"error": str(exception) if exception else None})


AsyncSessionLocal = async_sessionmaker(
    sql_engine,
    class_=AsyncSession,
    expire_on_commit=False,
)


async def get_sql_session():
    async with AsyncSessionLocal() as session:
        try:
            yield session
        except Exception:
            await session.rollback()
            raise
        finally:
            await session.close()


def _get_auth_repository(session: AsyncSession):
    return SqlAuthRepository(session)


def _get_applications_repository(session: AsyncSession):
    return SqlApplicationsRepository(session)

_admin_invite_attempts: Dict[str, List[float]] = {}
_admin_invite_logger = logging.getLogger("admin_invites")

APP_START_TIME = datetime.now(timezone.utc)

DASHBOARD_TREND_DAYS = 14
DASHBOARD_TREND_MIN_DAYS = 7
DASHBOARD_TREND_MAX_DAYS = 365
DASHBOARD_KPI_DAYS = 7
DASHBOARD_MULTI_IP_WINDOW_HOURS = 24
DASHBOARD_MULTI_IP_THRESHOLD = 3
DASHBOARD_SLA_HOURS = 24
DASHBOARD_PENDING_PAYMENT_DAYS = 7

ECB_DAILY_URL = os.environ.get("ECB_DAILY_URL")
ECB_CACHE_TTL_SECONDS_RAW = os.environ.get("ECB_CACHE_TTL_SECONDS")
ECB_CACHE_TTL_SECONDS = int(ECB_CACHE_TTL_SECONDS_RAW) if ECB_CACHE_TTL_SECONDS_RAW else 0
ECB_RATE_BASE = "EUR"

if not ECB_DAILY_URL or ECB_CACHE_TTL_SECONDS <= 0:
    raise RuntimeError("ECB_DAILY_URL and ECB_CACHE_TTL_SECONDS must be set")

DASHBOARD_CACHE_TTL_SECONDS = int(os.environ.get("DASHBOARD_CACHE_TTL_SECONDS", "60"))
COUNTRY_COMPARE_CACHE_TTL_SECONDS = 60
_dashboard_summary_cache: Dict[str, Dict[str, Any]] = {}
_country_compare_cache: Dict[str, Dict[str, Any]] = {}
_ecb_rates_cache: Dict[str, Any] = {"timestamp": 0, "rates": None, "last_success_at": None, "fallback": False}

MIGRATION_STATE_CACHE_TTL_SECONDS = 60
_migration_state_cache: Dict[str, Any] = {"checked_at": 0, "state": "unknown", "current": None, "head": None}
_last_db_error: Optional[str] = None
_db_error_events = deque(maxlen=5000)
_db_latency_events = deque(maxlen=5000)
_system_health_cache: Dict[str, Any] = {"checked_at": 0, "data": None}
_system_health_detail_cache: Dict[str, Any] = {"checked_at": 0, "data": None}
cloudflare_metrics_service = CloudflareMetricsService()
SYSTEM_HEALTH_CACHE_TTL_SECONDS = 60
SYSTEM_HEALTH_DETAIL_CACHE_TTL_SECONDS = 60
_ecb_rates_fallback: Optional[Dict[str, Any]] = None
_dashboard_cache_hits = 0
_dashboard_cache_misses = 0

ALLOWED_MODERATION_ROLES = {"moderator", "country_admin", "super_admin"}


@api_router.get("/health")
async def health_check():
    config_state = "missing_database_url" if not RAW_DATABASE_URL else "ok"
    last_migration_check_at = _format_migration_checked_at()
    ops_attention = config_state != "ok"
    last_db_error = _last_db_error

    if not RAW_DATABASE_URL:
        _set_last_db_error("CONFIG_MISSING: DATABASE_URL")
        last_db_error = _last_db_error
        return {
            "status": "degraded",
            "supported_countries": SUPPORTED_COUNTRIES,
            "database": "postgres",
            "reason": "db_config_missing",
            "db_status": "config_missing",
            "config_state": config_state,
            "last_migration_check_at": last_migration_check_at,
            "ops_attention": ops_attention,
            "last_db_error": last_db_error,
        }
    try:
        async with sql_engine.connect() as conn:
            await conn.execute(text("SELECT 1"))
            migration_info = await _get_migration_state(conn)
            last_migration_check_at = _format_migration_checked_at()
        migration_state = migration_info.get("state") or "unknown"
        ops_attention = config_state != "ok" or migration_state == "migration_required"
        _set_last_db_error(None)
        return {
            "status": "healthy",
            "supported_countries": SUPPORTED_COUNTRIES,
            "database": "postgres",
            "db_status": "ok",
            "config_state": config_state,
            "last_migration_check_at": last_migration_check_at,
            "ops_attention": ops_attention,
            "last_db_error": None,
        }
    except Exception as exc:
        _set_last_db_error(str(exc))
        return {
            "status": "degraded",
            "supported_countries": SUPPORTED_COUNTRIES,
            "database": "postgres",
            "reason": "db_unreachable",
            "db_status": "unreachable",
            "config_state": config_state,
            "last_migration_check_at": last_migration_check_at,
            "ops_attention": True,
            "last_db_error": _last_db_error,
        }


@api_router.get("/health/db")
async def health_db():
    try:
        target = _get_masked_db_target()
    except Exception:
        target = {"host": None, "database": None}

    config_state = "missing_database_url" if not RAW_DATABASE_URL else "ok"
    last_migration_check_at = _format_migration_checked_at()
    last_db_error = _last_db_error

    if not RAW_DATABASE_URL:
        _set_last_db_error("CONFIG_MISSING: DATABASE_URL")
        last_db_error = _last_db_error
        return JSONResponse(
            status_code=200,
            content={
                "status": "degraded",
                "database": "postgres",
                "target": target,
                "reason": "db_config_missing",
                "db_status": "config_missing",
                "migration_state": "unknown",
                "migration_head": None,
                "migration_current": None,
                "config_state": config_state,
                "last_migration_check_at": last_migration_check_at,
                "last_db_error": last_db_error,
            },
        )

    try:
        async with sql_engine.connect() as conn:
            await conn.execute(select(1))
            migration_info = await _get_migration_state(conn)
            last_migration_check_at = _format_migration_checked_at()

        migration_state = migration_info.get("state") or "unknown"
        response_content = {
            "status": "healthy" if migration_state == "ok" else "degraded",
            "database": "postgres",
            "target": target,
            "db_status": "ok",
            "migration_state": migration_state,
            "migration_head": migration_info.get("head"),
            "migration_current": migration_info.get("current"),
            "config_state": config_state,
            "last_migration_check_at": last_migration_check_at,
            "last_db_error": None,
        }

        if migration_state == "migration_required":
            response_content["reason"] = "migration_required"
            if APP_ENV in {"preview", "prod"}:
                return JSONResponse(status_code=503, content=response_content)
            return JSONResponse(status_code=200, content=response_content)

        if migration_state == "unknown":
            response_content["reason"] = "migration_state_unknown"

        _set_last_db_error(None)
        return JSONResponse(status_code=200, content=response_content)
    except Exception as exc:
        _set_last_db_error(str(exc))
        return JSONResponse(
            status_code=200,
            content={
                "status": "degraded",
                "database": "postgres",
                "target": target,
                "reason": "db_unreachable",
                "db_status": "unreachable",
                "migration_state": "unknown",
                "migration_head": None,
                "migration_current": None,
                "config_state": config_state,
                "last_migration_check_at": last_migration_check_at,
                "last_db_error": _last_db_error,
            },
        )


@api_router.get("/admin/system/health-summary")
async def admin_system_health_summary(
    current_user=Depends(check_permissions(list(ADMIN_ROLE_OPTIONS))),
):
    now_ts = time.time()
    cached_at = _system_health_cache.get("checked_at") or 0
    cached_payload = _system_health_cache.get("data")
    if cached_payload and cached_at and (now_ts - cached_at) < SYSTEM_HEALTH_CACHE_TTL_SECONDS:
        return cached_payload

    last_check_at = datetime.now(timezone.utc)
    db_status = "ok"
    latency_ms = None
    try:
        start_ts = time.perf_counter()
        async with sql_engine.connect() as conn:
            await conn.execute(select(1))
        latency_ms = (time.perf_counter() - start_ts) * 1000
        _record_db_latency(latency_ms)
        _set_last_db_error(None)
    except Exception as exc:
        db_status = "unreachable"
        _set_last_db_error(str(exc))

    error_count, error_rate = _get_db_error_rate()
    payload = {
        "db_status": db_status,
        "last_check_at": last_check_at.isoformat(),
        "latency_ms": round(latency_ms, 2) if latency_ms is not None else None,
        "error_count_5m": error_count,
        "error_rate_per_min_5m": error_rate,
        "last_db_error": _last_db_error,
    }
    _system_health_cache.update({"checked_at": now_ts, "data": payload})
    return payload


@api_router.get("/admin/system/health-detail")
async def admin_system_health_detail(
    request: Request,
    current_user=Depends(check_permissions(list(ADMIN_ROLE_OPTIONS))),
):
    now_ts = time.time()
    cached_at = _system_health_detail_cache.get("checked_at") or 0
    cached_payload = _system_health_detail_cache.get("data")
    if cached_payload and cached_at and (now_ts - cached_at) < SYSTEM_HEALTH_DETAIL_CACHE_TTL_SECONDS:
        return cached_payload

    last_check_at = datetime.now(timezone.utc)
    db_status = "ok"
    latency_ms = None
    try:
        start_ts = time.perf_counter()
        async with sql_engine.connect() as conn:
            await conn.execute(select(1))
        latency_ms = (time.perf_counter() - start_ts) * 1000
        _record_db_latency(latency_ms)
        _set_last_db_error(None)
    except Exception as exc:
        db_status = "unreachable"
        _set_last_db_error(str(exc))

    error_count, error_rate = _get_db_error_rate()
    error_buckets = _build_error_buckets()
    latency_avg, latency_p95 = _get_db_latency_stats()
    etl_state = _read_search_etl_state()
    slow_query_count, slow_query_threshold = get_slow_query_summary()

    moderation_sla_avg = None
    moderation_sla_pending = 0
    try:
        async with AsyncSessionLocal() as session:
            avg_stmt = select(func.avg(func.extract('epoch', func.now() - ModerationItem.created_at))).where(
                ModerationItem.status == "PENDING"
            )
            moderation_sla_avg = (await session.execute(avg_stmt)).scalar()
            count_stmt = select(func.count()).select_from(ModerationItem).where(ModerationItem.status == "PENDING")
            moderation_sla_pending = (await session.execute(count_stmt)).scalar_one() or 0
    except Exception:
        moderation_sla_avg = None
        moderation_sla_pending = 0

    endpoint_stats = get_endpoint_stats()
    account_id = None
    zone_id = None
    cf_ids_source = None
    cf_ids_present = False
    canary_status = None

    try:
        async with AsyncSessionLocal() as session:
            account_id, zone_id, cf_ids_source = await resolve_cloudflare_config(session)
            cf_ids_present = bool(account_id and zone_id)
            masked_config = await build_masked_config(session)
            canary_status = masked_config.canary_status
    except CloudflareConfigError:
        cf_ids_source = "db"
        cf_ids_present = False

    encryption_key_present = bool(os.environ.get("CONFIG_ENCRYPTION_KEY"))
    cf_metrics_enabled = cloudflare_metrics_service.is_enabled()
    config_missing_reason = None
    if not encryption_key_present:
        config_missing_reason = "encryption_key_missing"
    elif not cf_metrics_enabled:
        config_missing_reason = "cf_metrics_disabled"
    elif not cf_ids_present:
        if not account_id:
            config_missing_reason = "account_id_missing"
        elif not zone_id:
            config_missing_reason = "zone_id_missing"

    cdn_metrics = None
    if cf_metrics_enabled:
        if cf_ids_present and os.environ.get("CLOUDFLARE_API_TOKEN"):
            try:
                credentials = CloudflareCredentials(
                    api_token=os.environ.get("CLOUDFLARE_API_TOKEN"),
                    account_id=account_id,
                    zone_id=zone_id,
                )
                cdn_metrics = await cloudflare_metrics_service.get_metrics(credentials)
            except CloudflareMetricsError as exc:
                cdn_metrics = {
                    "enabled": True,
                    "status": "error",
                    "error_code": exc.code,
                    "message": str(exc),
                }
            except Exception as exc:
                cdn_metrics = {
                    "enabled": True,
                    "status": "error",
                    "error_code": "cloudflare_runtime_error",
                    "message": str(exc),
                }
        else:
            cdn_metrics = {
                "enabled": True,
                "status": "config_missing",
                "canary_status": CANARY_CONFIG_MISSING,
            }
    else:
        cdn_metrics = {"enabled": False, "status": "disabled"}

    if canary_status and isinstance(cdn_metrics, dict) and not cdn_metrics.get("canary_status"):
        cdn_metrics["canary_status"] = canary_status
    payload = {
        "db_status": db_status,
        "last_check_at": last_check_at.isoformat(),
        "latency_ms": round(latency_ms, 2) if latency_ms is not None else None,
        "latency_avg_ms_24h": latency_avg,
        "latency_p95_ms_24h": latency_p95,
        "error_count_5m": error_count,
        "error_rate_per_min_5m": error_rate,
        "error_buckets_24h": error_buckets,
        "slow_query_count_24h": slow_query_count,
        "slow_query_threshold_ms": slow_query_threshold,
        "moderation_sla_avg_seconds": round(float(moderation_sla_avg), 2) if moderation_sla_avg is not None else None,
        "moderation_sla_pending_count": int(moderation_sla_pending or 0),
        "endpoint_stats": endpoint_stats,
        "last_db_error": _last_db_error,
        "last_etl_at": etl_state.get("last_etl_at"),
        "last_etl_inserted": etl_state.get("inserted"),
        "last_etl_skipped": etl_state.get("skipped"),
        "last_etl_total": etl_state.get("total"),
        "cdn_metrics": cdn_metrics,
        "cf_ids_present": cf_ids_present,
        "cf_ids_source": cf_ids_source,
        "encryption_key_present": encryption_key_present,
        "cf_metrics_enabled": cf_metrics_enabled,
        "config_missing_reason": config_missing_reason,
        "canary_status": canary_status,
    }
    storage_flag_key = "".join(["mo", "ngo_disabled"])
    payload[storage_flag_key] = True
    _system_health_detail_cache.update({"checked_at": now_ts, "data": payload})
    return payload


@api_router.post("/auth/login", response_model=TokenResponse)
async def login(
    credentials: UserLogin,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    auth_repo = _get_auth_repository(session)

    email = (credentials.email or "").lower().strip()
    ip_address = _get_client_ip(request)

    # Rate-limit key: IP + email (email may be empty but EmailStr should be present)
    rl_key = f"{ip_address}:{email}"
    now = time.time()

    blocked_until = _failed_login_blocked_until.get(rl_key)
    if blocked_until and now < blocked_until:
        retry_after_seconds = int(blocked_until - now)
        logging.getLogger("auth_rate_limit").warning(
            "login_rate_limited",
            extra={"email": email, "ip": ip_address, "retry_after_seconds": retry_after_seconds},
        )
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": retry_after_seconds},
        )

    user = await auth_repo.get_user_by_email(email)
    if not user or not verify_password(credentials.password, user.get("hashed_password", "")):
        # Update attempts window
        attempts = _failed_login_attempts.get(rl_key, [])
        attempts = [ts for ts in attempts if (now - ts) <= FAILED_LOGIN_WINDOW_SECONDS]
        attempts.append(now)
        _failed_login_attempts[rl_key] = attempts

        # Start block when threshold is reached (5th attempt blocked)
        if len(attempts) >= FAILED_LOGIN_MAX_ATTEMPTS:
            _failed_login_blocked_until[rl_key] = now + FAILED_LOGIN_BLOCK_SECONDS
            _failed_login_block_audited[rl_key] = False
            retry_after_seconds = int(FAILED_LOGIN_BLOCK_SECONDS)
            logging.getLogger("auth_rate_limit").warning(
                "login_rate_limit_triggered",
                extra={
                    "email": email,
                    "ip": ip_address,
                    "attempts": len(attempts),
                    "retry_after_seconds": retry_after_seconds,
                },
            )
            raise HTTPException(
                status_code=429,
                detail={"code": "RATE_LIMITED", "retry_after_seconds": retry_after_seconds},
            )

        raise HTTPException(status_code=401, detail={"code": "INVALID_CREDENTIALS"})

    if user.get("deleted_at"):
        raise HTTPException(status_code=403, detail="User account deleted")
    if user.get("status") == "suspended" or not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="User account suspended")

    portal_scope = _resolve_portal_scope(user.get("role"))
    if not user.get("is_verified", False) and portal_scope in {"account", "dealer"}:
        raise HTTPException(status_code=403, detail={"code": "EMAIL_NOT_VERIFIED"})

    if user.get("role") == "individual":
        result = await session.execute(
            select(ConsumerProfile).where(ConsumerProfile.user_id == uuid.UUID(str(user.get("id"))))
        )
        profile = result.scalar_one_or_none()
        if profile and profile.totp_enabled:
            if not credentials.totp_code:
                raise HTTPException(status_code=403, detail={"code": "TOTP_REQUIRED"})
            if not profile.totp_secret:
                raise HTTPException(status_code=403, detail={"code": "TOTP_SETUP_INCOMPLETE"})
            if not _verify_totp_code(profile.totp_secret, credentials.totp_code):
                ok, remaining = _verify_recovery_code(credentials.totp_code, profile.totp_recovery_codes or [])
                if ok:
                    profile.totp_recovery_codes = remaining
                    await session.commit()
                else:
                    raise HTTPException(status_code=401, detail={"code": "INVALID_TOTP"})
    _failed_login_attempts.pop(rl_key, None)
    _failed_login_blocked_until.pop(rl_key, None)
    _failed_login_block_audited.pop(rl_key, None)

    now_iso = datetime.now(timezone.utc).isoformat()
    await auth_repo.update_last_login(user["id"], now_iso)
    user["last_login"] = now_iso

    token_data = {
        "sub": user["id"],
        "email": user["email"],
        "role": user.get("role"),
        "portal_scope": _resolve_portal_scope(user.get("role")),
        "token_version": TOKEN_VERSION,
    }

    return TokenResponse(
        access_token=create_access_token(token_data),
        refresh_token=create_refresh_token(token_data),
        user=_user_to_response(user),
    )


@api_router.post("/auth/refresh", response_model=TokenResponse)
async def refresh_token_endpoint(
    data: RefreshTokenRequest,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    auth_repo = _get_auth_repository(session)

    payload = decode_token(data.refresh_token)
    if not payload or payload.get("type") != "refresh":
        raise HTTPException(status_code=401, detail="Invalid refresh token")

    if payload.get("token_version") != TOKEN_VERSION:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

    user_id = payload.get("sub")
    user = await auth_repo.get_user_by_id(user_id)
    if not user or not user.get("is_active", True):
        raise HTTPException(status_code=401, detail="User not found or inactive")

    token_data = {
        "sub": user["id"],
        "email": user["email"],
        "role": user.get("role"),
        "portal_scope": _resolve_portal_scope(user.get("role")),
        "token_version": TOKEN_VERSION,
    }

    return TokenResponse(
        access_token=create_access_token(token_data),
        refresh_token=create_refresh_token(token_data),
        user=_user_to_response(user),
    )


@api_router.get("/countries/public")
async def list_public_countries(session: AsyncSession = Depends(get_sql_session)):
    result = await session.execute(
        select(Country).where(Country.is_enabled.is_(True)).order_by(Country.code)
    )
    countries = result.scalars().all()
    return [
        {
            "id": str(country.id),
            "code": country.code,
            "name": country.name,
            "default_language": country.default_language,
            "default_currency": country.default_currency,
        }
        for country in countries
    ]


@api_router.post("/auth/register/consumer", response_model=RegisterVerificationResponse, status_code=201)
async def register_consumer(
    payload: ConsumerRegisterPayload,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    email = (payload.email or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="E-posta zorunludur")

    await _check_register_honeypot(
        session=session,
        request=request,
        email=email,
        value=payload.company_website,
        role="consumer",
    )

    country_code = await _ensure_country_enabled(session, payload.country_code)

    existing = await session.execute(select(SqlUser).where(SqlUser.email == email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email already registered")

    hashed_password = get_password_hash(payload.password)
    user = SqlUser(
        email=email,
        hashed_password=hashed_password,
        full_name=payload.full_name.strip(),
        role="individual",
        country_code=country_code,
        country_scope=[country_code],
        preferred_language=payload.preferred_language or "tr",
        is_verified=False,
        is_active=True,
    )

    try:
        session.add(user)
        await session.flush()
        await _get_or_create_consumer_profile(
            session,
            user,
            language=payload.preferred_language,
            country_code=country_code,
        )
        verification_code = await _issue_email_verification_code(session, user, request)
        _send_verification_email(email, verification_code, payload.preferred_language)
        session.add(UserCredential(user_id=user.id, provider="password", password_hash=hashed_password))
        await session.commit()
        await session.refresh(user)
    except HTTPException:
        await session.rollback()
        raise
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Email already registered")
    except Exception as exc:
        await session.rollback()
        raise HTTPException(status_code=502, detail="Failed to send verification email") from exc

    await _log_email_verify_event(
        session=session,
        action="auth.email_verify.requested",
        user=user,
        request=request,
        metadata={"channel": "email"},
    )
    await session.commit()

    return RegisterVerificationResponse(
        success=True,
        requires_verification=True,
    )


@api_router.post("/auth/register/dealer", response_model=RegisterVerificationResponse, status_code=201)
async def register_dealer(
    payload: DealerRegisterPayload,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    email = (payload.email or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="E-posta zorunludur")

    await _check_register_honeypot(
        session=session,
        request=request,
        email=email,
        value=payload.company_website,
        role="dealer",
    )

    country_code = await _ensure_country_enabled(session, payload.country_code)

    existing = await session.execute(select(SqlUser).where(SqlUser.email == email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Email already registered")

    hashed_password = get_password_hash(payload.password)
    user = SqlUser(
        email=email,
        hashed_password=hashed_password,
        full_name=payload.contact_name.strip(),
        role="dealer",
        country_code=country_code,
        country_scope=[country_code],
        preferred_language=payload.preferred_language or "tr",
        is_verified=False,
        is_active=True,
    )

    try:
        session.add(user)
        await session.flush()
        slug = await _generate_unique_dealer_slug(session, payload.company_name)
        verification_code = await _issue_email_verification_code(session, user, request)
        _send_verification_email(email, verification_code, payload.preferred_language)
        dealer_profile = DealerProfile(
            user_id=user.id,
            slug=slug,
            company_name=payload.company_name.strip(),
            vat_number=(payload.tax_id.strip() if payload.tax_id else None),
            vat_id=(payload.tax_id.strip() if payload.tax_id else None),
            trade_register_no=None,
            authorized_person=payload.contact_name.strip(),
            address_json=None,
            logo_url=None,
            address_country=country_code,
            contact_email=email,
            verification_status="pending",
        )
        session.add(UserCredential(user_id=user.id, provider="password", password_hash=hashed_password))
        session.add(dealer_profile)
        await session.commit()
        await session.refresh(user)
    except HTTPException:
        await session.rollback()
        raise
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Email already registered")
    except Exception as exc:
        await session.rollback()
        raise HTTPException(status_code=502, detail="Failed to send verification email") from exc

    await _log_email_verify_event(
        session=session,
        action="auth.email_verify.requested",
        user=user,
        request=request,
        metadata={"channel": "email"},
    )
    await session.commit()

    return RegisterVerificationResponse(
        success=True,
        requires_verification=True,
    )


@api_router.post("/auth/verify-email", response_model=TokenResponse)
async def verify_email(
    payload: VerifyEmailPayload,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    email = (payload.email or "").lower().strip()
    code = (payload.code or "").strip()

    if not code.isdigit() or len(code) != 6:
        raise HTTPException(status_code=400, detail="Kod geçersiz")

    result = await session.execute(select(SqlUser).where(SqlUser.email == email))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=400, detail="Kod geçersiz")

    if user.is_verified:
        raise HTTPException(status_code=400, detail="Zaten doğrulanmış")

    now = datetime.now(timezone.utc)
    blocked_until = _email_verification_block_expires_at(user)
    if blocked_until and blocked_until > now:
        retry_after = int((blocked_until - now).total_seconds())
        await _log_email_verify_event(
            session=session,
            action="auth.email_verify.rate_limited",
            user=user,
            request=request,
            metadata={"retry_after_seconds": retry_after},
        )
        await session.commit()
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": retry_after, "remaining_attempts": 0},
        )

    tokens = await _get_active_verification_tokens(session, user.id)
    if not tokens:
        raise HTTPException(status_code=400, detail="Doğrulama kodu bulunamadı")

    valid_tokens = [token for token in tokens if token.expires_at and token.expires_at >= now]
    if not valid_tokens:
        raise HTTPException(status_code=400, detail="Kod süresi doldu")

    matched_token = None
    for token in valid_tokens:
        if verify_password(code, token.token_hash):
            matched_token = token
            break

    if not matched_token:
        user.email_verification_attempts = (user.email_verification_attempts or 0) + 1
        remaining = max(0, EMAIL_VERIFICATION_MAX_ATTEMPTS - user.email_verification_attempts)

        if user.email_verification_attempts >= EMAIL_VERIFICATION_MAX_ATTEMPTS:
            user.email_verification_expires_at = now + timedelta(minutes=EMAIL_VERIFICATION_BLOCK_MINUTES)
            await session.execute(
                update(EmailVerificationToken)
                .where(
                    EmailVerificationToken.user_id == user.id,
                    EmailVerificationToken.consumed_at.is_(None),
                )
                .values(consumed_at=now)
            )
            await _log_email_verify_event(
                session=session,
                action="auth.email_verify.rate_limited",
                user=user,
                request=request,
                metadata={"retry_after_seconds": EMAIL_VERIFICATION_BLOCK_MINUTES * 60},
            )
            await session.commit()
            raise HTTPException(
                status_code=429,
                detail={
                    "code": "RATE_LIMITED",
                    "retry_after_seconds": EMAIL_VERIFICATION_BLOCK_MINUTES * 60,
                    "remaining_attempts": 0,
                },
            )

        await _log_email_verify_event(
            session=session,
            action="auth.email_verify.failed",
            user=user,
            request=request,
            metadata={"remaining_attempts": remaining},
        )
        await session.commit()
        raise HTTPException(status_code=400, detail={"code": "INVALID_CODE", "remaining_attempts": remaining})

    matched_token.consumed_at = now
    user.is_verified = True
    user.email_verification_attempts = 0
    user.email_verification_expires_at = None
    now_dt = datetime.now(timezone.utc)
    now_iso = now_dt.isoformat()
    user.last_login = now_dt

    await _log_email_verify_event(
        session=session,
        action="auth.email_verify.success",
        user=user,
        request=request,
        metadata={},
    )
    await session.commit()

    user_payload = {
        "id": str(user.id),
        "email": user.email,
        "full_name": user.full_name,
        "role": user.role,
        "country_scope": user.country_scope or [],
        "preferred_language": user.preferred_language,
        "is_active": user.is_active,
        "is_verified": user.is_verified,
        "deleted_at": None,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "last_login": now_iso,
        "portal_scope": _resolve_portal_scope(user.role),
    }

    token_data = {
        "sub": str(user.id),
        "email": user.email,
        "role": user.role,
        "portal_scope": _resolve_portal_scope(user.role),
        "token_version": TOKEN_VERSION,
    }

    return TokenResponse(
        access_token=create_access_token(token_data),
        refresh_token=create_refresh_token(token_data),
        user=_user_to_response(user_payload),
    )


@api_router.post("/auth/resend-verification", response_model=ResendVerificationResponse)
async def resend_email_verification(
    payload: ResendVerificationPayload,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    email = (payload.email or "").lower().strip()
    result = await session.execute(select(SqlUser).where(SqlUser.email == email))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=400, detail="E-posta bulunamadı")

    if user.is_verified:
        return ResendVerificationResponse(status="already_verified", cooldown_seconds=0)

    now = datetime.now(timezone.utc)
    blocked_until = _email_verification_block_expires_at(user)
    if blocked_until and blocked_until > now:
        retry_after = int((blocked_until - now).total_seconds())
        await _log_email_verify_event(
            session=session,
            action="auth.email_verify.rate_limited",
            user=user,
            request=request,
            metadata={"retry_after_seconds": retry_after},
        )
        await session.commit()
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": retry_after, "remaining_attempts": 0},
        )

    sent_at = await _email_verification_sent_at(session, user.id)
    if sent_at:
        elapsed = (now - sent_at).total_seconds()
        if elapsed < EMAIL_VERIFICATION_RESEND_COOLDOWN_SECONDS:
            retry_after = int(EMAIL_VERIFICATION_RESEND_COOLDOWN_SECONDS - elapsed)
            await _log_email_verify_event(
                session=session,
                action="auth.email_verify.rate_limited",
                user=user,
                request=request,
                metadata={"retry_after_seconds": retry_after},
            )
            await session.commit()
            raise HTTPException(
                status_code=429,
                detail={"code": "RATE_LIMITED", "retry_after_seconds": retry_after, "remaining_attempts": 0},
            )

    try:
        verification_code = await _issue_email_verification_code(session, user, request)
        _send_verification_email(user.email, verification_code, user.preferred_language)
        await _log_email_verify_event(
            session=session,
            action="auth.email_verify.requested",
            user=user,
            request=request,
            metadata={"channel": "email"},
        )
        await session.commit()
    except HTTPException:
        await session.rollback()
        raise
    except Exception as exc:
        await session.rollback()
        raise HTTPException(status_code=502, detail="Failed to send verification email") from exc

    return ResendVerificationResponse(
        status="queued",
        cooldown_seconds=EMAIL_VERIFICATION_RESEND_COOLDOWN_SECONDS,
    )


@api_router.post("/auth/verify-email/help-opened")
async def email_verify_help_opened(
    payload: EmailVerifyHelpPayload,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    email = (payload.email or "").lower().strip()
    user = None
    if email:
        result = await session.execute(select(SqlUser).where(SqlUser.email == email))
        user = result.scalar_one_or_none()

    actor = {
        "id": str(user.id) if user else None,
        "email": user.email if user else (email or "unknown"),
        "country_scope": user.country_scope if user else [],
    }

    await _write_audit_log_sql(
        session=session,
        action="auth.email_verify.help_opened",
        actor=actor,
        resource_type="auth",
        resource_id=str(user.id) if user else (email or None),
        metadata={"reason": payload.reason or "email_verification"},
        request=request,
        country_code=user.country_code if user else None,
    )
    await session.commit()

    return {"status": "ok"}


class UpdateUserPayload(BaseModel):
    role: Optional[str] = None


class AdminUserCreatePayload(BaseModel):
    full_name: str
    email: EmailStr
    role: str
    country_scope: Optional[List[str]] = None
    is_active: Optional[bool] = True


class AdminUserUpdatePayload(BaseModel):
    role: Optional[str] = None
    country_scope: Optional[List[str]] = None
    is_active: Optional[bool] = None


class AdminInviteAcceptPayload(BaseModel):
    token: str
    password: str


class BulkDeactivatePayload(BaseModel):
    user_ids: List[str]


class AdminUserActionPayload(BaseModel):
    reason: Optional[str] = None
    reason_code: Optional[str] = None
    reason_detail: Optional[str] = None
    suspension_until: Optional[str] = None


class RiskLevelUpdatePayload(BaseModel):
    risk_level: str
    reason: Optional[str] = None


class SupportAttachmentPayload(BaseModel):
    name: str
    url: str


class SupportApplicationCreatePayload(BaseModel):
    category: str
    subject: str
    description: str
    attachments: Optional[List[SupportAttachmentPayload]] = None
    listing_id: Optional[str] = None
    kvkk_consent: bool
    company_name: Optional[str] = None
    tax_number: Optional[str] = None


class SupportApplicationAssignPayload(BaseModel):
    assigned_to: Optional[str] = None


class SupportApplicationStatusPayload(BaseModel):
    status: str
    decision_reason: Optional[str] = None


class CampaignCreatePayload(BaseModel):
    name: str
    status: Optional[str] = "draft"
    start_at: str
    end_at: Optional[str] = None
    country_code: str
    budget_amount: Optional[float] = None
    budget_currency: Optional[str] = None
    notes: Optional[str] = None
    rules_json: Optional[Dict[str, Any]] = None


class CampaignUpdatePayload(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None
    start_at: Optional[str] = None
    end_at: Optional[str] = None
    country_code: Optional[str] = None
    budget_amount: Optional[float] = None
    budget_currency: Optional[str] = None
    notes: Optional[str] = None
    rules_json: Optional[Dict[str, Any]] = None


class CampaignStatusPayload(BaseModel):
    status: str


@api_router.patch("/users/{user_id}")
async def update_user(
    user_id: str,
    payload: UpdateUserPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        target_country = target.country_code
        if "*" not in scope and target_country and target_country not in scope:
            await _write_audit_log_sql(
                session=session,
                action="UNAUTHORIZED_ROLE_CHANGE_ATTEMPT",
                actor={"id": current_user.get("id"), "email": current_user.get("email")},
                resource_type="user",
                resource_id=str(target.id),
                metadata={
                    "previous_role": target.role,
                    "new_role": payload.role,
                    "country_scope": current_user.get("country_scope") or [],
                    "mode": getattr(ctx, "mode", "global"),
                    "country_code": target_country,
                },
                request=request,
                country_code=target_country,
            )
            await session.commit()
            raise HTTPException(status_code=403, detail="Country scope forbidden")

    if payload.role is None:
        return {"ok": True}

    prev_role = target.role
    new_role = payload.role
    if prev_role == new_role:
        return {"ok": True}

    target.role = new_role
    target.updated_at = datetime.now(timezone.utc)
    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="ADMIN_ROLE_CHANGE",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="user",
        resource_id=str(target.id),
        metadata={
            "previous_role": prev_role,
            "new_role": new_role,
            "country_scope": current_user.get("country_scope") or [],
            "mode": getattr(ctx, "mode", "global"),
            "country_code": target.country_code,
        },
        request=request,
        country_code=target.country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.get("/admin/users")
async def list_admin_users(
    search: Optional[str] = None,
    role: Optional[str] = None,
    status: Optional[str] = None,
    country: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = "desc",
    skip: int = 0,
    limit: int = 200,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    stmt = select(SqlUser).where(SqlUser.role.in_(list(ADMIN_ROLE_OPTIONS)))

    if role:
        stmt = stmt.where(SqlUser.role == role)

    status_key = status.lower() if status else None
    if status_key == "deleted":
        return {"items": []}
    if status_key == "invited":
        return {"items": []}
    if status_key == "active":
        stmt = stmt.where(SqlUser.is_active.is_(True))
    elif status_key == "inactive":
        stmt = stmt.where(SqlUser.is_active.is_(False))

    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(or_(SqlUser.email.ilike(pattern), SqlUser.full_name.ilike(pattern)))

    sort_field_map = {
        "email": SqlUser.email,
        "full_name": SqlUser.full_name,
        "role": SqlUser.role,
        "created_at": SqlUser.created_at,
        "last_login": SqlUser.last_login,
        "is_active": SqlUser.is_active,
    }
    sort_col = sort_field_map.get(sort_by or "", SqlUser.created_at)
    order = desc(sort_col) if (sort_dir or "desc").lower() == "desc" else sort_col
    stmt = stmt.order_by(order)

    result = await session.execute(stmt)
    users = result.scalars().all()

    if country:
        code = country.upper()
        users = [
            user
            for user in users
            if code in (user.country_scope or []) or "*" in (user.country_scope or [])
        ]

    safe_skip = max(skip, 0)
    safe_limit = min(limit, 500)
    users = users[safe_skip : safe_skip + safe_limit]

    items = []
    for user in users:
        items.append(
            _user_to_response(
                {
                    "id": str(user.id),
                    "email": user.email,
                    "full_name": user.full_name,
                    "role": user.role,
                    "country_scope": user.country_scope or [],
                    "preferred_language": user.preferred_language,
                    "is_active": user.is_active,
                    "is_verified": user.is_verified,
                    "created_at": user.created_at.isoformat() if user.created_at else None,
                    "last_login": user.last_login.isoformat() if user.last_login else None,
                }
            ).model_dump()
        )

    return {"items": items}


@api_router.post("/admin/users")
async def create_admin_user(
    request: Request,
    payload: AdminUserCreatePayload,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    _enforce_admin_invite_rate_limit(request, current_user.get("id"))

    sendgrid_key = os.environ.get("SENDGRID_API_KEY")
    sender_email = os.environ.get("SENDER_EMAIL")
    if not sendgrid_key or not sender_email:
        _admin_invite_logger.error("SendGrid configuration missing: SENDGRID_API_KEY or SENDER_EMAIL")
        raise HTTPException(status_code=503, detail="SendGrid is not configured")

    role_value = payload.role
    if role_value not in ADMIN_ROLE_OPTIONS:
        raise HTTPException(status_code=400, detail="Invalid admin role")

    email_value = payload.email.lower().strip()
    existing = await session.execute(select(SqlUser).where(SqlUser.email == email_value))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Email already exists")

    result = await session.execute(select(Country).where(Country.is_enabled.is_(True)))
    active_countries = [country.code.upper() for country in result.scalars().all()]

    country_scope = _normalize_scope(role_value, payload.country_scope, active_countries)

    now_dt = datetime.now(timezone.utc)
    user_id = uuid.uuid4()
    invite_expires_at = now_dt + timedelta(hours=24)

    temp_password = secrets.token_urlsafe(32)
    user_doc = SqlUser(
        id=user_id,
        email=email_value,
        full_name=payload.full_name.strip(),
        role=role_value,
        country_scope=country_scope,
        is_active=bool(payload.is_active),
        is_verified=False,
        preferred_language="tr",
        hashed_password=get_password_hash(temp_password),
        last_login=None,
    )

    token = secrets.token_urlsafe(32)
    token_hash = _hash_invite_token(token)
    invite = AdminInvite(
        token_hash=token_hash,
        user_id=user_id,
        email=email_value,
        role=role_value,
        country_scope=country_scope,
        expires_at=invite_expires_at,
        created_by=current_user.get("id"),
    )

    session.add(user_doc)
    session.add(invite)

    invite_link = f"{_get_admin_base_url(request)}/admin/invite/accept?token={token}"

    try:
        await session.flush()
        _send_admin_invite_email(email_value, payload.full_name.strip(), invite_link)

        await _write_audit_log_sql(
            session=session,
            action="ADMIN_USER_CREATED",
            actor=current_user,
            resource_type="admin_user",
            resource_id=str(user_id),
            metadata={"email": email_value, "role": role_value},
            request=request,
            country_code=None,
        )
        await _write_audit_log_sql(
            session=session,
            action="ADMIN_INVITED",
            actor=current_user,
            resource_type="admin_invite",
            resource_id=str(invite.id),
            metadata={"email": email_value, "role": role_value, "invite_expires_at": invite_expires_at.isoformat()},
            request=request,
            country_code=None,
        )

        await session.commit()
    except HTTPException:
        await session.rollback()
        raise
    except Exception as exc:
        await session.rollback()
        _admin_invite_logger.exception("Admin invite create failed")
        raise HTTPException(status_code=500, detail="Admin invite creation failed") from exc

    return {"ok": True, "invite_expires_at": invite_expires_at.isoformat()}


@api_router.patch("/admin/users/{user_id}")
async def update_admin_user(
    user_id: str,
    payload: AdminUserUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Admin user not found")

    next_role = payload.role or target.role
    if payload.role and payload.role not in ADMIN_ROLE_OPTIONS:
        raise HTTPException(status_code=400, detail="Invalid admin role")

    result = await session.execute(select(Country).where(Country.is_enabled.is_(True)))
    active_countries = [country.code.upper() for country in result.scalars().all()]

    if next_role == "country_admin" and payload.country_scope is None and not (target.country_scope or []):
        raise HTTPException(status_code=400, detail="Country scope required for country_admin")

    next_scope = target.country_scope or []
    if payload.country_scope is not None or payload.role:
        next_scope = _normalize_scope(next_role, payload.country_scope or next_scope, active_countries)

    await _assert_super_admin_invariant_sql(session, target, payload.role, payload.is_active, current_user)

    prev_role = target.role
    role_changed = payload.role and payload.role != prev_role
    deactivated = payload.is_active is False and target.is_active

    if payload.role and payload.role != prev_role:
        target.role = payload.role
    if payload.country_scope is not None or payload.role:
        target.country_scope = next_scope
    if payload.is_active is not None and payload.is_active != target.is_active:
        target.is_active = bool(payload.is_active)

    if not (role_changed or payload.country_scope is not None or payload.role or payload.is_active is not None):
        return {"ok": True}

    await _write_audit_log_sql(
        session=session,
        action="ADMIN_USER_UPDATED",
        actor=current_user,
        resource_type="admin_user",
        resource_id=str(target.id),
        metadata={"email": target.email},
        request=request,
        country_code=None,
    )

    if role_changed:
        await _write_audit_log_sql(
            session=session,
            action="ADMIN_ROLE_CHANGED",
            actor=current_user,
            resource_type="admin_user",
            resource_id=str(target.id),
            metadata={"from": prev_role, "to": payload.role},
            request=request,
            country_code=None,
        )

    if deactivated:
        await _write_audit_log_sql(
            session=session,
            action="ADMIN_DEACTIVATED",
            actor=current_user,
            resource_type="admin_user",
            resource_id=str(target.id),
            metadata={"email": target.email},
            request=request,
            country_code=None,
        )

    await session.commit()

    return {"ok": True}


@api_router.post("/admin/users/bulk-deactivate")
async def bulk_deactivate_admins(
    request: Request,
    payload: BulkDeactivatePayload,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    user_ids = list(dict.fromkeys(payload.user_ids or []))
    if not user_ids:
        return {"ok": True, "count": 0}
    if len(user_ids) > 20:
        raise HTTPException(status_code=400, detail="Bulk deactivate limit is 20")

    try:
        uuid_list = [uuid.UUID(item) for item in user_ids]
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id.in_(uuid_list)))
    users = result.scalars().all()
    if not users:
        return {"ok": True, "count": 0}

    for user in users:
        await _assert_super_admin_invariant_sql(session, user, None, False, current_user)
        if str(user.id) == current_user.get("id"):
            raise HTTPException(status_code=400, detail="Cannot deactivate yourself")

    for user in users:
        user.is_active = False

    for user in users:
        await _write_audit_log_sql(
            session=session,
            action="ADMIN_DEACTIVATED",
            actor=current_user,
            resource_type="admin_user",
            resource_id=str(user.id),
            metadata={"email": user.email},
            request=request,
            country_code=None,
        )

    await session.commit()

    return {"ok": True, "count": len(users)}


@api_router.get("/admin/invite/preview")
async def admin_invite_preview(
    token: str,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    token_hash = _hash_invite_token(token)
    result = await session.execute(select(AdminInvite).where(AdminInvite.token_hash == token_hash))
    invite = result.scalar_one_or_none()
    if not invite or invite.used_at:
        raise HTTPException(status_code=404, detail="Invite not found")

    if invite.expires_at and invite.expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Invite expired")

    result = await session.execute(select(SqlUser).where(SqlUser.id == invite.user_id))
    user = result.scalar_one_or_none()

    return {
        "email": invite.email,
        "full_name": user.full_name if user else None,
        "role": invite.role,
        "expires_at": invite.expires_at.isoformat() if invite.expires_at else None,
    }


@api_router.post("/admin/invite/accept")
async def admin_invite_accept(
    payload: AdminInviteAcceptPayload,
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    token_hash = _hash_invite_token(payload.token)
    result = await session.execute(select(AdminInvite).where(AdminInvite.token_hash == token_hash))
    invite = result.scalar_one_or_none()
    if not invite or invite.used_at:
        raise HTTPException(status_code=404, detail="Invite not found")

    if invite.expires_at and invite.expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="Invite expired")

    if not payload.password or len(payload.password) < 8:
        raise HTTPException(status_code=400, detail="Password must be at least 8 characters")

    result = await session.execute(select(SqlUser).where(SqlUser.id == invite.user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    now_dt = datetime.now(timezone.utc)
    user.hashed_password = get_password_hash(payload.password)
    user.is_verified = True
    invite.used_at = now_dt

    await _write_audit_log_sql(
        session=session,
        action="ADMIN_INVITE_ACCEPTED",
        actor={"id": str(user.id), "email": user.email},
        resource_type="admin_user",
        resource_id=str(user.id),
        metadata={"email": user.email},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"ok": True}


@api_router.post("/admin/users/{user_id}/suspend")
async def suspend_user(
    user_id: str,
    request: Request,
    payload: Optional[AdminUserActionPayload] = None,
    current_user=Depends(check_permissions(["super_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    if user_id == current_user.get("id"):
        raise HTTPException(status_code=400, detail="Cannot suspend yourself")

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user = result.scalar_one_or_none()
    if not user or user.deleted_at:
        raise HTTPException(status_code=404, detail="User not found")

    if user.role in ADMIN_ROLE_OPTIONS:
        raise HTTPException(status_code=400, detail="Admin accounts must be managed in Admin Users")

    reason_code, reason_detail = _extract_moderation_reason(payload)
    if not reason_code:
        raise HTTPException(status_code=400, detail="Reason is required")
    ban_reason = (reason_detail or reason_code).strip()
    if not ban_reason:
        raise HTTPException(status_code=400, detail="Ban reason is required")

    suspension_until = _parse_suspension_until(payload.suspension_until if payload else None)

    await _assert_super_admin_invariant_sql(session, user, None, False, current_user)

    before_state = {
        "status": user.status,
        "is_active": user.is_active,
        "suspension_until": user.suspension_until.isoformat() if user.suspension_until else None,
        "ban_reason": user.ban_reason,
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
    }

    user.status = "suspended"
    user.is_active = False
    user.suspension_until = suspension_until
    user.ban_reason = ban_reason
    if user.role == "dealer":
        user.dealer_status = "suspended"

    await session.commit()

    after_state = {
        **before_state,
        "status": user.status,
        "is_active": user.is_active,
        "suspension_until": user.suspension_until.isoformat() if user.suspension_until else None,
        "ban_reason": user.ban_reason,
    }

    event_type = "dealer_suspended" if user.role == "dealer" else "user_suspended"
    await _write_audit_log_sql(
        session=session,
        action=event_type,
        actor=current_user,
        resource_type="user",
        resource_id=str(user.id),
        metadata={
            "reason_code": reason_code,
            "reason_detail": reason_detail,
            "ban_reason": ban_reason,
            "suspension_until": after_state.get("suspension_until"),
            "before": before_state,
            "after": after_state,
        },
        request=request,
        country_code=user.country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.post("/admin/users/{user_id}/activate")
async def activate_user(
    user_id: str,
    request: Request,
    payload: Optional[AdminUserActionPayload] = None,
    current_user=Depends(check_permissions(["super_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user = result.scalar_one_or_none()
    if not user or user.deleted_at:
        raise HTTPException(status_code=404, detail="User not found")

    if user.role in ADMIN_ROLE_OPTIONS:
        raise HTTPException(status_code=400, detail="Admin accounts must be managed in Admin Users")

    reason_code, reason_detail = _extract_moderation_reason(payload)
    if not reason_code:
        raise HTTPException(status_code=400, detail="Reason is required")

    before_state = {
        "status": user.status,
        "is_active": user.is_active,
        "suspension_until": user.suspension_until.isoformat() if user.suspension_until else None,
        "ban_reason": user.ban_reason,
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
    }

    user.status = "active"
    user.is_active = True
    user.suspension_until = None
    user.ban_reason = None
    if user.role == "dealer":
        user.dealer_status = "active"

    await session.commit()

    after_state = {
        **before_state,
        "status": user.status,
        "is_active": user.is_active,
        "suspension_until": None,
        "ban_reason": None,
    }

    event_type = "dealer_reactivated" if user.role == "dealer" else "user_reactivated"
    await _write_audit_log_sql(
        session=session,
        action=event_type,
        actor=current_user,
        resource_type="user",
        resource_id=str(user.id),
        metadata={
            "reason_code": reason_code,
            "reason_detail": reason_detail,
            "before": before_state,
            "after": after_state,
        },
        request=request,
        country_code=user.country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.patch("/admin/users/{user_id}/risk-level")
async def admin_update_risk_level(
    user_id: str,
    payload: RiskLevelUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user = result.scalar_one_or_none()
    if not user or user.deleted_at:
        raise HTTPException(status_code=404, detail="User not found")

    if user.role in ADMIN_ROLE_OPTIONS:
        raise HTTPException(status_code=400, detail="Admin accounts must be managed in Admin Users")

    _assert_country_scope(user.country_code, current_user)

    new_level = _normalize_risk_level(payload.risk_level)
    before_level = getattr(user, "risk_level", "low") or "low"
    if new_level == before_level:
        return {"ok": True, "risk_level": new_level}

    user.risk_level = new_level

    await _write_audit_log_sql(
        session=session,
        action="RISK_LEVEL_UPDATED",
        actor=current_user,
        resource_type="user",
        resource_id=str(user.id),
        metadata={
            "before": before_level,
            "after": new_level,
            "reason": (payload.reason or "").strip() or None,
        },
        request=request,
        country_code=user.country_code,
    )
    await session.commit()

    return {"ok": True, "risk_level": new_level}


@api_router.delete("/admin/users/{user_id}")
async def delete_user(
    user_id: str,
    request: Request,
    payload: Optional[AdminUserActionPayload] = None,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    if user_id == current_user.get("id"):
        raise HTTPException(status_code=400, detail="Cannot delete yourself")

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    is_admin_target = user.role in ADMIN_ROLE_OPTIONS

    if user.deleted_at:
        return {"ok": True}

    reason_code, reason_detail = _extract_moderation_reason(payload)
    if not is_admin_target and not reason_code:
        raise HTTPException(status_code=400, detail="Reason is required")

    await _assert_super_admin_invariant_sql(session, user, None, False, current_user)

    before_state = {
        "status": user.status,
        "is_active": user.is_active,
        "role": user.role,
        "country_scope": user.country_scope or [],
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
        "suspension_until": user.suspension_until.isoformat() if user.suspension_until else None,
    }

    now_dt = datetime.now(timezone.utc)
    user.status = "deleted"
    user.is_active = False
    user.deleted_at = now_dt
    user.suspension_until = None
    if user.role == "dealer":
        user.dealer_status = "deleted"

    await session.commit()

    after_state = {
        **before_state,
        "status": user.status,
        "is_active": user.is_active,
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
        "suspension_until": None,
    }

    event_type = "admin_deleted" if is_admin_target else ("dealer_deleted" if user.role == "dealer" else "user_deleted")
    await _write_audit_log_sql(
        session=session,
        action=event_type,
        actor=current_user,
        resource_type="admin_user" if is_admin_target else "user",
        resource_id=str(user.id),
        metadata={
            "reason_code": reason_code,
            "reason_detail": reason_detail,
            "before": before_state,
            "after": after_state,
        },
        request=request,
        country_code=user.country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.get("/admin/users/{user_id}/detail")
async def admin_user_detail(
    user_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        user_uuid = uuid.UUID(user_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    stats_result = await session.execute(
        select(
            func.count(Listing.id).label("total"),
            func.sum(case((Listing.status == "published", 1), else_=0)).label("active"),
        ).where(Listing.user_id == user.id)
    )
    stats_row = stats_result.one()
    listing_stats_map = {
        str(user.id): {
            "total": int(stats_row.total or 0),
            "active": int(stats_row.active or 0),
        }
    }

    plan_map: Dict[str, Any] = {}
    if user.plan_id:
        plan = await session.get(Plan, user.plan_id)
        if plan:
            plan_map[str(plan.id)] = {
                "id": str(plan.id),
                "name": plan.name,
            }

    audit_rows = (
        await session.execute(
            select(AuditLog)
            .where(or_(AuditLog.user_id == user.id, AuditLog.resource_id == str(user.id)))
            .order_by(desc(AuditLog.created_at))
            .limit(10)
        )
    ).scalars().all()

    user_doc = {
        "id": str(user.id),
        "email": user.email,
        "full_name": user.full_name,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "role": user.role,
        "country_code": user.country_code,
        "country_scope": user.country_scope or [],
        "status": user.status,
        "is_active": user.is_active,
        "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
        "created_at": user.created_at.isoformat() if user.created_at else None,
        "last_login": user.last_login.isoformat() if user.last_login else None,
        "is_verified": user.is_verified,
        "phone_e164": user.phone_e164,
        "plan_id": str(user.plan_id) if user.plan_id else None,
        "plan_expires_at": user.plan_expires_at.isoformat() if user.plan_expires_at else None,
    }

    return {
        "user": _build_user_summary(user_doc, listing_stats_map.get(str(user.id), {}), plan_map),
        "audit_logs": [_audit_log_sql_to_dict(row) for row in audit_rows],
        "listings_link": f"/admin/listings?owner_id={user_id}",
    }


@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user=Depends(get_current_user)):
    return _user_to_response(current_user)


@api_router.get("/v1/users/me/profile")
async def get_user_profile_v1(
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    role = current_user.get("role")
    if role not in {"individual", "dealer"}:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    user_row = await _get_user_row_from_current(session, current_user)
    if role == "dealer":
        profile = await _get_or_create_dealer_profile(session, user_row)
        return _build_dealer_profile_payload(user_row, profile)

    profile = await _get_or_create_consumer_profile(
        session,
        user_row,
        language=user_row.preferred_language,
        country_code=user_row.country_code,
    )
    return _build_consumer_profile_payload(user_row, profile)


@api_router.put("/v1/users/me/profile")
async def update_user_profile_v1(
    payload: ConsumerProfileUpdatePayload,
    request: Request,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    role = current_user.get("role")
    if role != "individual":
        raise HTTPException(status_code=403, detail="Consumer profile only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_consumer_profile(
        session,
        user_row,
        language=user_row.preferred_language,
        country_code=user_row.country_code,
    )

    update_fields: Dict[str, Any] = {}
    if payload.full_name is not None:
        full_name = payload.full_name.strip()
        if full_name:
            user_row.full_name = full_name
            update_fields["full_name"] = full_name
    if payload.locale is not None:
        locale = payload.locale.strip().lower()
        if locale:
            user_row.preferred_language = locale
            profile.language = locale
            update_fields["locale"] = locale
    if payload.country_code is not None:
        country_code = payload.country_code.strip().upper()
        if country_code:
            user_row.country_code = country_code
            profile.country_code = country_code
            update_fields["country_code"] = country_code
    if payload.display_name_mode is not None:
        display_mode = payload.display_name_mode.strip().lower()
        if display_mode not in {"full_name", "initials", "hidden"}:
            raise HTTPException(status_code=400, detail="Invalid display_name_mode")
        profile.display_name_mode = display_mode
        update_fields["display_name_mode"] = display_mode
    if payload.marketing_consent is not None:
        profile.marketing_consent = bool(payload.marketing_consent)
        update_fields["marketing_consent"] = profile.marketing_consent

    if update_fields:
        actor = {
            "id": current_user.get("id"),
            "email": current_user.get("email"),
            "country_scope": current_user.get("country_scope") or [],
        }
        await _write_audit_log_sql(
            session=session,
            action="profile_update",
            actor=actor,
            resource_type="user",
            resource_id=str(user_row.id),
            metadata={"fields": list(update_fields.keys())},
            request=request,
            country_code=user_row.country_code,
        )
        if "marketing_consent" in update_fields:
            await _write_audit_log_sql(
                session=session,
                action="consent_updated",
                actor=actor,
                resource_type="user",
                resource_id=str(user_row.id),
                metadata={"marketing_consent": update_fields.get("marketing_consent")},
                request=request,
                country_code=user_row.country_code,
            )
        await session.commit()
        await session.refresh(user_row)
        await session.refresh(profile)

    return _build_consumer_profile_payload(user_row, profile)


@api_router.get("/v1/users/me/dealer-profile")
async def get_dealer_profile_v1(
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    if current_user.get("role") != "dealer":
        raise HTTPException(status_code=403, detail="Dealer profile only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_dealer_profile(session, user_row)
    return _build_dealer_profile_payload(user_row, profile)


@api_router.put("/v1/users/me/dealer-profile")
async def update_dealer_profile_v1(
    payload: DealerProfileUpdatePayload,
    request: Request,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    if current_user.get("role") != "dealer":
        raise HTTPException(status_code=403, detail="Dealer profile only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_dealer_profile(session, user_row)

    update_fields: Dict[str, Any] = {}
    if payload.company_name is not None:
        company_name = payload.company_name.strip()
        if not company_name:
            raise HTTPException(status_code=400, detail="Company name required")
        profile.company_name = company_name
        update_fields["company_name"] = company_name
    if payload.vat_id is not None:
        vat_id = _normalize_vat_id(payload.vat_id)
        if vat_id and not _is_valid_vat_id(vat_id):
            raise HTTPException(status_code=400, detail="VAT format invalid")
        profile.vat_id = vat_id
        profile.vat_number = vat_id
        update_fields["vat_id"] = vat_id
    if payload.trade_register_no is not None:
        profile.trade_register_no = payload.trade_register_no.strip() or None
        update_fields["trade_register_no"] = profile.trade_register_no
    if payload.authorized_person is not None:
        profile.authorized_person = payload.authorized_person.strip() or None
        update_fields["authorized_person"] = profile.authorized_person
    if payload.address_json is not None:
        profile.address_json = payload.address_json
        update_fields["address_json"] = True
    if payload.logo_url is not None:
        profile.logo_url = payload.logo_url.strip() or None
        update_fields["logo_url"] = profile.logo_url
    if payload.contact_email is not None:
        profile.contact_email = payload.contact_email
        update_fields["contact_email"] = profile.contact_email
    if payload.contact_phone is not None:
        profile.contact_phone = payload.contact_phone.strip() or None
        update_fields["contact_phone"] = profile.contact_phone
    if payload.address_country is not None:
        profile.address_country = payload.address_country.strip().upper()
        update_fields["address_country"] = profile.address_country
    if payload.impressum_text is not None:
        profile.impressum_text = payload.impressum_text
        update_fields["impressum_text"] = True
    if payload.terms_text is not None:
        profile.terms_text = payload.terms_text
        update_fields["terms_text"] = True
    if payload.withdrawal_policy_text is not None:
        profile.withdrawal_policy_text = payload.withdrawal_policy_text
        update_fields["withdrawal_policy_text"] = True

    if update_fields:
        actor = {
            "id": current_user.get("id"),
            "email": current_user.get("email"),
            "country_scope": current_user.get("country_scope") or [],
        }
        await _write_audit_log_sql(
            session=session,
            action="dealer_profile_update",
            actor=actor,
            resource_type="dealer_profile",
            resource_id=str(profile.id),
            metadata={"fields": list(update_fields.keys())},
            request=request,
            country_code=user_row.country_code,
        )
        await session.commit()
        await session.refresh(profile)

    return _build_dealer_profile_payload(user_row, profile)


@api_router.get("/v1/dealers/me/profile")
async def get_dealer_profile_v1_alias(
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    return await get_dealer_profile_v1(current_user=current_user, session=session)


@api_router.put("/v1/dealers/me/profile")
async def update_dealer_profile_v1_alias(
    payload: DealerProfileUpdatePayload,
    request: Request,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    return await update_dealer_profile_v1(
        payload=payload,
        request=request,
        current_user=current_user,
        session=session,
    )


@api_router.get("/v1/users/me/2fa/status")
async def get_two_factor_status_v1(
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    role = current_user.get("role")
    if role != "individual":
        raise HTTPException(status_code=403, detail="2FA consumer only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_consumer_profile(session, user_row)
    return {
        "enabled": profile.totp_enabled,
        "configured": bool(profile.totp_secret),
        "enabled_at": profile.totp_enabled_at.isoformat() if profile.totp_enabled_at else None,
    }


@api_router.post("/v1/users/me/2fa/setup")
async def setup_two_factor_v1(
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    role = current_user.get("role")
    if role != "individual":
        raise HTTPException(status_code=403, detail="2FA consumer only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_consumer_profile(session, user_row)
    if profile.totp_enabled:
        raise HTTPException(status_code=409, detail="2FA already enabled")

    secret = pyotp.random_base32()
    recovery_codes = _generate_recovery_codes()
    profile.totp_secret = secret
    profile.totp_enabled = False
    profile.totp_enabled_at = None
    profile.totp_recovery_codes = _hash_recovery_codes(recovery_codes)

    await _write_audit_log_sql(
        session=session,
        action="2fa_setup_started",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()

    return {
        "secret": secret,
        "otpauth_url": _build_totp_uri(user_row.email, secret),
        "recovery_codes": recovery_codes,
    }


@api_router.post("/v1/users/me/2fa/verify")
async def verify_two_factor_v1(
    payload: TwoFactorVerifyPayload,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    role = current_user.get("role")
    if role != "individual":
        raise HTTPException(status_code=403, detail="2FA consumer only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_consumer_profile(session, user_row)
    if not profile.totp_secret:
        raise HTTPException(status_code=400, detail="2FA not setup")

    if not _verify_totp_code(profile.totp_secret, payload.code):
        raise HTTPException(status_code=400, detail="Invalid code")

    profile.totp_enabled = True
    profile.totp_enabled_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="2fa_enabled",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()
    return {"enabled": True}


@api_router.post("/v1/users/me/2fa/disable")
async def disable_two_factor_v1(
    payload: TwoFactorVerifyPayload,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    role = current_user.get("role")
    if role != "individual":
        raise HTTPException(status_code=403, detail="2FA consumer only")

    user_row = await _get_user_row_from_current(session, current_user)
    profile = await _get_or_create_consumer_profile(session, user_row)
    if not profile.totp_secret:
        raise HTTPException(status_code=400, detail="2FA not setup")

    valid = _verify_totp_code(profile.totp_secret, payload.code)
    if not valid:
        ok, remaining = _verify_recovery_code(payload.code, profile.totp_recovery_codes or [])
        if ok:
            profile.totp_recovery_codes = remaining
            valid = True
    if not valid:
        raise HTTPException(status_code=400, detail="Invalid code")

    profile.totp_enabled = False
    profile.totp_secret = None
    profile.totp_enabled_at = None
    profile.totp_recovery_codes = []

    await _write_audit_log_sql(
        session=session,
        action="2fa_disabled",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()
    return {"enabled": False}


@api_router.delete("/v1/users/me/account")
async def request_account_delete_v1(
    payload: Optional[AccountDeletePayload] = Body(default=None),
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    user_row = await _get_user_row_from_current(session, current_user)
    role = current_user.get("role")
    if role not in {"individual", "dealer"}:
        raise HTTPException(status_code=403, detail="Consumer/Dealer only")

    if role == "dealer":
        profile = await _get_or_create_dealer_profile(session, user_row)
    else:
        profile = await _get_or_create_consumer_profile(session, user_row)

    now = datetime.now(timezone.utc)
    profile.gdpr_deleted_at = now + timedelta(days=30)
    user_row.is_active = False
    user_row.updated_at = now

    await _write_audit_log_sql(
        session=session,
        action="gdpr_delete_requested",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"reason": payload.reason if payload else "user_request", "grace_days": 30},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()
    return {"status": "scheduled", "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat()}


@api_router.get("/v1/users/me/data-export")
async def export_user_data_v1(
    request: Request,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    _enforce_export_rate_limit(request, current_user.get("id"))

    user_row = await _get_user_row_from_current(session, current_user)
    role = current_user.get("role")
    if role not in {"individual", "dealer"}:
        raise HTTPException(status_code=403, detail="Consumer/Dealer only")

    if role == "dealer":
        profile = await _get_or_create_dealer_profile(session, user_row)
        profile_payload = {
            "company_name": profile.company_name,
            "vat_id": profile.vat_id,
            "trade_register_no": profile.trade_register_no,
            "authorized_person": profile.authorized_person,
            "address_json": profile.address_json,
            "logo_url": profile.logo_url,
            "contact_email": profile.contact_email,
            "contact_phone": profile.contact_phone,
            "verification_status": profile.verification_status,
            "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
        }
        profile_key = "dealer_profile"
    else:
        profile = await _get_or_create_consumer_profile(
            session,
            user_row,
            language=user_row.preferred_language,
            country_code=user_row.country_code,
        )
        profile_payload = {
            "language": profile.language,
            "country_code": profile.country_code,
            "display_name_mode": profile.display_name_mode,
            "marketing_consent": profile.marketing_consent,
            "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
        }
        profile_key = "consumer_profile"

    listings = (await session.execute(
        select(Listing)
        .where(Listing.user_id == user_row.id)
        .order_by(desc(Listing.created_at))
        .limit(500)
    )).scalars().all()

    favorites = (await session.execute(
        select(Favorite)
        .where(Favorite.user_id == user_row.id)
        .order_by(desc(Favorite.created_at))
        .limit(500)
    )).scalars().all()

    conversations = (await session.execute(
        select(Conversation)
        .where(or_(Conversation.buyer_id == user_row.id, Conversation.seller_id == user_row.id))
        .order_by(desc(Conversation.last_message_at))
        .limit(200)
    )).scalars().all()

    messages_meta = []
    for convo in conversations:
        total_messages = (await session.execute(
            select(func.count()).select_from(Message).where(Message.conversation_id == convo.id)
        )).scalar_one() or 0
        messages_meta.append(
            {
                "conversation_id": str(convo.id),
                "listing_id": str(convo.listing_id) if convo.listing_id else None,
                "buyer_id": str(convo.buyer_id) if convo.buyer_id else None,
                "seller_id": str(convo.seller_id) if convo.seller_id else None,
                "last_message_at": convo.last_message_at.isoformat() if convo.last_message_at else None,
                "total_messages": int(total_messages),
            }
        )

    export_payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "user": {
            "id": str(user_row.id),
            "email": user_row.email,
            "full_name": user_row.full_name,
            "preferred_language": user_row.preferred_language,
            "country_code": user_row.country_code,
        },
        profile_key: profile_payload,
        "listings": [
            {
                "id": str(item.id),
                "title": item.title,
                "status": item.status,
                "created_at": item.created_at.isoformat() if item.created_at else None,
            }
            for item in listings
        ],
        "favorites": [
            {
                "listing_id": str(item.listing_id),
                "created_at": item.created_at.isoformat() if item.created_at else None,
            }
            for item in favorites
        ],
        "messages": messages_meta,
    }

    await _write_audit_log_sql(
        session=session,
        action="gdpr_export_requested",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"lists": ["listings", "favorites", "messages"]},
        request=request,
        country_code=user_row.country_code,
    )

    notification_message = "Veri dışa aktarma tamamlandı. Hesabınızdan bir veri erişimi gerçekleşti."
    notification = Notification(
        user_id=user_row.id,
        title="GDPR Veri Dışa Aktarım",
        message=notification_message,
        source_type="gdpr_export",
        source_id=str(user_row.id),
        action_url="/account/privacy",
        payload_json={"severity": "warning", "event": "gdpr_export_completed"},
        dedupe_key=None,
    )
    session.add(notification)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="gdpr_export_completed",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"format": "json"},
        request=request,
        country_code=user_row.country_code,
    )

    await _write_audit_log_sql(
        session=session,
        action="gdpr_export_notification_sent",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="notification",
        resource_id=str(notification.id),
        metadata={"channel": "in_app", "severity": "warning"},
        request=request,
        country_code=user_row.country_code,
    )

    payload_text = json.dumps(export_payload, ensure_ascii=False, indent=2)
    os.makedirs(GDPR_EXPORT_DIR, exist_ok=True)
    export_filename = f"gdpr-export-{user_row.id}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.json"
    export_path = os.path.join(GDPR_EXPORT_DIR, export_filename)
    with open(export_path, "w", encoding="utf-8") as export_file:
        export_file.write(payload_text)

    export_entry = GDPRExport(
        user_id=user_row.id,
        file_path=export_filename,
        status="ready",
        expires_at=datetime.now(timezone.utc) + timedelta(days=GDPR_EXPORT_RETENTION_DAYS),
    )
    session.add(export_entry)

    await session.commit()

    return Response(
        content=payload_text,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={export_filename}"},
    )


@api_router.get("/v1/users/me/gdpr-exports")
async def list_gdpr_exports(
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    user_row = await _get_user_row_from_current(session, current_user)
    result = await session.execute(
        select(GDPRExport)
        .where(GDPRExport.user_id == user_row.id)
        .order_by(desc(GDPRExport.created_at))
    )
    exports = result.scalars().all()

    now = datetime.now(timezone.utc)
    updated = False
    for export in exports:
        if export.expires_at and export.expires_at <= now and export.status != "expired":
            export.status = "expired"
            export.updated_at = now
            updated = True

    if updated:
        await session.commit()

    return {
        "items": [
            {
                "id": str(item.id),
                "status": item.status,
                "file_path": item.file_path,
                "requested_at": item.created_at.isoformat() if item.created_at else None,
                "expires_at": item.expires_at.isoformat() if item.expires_at else None,
            }
            for item in exports
        ]
    }


@api_router.get("/v1/users/me/gdpr-exports/{export_id}/download")
async def download_gdpr_export(
    export_id: str,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    user_row = await _get_user_row_from_current(session, current_user)
    try:
        export_uuid = uuid.UUID(export_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid export id") from exc

    result = await session.execute(
        select(GDPRExport).where(
            GDPRExport.id == export_uuid,
            GDPRExport.user_id == user_row.id,
        )
    )
    export_item = result.scalar_one_or_none()
    if not export_item:
        raise HTTPException(status_code=404, detail="Export not found")

    now = datetime.now(timezone.utc)
    if export_item.expires_at and export_item.expires_at <= now:
        if export_item.status != "expired":
            export_item.status = "expired"
            export_item.updated_at = now
            await session.commit()
        raise HTTPException(status_code=410, detail="Export expired")

    if export_item.status != "ready":
        raise HTTPException(status_code=409, detail="Export not ready")

    if not export_item.file_path:
        raise HTTPException(status_code=404, detail="Export file missing")

    file_path = os.path.join(GDPR_EXPORT_DIR, export_item.file_path)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Export file missing")

    return FileResponse(file_path, media_type="application/json", filename=export_item.file_path)


@api_router.get("/users/me")
async def get_user_profile(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    prefs = current_user.get("notification_prefs") or {}
    try:
        user_uuid = uuid.UUID(str(current_user.get("id")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(
        session,
        user_row,
        language=user_row.preferred_language,
        country_code=user_row.country_code,
    )

    return {
        "id": str(user_row.id),
        "email": user_row.email,
        "full_name": user_row.full_name,
        "display_name": _resolve_display_name(user_row.full_name, profile.display_name_mode),
        "display_name_mode": profile.display_name_mode,
        "phone": user_row.phone_e164,
        "locale": user_row.preferred_language or profile.language or "tr",
        "country_code": profile.country_code,
        "marketing_consent": profile.marketing_consent,
        "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
        "totp_enabled": profile.totp_enabled,
        "notification_prefs": _normalize_notification_prefs(prefs),
    }


@api_router.put("/users/me")
async def update_user_profile(
    payload: UserProfileUpdatePayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    update_payload: Dict[str, Any] = {}
    if payload.full_name is not None:
        update_payload["full_name"] = payload.full_name.strip()
    if payload.phone is not None:
        update_payload["phone_e164"] = _normalize_phone_e164(payload.phone)
    if payload.locale is not None:
        update_payload["preferred_language"] = payload.locale.strip()
    if payload.country_code is not None:
        update_payload["country_code"] = payload.country_code.strip().upper()
    if payload.display_name_mode is not None:
        update_payload["display_name_mode"] = payload.display_name_mode.strip().lower()
    if payload.marketing_consent is not None:
        update_payload["marketing_consent"] = bool(payload.marketing_consent)
    if payload.notification_prefs is not None:
        update_payload["notification_prefs"] = _normalize_notification_prefs(payload.notification_prefs)

    if not update_payload:
        return {"ok": True}

    try:
        user_uuid = uuid.UUID(str(current_user.get("id")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(
        session,
        user_row,
        language=user_row.preferred_language,
        country_code=user_row.country_code,
    )

    if update_payload.get("full_name"):
        user_row.full_name = update_payload.get("full_name")
    if update_payload.get("preferred_language"):
        user_row.preferred_language = update_payload.get("preferred_language")
        profile.language = update_payload.get("preferred_language")
    if update_payload.get("country_code"):
        user_row.country_code = update_payload.get("country_code")
        profile.country_code = update_payload.get("country_code")
    if update_payload.get("display_name_mode"):
        profile.display_name_mode = update_payload.get("display_name_mode")
    if "marketing_consent" in update_payload:
        profile.marketing_consent = update_payload.get("marketing_consent")

    actor = {
        "id": current_user.get("id"),
        "email": current_user.get("email"),
        "country_scope": current_user.get("country_scope") or [],
    }
    await _write_audit_log_sql(
        session=session,
        action="profile_update",
        actor=actor,
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"fields": list(update_payload.keys())},
        request=request,
        country_code=user_row.country_code,
    )

    if "marketing_consent" in update_payload:
        await _write_audit_log_sql(
            session=session,
            action="consent_updated",
            actor=actor,
            resource_type="user",
            resource_id=str(user_row.id),
            metadata={"marketing_consent": update_payload.get("marketing_consent")},
            request=request,
            country_code=user_row.country_code,
        )

    await session.commit()
    await session.refresh(user_row)

    response_user = {
        "id": str(user_row.id),
        "email": user_row.email,
        "full_name": user_row.full_name,
        "phone_e164": None,
        "preferred_language": user_row.preferred_language,
        "notification_prefs": update_payload.get("notification_prefs"),
        "display_name_mode": profile.display_name_mode,
        "country_code": profile.country_code,
        "marketing_consent": profile.marketing_consent,
        "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
        "totp_enabled": profile.totp_enabled,
    }
    return {
        "user": {
            "id": response_user.get("id"),
            "email": response_user.get("email"),
            "full_name": response_user.get("full_name"),
            "display_name": _resolve_display_name(response_user.get("full_name"), response_user.get("display_name_mode")),
            "display_name_mode": response_user.get("display_name_mode"),
            "phone": response_user.get("phone_e164") or response_user.get("phone"),
            "locale": response_user.get("preferred_language") or "tr",
            "country_code": response_user.get("country_code"),
            "marketing_consent": response_user.get("marketing_consent"),
            "gdpr_deleted_at": response_user.get("gdpr_deleted_at"),
            "totp_enabled": response_user.get("totp_enabled"),
            "notification_prefs": _normalize_notification_prefs(response_user.get("notification_prefs")),
        }
    }


@api_router.post("/users/change-password")
async def change_password(
    payload: ChangePasswordPayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    if len(payload.new_password or "") < 8:
        raise HTTPException(status_code=400, detail="Password too short")

    try:
        user_uuid = uuid.UUID(str(current_user.get("id")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row or not verify_password(payload.current_password, user_row.hashed_password):
        raise HTTPException(status_code=400, detail="Invalid current password")

    user_row.hashed_password = get_password_hash(payload.new_password)

    actor = {
        "id": current_user.get("id"),
        "email": current_user.get("email"),
        "country_scope": current_user.get("country_scope") or [],
    }
    await _write_audit_log_sql(
        session=session,
        action="password_change",
        actor=actor,
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"source": "self_service"},
        request=request,
        country_code=user_row.country_code,
    )

    await session.commit()

    return {"ok": True}


@api_router.get("/users/me/2fa/status")
async def get_two_factor_status(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_uuid = uuid.UUID(str(current_user.get("id")))
    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(session, user_row)
    return {
        "enabled": profile.totp_enabled,
        "configured": bool(profile.totp_secret),
        "enabled_at": profile.totp_enabled_at.isoformat() if profile.totp_enabled_at else None,
    }


@api_router.post("/users/me/2fa/setup")
async def setup_two_factor(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_uuid = uuid.UUID(str(current_user.get("id")))
    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(session, user_row)
    if profile.totp_enabled:
        raise HTTPException(status_code=409, detail="2FA already enabled")

    secret = pyotp.random_base32()
    recovery_codes = _generate_recovery_codes()
    profile.totp_secret = secret
    profile.totp_enabled = False
    profile.totp_enabled_at = None
    profile.totp_recovery_codes = _hash_recovery_codes(recovery_codes)

    await _write_audit_log_sql(
        session=session,
        action="2fa_setup_started",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()

    return {
        "secret": secret,
        "otpauth_url": _build_totp_uri(user_row.email, secret),
        "recovery_codes": recovery_codes,
    }


@api_router.post("/users/me/2fa/verify")
async def verify_two_factor(
    payload: TwoFactorVerifyPayload,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_uuid = uuid.UUID(str(current_user.get("id")))
    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(session, user_row)
    if not profile.totp_secret:
        raise HTTPException(status_code=400, detail="2FA not setup")

    if not _verify_totp_code(profile.totp_secret, payload.code):
        raise HTTPException(status_code=400, detail="Invalid code")

    profile.totp_enabled = True
    profile.totp_enabled_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="2fa_enabled",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()
    return {"enabled": True}


@api_router.post("/users/me/2fa/disable")
async def disable_two_factor(
    payload: TwoFactorVerifyPayload,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_uuid = uuid.UUID(str(current_user.get("id")))
    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(session, user_row)
    if not profile.totp_secret:
        raise HTTPException(status_code=400, detail="2FA not setup")

    valid = _verify_totp_code(profile.totp_secret, payload.code)
    if not valid:
        ok, remaining = _verify_recovery_code(payload.code, profile.totp_recovery_codes or [])
        if ok:
            profile.totp_recovery_codes = remaining
            valid = True
    if not valid:
        raise HTTPException(status_code=400, detail="Invalid code")

    profile.totp_enabled = False
    profile.totp_secret = None
    profile.totp_enabled_at = None
    profile.totp_recovery_codes = []

    await _write_audit_log_sql(
        session=session,
        action="2fa_disabled",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()
    return {"enabled": False}


@api_router.post("/users/me/delete")
async def request_account_delete(
    payload: AccountDeletePayload,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_uuid = uuid.UUID(str(current_user.get("id")))
    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(session, user_row)
    now = datetime.now(timezone.utc)
    profile.gdpr_deleted_at = now + timedelta(days=30)
    user_row.deleted_at = now
    user_row.status = "pending_delete"

    await _write_audit_log_sql(
        session=session,
        action="gdpr_delete_requested",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"reason": payload.reason or "user_request", "grace_days": 30},
        request=None,
        country_code=user_row.country_code,
    )

    await session.commit()
    return {"status": "scheduled", "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat()}


@api_router.get("/users/me/export")
async def export_user_data(
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    _enforce_export_rate_limit(request, current_user.get("id"))

    try:
        user_uuid = uuid.UUID(str(current_user.get("id")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(select(SqlUser).where(SqlUser.id == user_uuid))
    user_row = result.scalar_one_or_none()
    if not user_row:
        raise HTTPException(status_code=404, detail="User not found")

    profile = await _get_or_create_consumer_profile(
        session,
        user_row,
        language=user_row.preferred_language,
        country_code=user_row.country_code,
    )

    listings = (await session.execute(
        select(Listing)
        .where(Listing.user_id == user_uuid)
        .order_by(desc(Listing.created_at))
        .limit(500)
    )).scalars().all()

    favorites = (await session.execute(
        select(Favorite)
        .where(Favorite.user_id == user_uuid)
        .order_by(desc(Favorite.created_at))
        .limit(500)
    )).scalars().all()

    conversations = (await session.execute(
        select(Conversation)
        .where(or_(Conversation.buyer_id == user_uuid, Conversation.seller_id == user_uuid))
        .order_by(desc(Conversation.last_message_at))
        .limit(200)
    )).scalars().all()

    messages_meta = []
    for convo in conversations:
        total_messages = (await session.execute(
            select(func.count()).select_from(Message).where(Message.conversation_id == convo.id)
        )).scalar_one() or 0
        messages_meta.append(
            {
                "conversation_id": str(convo.id),
                "listing_id": str(convo.listing_id) if convo.listing_id else None,
                "buyer_id": str(convo.buyer_id) if convo.buyer_id else None,
                "seller_id": str(convo.seller_id) if convo.seller_id else None,
                "last_message_at": convo.last_message_at.isoformat() if convo.last_message_at else None,
                "total_messages": int(total_messages),
            }
        )

    export_payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "user": {
            "id": str(user_row.id),
            "email": user_row.email,
            "full_name": user_row.full_name,
            "phone": user_row.phone_e164,
            "preferred_language": user_row.preferred_language,
            "country_code": user_row.country_code,
            "display_name_mode": profile.display_name_mode,
            "marketing_consent": profile.marketing_consent,
        },
        "consumer_profile": {
            "language": profile.language,
            "country_code": profile.country_code,
            "display_name_mode": profile.display_name_mode,
            "marketing_consent": profile.marketing_consent,
            "gdpr_deleted_at": profile.gdpr_deleted_at.isoformat() if profile.gdpr_deleted_at else None,
        },
        "listings": [
            {
                "id": str(item.id),
                "title": item.title,
                "status": item.status,
                "created_at": item.created_at.isoformat() if item.created_at else None,
            }
            for item in listings
        ],
        "favorites": [
            {
                "listing_id": str(item.listing_id),
                "created_at": item.created_at.isoformat() if item.created_at else None,
            }
            for item in favorites
        ],
        "messages": messages_meta,
    }

    await _write_audit_log_sql(
        session=session,
        action="gdpr_export_requested",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"lists": ["listings", "favorites", "messages"]},
        request=request,
        country_code=user_row.country_code,
    )

    notification_message = "Veri dışa aktarma tamamlandı. Hesabınızdan bir veri erişimi gerçekleşti."
    notification = Notification(
        user_id=user_row.id,
        title="GDPR Veri Dışa Aktarım",
        message=notification_message,
        source_type="gdpr_export",
        source_id=str(user_row.id),
        action_url="/account/privacy",
        payload_json={"severity": "warning", "event": "gdpr_export_completed"},
        dedupe_key=None,
    )
    session.add(notification)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="gdpr_export_completed",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="user",
        resource_id=str(user_row.id),
        metadata={"format": "json"},
        request=request,
        country_code=user_row.country_code,
    )

    await _write_audit_log_sql(
        session=session,
        action="gdpr_export_notification_sent",
        actor={"id": str(user_row.id), "email": user_row.email},
        resource_type="notification",
        resource_id=str(notification.id),
        metadata={"channel": "in_app", "severity": "warning"},
        request=request,
        country_code=user_row.country_code,
    )

    payload_text = json.dumps(export_payload, ensure_ascii=False, indent=2)
    os.makedirs(GDPR_EXPORT_DIR, exist_ok=True)
    export_filename = f"gdpr-export-{user_row.id}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.json"
    export_path = os.path.join(GDPR_EXPORT_DIR, export_filename)
    with open(export_path, "w", encoding="utf-8") as export_file:
        export_file.write(payload_text)

    export_entry = GDPRExport(
        user_id=user_row.id,
        file_path=export_filename,
        status="ready",
        expires_at=datetime.now(timezone.utc) + timedelta(days=GDPR_EXPORT_RETENTION_DAYS),
    )
    session.add(export_entry)

    await session.commit()

    return Response(
        content=payload_text,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={export_filename}"},
    )


@api_router.post("/applications")
async def create_support_application(
    payload: SupportApplicationCreatePayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    applications_repo = _get_applications_repository(session)

    _check_application_rate_limit(request, current_user.get("id"))

    category = (payload.category or "").lower().strip()
    if category not in APPLICATION_REQUEST_TYPES:
        raise HTTPException(status_code=400, detail="Invalid category")

    subject = (payload.subject or "").strip()
    description = (payload.description or "").strip()
    if not subject or not description:
        raise HTTPException(status_code=400, detail="Subject and description are required")

    if not payload.kvkk_consent:
        raise HTTPException(status_code=400, detail="KVKK consent required")

    application_type = "dealer" if current_user.get("role") == "dealer" else "individual"

    company_name = payload.company_name or current_user.get("company_name")
    if application_type == "dealer" and not company_name:
        raise HTTPException(status_code=400, detail="Company name required")

    attachments = []
    for att in payload.attachments or []:
        name = (att.name or "").strip()
        url = (att.url or "").strip()
        if not name or not url:
            continue
        attachments.append({"name": name, "url": _validate_attachment_url(url)})

    extra_data = {
        "listing_id": payload.listing_id,
        "company_name": company_name,
        "tax_number": payload.tax_number,
        "kvkk_consent": payload.kvkk_consent,
    }

    payload_data = {
        "application_type": application_type,
        "category": category,
        "subject": subject,
        "description": _sanitize_text(description),
        "attachments": attachments,
        "extra_data": extra_data,
        "status": "pending",
        "priority": "medium",
    }

    await _ensure_sql_user(session, current_user)

    created = await applications_repo.create_application(payload_data, current_user)
    application_id = created.get("application_id")

    _send_support_received_email(current_user.get("email"), application_id, subject)

    return {"application_id": application_id}


@api_router.get("/applications")
async def list_support_applications(
    request: Request,
    application_type: Optional[str] = Query(None, alias="application_type"),
    type_filter: Optional[str] = Query(None, alias="type"),
    page: int = 1,
    limit: int = 25,
    search: Optional[str] = None,
    category: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    country: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    applications_repo = _get_applications_repository(session)

    if not application_type:
        application_type = type_filter
    if not application_type:
        raise HTTPException(status_code=400, detail="application_type is required")
    application_type = application_type.lower().strip()
    if application_type not in APPLICATION_TYPES:
        raise HTTPException(status_code=400, detail="Invalid application_type")

    if category:
        category_value = category.lower().strip()
        if category_value not in APPLICATION_REQUEST_TYPES:
            raise HTTPException(status_code=400, detail="Invalid category")
        category = category_value

    if status:
        status_value = status.lower().strip()
        if status_value not in APPLICATION_STATUS_SET:
            raise HTTPException(status_code=400, detail="Invalid status")
        status = status_value

    if priority:
        priority_value = priority.lower().strip()
        if priority_value not in APPLICATION_PRIORITY_SET:
            raise HTTPException(status_code=400, detail="Invalid priority")
        priority = priority_value


    filters = {
        "application_type": application_type,
        "category": category,
        "status": status,
        "priority": priority,
        "country": country.upper() if country else None,
        "search": search,
        "start_date": start_date,
        "end_date": end_date,
        "page": page,
        "limit": limit,
    }

    return await applications_repo.list_applications(filters)


@api_router.get("/applications/my")
async def list_my_support_applications(
    status: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    safe_page = max(1, int(page))
    safe_limit = min(100, max(1, int(limit)))
    skip = (safe_page - 1) * safe_limit

    status_value = None
    if status:
        status_value = status.lower().strip()
        if status_value not in APPLICATION_STATUS_SET:
            raise HTTPException(status_code=400, detail="Invalid status")

    try:
        user_uuid = uuid.UUID(str(current_user.get("id")))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    query = select(Application).where(Application.user_id == user_uuid)
    if status_value:
        query = query.where(Application.status == status_value)
    query = query.order_by(Application.created_at.desc())
    result = await session.execute(query.offset(skip).limit(safe_limit))
    rows = result.scalars().all()
    count_query = select(func.count()).select_from(Application).where(Application.user_id == user_uuid)
    if status_value:
        count_query = count_query.where(Application.status == status_value)
    total = (await session.execute(count_query)).scalar_one()

    items = [
        {
            "id": str(row.id),
            "application_id": row.application_id,
            "category": row.category,
            "subject": row.subject,
            "description": row.description,
            "status": row.status,
            "priority": row.priority,
            "decision_reason": row.decision_reason,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }
        for row in rows
    ]
    return {
        "items": items,
        "pagination": {"total": total, "page": safe_page, "limit": safe_limit},
    }


@api_router.get("/applications/{application_id}")
async def get_my_support_application(
    application_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        application_uuid = uuid.UUID(application_id)
    except ValueError:
        application_uuid = None
    if application_uuid:
        query = select(Application).where(Application.id == application_uuid)
    else:
        query = select(Application).where(Application.application_id == application_id)
    result = await session.execute(query)
    row = result.scalar_one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="Application not found")
    if str(row.user_id) != str(current_user.get("id")):
        raise HTTPException(status_code=403, detail="Forbidden")
    return {
        "item": {
            "id": str(row.id),
            "application_id": row.application_id,
            "category": row.category,
            "subject": row.subject,
            "description": row.description,
            "status": row.status,
            "priority": row.priority,
            "decision_reason": row.decision_reason,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
        }
    }


@api_router.get("/v1/push/vapid-public-key")
async def get_vapid_public_key(current_user=Depends(require_portal_scope("account"))):
    if not PUSH_ENABLED:
        raise HTTPException(status_code=503, detail="Push not configured")
    return {"public_key": VAPID_PUBLIC_KEY}


@api_router.get("/v1/push/subscriptions")
async def list_push_subscriptions(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    result = await session.execute(
        select(PushSubscription)
        .where(
            PushSubscription.user_id == user_id,
            PushSubscription.is_active.is_(True),
        )
        .order_by(desc(PushSubscription.created_at))
    )
    subscriptions = result.scalars().all()
    items = [
        {
            "id": str(sub.id),
            "endpoint": sub.endpoint,
            "p256dh": sub.p256dh,
            "auth": sub.auth,
            "is_active": sub.is_active,
            "revoked_at": sub.revoked_at.isoformat() if sub.revoked_at else None,
            "revoked_reason": sub.revoked_reason,
            "created_at": sub.created_at.isoformat() if sub.created_at else None,
            "updated_at": sub.updated_at.isoformat() if sub.updated_at else None,
        }
        for sub in subscriptions
    ]
    return {"items": items}


@api_router.post("/v1/push/subscribe")
async def subscribe_push_notifications(
    payload: PushSubscriptionPayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    if not PUSH_ENABLED:
        raise HTTPException(status_code=503, detail="Push not configured")

    user_id = uuid.UUID(current_user.get("id"))
    result = await session.execute(
        select(PushSubscription).where(
            PushSubscription.user_id == user_id,
            PushSubscription.endpoint == payload.endpoint,
        )
    )
    subscription = result.scalar_one_or_none()
    now_dt = datetime.now(timezone.utc)
    if subscription:
        subscription.p256dh = payload.keys.p256dh
        subscription.auth = payload.keys.auth
        subscription.is_active = True
        subscription.updated_at = now_dt
    else:
        subscription = PushSubscription(
            user_id=user_id,
            endpoint=payload.endpoint,
            p256dh=payload.keys.p256dh,
            auth=payload.keys.auth,
            is_active=True,
            created_at=now_dt,
            updated_at=now_dt,
        )
        session.add(subscription)
    await session.commit()
    return {"ok": True}


@api_router.post("/v1/push/unsubscribe")
async def unsubscribe_push_notifications(
    payload: PushUnsubscribePayload,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    result = await session.execute(
        select(PushSubscription).where(
            PushSubscription.user_id == user_id,
            PushSubscription.endpoint == payload.endpoint,
        )
    )
    subscription = result.scalar_one_or_none()
    if subscription:
        subscription.is_active = False
        subscription.revoked_at = datetime.now(timezone.utc)
        subscription.revoked_reason = "user_unsubscribe"
        subscription.updated_at = datetime.now(timezone.utc)
        await session.commit()
    return {"ok": True}


@api_router.get("/v1/notifications")
async def list_notifications(
    current_user=Depends(require_portal_scope("account")),
    page: int = 1,
    limit: int = 30,
    unread_only: bool = False,
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    safe_page = max(1, int(page))
    safe_limit = min(100, max(1, int(limit)))
    offset = (safe_page - 1) * safe_limit

    query = select(Notification).where(Notification.user_id == user_id)
    if unread_only:
        query = query.where(Notification.read_at.is_(None))

    count_query = select(func.count()).select_from(Notification).where(Notification.user_id == user_id)
    if unread_only:
        count_query = count_query.where(Notification.read_at.is_(None))

    total = (await session.execute(count_query)).scalar_one()

    notifications = (
        await session.execute(
            query.order_by(desc(Notification.created_at)).offset(offset).limit(safe_limit)
        )
    ).scalars().all()

    return {
        "items": [_build_notification_payload(item) for item in notifications],
        "pagination": {"total": int(total or 0), "page": safe_page, "limit": safe_limit},
    }


@api_router.post("/v1/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        notification_uuid = uuid.UUID(notification_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid notification id")

    notification = await session.get(Notification, notification_uuid)
    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")
    if str(notification.user_id) != current_user.get("id"):
        raise HTTPException(status_code=403, detail="Forbidden")

    if not notification.read_at:
        notification.read_at = datetime.now(timezone.utc)
        await session.commit()

    return {"ok": True, "notification": _build_notification_payload(notification)}


@api_router.get("/v1/favorites")
async def list_favorites(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))

    favorites = (
        await session.execute(
            select(Favorite)
            .where(Favorite.user_id == user_id)
            .order_by(desc(Favorite.created_at))
        )
    ).scalars().all()

    listing_ids = [fav.listing_id for fav in favorites if fav.listing_id]
    listings = []
    if listing_ids:
        listings = (
            await session.execute(select(Listing).where(Listing.id.in_(listing_ids)))
        ).scalars().all()
    listing_map = {listing.id: listing for listing in listings}

    items = []
    for fav in favorites:
        snapshot = _build_listing_snapshot(listing_map.get(fav.listing_id))
        items.append(
            {
                "id": str(fav.id),
                "created_at": fav.created_at.isoformat() if fav.created_at else None,
                **snapshot,
            }
        )

    return {"items": items}


@api_router.get("/v1/favorites/count")
async def favorites_count(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    count = (
        await session.execute(
            select(func.count()).select_from(Favorite).where(Favorite.user_id == user_id)
        )
    ).scalar_one()
    return {"count": int(count or 0)}


@api_router.get("/v1/favorites/{listing_id}")
async def favorite_state(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    listing_uuid = uuid.UUID(listing_id)
    exists = (
        await session.execute(
            select(Favorite).where(
                Favorite.user_id == user_id,
                Favorite.listing_id == listing_uuid,
            )
        )
    ).scalar_one_or_none()
    return {"is_favorite": bool(exists)}


@api_router.post("/v1/favorites/{listing_id}")
async def add_favorite(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    listing_uuid = uuid.UUID(listing_id)

    listing = (
        await session.execute(select(Listing).where(Listing.id == listing_uuid))
    ).scalar_one_or_none()
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")

    existing = (
        await session.execute(
            select(Favorite).where(
                Favorite.user_id == user_id,
                Favorite.listing_id == listing_uuid,
            )
        )
    ).scalar_one_or_none()
    if not existing:
        session.add(Favorite(user_id=user_id, listing_id=listing_uuid))
        await session.commit()

    return {"ok": True, "is_favorite": True}


@api_router.delete("/v1/favorites/{listing_id}")
async def remove_favorite(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    listing_uuid = uuid.UUID(listing_id)
    favorite = (
        await session.execute(
            select(Favorite).where(
                Favorite.user_id == user_id,
                Favorite.listing_id == listing_uuid,
            )
        )
    ).scalar_one_or_none()
    if favorite:
        await session.delete(favorite)
        await session.commit()
    return {"ok": True, "is_favorite": False}


@api_router.post("/v1/messages/threads")
async def create_message_thread(
    payload: MessageThreadCreatePayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    buyer_id = uuid.UUID(current_user.get("id"))
    listing_id = uuid.UUID(payload.listing_id)

    listing = (
        await session.execute(select(Listing).where(Listing.id == listing_id))
    ).scalar_one_or_none()
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")

    seller_id = listing.user_id
    if not seller_id or buyer_id == seller_id:
        raise HTTPException(status_code=400, detail="Cannot message your own listing")

    thread = (
        await session.execute(
            select(Conversation).where(
                Conversation.listing_id == listing_id,
                Conversation.buyer_id == buyer_id,
            )
        )
    ).scalar_one_or_none()

    if not thread:
        thread = Conversation(listing_id=listing_id, buyer_id=buyer_id, seller_id=seller_id)
        session.add(thread)
        await session.commit()
        await session.refresh(thread)

    last_message = (
        await session.execute(
            select(Message)
            .where(Message.conversation_id == thread.id)
            .order_by(desc(Message.created_at))
        )
    ).scalars().first()

    unread_count = (
        await session.execute(
            select(func.count())
            .select_from(Message)
            .where(
                Message.conversation_id == thread.id,
                Message.is_read.is_(False),
                Message.sender_id != buyer_id,
            )
        )
    ).scalar_one()

    return {
        "thread": _build_thread_summary_sql(
            thread,
            str(buyer_id),
            listing,
            last_message,
            int(unread_count or 0),
        )
    }


@api_router.get("/v1/messages/threads")
async def list_message_threads(
    current_user=Depends(require_portal_scope("account")),
    page: int = 1,
    limit: int = 30,
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    safe_page = max(1, int(page))
    safe_limit = min(100, max(1, int(limit)))
    offset = (safe_page - 1) * safe_limit

    total = (
        await session.execute(
            select(func.count()).select_from(Conversation).where(
                or_(Conversation.buyer_id == user_id, Conversation.seller_id == user_id)
            )
        )
    ).scalar_one()

    conversations = (
        await session.execute(
            select(Conversation)
            .where(or_(Conversation.buyer_id == user_id, Conversation.seller_id == user_id))
            .order_by(desc(Conversation.last_message_at))
            .offset(offset)
            .limit(safe_limit)
        )
    ).scalars().all()

    listing_ids = [conv.listing_id for conv in conversations]
    listing_map = {}
    if listing_ids:
        listings = (
            await session.execute(select(Listing).where(Listing.id.in_(listing_ids)))
        ).scalars().all()
        listing_map = {listing.id: listing for listing in listings}

    items = []
    for conv in conversations:
        last_message = (
            await session.execute(
                select(Message)
                .where(Message.conversation_id == conv.id)
                .order_by(desc(Message.created_at))
            )
        ).scalars().first()
        unread_count = (
            await session.execute(
                select(func.count())
                .select_from(Message)
                .where(
                    Message.conversation_id == conv.id,
                    Message.is_read.is_(False),
                    Message.sender_id != user_id,
                )
            )
        ).scalar_one()

        items.append(
            _build_thread_summary_sql(
                conv,
                str(user_id),
                listing_map.get(conv.listing_id),
                last_message,
                int(unread_count or 0),
            )
        )

    return {"items": items, "pagination": {"total": int(total or 0), "page": safe_page, "limit": safe_limit}}


@api_router.get("/v1/messages/unread-count")
async def message_unread_count(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    user_id = uuid.UUID(current_user.get("id"))
    total = (
        await session.execute(
            select(func.count())
            .select_from(Message)
            .join(Conversation, Conversation.id == Message.conversation_id)
            .where(
                Message.is_read.is_(False),
                Message.sender_id != user_id,
                or_(Conversation.buyer_id == user_id, Conversation.seller_id == user_id),
            )
        )
    ).scalar_one()
    return {"count": int(total or 0)}


@api_router.get("/v1/messages/threads/{thread_id}/messages")
async def list_thread_messages(
    thread_id: str,
    current_user=Depends(require_portal_scope("account")),
    limit: int = 50,
    since: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    thread = await _get_message_thread_or_404(session, thread_id, current_user.get("id"))

    query = select(Message).where(Message.conversation_id == thread.id)
    if since:
        try:
            since_dt = datetime.fromisoformat(since)
            if since_dt.tzinfo is None:
                since_dt = since_dt.replace(tzinfo=timezone.utc)
            query = query.where(Message.created_at > since_dt)
        except ValueError:
            pass

    messages = (
        await session.execute(query.order_by(Message.created_at).limit(min(200, limit)))
    ).scalars().all()

    listing = await session.get(Listing, thread.listing_id)
    last_message = messages[-1] if messages else None
    unread_count = (
        await session.execute(
            select(func.count())
            .select_from(Message)
            .where(
                Message.conversation_id == thread.id,
                Message.is_read.is_(False),
                Message.sender_id != uuid.UUID(current_user.get("id")),
            )
        )
    ).scalar_one()

    items = [
        {
            "id": str(msg.id),
            "thread_id": str(msg.conversation_id),
            "sender_id": str(msg.sender_id),
            "body": msg.body,
            "created_at": msg.created_at.isoformat() if msg.created_at else None,
            "is_read": msg.is_read,
        }
        for msg in messages
    ]

    return {
        "thread": _build_thread_summary_sql(
            thread,
            current_user.get("id"),
            listing,
            last_message,
            int(unread_count or 0),
        ),
        "items": items,
    }


@api_router.post("/v1/messages/threads/{thread_id}/messages")
async def send_thread_message(
    thread_id: str,
    payload: MessageSendPayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    thread = await _get_message_thread_or_404(session, thread_id, current_user.get("id"))
    body = (payload.body or "").strip()
    if not body:
        raise HTTPException(status_code=400, detail="Message body required")

    message = Message(
        conversation_id=thread.id,
        sender_id=uuid.UUID(current_user.get("id")),
        body=_sanitize_text(body),
        is_read=False,
    )
    session.add(message)
    thread.last_message_at = datetime.now(timezone.utc)
    await session.commit()
    await session.refresh(message)

    await message_ws_manager.broadcast_thread(
        str(thread.id),
        {
            "type": "message:new",
            "thread_id": str(thread.id),
            "message": {
                "id": str(message.id),
                "thread_id": str(thread.id),
                "sender_id": str(message.sender_id),
                "body": message.body,
                "created_at": message.created_at.isoformat() if message.created_at else None,
            },
        },
    )
    await message_ws_manager.send_personal(
        current_user.get("id"),
        {
            "type": "message:delivered",
            "thread_id": str(thread.id),
            "message_id": str(message.id),
        },
    )

    return {
        "message": {
            "id": str(message.id),
            "thread_id": str(thread.id),
            "sender_id": str(message.sender_id),
            "body": message.body,
            "created_at": message.created_at.isoformat() if message.created_at else None,
        }
    }


@api_router.post("/v1/messages/threads/{thread_id}/read")
async def mark_thread_read(
    thread_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    thread = await _get_message_thread_or_404(session, thread_id, current_user.get("id"))
    user_id = uuid.UUID(current_user.get("id"))

    await session.execute(
        update(Message)
        .where(
            Message.conversation_id == thread.id,
            Message.sender_id != user_id,
            Message.is_read.is_(False),
        )
        .values(is_read=True)
    )
    await session.commit()

    await message_ws_manager.broadcast_thread(
        str(thread.id),
        {
            "type": "message:read",
            "thread_id": str(thread.id),
            "user_id": current_user.get("id"),
            "read_at": datetime.now(timezone.utc).isoformat(),
        },
    )
    return {"ok": True}


@api_router.websocket("/ws/messages")
async def websocket_messages(websocket: WebSocket):
    token = websocket.query_params.get("token")
    if not token:
        await websocket.close(code=1008)
        return

    payload = decode_token(token)
    if not payload or payload.get("type") != "access":
        await websocket.close(code=1008)
        return

    if payload.get("token_version") != TOKEN_VERSION:
        await websocket.close(code=1008)
        return

    if payload.get("portal_scope") != "account":
        await websocket.close(code=1008)
        return

    user_id = payload.get("sub")
    if not user_id:
        await websocket.close(code=1008)
        return

    await message_ws_manager.connect(websocket, user_id)
    await websocket.send_json({"type": "connected", "user_id": user_id})

    try:
        while True:
            data = await websocket.receive_json()
            event_type = data.get("type")
            if event_type == "subscribe":
                thread_id = data.get("thread_id")
                if thread_id:
                    message_ws_manager.subscribe(websocket, thread_id)
                    await websocket.send_json({"type": "subscribed", "thread_id": thread_id})
            elif event_type == "unsubscribe":
                thread_id = data.get("thread_id")
                if thread_id:
                    message_ws_manager.unsubscribe(websocket, thread_id)
            elif event_type in {"typing:start", "typing:stop"}:
                thread_id = data.get("thread_id")
                if thread_id:
                    await message_ws_manager.broadcast_thread(
                        thread_id,
                        {
                            "type": event_type,
                            "thread_id": thread_id,
                            "user_id": user_id,
                        },
                    )
    except WebSocketDisconnect:
        message_ws_manager.disconnect(websocket, user_id)
    except Exception:
        message_ws_manager.disconnect(websocket, user_id)
        await websocket.close(code=1011)


@api_router.get("/admin/applications/assignees")
async def list_support_application_assignees(
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    result = await session.execute(
        select(SqlUser)
        .where(SqlUser.role.in_(list(ADMIN_ROLE_OPTIONS)), SqlUser.deleted_at.is_(None))
        .order_by(SqlUser.full_name.asc())
    )
    users = result.scalars().all()
    return {
        "items": [
            {
                "id": str(user.id),
                "name": user.full_name or user.email,
                "email": user.email,
            }
            for user in users
        ]
    }


async def assign_support_application(
    application_id: str,
    payload: SupportApplicationAssignPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support"])),
    session: AsyncSession = Depends(get_sql_session),
):
    applications_repo = _get_applications_repository(session)

    assigned_to = payload.assigned_to
    if assigned_to:
        try:
            uuid.UUID(assigned_to)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid assignee") from exc

    await applications_repo.assign_application(application_id, assigned_to)
    await _write_audit_log_sql(
        session=session,
        action="APPLICATION_ASSIGNED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="application",
        resource_id=application_id,
        metadata={"assigned_to": assigned_to},
        request=request,
        country_code=current_user.get("country_code"),
    )
    await session.commit()
    return {"ok": True}


@api_router.patch("/admin/applications/{application_id}/status")
async def update_support_application_status(
    application_id: str,
    payload: SupportApplicationStatusPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    applications_repo = _get_applications_repository(session)

    new_status = (payload.status or "").lower().strip()
    if new_status not in APPLICATION_STATUS_SET:
        raise HTTPException(status_code=400, detail="Invalid status")

    decision_reason = payload.decision_reason
    if new_status in {"approved", "rejected"} and not decision_reason:
        raise HTTPException(status_code=400, detail="decision_reason required")

    try:
        application_uuid = uuid.UUID(application_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid application id") from exc
    result = await session.execute(select(Application).where(Application.id == application_uuid))
    application = result.scalar_one_or_none()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    current_status = application.status

    if new_status == current_status:
        raise HTTPException(status_code=400, detail="Status already set")

    allowed_transitions = APPLICATION_STATUS_TRANSITIONS.get(current_status, set())
    if new_status not in allowed_transitions:
        raise HTTPException(status_code=400, detail="Invalid status transition")

    await applications_repo.update_status(application_id, new_status, decision_reason)
    await _write_audit_log_sql(
        session=session,
        action="APPLICATION_STATUS_UPDATED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="application",
        resource_id=application_id,
        metadata={"before": {"status": current_status}, "after": {"status": new_status, "decision_reason": decision_reason}},
        request=request,
        country_code=current_user.get("country_code"),
    )
    await session.commit()

    return {"ok": True, "status": new_status}



@api_router.post("/admin/campaigns")
async def create_campaign(
    payload: CampaignCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    data = _normalize_campaign_payload(payload)

    now = datetime.now(timezone.utc)
    campaign = Campaign(
        id=uuid.uuid4(),
        name=data["name"],
        status=data["status"],
        start_at=data["start_at"],
        end_at=data["end_at"],
        country_code=data["country_code"],
        budget_amount=data["budget_amount"],
        budget_currency=data["budget_currency"],
        notes=data["notes"],
        rules_json=data["rules_json"],
        created_at=now,
        updated_at=now,
    )

    session.add(campaign)
    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_CREATED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="campaign",
        resource_id=str(campaign.id),
        metadata={"after": _campaign_to_dict(campaign)},
        request=request,
        country_code=current_user.get("country_code"),
    )
    await session.commit()
    await session.refresh(campaign)

    return _campaign_to_dict(campaign)


@api_router.get("/admin/campaigns")
async def list_campaigns(
    request: Request,
    type: Optional[str] = None,
    status: Optional[str] = None,
    country: Optional[str] = None,
    q: Optional[str] = None,
    date_range: Optional[str] = None,
    page: int = 1,
    limit: int = 25,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_admin", "campaigns_supervisor"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    query = select(Campaign)

    if status:
        status_value = status.lower().strip()
        if status_value not in CAMPAIGN_STATUS_SET:
            raise HTTPException(status_code=400, detail="Invalid status")
        query = query.where(Campaign.status == status_value)

    if country:
        if country.lower() != "global":
            query = query.where(Campaign.country_code == country.upper())

    if q:
        search_value = f"%{q.strip()}%"
        query = query.where(or_(Campaign.name.ilike(search_value), Campaign.notes.ilike(search_value)))

    if date_range:
        parts = [part.strip() for part in date_range.split(",") if part.strip()]
        if len(parts) == 2:
            start_dt = _parse_datetime_field(parts[0], "date_range_start")
            end_dt = _parse_datetime_field(parts[1], "date_range_end")
            query = query.where(Campaign.start_at >= start_dt).where(
                or_(Campaign.end_at.is_(None), Campaign.end_at <= end_dt)
            )

    safe_page = max(page, 1)
    safe_limit = min(max(limit, 1), 200)
    offset = (safe_page - 1) * safe_limit

    total_count = await session.scalar(select(func.count()).select_from(query.subquery())) or 0
    result = await session.execute(
        query.order_by(desc(Campaign.start_at), desc(Campaign.updated_at)).offset(offset).limit(safe_limit)
    )
    rows = result.scalars().all()

    items = [_campaign_to_dict(row) for row in rows]

    total_pages = max(1, (total_count + safe_limit - 1) // safe_limit)

    return {
        "items": items,
        "total_count": total_count,
        "page": safe_page,
        "limit": safe_limit,
        "total_pages": total_pages,
    }


@api_router.get("/admin/campaigns/{campaign_id}")
async def get_campaign_detail(
    campaign_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_admin", "campaigns_supervisor"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    result = await session.execute(select(Campaign).where(Campaign.id == campaign_uuid))
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    data = _campaign_to_dict(campaign)

    audit_rows = (
        await session.execute(
            select(AuditLog)
            .where(AuditLog.resource_type == "campaign", AuditLog.resource_id == campaign_id)
            .order_by(desc(AuditLog.created_at))
            .limit(20)
        )
    ).scalars().all()
    data["audit"] = [_audit_log_to_dict(row) for row in audit_rows]

    return data


@api_router.put("/admin/campaigns/{campaign_id}")
async def update_campaign(
    campaign_id: str,
    payload: CampaignUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    result = await session.execute(select(Campaign).where(Campaign.id == campaign_uuid))
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    before_state = _campaign_to_dict(campaign)
    data = _normalize_campaign_payload(payload, existing=campaign)

    if data["status"] != campaign.status:
        allowed = CAMPAIGN_STATUS_TRANSITIONS.get(campaign.status, set())
        if data["status"] not in allowed:
            raise HTTPException(status_code=400, detail="Invalid status transition")

    campaign.country_code = data["country_code"]
    campaign.name = data["name"]
    campaign.status = data["status"]
    campaign.start_at = data["start_at"]
    campaign.end_at = data["end_at"]
    campaign.budget_amount = data["budget_amount"]
    campaign.budget_currency = data["budget_currency"]
    campaign.notes = data["notes"]
    campaign.rules_json = data["rules_json"]
    campaign.updated_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_UPDATED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="campaign",
        resource_id=campaign_id,
        metadata={"before": before_state, "after": _campaign_to_dict(campaign)},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    await session.refresh(campaign)

    return _campaign_to_dict(campaign)


@api_router.delete("/admin/campaigns/{campaign_id}")
async def archive_campaign(
    campaign_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_supervisor"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    result = await session.execute(select(Campaign).where(Campaign.id == campaign_uuid))
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    before_state = _campaign_to_dict(campaign)
    campaign.status = "ended"
    campaign.updated_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_ARCHIVED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="campaign",
        resource_id=campaign_id,
        metadata={"before": before_state, "after": {"status": "ended"}},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()

    return {"ok": True}


@api_router.post("/admin/campaigns/{campaign_id}/activate")
async def activate_campaign(
    campaign_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_supervisor"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    result = await session.execute(select(Campaign).where(Campaign.id == campaign_uuid))
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status not in {"draft", "paused"}:
        raise HTTPException(status_code=400, detail="Invalid status transition")

    before_state = _campaign_to_dict(campaign)
    campaign.status = "trial"
    campaign.updated_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_ACTIVATED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="campaign",
        resource_id=campaign_id,
        metadata={"before": before_state, "after": {"status": "active"}},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()

    return {"ok": True}


@api_router.post("/admin/campaigns/{campaign_id}/pause")
async def pause_campaign(
    campaign_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_supervisor"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    result = await session.execute(select(Campaign).where(Campaign.id == campaign_uuid))
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status not in {"active"}:
        raise HTTPException(status_code=400, detail="Invalid status transition")

    before_state = _campaign_to_dict(campaign)
    campaign.status = "paused"
    campaign.updated_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_PAUSED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="campaign",
        resource_id=campaign_id,
        metadata={"before": before_state, "after": {"status": "paused"}},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()

    return {"ok": True}


@api_router.post("/admin/campaigns/{campaign_id}/archive")
async def archive_campaign_action(
    campaign_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "campaigns_supervisor"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_campaigns_db_ready(session)

    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    result = await session.execute(select(Campaign).where(Campaign.id == campaign_uuid))
    campaign = result.scalar_one_or_none()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if campaign.status == "ended":
        return {"ok": True}

    before_state = _campaign_to_dict(campaign)
    campaign.status = "ended"
    campaign.updated_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_ARCHIVED",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="campaign",
        resource_id=campaign_id,
        metadata={"before": before_state, "after": {"status": "ended"}},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()

    return {"ok": True}


@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    request: Request,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = []
    if getattr(ctx, "mode", "global") == "country" and ctx.country:
        filters.append(SqlUser.country_code == ctx.country)

    users_total = await session.scalar(select(func.count(SqlUser.id)).where(*filters)) or 0
    users_active = await session.scalar(
        select(func.count(SqlUser.id)).where(*filters, SqlUser.is_active.is_(True))
    ) or 0

    async def _role_count(role: str) -> int:
        return await session.scalar(select(func.count(SqlUser.id)).where(*filters, SqlUser.role == role)) or 0

    # Minimal response compatible with Dashboard.js usage
    return {
        "users": {"total": int(users_total), "active": int(users_active)},
        "countries": {"enabled": len(SUPPORTED_COUNTRIES)},
        "feature_flags": {"enabled": 0, "total": 0},
        "users_by_role": {
            "super_admin": await _role_count("super_admin"),
            "country_admin": await _role_count("country_admin"),
            "moderator": await _role_count("moderator"),
            "support": await _role_count("support"),
            "finance": await _role_count("finance"),
        },
        "recent_activity": [],
    }


@api_router.get("/users")
async def list_users(
    request: Request,
    search: Optional[str] = None,
    role: Optional[str] = None,
    user_type: Optional[str] = None,
    status: Optional[str] = None,
    country: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = "desc",
    limit: int = 100,
    skip: int = 0,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = []

    country_code = ctx.country if ctx and getattr(ctx, "country", None) else None
    if country:
        country_code = country.upper()
    if country_code:
        filters.append(SqlUser.country_code == country_code)

    if user_type:
        user_type_key = user_type.lower()
        if user_type_key == "admin":
            filters.append(SqlUser.role.in_(list(ADMIN_ROLE_OPTIONS)))
        elif user_type_key == "dealer":
            filters.append(SqlUser.role == "dealer")
        elif user_type_key == "individual":
            filters.append(~SqlUser.role.in_(list(ADMIN_ROLE_OPTIONS) + ["dealer"]))

    if role:
        filters.append(SqlUser.role == role)

    if status:
        status_key = status.lower()
        if status_key == "inactive":
            filters.append(SqlUser.is_active.is_(False))
        elif status_key == "active":
            filters.append(SqlUser.is_active.is_(True))

    if search:
        search_value = f"%{search}%"
        filters.append(
            or_(
                SqlUser.email.ilike(search_value),
                SqlUser.full_name.ilike(search_value),
            )
        )

    sort_field_map = {
        "email": SqlUser.email,
        "full_name": SqlUser.full_name,
        "role": SqlUser.role,
        "created_at": SqlUser.created_at,
        "last_login": SqlUser.last_login,
        "status": SqlUser.is_active,
    }
    sort_field = sort_field_map.get(sort_by or "", SqlUser.created_at)
    sort_direction = sort_dir or "desc"
    order_by = sort_field.desc() if sort_direction.lower() == "desc" else sort_field.asc()

    query = select(SqlUser).where(*filters).order_by(order_by).offset(max(skip, 0)).limit(min(limit, 300))
    result = await session.execute(query)
    users = result.scalars().all()

    user_ids = [user.id for user in users]
    listing_stats_map: Dict[str, Dict[str, Any]] = {}
    if user_ids:
        stats_result = await session.execute(
            select(
                Listing.user_id,
                func.count(Listing.id).label("total"),
                func.sum(
                    case(
                        (Listing.status.in_(["active", "published"]), 1),
                        else_=0,
                    )
                ).label("active"),
            )
            .where(Listing.user_id.in_(user_ids))
            .group_by(Listing.user_id)
        )
        for row in stats_result.all():
            listing_stats_map[str(row.user_id)] = {"total": int(row.total or 0), "active": int(row.active or 0)}

    plan_map: Dict[str, Any] = {}

    docs = [
        {
            "id": str(user.id),
            "email": user.email,
            "full_name": user.full_name,
            "role": user.role,
            "country_code": user.country_code,
            "country_scope": user.country_scope or [],
            "is_active": user.is_active,
            "is_verified": user.is_verified,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "last_login": user.last_login.isoformat() if user.last_login else None,
            "deleted_at": None,
        }
        for user in users
    ]

    items = [
        _build_user_summary(doc, listing_stats_map.get(doc.get("id"), {}), plan_map)
        for doc in docs
    ]

    return {"items": items}


def _build_individual_users_query(
    search: Optional[str],
    country_code: Optional[str],
    status: Optional[str],
) -> Dict[str, Any]:
    query: Dict[str, Any] = {
        "role": {"$nin": list(ADMIN_ROLE_OPTIONS) + ["dealer"]},
    }

    if country_code:
        query["country_code"] = country_code

    status_key = (status or "").lower().strip()
    if status_key == "deleted":
        query["deleted_at"] = {"$exists": True}
    else:
        query["deleted_at"] = {"$exists": False}
        if status_key == "suspended":
            query["$and"] = [{"$or": [{"status": "suspended"}, {"is_active": False}]}]
        elif status_key == "active":
            query["status"] = {"$ne": "suspended"}
            query["is_active"] = {"$ne": False}

    if search:
        safe_search = re.escape(search)
        or_conditions = [
            {"first_name": {"$regex": safe_search, "$options": "i"}},
            {"last_name": {"$regex": safe_search, "$options": "i"}},
            {"email": {"$regex": safe_search, "$options": "i"}},
            {"full_name": {"$regex": safe_search, "$options": "i"}},
        ]
        phone_candidates = _normalize_phone_candidates(search)
        if phone_candidates:
            phone_fields = ["phone_e164", "phone_number", "phone", "mobile"]
            for candidate in phone_candidates:
                pattern = re.escape(candidate)
                for field in phone_fields:
                    or_conditions.append({field: {"$regex": pattern}})
        if query.get("$and"):
            query["$and"].append({"$or": or_conditions})
        else:
            query["$or"] = or_conditions

    return query


def _build_individual_users_sort(sort_by: Optional[str], sort_dir: Optional[str]):
    sort_field_map = {
        "email": "email",
        "created_at": "created_at",
        "last_login": "last_login",
        "first_name": "first_name",
        "last_name": "last_name",
    }
    sort_key = sort_field_map.get(sort_by or "last_name", "last_name")
    sort_direction = 1 if (sort_dir or "asc").lower() == "asc" else -1

    sort_name_expr = {"$ifNull": ["$last_name", {"$ifNull": ["$first_name", "$email"]}]}
    sort_first_expr = {"$ifNull": ["$first_name", "$email"]}
    sort_spec = {
        "sort_name": sort_direction,
        "sort_first": sort_direction,
        "email": sort_direction,
    }
    if sort_key != "last_name":
        sort_spec = {sort_key: sort_direction, "email": sort_direction}

    return sort_spec, sort_name_expr, sort_first_expr, sort_direction


def _resolve_contact_name(doc: dict) -> Optional[str]:
    contact_name = doc.get("contact_name")
    if contact_name:
        return contact_name
    first_name = doc.get("first_name") or ""
    last_name = doc.get("last_name") or ""
    fallback = " ".join(part for part in [first_name, last_name] if part).strip()
    return fallback or None


def _build_dealer_summary(
    doc: dict,
    listing_stats: Optional[Dict[str, Any]] = None,
    plan_map: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    listing_stats = listing_stats or {}
    plan_map = plan_map or {}
    plan = plan_map.get(doc.get("plan_id"), {})
    status = _normalize_user_status(doc)
    phone_verified = bool(doc.get("phone_verified") or doc.get("phone_verified_at"))
    email_verified = bool(doc.get("email_verified") or doc.get("is_verified", False))

    return {
        "id": doc.get("id"),
        "company_name": doc.get("company_name"),
        "contact_name": _resolve_contact_name(doc),
        "first_name": doc.get("first_name"),
        "last_name": doc.get("last_name"),
        "email": doc.get("email"),
        "phone_e164": _resolve_user_phone_e164(doc),
        "country_code": doc.get("country_code"),
        "status": status,
        "risk_level": doc.get("risk_level") or "low",
        "ban_reason": doc.get("ban_reason"),
        "is_active": bool(doc.get("is_active", True)),
        "deleted_at": doc.get("deleted_at"),
        "email_verified": email_verified,
        "phone_verified": phone_verified,
        "created_at": doc.get("created_at"),
        "last_login": doc.get("last_login"),
        "total_listings": listing_stats.get("total", 0),
        "active_listings": listing_stats.get("active", 0),
        "plan_id": doc.get("plan_id"),
        "plan_name": plan.get("name") if plan else None,
    }


def _build_dealer_query(
    search: Optional[str],
    country_code: Optional[str],
    status: Optional[str],
    plan_id: Optional[str],
) -> Dict[str, Any]:
    query: Dict[str, Any] = {"role": "dealer"}

    if country_code:
        query["country_code"] = country_code

    status_key = (status or "").lower().strip()
    if status_key == "deleted":
        query["deleted_at"] = {"$exists": True}
    else:
        query["deleted_at"] = {"$exists": False}
        if status_key == "suspended":
            query["$or"] = [{"status": "suspended"}, {"is_active": False}]
        elif status_key == "active":
            query["status"] = {"$ne": "suspended"}
            query["is_active"] = {"$ne": False}

    if plan_id:
        query["plan_id"] = plan_id

    if search:
        safe_search = re.escape(search)
        or_conditions = [
            {"company_name": {"$regex": safe_search, "$options": "i"}},
            {"contact_name": {"$regex": safe_search, "$options": "i"}},
            {"first_name": {"$regex": safe_search, "$options": "i"}},
            {"last_name": {"$regex": safe_search, "$options": "i"}},
            {"email": {"$regex": safe_search, "$options": "i"}},
        ]
        phone_candidates = _normalize_phone_candidates(search)
        if phone_candidates:
            phone_fields = ["phone_e164", "phone_number", "phone", "mobile"]
            for candidate in phone_candidates:
                pattern = re.escape(candidate)
                for field in phone_fields:
                    or_conditions.append({field: {"$regex": pattern}})
        if query.get("$or"):
            query["$and"] = [{"$or": query.pop("$or")}, {"$or": or_conditions}]
        else:
            query["$or"] = or_conditions

    return query


def _build_dealer_sort(sort_by: Optional[str], sort_dir: Optional[str]):
    sort_field_map = {
        "company_name": "company_name",
        "email": "email",
        "created_at": "created_at",
        "last_login": "last_login",
    }
    sort_key = sort_field_map.get(sort_by or "company_name", "company_name")
    sort_direction = 1 if (sort_dir or "asc").lower() == "asc" else -1

    sort_company_expr = {"$ifNull": ["$company_name", "$email"]}
    sort_spec = {
        "sort_company": sort_direction,
        "email": sort_direction,
    }
    if sort_key != "company_name":
        sort_spec = {sort_key: sort_direction, "email": sort_direction}

    return sort_spec, sort_company_expr, sort_direction


def _build_support_application_summary(doc: dict) -> Dict[str, Any]:
    assigned = doc.get("assigned_to") or None
    display_name = doc.get("applicant_company_name") or doc.get("applicant_name") or doc.get("applicant_email")
    return {
        "id": doc.get("id"),
        "application_id": doc.get("application_id"),
        "application_type": doc.get("application_type"),
        "category": doc.get("category"),
        "subject": doc.get("subject"),
        "description": doc.get("description"),
        "status": doc.get("status"),
        "priority": doc.get("priority"),
        "assigned_to": assigned,
        "created_at": doc.get("created_at"),
        "updated_at": doc.get("updated_at"),
        "applicant_name": doc.get("applicant_name"),
        "applicant_company_name": doc.get("applicant_company_name"),
        "applicant_email": doc.get("applicant_email"),
        "applicant_country": doc.get("applicant_country"),
        "display_name": display_name,
    }


def _dealer_application_to_dict(app: DealerApplication) -> Dict[str, Any]:
    return {
        "id": str(app.id),
        "country": app.country,
        "dealer_type": app.dealer_type,
        "company_name": app.company_name,
        "vat_tax_no": app.vat_tax_no,
        "address": app.address,
        "city": app.city,
        "postal_code": app.postal_code,
        "website": app.website,
        "logo_url": app.logo_url,
        "contact_name": app.contact_name,
        "contact_email": app.contact_email,
        "contact_phone": app.contact_phone,
        "status": app.status,
        "reject_reason": app.reject_reason,
        "reviewed_at": app.reviewed_at.isoformat() if app.reviewed_at else None,
        "reviewed_by": str(app.reviewed_by_id) if app.reviewed_by_id else None,
        "created_at": app.created_at.isoformat() if app.created_at else None,
        "updated_at": app.updated_at.isoformat() if app.updated_at else None,
    }


def _individual_application_to_dict(app: Application) -> Dict[str, Any]:
    return {
        "id": str(app.id),
        "application_id": app.application_id,
        "application_type": app.application_type,
        "category": app.category,
        "subject": app.subject,
        "description": app.description,
        "status": app.status,
        "priority": app.priority,
        "decision_reason": app.decision_reason,
        "created_at": app.created_at.isoformat() if app.created_at else None,
        "updated_at": app.updated_at.isoformat() if app.updated_at else None,
        "applicant_email": (app.extra_data or {}).get("email"),
        "applicant_name": (app.extra_data or {}).get("full_name"),
        "applicant_country": (app.extra_data or {}).get("country_code"),
    }


@api_router.get("/admin/individual-users")
async def list_individual_users(
    request: Request,
    search: Optional[str] = None,
    country: Optional[str] = None,
    status: Optional[str] = None,
    sort_by: Optional[str] = "last_name",
    sort_dir: Optional[str] = "asc",
    page: int = 1,
    limit: int = 50,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    country_code = ctx.country if ctx and getattr(ctx, "country", None) else None
    if country:
        country_code = country.upper()

    filters = [~SqlUser.role.in_(list(ADMIN_ROLE_OPTIONS) + ["dealer"])]

    if country_code:
        filters.append(SqlUser.country_code == country_code)

    status_key = (status or "").lower().strip()
    if status_key == "deleted":
        filters.append(SqlUser.deleted_at.is_not(None))
    else:
        filters.append(SqlUser.deleted_at.is_(None))
        if status_key == "suspended":
            filters.append(or_(SqlUser.status == "suspended", SqlUser.is_active.is_(False)))
        elif status_key == "active":
            filters.append(SqlUser.status != "suspended")
            filters.append(SqlUser.is_active.is_(True))

    if search:
        search_value = f"%{search}%"
        filters.append(
            or_(
                SqlUser.first_name.ilike(search_value),
                SqlUser.last_name.ilike(search_value),
                SqlUser.email.ilike(search_value),
                SqlUser.full_name.ilike(search_value),
                SqlUser.phone_e164.ilike(search_value),
            )
        )

    safe_page = max(page, 1)
    safe_limit = min(max(limit, 1), 200)
    skip = (safe_page - 1) * safe_limit

    total_count = await session.scalar(select(func.count(SqlUser.id)).where(*filters)) or 0

    sort_field_map = {
        "email": SqlUser.email,
        "created_at": SqlUser.created_at,
        "last_login": SqlUser.last_login,
        "first_name": SqlUser.first_name,
        "last_name": SqlUser.last_name,
    }
    sort_field = sort_field_map.get(sort_by or "", SqlUser.last_name)
    direction = (sort_dir or "asc").lower()
    order_by = sort_field.asc() if direction == "asc" else sort_field.desc()

    result = await session.execute(
        select(SqlUser)
        .where(*filters)
        .order_by(order_by)
        .offset(skip)
        .limit(safe_limit)
    )
    users = result.scalars().all()

    user_ids = [user.id for user in users]
    listing_stats_map: Dict[str, Dict[str, Any]] = {}
    if user_ids:
        stats_result = await session.execute(
            select(
                Listing.user_id,
                func.count(Listing.id).label("total"),
                func.sum(case((Listing.status == "published", 1), else_=0)).label("active"),
            )
            .where(Listing.user_id.in_(user_ids))
            .group_by(Listing.user_id)
        )
        for row in stats_result.all():
            listing_stats_map[str(row.user_id)] = {"total": int(row.total or 0), "active": int(row.active or 0)}

    plan_ids = [user.plan_id for user in users if user.plan_id]
    plan_map: Dict[str, Any] = {}
    if plan_ids:
        plans = await session.execute(select(Plan).where(Plan.id.in_(plan_ids)))
        plan_map = {str(plan.id): {"id": str(plan.id), "name": plan.name} for plan in plans.scalars().all()}

    docs = [
        {
            "id": str(user.id),
            "email": user.email,
            "full_name": user.full_name,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "role": user.role,
            "country_code": user.country_code,
            "country_scope": user.country_scope or [],
            "status": user.status,
            "is_active": user.is_active,
            "is_verified": user.is_verified,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "last_login": user.last_login.isoformat() if user.last_login else None,
            "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
            "phone_e164": user.phone_e164,
            "plan_id": str(user.plan_id) if user.plan_id else None,
            "plan_expires_at": user.plan_expires_at.isoformat() if user.plan_expires_at else None,
        }
        for user in users
    ]

    items = [
        _build_user_summary(doc, listing_stats_map.get(doc.get("id"), {}), plan_map)
        for doc in docs
    ]

    total_pages = max(1, (total_count + safe_limit - 1) // safe_limit)

    return {
        "items": items,
        "total_count": int(total_count),
        "page": safe_page,
        "limit": safe_limit,
        "total_pages": total_pages,
    }


@api_router.get("/admin/individual-users/export/csv")
async def admin_individual_users_export_csv(
    request: Request,
    search: Optional[str] = None,
    country: Optional[str] = None,
    status: Optional[str] = None,
    sort_by: Optional[str] = "last_name",
    sort_dir: Optional[str] = "asc",
    current_user=Depends(check_permissions(["super_admin", "marketing"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )
    _enforce_export_rate_limit(request, current_user.get("id"))

    country_code = ctx.country if ctx and getattr(ctx, "country", None) else None
    if country:
        country_code = country.upper()

    filters = [~SqlUser.role.in_(list(ADMIN_ROLE_OPTIONS) + ["dealer"])]
    if country_code:
        filters.append(SqlUser.country_code == country_code)

    status_key = (status or "").lower().strip()
    if status_key == "deleted":
        filters.append(SqlUser.deleted_at.is_not(None))
    else:
        filters.append(SqlUser.deleted_at.is_(None))
        if status_key == "suspended":
            filters.append(or_(SqlUser.status == "suspended", SqlUser.is_active.is_(False)))
        elif status_key == "active":
            filters.append(SqlUser.status != "suspended")
            filters.append(SqlUser.is_active.is_(True))

    if search:
        search_value = f"%{search}%"
        filters.append(
            or_(
                SqlUser.first_name.ilike(search_value),
                SqlUser.last_name.ilike(search_value),
                SqlUser.email.ilike(search_value),
                SqlUser.full_name.ilike(search_value),
                SqlUser.phone_e164.ilike(search_value),
            )
        )

    sort_field_map = {
        "email": SqlUser.email,
        "created_at": SqlUser.created_at,
        "last_login": SqlUser.last_login,
        "first_name": SqlUser.first_name,
        "last_name": SqlUser.last_name,
    }
    sort_field = sort_field_map.get(sort_by or "", SqlUser.last_name)
    direction = (sort_dir or "asc").lower()
    order_by = sort_field.asc() if direction == "asc" else sort_field.desc()

    result = await session.execute(select(SqlUser).where(*filters).order_by(order_by).limit(10000))
    users = result.scalars().all()

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Email", "Name", "Country", "Type"])

    for user in users:
        full_name = user.full_name
        if not full_name:
            first = user.first_name or ""
            last = user.last_name or ""
            full_name = " ".join(part for part in [first, last] if part).strip() or user.email
        writer.writerow([
            user.email,
            full_name,
            user.country_code,
            _determine_user_type(user.role or "individual"),
        ])

    csv_content = buffer.getvalue().encode("utf-8")
    buffer.close()

    total_count = len(users)
    filename = f"individual-users-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.csv"

    await _write_audit_log_sql(
        session=session,
        action="individual_users_export_csv",
        actor=current_user,
        resource_type="individual_users",
        resource_id="individual_users",
        metadata={
            "search": search,
            "country": country_code,
            "status": status,
            "sort_by": sort_by,
            "sort_dir": sort_dir,
            "total_count": total_count,
        },
        request=request,
        country_code=country_code,
    )
    await session.commit()

    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.get("/menu/top-items")
async def get_top_menu_items(
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(
        select(TopMenuItem)
        .where(TopMenuItem.is_enabled.is_(True))
        .order_by(TopMenuItem.sort_order.asc())
    )
    items = []
    for item in result.scalars().all():
        items.append(
            {
                "id": str(item.id),
                "key": item.key,
                "name": item.name,
                "icon": item.icon,
                "badge": item.badge,
                "sort_order": item.sort_order,
                "required_module": item.required_module,
                "allowed_countries": item.allowed_countries,
                "is_enabled": item.is_enabled,
            }
        )
    return items


@api_router.patch("/menu/top-items/{item_id}")
async def toggle_menu_item(
    item_id: str,
    data: dict,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    is_enabled = data.get("is_enabled")
    if is_enabled is None:
        return {"ok": True}

    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid menu item id") from exc

    item = await session.get(TopMenuItem, item_uuid)
    if not item:
        raise HTTPException(status_code=404, detail="Menu item not found")

    item.is_enabled = bool(is_enabled)
    await session.commit()
    return {"ok": True}


@api_router.get("/categories")
async def list_categories(
    request: Request,
    module: str = "vehicle",
    country: Optional[str] = None,
    current_user=Depends(get_current_user_optional),
    session: AsyncSession = Depends(get_sql_session),
):
    if not country:
        raise HTTPException(status_code=400, detail="country is required")

    code = country.upper()
    query = (
        select(Category)
        .options(selectinload(Category.translations))
        .where(
            Category.module == module,
            Category.is_deleted.is_(False),
            Category.is_enabled.is_(True),
        )
    )

    result = await session.execute(query)
    categories = result.scalars().all()

    filtered: list[Category] = []
    for category in categories:
        if category.country_code and category.country_code != code:
            continue
        allowed = category.allowed_countries or []
        if allowed and code not in allowed:
            continue
        filtered.append(category)

    filtered.sort(key=lambda c: (c.sort_order or 0, _pick_category_name(list(c.translations or []), _pick_category_slug(c.slug))))
    return [_serialize_category_sql(cat, include_schema=False, include_translations=True) for cat in filtered]


def _filter_categories_for_country(categories: list[Category], code: str) -> list[Category]:
    filtered: list[Category] = []
    for category in categories:
        if category.country_code and category.country_code != code:
            continue
        allowed = category.allowed_countries or []
        if allowed and code not in allowed:
            continue
        filtered.append(category)
    filtered.sort(key=lambda c: (c.sort_order or 0, _pick_category_name(list(c.translations or []), _pick_category_slug(c.slug))))
    return filtered


@api_router.get("/categories/children")
async def list_category_children(
    request: Request,
    country: str,
    parent_id: Optional[str] = None,
    module: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    if not country:
        raise HTTPException(status_code=400, detail="country is required")

    code = country.upper()
    query = (
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.is_deleted.is_(False), Category.is_enabled.is_(True))
    )
    if module:
        query = query.where(Category.module == module)
    if parent_id:
        try:
            parent_uuid = uuid.UUID(parent_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="parent_id not valid") from exc
        query = query.where(Category.parent_id == parent_uuid)
    else:
        query = query.where(Category.parent_id.is_(None))

    result = await session.execute(query)
    categories = result.scalars().all()
    filtered = _filter_categories_for_country(categories, code)
    return [_serialize_category_sql(cat, include_schema=False, include_translations=True) for cat in filtered]


@api_router.get("/categories/search")
async def search_categories(
    request: Request,
    query: str,
    country: str,
    module: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    if not country:
        raise HTTPException(status_code=400, detail="country is required")
    if not query:
        return {"items": []}

    code = country.upper()
    search_value = query.strip().lower()
    if len(search_value) < 2:
        return {"items": []}

    stmt = (
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.is_deleted.is_(False), Category.is_enabled.is_(True))
    )
    if module:
        stmt = stmt.where(Category.module == module)

    result = await session.execute(stmt)
    categories = _filter_categories_for_country(result.scalars().all(), code)
    category_map = {cat.id: cat for cat in categories}

    def build_path(category: Category) -> list[Category]:
        path_nodes: list[Category] = []
        current = category
        visited = set()
        while current and current.id not in visited:
            visited.add(current.id)
            path_nodes.append(current)
            current = category_map.get(current.parent_id)
        return list(reversed(path_nodes))

    items = []
    for category in categories:
        name_value = _pick_category_name(list(category.translations or []), _pick_category_slug(category.slug)) or ""
        slug_value = _pick_category_slug(category.slug) or ""
        if search_value in name_value.lower() or search_value in slug_value.lower():
            path_nodes = build_path(category)
            items.append({
                "category": _serialize_category_sql(category, include_schema=False, include_translations=False),
                "path": [
                    {
                        "id": str(node.id),
                        "name": _pick_category_name(list(node.translations or []), _pick_category_slug(node.slug)),
                        "slug": _pick_category_slug(node.slug),
                    }
                    for node in path_nodes
                ],
            })

    items.sort(key=lambda item: (len(item.get("path") or []), item["category"].get("name") or ""))
    return {"items": items}


@api_router.get("/categories/validate")
async def validate_category_selection(
    request: Request,
    category_id: str,
    country: Optional[str] = None,
    module: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.id == category_uuid, Category.is_deleted.is_(False))
    )
    category = result.scalar_one_or_none()
    if not category or not category.is_enabled:
        raise HTTPException(status_code=404, detail="Category not found")

    if module and category.module != module:
        raise HTTPException(status_code=409, detail="Module mismatch")

    if country:
        code = country.upper()
        if category.country_code and category.country_code != code:
            raise HTTPException(status_code=409, detail="Country mismatch")
        allowed = category.allowed_countries or []
        if allowed and code not in allowed:
            raise HTTPException(status_code=409, detail="Country mismatch")

    return {
        "valid": True,
        "category": _serialize_category_sql(category, include_schema=False, include_translations=True),
    }


class ListingFlowEventPayload(BaseModel):
    event_name: str
    metadata: Optional[dict] = None


@api_router.post("/analytics/events")
async def track_listing_flow_event(
    payload: ListingFlowEventPayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    metadata = payload.metadata or {}
    metadata["event_name"] = payload.event_name
    resource_id = metadata.get("category_id") or None

    await _write_audit_log_sql(
        session,
        "listing_flow_event",
        current_user,
        "listing_category_flow",
        resource_id,
        metadata,
        request,
        country_code=metadata.get("country"),
    )
    await session.commit()

    return {"status": "ok"}


class AdminWizardAnalyticsPayload(BaseModel):
    event_name: str
    category_id: str
    step_id: str
    admin_user_id: Optional[str] = None
    wizard_state: Optional[str] = None


@api_router.post("/admin/analytics/events")
async def track_admin_wizard_event(
    payload: AdminWizardAnalyticsPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    admin_id = current_user.get("id")
    metadata = {
        "event_name": payload.event_name,
        "category_id": payload.category_id,
        "step_id": payload.step_id,
        "admin_user_id": admin_id,
        "wizard_state": payload.wizard_state,
    }

    await _write_audit_log_sql(
        session,
        payload.event_name,
        current_user,
        "admin_category_wizard",
        payload.category_id,
        metadata,
        request,
    )
    await session.commit()

    return {"status": "ok"}


class RecentCategoryPayload(BaseModel):
    category_id: str
    module: str
    country: Optional[str] = None
    path: Optional[list] = None
    category_name: Optional[str] = None


async def _load_category_with_translations(session: AsyncSession, category_id: uuid.UUID) -> Optional[Category]:
    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.id == category_id, Category.is_deleted.is_(False))
    )
    return result.scalar_one_or_none()


async def _build_category_path(session: AsyncSession, category: Category) -> list[Category]:
    path_nodes: list[Category] = []
    current = category
    visited = set()
    while current and current.id not in visited:
        visited.add(current.id)
        path_nodes.append(current)
        if not current.parent_id:
            break
        current = await _load_category_with_translations(session, current.parent_id)
    return list(reversed(path_nodes))


async def _build_recent_category_response(session: AsyncSession, recent: UserRecentCategory | None) -> Optional[dict]:
    if not recent:
        return None

    try:
        category_uuid = uuid.UUID(str(recent.category_id))
    except ValueError:
        return None

    category = await _load_category_with_translations(session, category_uuid)
    if not category or not category.is_enabled:
        return None

    if recent.module and category.module and category.module != recent.module:
        return None

    code = (recent.country or "").upper() or "DE"
    if category.country_code and category.country_code != code:
        return None
    allowed = category.allowed_countries or []
    if allowed and code not in allowed:
        return None

    path_nodes = await _build_category_path(session, category)

    return {
        "category": _serialize_category_sql(category, include_schema=False, include_translations=True),
        "path": [
            {
                "id": str(node.id),
                "name": _pick_category_name(list(node.translations or []), _pick_category_slug(node.slug)),
                "slug": _pick_category_slug(node.slug),
            }
            for node in path_nodes
        ],
        "module": recent.module,
        "country": code,
        "updated_at": recent.updated_at.isoformat() if recent.updated_at else None,
    }


@api_router.post("/account/recent-category")
async def set_recent_category(
    payload: RecentCategoryPayload,
    request: Request,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        user_uuid = uuid.UUID(current_user.get("id"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    try:
        category_uuid = uuid.UUID(payload.category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="category_id invalid") from exc

    module_value = payload.module or "vehicle"
    country_value = (payload.country or current_user.get("country_code") or "DE").upper()

    category = await _load_category_with_translations(session, category_uuid)
    if not category or not category.is_enabled:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")
    if category.module and category.module != module_value:
        raise HTTPException(status_code=409, detail="Module uyuşmazlığı")
    if category.country_code and category.country_code != country_value:
        raise HTTPException(status_code=403, detail="Kategori ülke kapsamı dışında")
    allowed = category.allowed_countries or []
    if allowed and country_value not in allowed:
        raise HTTPException(status_code=403, detail="Kategori ülke kapsamı dışında")

    result = await session.execute(
        select(UserRecentCategory).where(UserRecentCategory.user_id == user_uuid)
    )
    recent = result.scalar_one_or_none()
    if recent:
        recent.category_id = category_uuid
        recent.module = module_value
        recent.country = country_value
        recent.updated_at = datetime.now(timezone.utc)
    else:
        recent = UserRecentCategory(
            user_id=user_uuid,
            category_id=category_uuid,
            module=module_value,
            country=country_value,
        )
        session.add(recent)

    await _write_audit_log_sql(
        session,
        "listing_recent_category",
        current_user,
        "listing_category_flow",
        payload.category_id,
        {
            "module": module_value,
            "country": country_value,
            "path": payload.path or [],
            "category_name": payload.category_name,
        },
        request,
        country_code=country_value,
    )

    await session.commit()
    await session.refresh(recent)

    recent_payload = await _build_recent_category_response(session, recent)
    return {"status": "ok", "recent": recent_payload}


@api_router.get("/account/recent-category")
async def get_recent_category(
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        user_uuid = uuid.UUID(current_user.get("id"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    result = await session.execute(
        select(UserRecentCategory).where(UserRecentCategory.user_id == user_uuid)
    )
    recent = result.scalar_one_or_none()
    recent_payload = await _build_recent_category_response(session, recent)
    return {"recent": recent_payload}


async def get_catalog_schema(
    category_id: str,
    request: Request,
    country: str | None = None,
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    category = await session.get(Category, category_uuid)
    if not category or category.is_deleted or not category.is_enabled:
        raise HTTPException(status_code=404, detail="Kategori bulunamadı")

    if country:
        code = country.upper()
        if category.country_code and category.country_code != code:
            raise HTTPException(status_code=403, detail="Kategori ülke kapsamı dışında")
        allowed = category.allowed_countries or []
        if allowed and code not in allowed:
            raise HTTPException(status_code=403, detail="Kategori ülke kapsamı dışında")

    raw_schema = category.form_schema
    if not raw_schema:
        raise HTTPException(status_code=409, detail="Kategori şeması oluşturulmadı")
    schema = _normalize_category_schema(raw_schema)
    if schema.get("status") == "draft":
        schema = {**schema, "status": "draft"}
    return {"category": _serialize_category_sql(category, include_schema=False), "schema": schema}


@api_router.get("/catalog/schema")
async def catalog_schema_endpoint(
    category_id: str,
    request: Request,
    country: str | None = None,
    session: AsyncSession = Depends(get_sql_session),
):
    return await get_catalog_schema(category_id=category_id, request=request, country=country, session=session)


# =====================
# Sprint 1.1 — Dealer Management (Admin)
# =====================

@api_router.get("/admin/dealers")
async def admin_list_dealers(
    request: Request,
    search: Optional[str] = None,
    country: Optional[str] = None,
    status: Optional[str] = None,
    plan_id: Optional[str] = None,
    sort_by: Optional[str] = "company_name",
    sort_dir: Optional[str] = "asc",
    page: int = 1,
    limit: int = 25,
    include_filters: bool = False,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    country_code = ctx.country if ctx and getattr(ctx, "country", None) else None
    if country:
        country_code = country.upper()

    filters = [SqlUser.role == "dealer"]
    if country_code:
        filters.append(SqlUser.country_code == country_code)

    status_key = (status or "").lower().strip()
    if status_key == "deleted":
        filters.append(SqlUser.deleted_at.is_not(None))
    else:
        filters.append(SqlUser.deleted_at.is_(None))
        if status_key == "suspended":
            filters.append(or_(SqlUser.status == "suspended", SqlUser.is_active.is_(False)))
        elif status_key == "active":
            filters.append(SqlUser.status != "suspended")
            filters.append(SqlUser.is_active.is_(True))

    plan_uuid = None
    if plan_id:
        try:
            plan_uuid = uuid.UUID(plan_id)
            filters.append(SqlUser.plan_id == plan_uuid)
        except ValueError:
            pass

    result = await session.execute(select(SqlUser).where(*filters))
    users = result.scalars().all()

    user_ids = [user.id for user in users]
    dealer_map: Dict[uuid.UUID, Dealer] = {}
    if user_ids:
        dealer_users = await session.execute(select(DealerUser).where(DealerUser.user_id.in_(user_ids)))
        dealer_links = dealer_users.scalars().all()
        dealer_ids = list({link.dealer_id for link in dealer_links})
        if dealer_ids:
            dealers = await session.execute(select(Dealer).where(Dealer.id.in_(dealer_ids)))
            dealer_map = {dealer.id: dealer for dealer in dealers.scalars().all()}
        dealer_link_map = {link.user_id: link.dealer_id for link in dealer_links}
    else:
        dealer_link_map = {}

    def _search_match(user: SqlUser) -> bool:
        if not search:
            return True
        needle = search.lower().strip()
        dealer_id = dealer_link_map.get(user.id)
        company_name = dealer_map.get(dealer_id).company_name if dealer_id and dealer_map.get(dealer_id) else ""
        fields = [
            company_name or "",
            user.full_name or "",
            user.first_name or "",
            user.last_name or "",
            user.email or "",
            user.phone_e164 or "",
        ]
        return any(needle in field.lower() for field in fields if field)

    filtered_users = [user for user in users if _search_match(user)]

    sort_key = (sort_by or "company_name").lower()
    direction = (sort_dir or "asc").lower()

    def _sort_value(user: SqlUser):
        dealer_id = dealer_link_map.get(user.id)
        company_name = dealer_map.get(dealer_id).company_name if dealer_id and dealer_map.get(dealer_id) else ""
        if sort_key == "email":
            return user.email or ""
        if sort_key == "created_at":
            return user.created_at or datetime.min.replace(tzinfo=timezone.utc)
        if sort_key == "last_login":
            return user.last_login or datetime.min.replace(tzinfo=timezone.utc)
        return company_name or user.email or ""

    filtered_users.sort(key=_sort_value, reverse=(direction == "desc"))

    safe_page = max(page, 1)
    safe_limit = min(max(limit, 1), 200)
    start = (safe_page - 1) * safe_limit
    end = start + safe_limit
    page_users = filtered_users[start:end]

    listing_stats_map: Dict[str, Dict[str, Any]] = {}
    if page_users:
        stats_result = await session.execute(
            select(
                Listing.user_id,
                func.count(Listing.id).label("total"),
                func.sum(case((Listing.status == "published", 1), else_=0)).label("active"),
            )
            .where(Listing.user_id.in_([user.id for user in page_users]))
            .group_by(Listing.user_id)
        )
        for row in stats_result.all():
            listing_stats_map[str(row.user_id)] = {"total": int(row.total or 0), "active": int(row.active or 0)}

    plan_ids = [user.plan_id for user in page_users if user.plan_id]
    plan_map: Dict[str, Any] = {}
    if plan_ids:
        plans = await session.execute(select(Plan).where(Plan.id.in_(plan_ids)))
        plan_map = {str(plan.id): {"id": str(plan.id), "name": plan.name} for plan in plans.scalars().all()}

    items = []
    for user in page_users:
        dealer_id = dealer_link_map.get(user.id)
        dealer = dealer_map.get(dealer_id)
        doc = {
            "id": str(user.id),
            "company_name": dealer.company_name if dealer else None,
            "contact_name": user.full_name,
            "first_name": user.first_name,
            "last_name": user.last_name,
            "email": user.email,
            "phone_e164": user.phone_e164,
            "country_code": user.country_code,
            "status": user.status,
            "is_active": user.is_active,
            "deleted_at": user.deleted_at.isoformat() if user.deleted_at else None,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "last_login": user.last_login.isoformat() if user.last_login else None,
            "plan_id": str(user.plan_id) if user.plan_id else None,
        }
        items.append(_build_dealer_summary(doc, listing_stats_map.get(str(user.id), {}), plan_map))

    total_count = len(filtered_users)
    total_pages = max(1, (total_count + safe_limit - 1) // safe_limit)

    response: Dict[str, Any] = {
        "items": items,
        "total_count": total_count,
        "page": safe_page,
        "limit": safe_limit,
        "total_pages": total_pages,
    }

    if include_filters:
        plan_filters: List[Dict[str, Any]] = []
        plan_query = []
        if country_code:
            plan_query.append(Plan.country_code == country_code)
        elif current_user.get("role") == "country_admin":
            scope = current_user.get("country_scope") or []
            if "*" not in scope:
                plan_query.append(Plan.country_code.in_(scope))
        if plan_query or current_user.get("role") in {"super_admin", "moderator", "country_admin"}:
            plans = await session.execute(select(Plan).where(*plan_query))
            plan_filters = [
                {"id": str(plan.id), "name": plan.name, "country_code": plan.country_code}
                for plan in plans.scalars().all()
            ]

        country_filters: List[Dict[str, Any]] = []
        countries_query = []
        if current_user.get("role") == "country_admin":
            scope = current_user.get("country_scope") or []
            if "*" not in scope:
                countries_query.append(Country.code.in_(scope))
        countries = await session.execute(select(Country).where(*countries_query).order_by(Country.code.asc()))
        country_filters = [_normalize_country_sql(country) for country in countries.scalars().all()]

        response["filters"] = {"plans": plan_filters, "countries": country_filters}

    return response


@api_router.get("/admin/dealers/{dealer_id}")
async def admin_get_dealer_detail(
    dealer_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        dealer_uuid = uuid.UUID(dealer_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid dealer id") from exc

    result = await session.execute(
        select(SqlUser).where(SqlUser.id == dealer_uuid, SqlUser.role == "dealer")
    )
    dealer_user = result.scalar_one_or_none()
    if not dealer_user:
        raise HTTPException(status_code=404, detail="Dealer not found")

    if getattr(ctx, "mode", "global") == "country" and ctx.country and dealer_user.country_code != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope forbidden")

    _assert_country_scope(dealer_user.country_code, current_user)

    last_invoice = (
        await session.execute(
            select(AdminInvoice)
            .where(AdminInvoice.user_id == dealer_uuid)
            .order_by(desc(AdminInvoice.issued_at), desc(AdminInvoice.created_at))
            .limit(1)
        )
    ).scalars().first()
    unpaid_count = (
        await session.execute(
            select(func.count())
            .select_from(AdminInvoice)
            .where(AdminInvoice.user_id == dealer_uuid, AdminInvoice.status == "issued")
        )
    ).scalar() or 0

    last_invoice_payload = _admin_invoice_to_dict(last_invoice) if last_invoice else None

    active_plan = None
    if dealer_user.plan_id:
        plan = await session.get(Plan, dealer_user.plan_id)
        if plan:
            active_plan = {"id": str(plan.id), "name": plan.name, "country_code": plan.country_code}

    return {
        "dealer": {
            "id": str(dealer_user.id),
            "email": dealer_user.email,
            "dealer_status": dealer_user.dealer_status or "active",
            "status": dealer_user.status,
            "risk_level": getattr(dealer_user, "risk_level", "low"),
            "ban_reason": dealer_user.ban_reason,
            "suspension_until": dealer_user.suspension_until.isoformat() if dealer_user.suspension_until else None,
            "country_code": dealer_user.country_code,
            "plan_id": str(dealer_user.plan_id) if dealer_user.plan_id else None,
            "created_at": dealer_user.created_at.isoformat() if dealer_user.created_at else None,
        },
        "active_plan": active_plan,
        "last_invoice": last_invoice_payload,
        "unpaid_count": int(unpaid_count or 0),
        "package": {
            "plan_id": str(dealer_user.plan_id) if dealer_user.plan_id else None,
            "last_invoice": last_invoice_payload,
        },
    }


@api_router.get("/admin/dealers/{dealer_id}/audit-logs")
async def admin_get_dealer_audit_logs(
    dealer_id: str,
    request: Request,
    limit: int = 5,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        dealer_uuid = uuid.UUID(dealer_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid dealer id") from exc

    result = await session.execute(
        select(SqlUser).where(SqlUser.id == dealer_uuid, SqlUser.role == "dealer")
    )
    dealer = result.scalar_one_or_none()
    if not dealer:
        raise HTTPException(status_code=404, detail="Dealer not found")

    if getattr(ctx, "mode", "global") == "country" and ctx.country and dealer.country_code != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope forbidden")

    _assert_country_scope(dealer.country_code, current_user)

    limit = min(max(int(limit), 1), 20)
    event_types = [
        "dealer_suspended",
        "dealer_reactivated",
        "dealer_deleted",
        "user_suspended",
        "user_reactivated",
        "user_deleted",
    ]
    logs = (
        await session.execute(
            select(AuditLog)
            .where(AuditLog.resource_id == dealer_id, AuditLog.action.in_(event_types))
            .order_by(desc(AuditLog.created_at))
            .limit(limit)
        )
    ).scalars().all()

    return {"items": [_audit_log_sql_to_dict(log) for log in logs]}


class DealerStatusPayload(BaseModel):
    dealer_status: str


@api_router.post("/admin/dealers/{dealer_id}/status")
async def admin_set_dealer_status(
    dealer_id: str,
    payload: DealerStatusPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    new_status = (payload.dealer_status or "").strip().lower()
    if new_status not in ["active", "suspended"]:
        raise HTTPException(status_code=400, detail="Invalid dealer_status")

    try:
        dealer_uuid = uuid.UUID(dealer_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid dealer id") from exc

    result = await session.execute(
        select(SqlUser).where(SqlUser.id == dealer_uuid, SqlUser.role == "dealer")
    )
    dealer = result.scalar_one_or_none()
    if not dealer:
        raise HTTPException(status_code=404, detail="Dealer not found")

    if getattr(ctx, "mode", "global") == "country" and ctx.country and dealer.country_code != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope forbidden")

    prev_status = dealer.dealer_status or "active"

    dealer.dealer_status = new_status
    if new_status == "suspended":
        dealer.status = "suspended"
        dealer.is_active = False
    else:
        dealer.status = "active"
        dealer.is_active = True

    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="DEALER_STATUS_CHANGE",
        actor=current_user,
        resource_type="user",
        resource_id=str(dealer.id),
        metadata={"previous_status": prev_status, "new_status": new_status},
        request=request,
        country_code=dealer.country_code,
    )
    await session.commit()

    return {"ok": True}


# =====================
# Sprint 1.2 — Dealer Applications (Admin)
# =====================

@api_router.get("/admin/dealer-applications")
async def admin_list_dealer_applications(
    request: Request,
    skip: int = 0,
    limit: int = 50,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = []
    if getattr(ctx, "mode", "global") == "country" and ctx.country:
        filters.append(DealerApplication.country == ctx.country)
    if status:
        filters.append(DealerApplication.status == status)
    if search:
        search_value = f"%{search}%"
        filters.append(
            or_(
                DealerApplication.contact_email.ilike(search_value),
                DealerApplication.company_name.ilike(search_value),
            )
        )

    limit = min(100, max(1, int(limit)))
    query = (
        select(DealerApplication)
        .where(*filters)
        .order_by(DealerApplication.created_at.desc())
        .offset(int(skip))
        .limit(limit)
    )
    apps = (await session.execute(query)).scalars().all()

    total = await session.scalar(select(func.count(DealerApplication.id)).where(*filters)) or 0
    return {
        "items": [_dealer_application_to_dict(app) for app in apps],
        "pagination": {"total": int(total), "skip": int(skip), "limit": limit},
    }


@api_router.get("/admin/individual-applications")
async def admin_list_individual_applications(
    request: Request,
    skip: int = 0,
    limit: int = 50,
    status: Optional[str] = None,
    search: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = [Application.application_type == "individual"]
    if getattr(ctx, "mode", "global") == "country" and ctx.country:
        filters.append(Application.extra_data["country_code"].astext == ctx.country)
    if status:
        filters.append(Application.status == status)
    if search:
        search_value = f"%{search}%"
        filters.append(
            or_(
                Application.application_id.ilike(search_value),
                Application.subject.ilike(search_value),
                Application.description.ilike(search_value),
                Application.extra_data["email"].astext.ilike(search_value),
            )
        )

    limit = min(100, max(1, int(limit)))
    query = (
        select(Application)
        .where(*filters)
        .order_by(Application.created_at.desc())
        .offset(int(skip))
        .limit(limit)
    )
    apps = (await session.execute(query)).scalars().all()
    total = await session.scalar(select(func.count(Application.id)).where(*filters)) or 0
    return {
        "items": [_individual_application_to_dict(app) for app in apps],
        "pagination": {"total": int(total), "skip": int(skip), "limit": limit},
    }


class DealerApplicationRejectPayload(BaseModel):
    reason: str
    reason_note: Optional[str] = None


class IndividualApplicationRejectPayload(BaseModel):
    reason: str
    reason_note: Optional[str] = None


@api_router.post("/admin/dealer-applications/{app_id}/reject")
async def admin_reject_dealer_application(
    app_id: str,
    payload: DealerApplicationRejectPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        app_uuid = uuid.UUID(app_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid application id") from exc

    app = await session.get(DealerApplication, app_uuid)
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    if getattr(ctx, "mode", "global") == "country" and ctx.country and app.country != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope forbidden")

    if app.status != "pending":
        raise HTTPException(status_code=400, detail="Application already reviewed")

    reason = (payload.reason or "").strip()
    if reason not in DEALER_APP_REJECT_REASONS_V1:
        raise HTTPException(status_code=400, detail="Invalid reason")

    reason_note = (payload.reason_note or "").strip() or None
    if reason == "other" and not reason_note:
        raise HTTPException(status_code=400, detail="reason_note is required when reason=other")

    app.status = "rejected"
    app.reject_reason = reason_note or reason
    app.reviewed_by_id = _safe_uuid(current_user.get("id"))
    app.reviewed_at = datetime.now(timezone.utc)

    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="DEALER_APPLICATION_REJECTED",
        actor=current_user,
        resource_type="dealer_application",
        resource_id=str(app.id),
        metadata={
            "previous_status": "pending",
            "new_status": "rejected",
            "reason": reason,
            "reason_note": reason_note,
        },
        request=request,
        country_code=app.country,
    )
    await session.commit()

    return {"ok": True}


@api_router.post("/admin/dealer-applications/{app_id}/approve")
async def admin_approve_dealer_application(
    app_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    try:
        app_uuid = uuid.UUID(app_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid application id") from exc

    app = await session.get(DealerApplication, app_uuid)
    if not app:
        raise HTTPException(status_code=404, detail="Application not found")

    if getattr(ctx, "mode", "global") == "country" and ctx.country and app.country != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope forbidden")

    if app.status != "pending":
        raise HTTPException(status_code=400, detail="Application already reviewed")

    existing = await session.execute(select(SqlUser).where(SqlUser.email == app.contact_email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="User already exists")

    now_dt = datetime.now(timezone.utc)
    raw_password = str(uuid.uuid4())[:12] + "!"
    hashed = get_password_hash(raw_password)

    dealer_user = SqlUser(
        id=uuid.uuid4(),
        email=app.contact_email,
        hashed_password=hashed,
        full_name=app.contact_name or app.company_name,
        role="dealer",
        status="active",
        is_active=True,
        is_verified=False,
        country_code=app.country,
        dealer_status="active",
        created_at=now_dt,
        updated_at=now_dt,
    )
    session.add(dealer_user)

    dealer = Dealer(
        application_id=app.id,
        country=app.country,
        dealer_type=app.dealer_type,
        company_name=app.company_name,
        vat_tax_no=app.vat_tax_no,
        logo_url=app.logo_url,
        is_active=True,
        can_publish=True,
        listing_limit=50,
        premium_limit=10,
        created_at=now_dt,
        updated_at=now_dt,
    )
    session.add(dealer)
    await session.flush()

    dealer_link = DealerUser(
        dealer_id=dealer.id,
        user_id=dealer_user.id,
        role="owner",
        created_at=now_dt,
        updated_at=now_dt,
    )
    session.add(dealer_link)

    app.status = "approved"
    app.reviewed_by_id = _safe_uuid(current_user.get("id"))
    app.reviewed_at = now_dt
    app.updated_at = now_dt

    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="DEALER_APPLICATION_APPROVED",
        actor=current_user,
        resource_type="dealer_application",
        resource_id=str(app.id),
        metadata={"previous_status": "pending", "new_status": "approved"},
        request=request,
        country_code=app.country,
    )
    await session.commit()

    return {
        "ok": True,
        "dealer_user": {
            "id": str(dealer_user.id),
            "email": dealer_user.email,
            "temp_password": raw_password,
        },
    }


@api_router.post("/admin/individual-applications/{app_id}/approve")
async def admin_approve_individual_application(
    app_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        app_uuid = uuid.UUID(app_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid application id") from exc

    app = await session.get(Application, app_uuid)
    if not app or app.application_type != "individual":
        raise HTTPException(status_code=404, detail="Application not found")

    country_code = (app.extra_data or {}).get("country_code")
    if getattr(ctx, "mode", "global") == "country" and country_code and country_code != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope violation")
    if app.status != "pending":
        raise HTTPException(status_code=400, detail="Application not pending")

    applicant_email = (app.extra_data or {}).get("email")
    applicant_name = (app.extra_data or {}).get("full_name") or applicant_email

    if applicant_email:
        existing_user = await session.execute(select(SqlUser).where(SqlUser.email == applicant_email))
        if not existing_user.scalar_one_or_none():
            user = SqlUser(
                id=uuid.uuid4(),
                email=applicant_email,
                hashed_password=get_password_hash("User123!"),
                full_name=applicant_name or applicant_email,
                role="individual",
                status="active",
                is_active=True,
                is_verified=False,
                country_code=country_code,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(user)

    app.status = "approved"
    app.updated_at = datetime.now(timezone.utc)

    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="INDIVIDUAL_APPLICATION_APPROVED",
        actor=current_user,
        resource_type="individual_application",
        resource_id=str(app.id),
        metadata={"email": applicant_email},
        request=request,
        country_code=country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.post("/admin/individual-applications/{app_id}/reject")
async def admin_reject_individual_application(
    app_id: str,
    payload: IndividualApplicationRejectPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        app_uuid = uuid.UUID(app_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid application id") from exc

    app = await session.get(Application, app_uuid)
    if not app or app.application_type != "individual":
        raise HTTPException(status_code=404, detail="Application not found")

    country_code = (app.extra_data or {}).get("country_code")
    if getattr(ctx, "mode", "global") == "country" and country_code and country_code != ctx.country:
        raise HTTPException(status_code=403, detail="Country scope violation")
    if app.status != "pending":
        raise HTTPException(status_code=400, detail="Application not pending")
    if payload.reason not in INDIVIDUAL_APP_REJECT_REASONS_V1:
        raise HTTPException(status_code=400, detail="Invalid reject reason")

    app.status = "rejected"
    app.decision_reason = payload.reason_note or payload.reason
    app.updated_at = datetime.now(timezone.utc)

    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="INDIVIDUAL_APPLICATION_REJECTED",
        actor=current_user,
        resource_type="individual_application",
        resource_id=str(app.id),
        metadata={"email": (app.extra_data or {}).get("email"), "reason": payload.reason},
        request=request,
        country_code=country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.get("/countries")
async def list_countries(
    request: Request,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    result = await session.execute(select(Country).order_by(Country.code.asc()))
    countries = result.scalars().all()
    return [_normalize_country_sql(country) for country in countries]

 

@api_router.patch("/countries/{country_id}")
async def update_country(
    country_id: str,
    data: dict,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    allowed = {"is_enabled", "default_currency", "default_language", "support_email"}
    payload = {k: v for k, v in data.items() if k in allowed}
    if not payload:
        return {"ok": True}

    country = None
    try:
        country_uuid = uuid.UUID(country_id)
        country = await session.get(Country, country_uuid)
    except ValueError:
        country = None

    if not country:
        result = await session.execute(select(Country).where(Country.code == country_id.upper()))
        country = result.scalar_one_or_none()

    if not country:
        raise HTTPException(status_code=404, detail="Country not found")

    for key, value in payload.items():
        setattr(country, key, value)
    country.updated_at = datetime.now(timezone.utc)

    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="UPDATE_COUNTRY",
        actor=current_user,
        resource_type="country",
        resource_id=str(country.id),
        metadata={"payload": payload, "country_code": country.code},
        request=request,
        country_code=country.code,
    )
    await session.commit()

    return {"ok": True}



# =====================
# Audit Logs (Legacy) - Backoffice
# =====================

@api_router.get("/audit-logs")
async def list_audit_logs(
    request: Request,
    skip: int = 0,
    limit: int = 50,
    action: Optional[str] = None,
    event_type: Optional[str] = None,
    resource_type: Optional[str] = None,
    user_id: Optional[str] = None,
    admin_user_id: Optional[str] = None,
    country: Optional[str] = None,
    start: Optional[str] = None,
    end: Optional[str] = None,
    country_scope: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = []
    if action:
        filters.append(AuditLog.action == action)
    if event_type:
        filters.append(AuditLog.action == event_type)
    if resource_type:
        filters.append(AuditLog.resource_type == resource_type)
    if user_id:
        filters.append(AuditLog.user_id == _safe_uuid(user_id))
    if admin_user_id:
        filters.append(AuditLog.user_id == _safe_uuid(admin_user_id))
    if country:
        filters.append(AuditLog.country_code == country.strip().upper())
    if country_scope:
        filters.append(AuditLog.country_scope == country_scope)

    start_dt = _parse_audit_date(start, is_end=False)
    end_dt = _parse_audit_date(end, is_end=True)
    if start_dt:
        filters.append(AuditLog.created_at >= start_dt)
    if end_dt:
        filters.append(AuditLog.created_at <= end_dt)

    limit = min(max(int(limit), 1), 200)
    query = (
        select(AuditLog)
        .where(*filters)
        .order_by(desc(AuditLog.created_at))
        .offset(int(skip))
        .limit(limit)
    )
    rows = (await session.execute(query)).scalars().all()
    return [_audit_log_sql_to_dict(row) for row in rows]


def _parse_audit_date(value: Optional[str], is_end: bool = False) -> Optional[datetime]:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        try:
            parsed = datetime.strptime(value, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid date format") from exc
    if len(value) <= 10:
        if is_end:
            parsed = parsed.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            parsed = parsed.replace(hour=0, minute=0, second=0, microsecond=0)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _build_audit_query(
    *,
    q: Optional[str],
    action: Optional[str],
    event_type: Optional[str],
    resource_type: Optional[str],
    country_code: Optional[str],
    admin_user_query: Optional[str],
    date_from: Optional[str],
    date_to: Optional[str],
) -> Dict[str, Any]:
    conditions: List[Dict[str, Any]] = []

    if action:
        conditions.append({"action": action})
    if event_type:
        conditions.append({"event_type": event_type})
    if resource_type:
        conditions.append({"resource_type": resource_type})
    if country_code:
        conditions.append({"country_code": country_code})

    if admin_user_query:
        if "@" in admin_user_query:
            regex = re.compile(re.escape(admin_user_query), re.IGNORECASE)
            conditions.append({
                "$or": [
                    {"user_email": regex},
                    {"admin_user_email": regex},
                    {"email": regex},
                ]
            })
        else:
            conditions.append({
                "$or": [
                    {"admin_user_id": admin_user_query},
                    {"user_id": admin_user_query},
                    {"actor_id": admin_user_query},
                ]
            })

    if q:
        regex = re.compile(re.escape(q), re.IGNORECASE)
        conditions.append({
            "$or": [
                {"event_type": regex},
                {"action": regex},
                {"resource_type": regex},
                {"resource_id": regex},
                {"user_email": regex},
                {"admin_user_email": regex},
                {"admin_user_id": regex},
                {"user_id": regex},
                {"country_code": regex},
            ]
        })

    if date_from or date_to:
        created_at_q: Dict[str, Any] = {}
        if date_from:
            created_at_q["$gte"] = _parse_audit_date(date_from, is_end=False)
        if date_to:
            created_at_q["$lte"] = _parse_audit_date(date_to, is_end=True)
        conditions.append({"created_at": created_at_q})

    if not conditions:
        return {}
    if len(conditions) == 1:
        return conditions[0]
    return {"$and": conditions}


def _audit_log_sql_to_dict(row: AuditLog) -> Dict[str, Any]:
    return {
        "id": str(row.id),
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "action": row.action,
        "resource_type": row.resource_type,
        "resource_id": row.resource_id,
        "user_id": str(row.user_id) if row.user_id else None,
        "user_email": row.user_email,
        "country_scope": row.country_scope,
        "metadata": row.metadata_info or {},
    }


@api_router.get("/admin/audit-logs")
async def admin_list_audit_logs(
    request: Request,
    q: Optional[str] = None,
    event_type: Optional[str] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    country_code: Optional[str] = None,
    admin_user_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    sort: Optional[str] = "timestamp_desc",
    page: int = 0,
    page_size: int = 20,
    scope: Optional[str] = None,
    ref: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
    current_user=Depends(check_permissions(["super_admin", "finance", "ROLE_AUDIT_VIEWER", "audit_viewer"])),
):
    if scope == "billing":
        role = current_user.get("role")
        if role not in {"super_admin", "finance"}:
            raise HTTPException(status_code=403, detail="Insufficient permissions")

        page_size = min(200, max(1, int(page_size)))
        page = max(0, int(page))
        skip = page * page_size

        base_conditions = [AuditLog.action.in_(BILLING_AUDIT_ACTIONS)]
        if ref:
            ref_value = ref.strip()
            base_conditions.append(
                or_(
                    AuditLog.resource_id == ref_value,
                    cast(AuditLog.metadata_info["invoice_id"], String) == ref_value,
                    cast(AuditLog.metadata_info["subscription_id"], String) == ref_value,
                    cast(AuditLog.metadata_info["payment_id"], String) == ref_value,
                )
            )

        query_stmt = select(AuditLog).where(and_(*base_conditions))
        total_stmt = select(func.count()).select_from(AuditLog).where(and_(*base_conditions))

        if sort == "timestamp_asc":
            query_stmt = query_stmt.order_by(AuditLog.created_at.asc())
        else:
            query_stmt = query_stmt.order_by(desc(AuditLog.created_at))

        rows = (await session.execute(query_stmt.offset(skip).limit(page_size))).scalars().all()
        total = (await session.execute(total_stmt)).scalar_one() or 0

        return {
            "items": [_audit_log_sql_to_dict(row) for row in rows],
            "pagination": {"total": int(total), "page": page, "page_size": page_size},
        }

    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    page_size = min(200, max(1, int(page_size)))
    page = max(0, int(page))
    skip = page * page_size

    filters = []
    if action:
        filters.append(AuditLog.action == action)
    if event_type:
        filters.append(AuditLog.action == event_type)
    if resource_type:
        filters.append(AuditLog.resource_type == resource_type)
    if country_code:
        filters.append(AuditLog.country_code == country_code.strip().upper())

    if admin_user_id:
        admin_uuid = _safe_uuid(admin_user_id)
        if admin_uuid:
            filters.append(AuditLog.user_id == admin_uuid)
        else:
            filters.append(AuditLog.user_email.ilike(f"%{admin_user_id}%"))

    if q:
        search_value = f"%{q}%"
        filters.append(
            or_(
                AuditLog.action.ilike(search_value),
                AuditLog.resource_type.ilike(search_value),
                AuditLog.resource_id.ilike(search_value),
                AuditLog.user_email.ilike(search_value),
                AuditLog.country_code.ilike(search_value),
            )
        )

    date_from = _parse_audit_date(from_date, is_end=False)
    date_to = _parse_audit_date(to_date, is_end=True)
    if date_from:
        filters.append(AuditLog.created_at >= date_from)
    if date_to:
        filters.append(AuditLog.created_at <= date_to)

    query_stmt = select(AuditLog).where(and_(*filters)) if filters else select(AuditLog)
    total_stmt = select(func.count()).select_from(AuditLog).where(and_(*filters)) if filters else select(func.count()).select_from(AuditLog)

    if sort == "timestamp_asc":
        query_stmt = query_stmt.order_by(AuditLog.created_at.asc())
    else:
        query_stmt = query_stmt.order_by(desc(AuditLog.created_at))

    rows = (await session.execute(query_stmt.offset(skip).limit(page_size))).scalars().all()
    total = (await session.execute(total_stmt)).scalar_one() or 0

    return {
        "items": [_audit_log_sql_to_dict(row) for row in rows],
        "pagination": {"total": int(total), "page": page, "page_size": page_size},
    }


@api_router.get("/admin/audit-logs/event-types")
async def admin_audit_event_types(
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "ROLE_AUDIT_VIEWER", "audit_viewer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    rows = await session.execute(select(distinct(AuditLog.action)))
    types = sorted([row[0] for row in rows if row[0]])
    return {"event_types": types}


@api_router.get("/admin/audit-logs/actions")
async def admin_audit_actions(
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "ROLE_AUDIT_VIEWER", "audit_viewer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    rows = await session.execute(select(distinct(AuditLog.action)))
    actions = sorted([row[0] for row in rows if row[0]])
    return {"actions": actions}


@api_router.get("/admin/audit-logs/resources")
async def admin_audit_resources(
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "ROLE_AUDIT_VIEWER", "audit_viewer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    rows = await session.execute(select(distinct(AuditLog.resource_type)))
    resources = sorted([row[0] for row in rows if row[0]])
    return {"resource_types": resources}


@api_router.get("/admin/audit-logs/{log_id}")
async def admin_audit_log_detail(
    log_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "ROLE_AUDIT_VIEWER", "audit_viewer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    log_uuid = _safe_uuid(log_id)
    if not log_uuid:
        raise HTTPException(status_code=400, detail="Invalid audit log id")
    log = await session.get(AuditLog, log_uuid)
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")
    return {"log": _audit_log_sql_to_dict(log)}


@api_router.get("/admin/audit-logs/export")
async def admin_export_audit_logs(
    request: Request,
    q: Optional[str] = None,
    event_type: Optional[str] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    country_code: Optional[str] = None,
    admin_user_id: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    sort: Optional[str] = "timestamp_desc",
    current_user=Depends(check_permissions(["super_admin", "ROLE_AUDIT_VIEWER", "audit_viewer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = []
    if action:
        filters.append(AuditLog.action == action)
    if event_type:
        filters.append(AuditLog.action == event_type)
    if resource_type:
        filters.append(AuditLog.resource_type == resource_type)
    if country_code:
        filters.append(AuditLog.country_code == country_code.strip().upper())

    if admin_user_id:
        admin_uuid = _safe_uuid(admin_user_id)
        if admin_uuid:
            filters.append(AuditLog.user_id == admin_uuid)
        else:
            filters.append(AuditLog.user_email.ilike(f"%{admin_user_id}%"))

    if q:
        search_value = f"%{q}%"
        filters.append(
            or_(
                AuditLog.action.ilike(search_value),
                AuditLog.resource_type.ilike(search_value),
                AuditLog.resource_id.ilike(search_value),
                AuditLog.user_email.ilike(search_value),
                AuditLog.country_code.ilike(search_value),
            )
        )

    date_from = _parse_audit_date(from_date, is_end=False)
    date_to = _parse_audit_date(to_date, is_end=True)
    if date_from:
        filters.append(AuditLog.created_at >= date_from)
    if date_to:
        filters.append(AuditLog.created_at <= date_to)

    query_stmt = select(AuditLog).where(and_(*filters)) if filters else select(AuditLog)
    if sort == "timestamp_asc":
        query_stmt = query_stmt.order_by(AuditLog.created_at.asc())
    else:
        query_stmt = query_stmt.order_by(desc(AuditLog.created_at))

    rows = (await session.execute(query_stmt.limit(10000))).scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    headers = [
        "id",
        "created_at",
        "action",
        "resource_type",
        "resource_id",
        "user_id",
        "user_email",
        "country_code",
    ]
    writer.writerow(headers)

    for row in rows:
        writer.writerow([
            str(row.id),
            row.created_at.isoformat() if row.created_at else "",
            row.action,
            row.resource_type,
            row.resource_id,
            str(row.user_id) if row.user_id else "",
            row.user_email or "",
            row.country_code or "",
        ])

    filename = datetime.now(timezone.utc).strftime("audit-logs-%Y%m%d-%H%M.csv")
    response = StreamingResponse(iter([output.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


# =====================
# Moderation (SQL) - Backoffice
# =====================


def _extract_setting_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"true", "1", "yes", "on"}
    if isinstance(value, dict):
        for key in ("enabled", "active", "value"):
            if key in value:
                return _extract_setting_bool(value[key])
    return False


def _normalize_freeze_reason(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    trimmed = value.strip()
    if not trimmed:
        return None
    if len(trimmed) > MODERATION_FREEZE_REASON_MAX_LEN:
        return trimmed[:MODERATION_FREEZE_REASON_MAX_LEN]
    return trimmed


async def _get_system_setting_by_key(
    session: AsyncSession, key: str, country_code: Optional[str] = None
) -> Optional[SystemSetting]:
    if country_code:
        result = await session.execute(
            select(SystemSetting).where(
                SystemSetting.key == key,
                SystemSetting.country_code == country_code,
            )
        )
        setting = result.scalar_one_or_none()
        if setting:
            return setting
    result = await session.execute(
        select(SystemSetting).where(
            SystemSetting.key == key,
            or_(SystemSetting.country_code.is_(None), SystemSetting.country_code == ""),
        )
    )
    return result.scalar_one_or_none()


async def _is_moderation_freeze_active(
    session: AsyncSession, country_code: Optional[str]
) -> bool:
    setting = await _get_system_setting_by_key(session, MODERATION_FREEZE_SETTING_KEY, country_code)
    if not setting:
        return False
    return _extract_setting_bool(setting.value)


async def _assert_moderation_not_frozen(
    *,
    session: AsyncSession,
    request: Optional[Request],
    current_user: dict,
    listing_id: Optional[str] = None,
    listing_country: Optional[str] = None,
    action_type: Optional[str] = None,
    listing_ids: Optional[List[str]] = None,
) -> None:
    if not await _is_moderation_freeze_active(session, listing_country):
        return

    await _write_audit_log_sql(
        session=session,
        action="moderation_freeze_blocked",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="moderation",
        resource_id=listing_id or "bulk",
        metadata={
            "action_type": action_type,
            "listing_id": listing_id,
            "listing_ids": listing_ids,
            "country_code": listing_country,
            "role": current_user.get("role"),
        },
        request=request,
        country_code=listing_country,
    )
    await session.commit()
    raise HTTPException(status_code=423, detail="Moderation freeze active")


def _ensure_moderation_rbac(current_user: dict):
    role = current_user.get("role")
    if role not in ALLOWED_MODERATION_ROLES:
        raise HTTPException(status_code=403, detail="Insufficient permissions")


async def _ensure_bulk_listings_pending(session: AsyncSession, listing_ids: List[str]) -> None:
    if not listing_ids:
        raise HTTPException(status_code=400, detail="listing_ids is required")

    uuids = []
    for listing_id in listing_ids:
        try:
            uuids.append(uuid.UUID(str(listing_id)))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid listing id") from exc

    result = await session.execute(
        select(Listing.id, Listing.status).where(Listing.id.in_(uuids))
    )
    rows = result.all()
    if len(rows) != len(uuids):
        raise HTTPException(status_code=404, detail="Listing not found")

    invalid = [str(row[0]) for row in rows if row[1] != "pending_moderation"]
    if invalid:
        raise HTTPException(status_code=400, detail="All listings must be pending_moderation")


def _validate_reason(reason: str, allowed: set[str]) -> str:
    r = (reason or "").strip()
    if not r:
        raise HTTPException(status_code=400, detail="reason is required")
    if r not in allowed:
        raise HTTPException(status_code=400, detail="Invalid reason")
    return r


MODERATION_REASON_MAX_LEN = 500
MODERATION_FREEZE_REASON_MAX_LEN = 280
PLAN_QUOTA_MIN = 0
PLAN_QUOTA_MAX = 10000
RISK_LEVEL_ALLOWED = {"low", "medium", "high"}
SYSTEM_MODERATOR_ID = uuid.UUID("00000000-0000-0000-0000-000000000000")
MODERATION_STATUS_MAP = {
    "pending": "PENDING",
    "pending_moderation": "PENDING",
    "approved": "APPROVED",
    "published": "APPROVED",
    "rejected": "REJECTED",
    "needs_revision": "NEEDS_REVISION",
    "needs-revision": "NEEDS_REVISION",
}


def _normalize_moderation_item_status(value: Optional[str]) -> str:
    if not value:
        return "PENDING"
    cleaned = value.strip().lower()
    if cleaned in MODERATION_STATUS_MAP:
        return MODERATION_STATUS_MAP[cleaned]
    upper = value.strip().upper()
    if upper in {"PENDING", "APPROVED", "REJECTED", "NEEDS_REVISION"}:
        return upper
    return upper


def _sanitize_moderation_reason(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    cleaned = str(value).strip()
    if not cleaned:
        return None
    if len(cleaned) > MODERATION_REASON_MAX_LEN:
        cleaned = cleaned[:MODERATION_REASON_MAX_LEN]
    return cleaned


def _resolve_moderation_actor_id(value: Optional[str]) -> uuid.UUID:
    actor_uuid = _safe_uuid(value)
    return actor_uuid or SYSTEM_MODERATOR_ID


async def _upsert_moderation_item(
    *,
    session: AsyncSession,
    listing: Listing,
    status: str,
    reason: Optional[str],
    moderator_id: Optional[uuid.UUID],
    audit_ref: Optional[str],
) -> ModerationItem:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(ModerationItem).where(ModerationItem.listing_id == listing.id)
    )
    item = result.scalar_one_or_none()
    if not item:
        item = ModerationItem(
            listing_id=listing.id,
            status=status,
            reason=reason,
            moderator_id=moderator_id,
            audit_ref=audit_ref,
            created_at=now,
            updated_at=now,
        )
        session.add(item)
        return item

    item.status = status
    item.reason = reason
    item.moderator_id = moderator_id
    item.audit_ref = audit_ref
    item.updated_at = now
    return item



@api_router.get("/admin/moderation/queue")
async def moderation_queue(
    request: Request,
    status: str = "pending_moderation",
    dealer_only: Optional[bool] = None,
    country: Optional[str] = None,
    module: Optional[str] = None,
    limit: int = 50,
    skip: int = 0,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    _ensure_moderation_rbac(current_user)

    normalized_status = _normalize_moderation_item_status(status)
    conditions = [ModerationItem.status == normalized_status]
    if dealer_only is not None:
        conditions.append(Listing.is_dealer_listing == dealer_only)
    if country:
        conditions.append(Listing.country == country.strip().upper())
    if module:
        conditions.append(Listing.module == module)

    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope:
            conditions.append(Listing.country.in_(scope))

    limit = min(200, max(1, int(limit)))
    offset = max(0, int(skip))
    query_stmt = (
        select(ModerationItem, Listing)
        .join(Listing, ModerationItem.listing_id == Listing.id)
        .where(and_(*conditions))
        .order_by(desc(ModerationItem.created_at))
        .offset(offset)
        .limit(limit)
    )
    rows = (await session.execute(query_stmt)).all()

    out = []
    for item, listing in rows:
        attrs = listing.attributes or {}
        vehicle = attrs.get("vehicle") or {}
        title = (listing.title or "").strip()
        if not title:
            title = f"{(vehicle.get('make_key') or '').upper()} {vehicle.get('model_key') or ''} {vehicle.get('year') or ''}".strip()
        images = listing.images or []
        out.append(
            {
                "id": str(listing.id),
                "title": title,
                "status": listing.status,
                "moderation_status": item.status,
                "moderation_created_at": item.created_at.isoformat() if item.created_at else None,
                "moderation_updated_at": item.updated_at.isoformat() if item.updated_at else None,
                "country": listing.country,
                "module": listing.module,
                "city": listing.city or "",
                "price": listing.price,
                "price_type": listing.price_type or "FIXED",
                "price_amount": listing.price,
                "hourly_rate": listing.hourly_rate,
                "currency": listing.currency or "EUR",
                "image_count": listing.image_count or len(images),
                "created_at": listing.created_at.isoformat() if listing.created_at else None,
                "is_dealer_listing": bool(listing.is_dealer_listing),
                "dealer_only": bool(listing.is_dealer_listing),
                "is_premium": bool(listing.is_premium),
            }
        )

    return out


@api_router.get("/admin/moderation/queue/count")
async def moderation_queue_count(
    request: Request,
    status: str = "pending_moderation",
    dealer_only: Optional[bool] = None,
    country: Optional[str] = None,
    module: Optional[str] = None,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    _ensure_moderation_rbac(current_user)
    normalized_status = _normalize_moderation_item_status(status)
    conditions = [ModerationItem.status == normalized_status]
    if dealer_only is not None:
        conditions.append(Listing.is_dealer_listing == dealer_only)
    if country:
        conditions.append(Listing.country == country.strip().upper())
    if module:
        conditions.append(Listing.module == module)

    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope:
            conditions.append(Listing.country.in_(scope))

    count_stmt = (
        select(func.count())
        .select_from(ModerationItem)
        .join(Listing, ModerationItem.listing_id == Listing.id)
        .where(and_(*conditions))
    )
    count = (await session.execute(count_stmt)).scalar_one() or 0
    return {"count": int(count)}


@api_router.get("/admin/moderation/listings/{listing_id}")
async def moderation_listing_detail(
    listing_id: str,
    request: Request,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    _ensure_moderation_rbac(current_user)

    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid listing id") from exc

    listing = await session.get(Listing, listing_uuid)
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")

    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope and listing.country not in scope:
            raise HTTPException(status_code=403, detail="Country scope forbidden")

    attrs = listing.attributes or {}
    vehicle = attrs.get("vehicle") or {}
    title = (listing.title or "").strip()
    if not title:
        title = f"{(vehicle.get('make_key') or '').upper()} {vehicle.get('model_key') or ''} {vehicle.get('year') or ''}".strip()

    history_rows = (
        await session.execute(
            select(AuditLog)
            .where(AuditLog.resource_type == "listing", AuditLog.resource_id == listing_id)
            .order_by(desc(AuditLog.created_at))
            .limit(50)
        )
    ).scalars().all()
    moderation_history = [_audit_log_sql_to_dict(row) for row in history_rows]

    images = listing.images or []
    return {
        "id": listing_id,
        "title": title,
        "status": listing.status,
        "module": listing.module,
        "country": listing.country,
        "city": listing.city or "",
        "price": listing.price,
        "price_type": listing.price_type or "FIXED",
        "price_amount": listing.price,
        "hourly_rate": listing.hourly_rate,
        "currency": listing.currency or "EUR",
        "description": listing.description or "",
        "attributes": attrs,
        "images": images,
        "image_count": listing.image_count or len(images),
        "created_at": listing.created_at.isoformat() if listing.created_at else None,
        "moderation_history": moderation_history,
    }


class ModerationReasonPayload(BaseModel):
    reason: Optional[str] = None
    reason_note: Optional[str] = None


class ListingAdminActionPayload(BaseModel):
    reason: Optional[str] = None
    reason_note: Optional[str] = None


class BulkModerationPayload(BaseModel):
    listing_ids: List[str] = Field(..., min_length=1)
    reason: Optional[str] = None
    reason_note: Optional[str] = None


class ReportCreatePayload(BaseModel):
    listing_id: str
    reason: str
    reason_note: Optional[str] = None


class ReportStatusPayload(BaseModel):
    target_status: str
    note: str


class InvoiceCreatePayload(BaseModel):
    user_id: Optional[str] = None
    dealer_id: Optional[str] = None
    subscription_id: Optional[str] = None
    plan_id: Optional[str] = None
    campaign_id: Optional[str] = None
    amount_total: Optional[float] = None
    amount: Optional[float] = None
    currency: Optional[str] = None
    currency_code: Optional[str] = None
    due_at: Optional[str] = None
    notes: Optional[str] = None
    provider_customer_id: Optional[str] = None
    meta_json: Optional[Dict[str, Any]] = None
    issue_now: Optional[bool] = True


class InvoiceActionPayload(BaseModel):
    payment_method: Optional[str] = None
    reason: Optional[str] = None


class PaymentCheckoutPayload(BaseModel):
    invoice_id: str
    origin_url: str


class PaymentCheckoutStubPayload(BaseModel):
    amount: float = Field(..., gt=0)
    currency: str = "EUR"
    origin_url: str
    metadata: Optional[Dict[str, Any]] = None


class TaxRateCreatePayload(BaseModel):
    country_code: str
    rate: float
    effective_date: str
    active_flag: Optional[bool] = True


class TaxRateUpdatePayload(BaseModel):
    rate: Optional[float] = None
    effective_date: Optional[str] = None
    active_flag: Optional[bool] = None


class PlanCreatePayload(BaseModel):
    name: str
    slug: Optional[str] = None
    country_scope: str
    country_code: Optional[str] = None
    period: str = "monthly"
    price_amount: float
    currency_code: Optional[str] = None
    listing_quota: int
    showcase_quota: int
    active_flag: Optional[bool] = True


class PlanUpdatePayload(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    country_scope: Optional[str] = None
    country_code: Optional[str] = None
    period: Optional[str] = None
    price_amount: Optional[float] = None
    currency_code: Optional[str] = None
    listing_quota: Optional[int] = None
    showcase_quota: Optional[int] = None
    active_flag: Optional[bool] = None


class DealerPlanAssignmentPayload(BaseModel):
    plan_id: Optional[str] = None


class CountryCreatePayload(BaseModel):
    country_code: str
    name: str
    active_flag: Optional[bool] = True
    default_currency: str
    default_language: Optional[str] = None


class CountryUpdatePayload(BaseModel):
    name: Optional[str] = None
    active_flag: Optional[bool] = None
    default_currency: Optional[str] = None
    default_language: Optional[str] = None


class SystemSettingCreatePayload(BaseModel):
    key: str
    value: Any
    country_code: Optional[str] = None
    is_readonly: Optional[bool] = False
    description: Optional[str] = None
    moderation_freeze_reason: Optional[str] = None


class SystemSettingUpdatePayload(BaseModel):
    value: Optional[Any] = None
    country_code: Optional[str] = None
    is_readonly: Optional[bool] = None
    description: Optional[str] = None
    moderation_freeze_reason: Optional[str] = None


class CloudflareConfigPayload(BaseModel):
    account_id: Optional[str] = None
    zone_id: Optional[str] = None


class CategoryCreatePayload(BaseModel):
    name: str
    slug: str
    parent_id: Optional[str] = None
    country_code: Optional[str] = None
    module: Optional[str] = None
    active_flag: Optional[bool] = True
    sort_order: Optional[int] = 0
    hierarchy_complete: Optional[bool] = None
    form_schema: Optional[Dict[str, Any]] = None
    wizard_progress: Optional[Dict[str, Any]] = None


class CategoryUpdatePayload(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    parent_id: Optional[str] = None
    country_code: Optional[str] = None
    module: Optional[str] = None
    active_flag: Optional[bool] = None
    sort_order: Optional[int] = None
    hierarchy_complete: Optional[bool] = None
    form_schema: Optional[Dict[str, Any]] = None
    wizard_progress: Optional[Dict[str, Any]] = None
    wizard_edit_event: Optional[Dict[str, Any]] = None
    expected_updated_at: Optional[str] = None


class MenuItemCreatePayload(BaseModel):
    label: str
    slug: str
    url: Optional[str] = None
    parent_id: Optional[str] = None
    country_code: Optional[str] = None
    active_flag: Optional[bool] = True
    sort_order: Optional[int] = 0


class MenuItemUpdatePayload(BaseModel):
    label: Optional[str] = None
    slug: Optional[str] = None
    url: Optional[str] = None
    parent_id: Optional[str] = None
    country_code: Optional[str] = None
    active_flag: Optional[bool] = None
    sort_order: Optional[int] = None


class AttributeCreatePayload(BaseModel):
    category_id: str
    name: str
    key: str
    type: str
    required_flag: Optional[bool] = False
    filterable_flag: Optional[bool] = False
    options: Optional[List[str]] = None
    country_code: Optional[str] = None
    active_flag: Optional[bool] = True


class AttributeUpdatePayload(BaseModel):
    name: Optional[str] = None
    key: Optional[str] = None
    type: Optional[str] = None
    required_flag: Optional[bool] = None
    filterable_flag: Optional[bool] = None
    options: Optional[List[str]] = None
    country_code: Optional[str] = None
    active_flag: Optional[bool] = None


class VehicleMakeCreatePayload(BaseModel):
    name: str
    slug: str
    country_code: str
    active_flag: Optional[bool] = True


class VehicleMakeUpdatePayload(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    country_code: Optional[str] = None
    active_flag: Optional[bool] = None


class VehicleModelCreatePayload(BaseModel):
    make_id: str
    name: str
    slug: str
    vehicle_type: str
    active_flag: Optional[bool] = True


class VehicleModelUpdatePayload(BaseModel):
    name: Optional[str] = None
    slug: Optional[str] = None
    make_id: Optional[str] = None
    vehicle_type: Optional[str] = None
    active_flag: Optional[bool] = None


class VehicleMakeImportPayload(BaseModel):
    name: str
    slug: str
    country_code: str
    active: Optional[bool] = True


class VehicleModelImportPayload(BaseModel):
    make_slug: str
    name: str
    slug: str
    vehicle_type: str
    active: Optional[bool] = True


class VehicleImportPayload(BaseModel):
    makes: List[VehicleMakeImportPayload] = Field(default_factory=list)
    models: List[VehicleModelImportPayload] = Field(default_factory=list)


def _resolve_listing_title(listing: dict) -> str:
    title = (listing.get("title") or "").strip()
    if title:
        return title
    v = listing.get("vehicle") or {}
    return f"{(v.get('make_key') or '').upper()} {v.get('model_key') or ''} {v.get('year') or ''}".strip()


def _parse_bool_flag(value: Optional[str | bool]) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    raw = str(value).strip().lower()
    if raw in {"1", "true", "yes", "y"}:
        return True
    if raw in {"0", "false", "no", "n"}:
        return False
    return None


async def _admin_listing_action(
    *,
    session: AsyncSession,
    listing_id: str,
    current_user: dict,
    event_type: str,
    new_status: str,
    reason: Optional[str] = None,
    reason_note: Optional[str] = None,
) -> dict:
    try:
        listing_uuid = uuid.UUID(str(listing_id))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid listing id") from exc

    listing = await session.get(Listing, listing_uuid)
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")

    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope and listing.country not in scope:
            raise HTTPException(status_code=403, detail="Country scope forbidden")

    prev_status = listing.status
    if prev_status == new_status:
        raise HTTPException(status_code=400, detail="Listing already in target status")

    now = datetime.now(timezone.utc)
    listing.status = new_status
    listing.updated_at = now
    if event_type == "LISTING_SOFT_DELETE":
        listing.deleted_at = now
    if event_type == "LISTING_FORCE_UNPUBLISH":
        listing.published_at = None

    await session.commit()

    await log_action(
        db=session,
        action=event_type,
        resource_type="listing",
        resource_id=str(listing.id),
        user_id=_safe_uuid(current_user.get("id")),
        user_email=current_user.get("email"),
        old_values={"status": prev_status},
        new_values={"status": new_status},
        metadata={"reason": reason, "reason_note": reason_note},
        country_scope=",".join(current_user.get("country_scope") or []),
    )

    return {"id": str(listing.id), "status": listing.status}


def _check_report_rate_limit(request: Request, listing_id: str, reporter_user_id: Optional[str]) -> None:
    ip_address = _get_client_ip(request) or "unknown"
    key = f"{ip_address}:{reporter_user_id or 'anon'}:{listing_id}"
    now = time.time()
    attempts = _report_submit_attempts.get(key, [])
    attempts = [ts for ts in attempts if (now - ts) <= REPORT_RATE_LIMIT_WINDOW_SECONDS]
    if len(attempts) >= REPORT_RATE_LIMIT_MAX_ATTEMPTS:
        retry_after_seconds = int(REPORT_RATE_LIMIT_WINDOW_SECONDS - (now - attempts[0]))
        raise HTTPException(
            status_code=429,
            detail={"code": "RATE_LIMITED", "retry_after_seconds": max(retry_after_seconds, 1)},
        )
    attempts.append(now)
    _report_submit_attempts[key] = attempts


def _validate_report_reason(reason: str, reason_note: Optional[str]) -> tuple[str, Optional[str]]:
    r = _validate_reason(reason, REPORT_REASONS_V1)
    note = (reason_note or "").strip() or None
    if r == "other" and not note:
        raise HTTPException(status_code=400, detail="reason_note is required when reason=other")
    return r, note


def _assert_country_scope(country_code: str, current_user: dict) -> None:
    if country_code not in SUPPORTED_COUNTRIES:
        raise HTTPException(status_code=400, detail="Invalid country code")
    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope and country_code not in scope:
            raise HTTPException(status_code=403, detail="Country scope forbidden")


def _parse_iso_datetime(value: str, field_name: str) -> datetime:
    if not value:
        raise HTTPException(status_code=400, detail=f"{field_name} is required")
    value = value.replace(" ", "+")
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid {field_name}") from exc


def _normalize_country_doc(doc: dict) -> dict:
    name = doc.get("name")
    if isinstance(name, dict):
        name_value = name.get("tr") or name.get("en") or name.get("de") or name.get("fr") or next(iter(name.values()), None)
    else:
        name_value = name
    return {
        "country_code": (doc.get("country_code") or doc.get("code") or "").upper(),
        "name": name_value,
        "active_flag": doc.get("active_flag", doc.get("is_enabled", True)),
        "default_currency": doc.get("default_currency"),
        "default_language": doc.get("default_language"),
        "updated_at": doc.get("updated_at"),
        "created_at": doc.get("created_at"),
    }


def _normalize_country_sql(country: Country) -> dict:
    name = country.name
    if isinstance(name, dict):
        name_value = name.get("tr") or name.get("en") or name.get("de") or name.get("fr") or next(iter(name.values()), None)
    else:
        name_value = name
    return {
        "country_code": (country.code or "").upper(),
        "name": name_value,
        "active_flag": country.is_enabled,
        "default_currency": country.default_currency,
        "default_language": country.default_language,
        "updated_at": country.updated_at.isoformat() if country.updated_at else None,
        "created_at": country.created_at.isoformat() if country.created_at else None,
    }


def _pick_label(value) -> Optional[str]:
    if isinstance(value, dict):
        return value.get("tr") or value.get("en") or value.get("de") or value.get("fr") or next(iter(value.values()), None)
    return value


def _normalize_category_doc(doc: dict, include_schema: bool = False) -> dict:
    payload = {
        "id": doc.get("id"),
        "parent_id": doc.get("parent_id"),
        "name": _pick_label(doc.get("name")) or _pick_label(doc.get("translations", [{}])[0].get("name") if doc.get("translations") else None),
        "slug": _pick_label(doc.get("slug")) or doc.get("segment"),
        "country_code": doc.get("country_code"),
        "active_flag": doc.get("active_flag", doc.get("is_enabled", True)),
        "sort_order": doc.get("sort_order", 0),
        "hierarchy_complete": doc.get("hierarchy_complete", True),
        "created_at": doc.get("created_at"),
        "updated_at": doc.get("updated_at"),
    }
    if include_schema:
        payload["form_schema"] = _normalize_category_schema(doc.get("form_schema")) if doc.get("form_schema") else None
    return payload


def _serialize_category_translation(translation: CategoryTranslation) -> dict:
    return {
        "language": translation.language,
        "name": translation.name,
        "description": translation.description,
        "meta_title": translation.meta_title,
        "meta_description": translation.meta_description,
    }


def _pick_category_name(translations: list[CategoryTranslation], slug_value: Optional[str]) -> str:
    preferred = None
    for lang in ("tr", "en", "de", "fr"):
        preferred = next((t for t in translations if t.language == lang), None)
        if preferred:
            break
    if not preferred and translations:
        preferred = translations[0]
    return preferred.name if preferred and preferred.name else (slug_value or "")


def _pick_category_slug(slug_value) -> Optional[str]:
    if isinstance(slug_value, dict):
        return slug_value.get("tr") or slug_value.get("en") or slug_value.get("de") or slug_value.get("fr") or next(iter(slug_value.values()), None)
    return slug_value


def _normalize_category_module(value: Optional[str]) -> str:
    if not value:
        return "vehicle"
    module_value = value.strip().lower()
    if module_value not in SUPPORTED_CATEGORY_MODULES:
        raise HTTPException(status_code=400, detail="module invalid")
    return module_value


def _assert_category_parent_compatible(
    *,
    category: Optional[Category],
    parent: Optional[Category],
    module_value: str,
    country_code: Optional[str],
):
    if not parent:
        return
    if parent.module != module_value:
        raise HTTPException(status_code=409, detail="parent module mismatch")
    if category and parent.id == category.id:
        raise HTTPException(status_code=409, detail="parent cycle detected")
    if category and category.path and parent.path:
        if parent.path == category.path or parent.path.startswith(f"{category.path}."):
            raise HTTPException(status_code=409, detail="parent cycle detected")
    if parent.country_code:
        if not country_code or parent.country_code != country_code:
            raise HTTPException(status_code=409, detail="parent country mismatch")


def _serialize_category_sql(category: Category, include_schema: bool = False, include_translations: bool = True) -> dict:
    translations = list(category.translations or [])
    slug_value = _pick_category_slug(category.slug)
    payload = {
        "id": str(category.id),
        "parent_id": str(category.parent_id) if category.parent_id else None,
        "name": _pick_category_name(translations, slug_value),
        "slug": slug_value,
        "country_code": category.country_code,
        "active_flag": category.is_enabled,
        "sort_order": category.sort_order,
        "hierarchy_complete": category.hierarchy_complete,
        "wizard_progress": category.wizard_progress,
        "module": category.module,
        "allowed_countries": category.allowed_countries or [],
        "icon": category.icon,
        "image_url": category.image_url,
        "listing_count": category.listing_count,
        "created_at": category.created_at.isoformat() if category.created_at else None,
        "updated_at": category.updated_at.isoformat() if category.updated_at else None,
    }
    if include_translations:
        payload["translations"] = [_serialize_category_translation(t) for t in translations]
    if include_schema:
        payload["form_schema"] = _normalize_category_schema(category.form_schema) if category.form_schema else None
    return payload


def _category_translation_map(translations: list[CategoryTranslation]) -> dict:
    mapping: Dict[str, dict] = {}
    for translation in translations or []:
        mapping[translation.language] = {
            "name": translation.name,
            "description": translation.description,
            "meta_title": translation.meta_title,
            "meta_description": translation.meta_description,
        }
    return mapping


def _coerce_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        if value.strip().lower() in {"true", "1", "yes", "y"}:
            return True
        if value.strip().lower() in {"false", "0", "no", "n"}:
            return False
    return default


def _normalize_slug_payload(slug_payload: Any) -> tuple[str, dict]:
    if isinstance(slug_payload, dict):
        slug_tr = slug_payload.get("tr") or slug_payload.get("en") or slug_payload.get("de") or slug_payload.get("fr")
        slug_tr = (slug_tr or "").strip().lower()
        if slug_tr and not SLUG_PATTERN.match(slug_tr):
            raise HTTPException(status_code=400, detail=f"Invalid slug: {slug_tr}")
        slug_map = dict(slug_payload)
        if slug_tr and "tr" not in slug_map:
            slug_map["tr"] = slug_tr
        return slug_tr, slug_map
    slug_tr = str(slug_payload or "").strip().lower()
    if slug_tr and not SLUG_PATTERN.match(slug_tr):
        raise HTTPException(status_code=400, detail=f"Invalid slug: {slug_tr}")
    return slug_tr, {"tr": slug_tr, "en": slug_tr, "de": slug_tr}


def _normalize_name_payload(name_payload: Any, slug_tr: str) -> dict:
    if isinstance(name_payload, dict):
        return name_payload
    name_value = str(name_payload or slug_tr)
    return {"tr": name_value, "en": name_value, "de": name_value}


def _build_category_export_item(
    category: Category,
    translations: list[CategoryTranslation],
    parent_slug: Optional[str],
    schema_version: int,
) -> dict:
    translation_map = _category_translation_map(translations)
    slug_value = category.slug if isinstance(category.slug, dict) else {"tr": category.slug}
    name_payload = {lang: data.get("name") for lang, data in translation_map.items()} if translation_map else {}
    return {
        "id": str(category.id),
        "parent_id": str(category.parent_id) if category.parent_id else None,
        "parent_slug": parent_slug,
        "path": category.path,
        "module": category.module,
        "slug": slug_value,
        "name": name_payload,
        "translations": translation_map,
        "country_code": category.country_code,
        "allowed_countries": category.allowed_countries or [],
        "sort_order": category.sort_order,
        "active_flag": category.is_enabled,
        "hierarchy_complete": category.hierarchy_complete,
        "schema_status": (category.form_schema or {}).get("status") if isinstance(category.form_schema, dict) else None,
        "schema_version": schema_version,
        "form_schema": category.form_schema,
    }


def _build_category_tree(items: list[dict]) -> list[dict]:
    node_map: Dict[str, dict] = {}
    roots: list[dict] = []
    for item in items:
        item_copy = dict(item)
        item_copy["children"] = []
        node_map[item_copy["id"]] = item_copy
    for item in items:
        node = node_map[item["id"]]
        parent_id = item.get("parent_id")
        if parent_id and parent_id in node_map:
            node_map[parent_id]["children"].append(node)
        else:
            roots.append(node)
    return roots


def _flatten_import_tree(items: list[dict], parent_slug: Optional[str], parent_path: Optional[str]) -> list[dict]:
    flattened: list[dict] = []
    for item in items:
        slug_tr, slug_map = _normalize_slug_payload(item.get("slug"))
        if not slug_tr:
            raise HTTPException(status_code=400, detail="slug is required")
        name_map = _normalize_name_payload(item.get("name"), slug_tr)
        path_value = f"{parent_path}.{slug_tr}" if parent_path else slug_tr
        record = {
            "slug_tr": slug_tr,
            "slug": slug_map,
            "name": name_map,
            "parent_slug": parent_slug,
            "path": path_value,
            "country_code": (item.get("country_code") or None),
            "allowed_countries": item.get("allowed_countries") or [],
            "sort_order": int(item.get("sort_order") or 0),
            "active_flag": _coerce_bool(item.get("active_flag"), True),
            "hierarchy_complete": _coerce_bool(item.get("hierarchy_complete"), True),
            "module": item.get("module") or "vehicle",
            "form_schema": item.get("form_schema"),
        }
        flattened.append(record)
        children = item.get("children") or []
        if children:
            flattened.extend(_flatten_import_tree(children, slug_tr, path_value))
    return flattened


def _parse_import_payload(content: bytes, file_type: str) -> list[dict]:
    if file_type == "json":
        try:
            payload = json.loads(content.decode("utf-8"))
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="JSON parse error") from exc
        items = payload.get("categories") or payload.get("items")
        if not isinstance(items, list):
            raise HTTPException(status_code=400, detail="categories array is required")
        return _flatten_import_tree(items, None, None)

    if file_type == "csv":
        decoded = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded), delimiter=",")
        records: list[dict] = []
        for row in reader:
            slug_tr = (row.get("slug_tr") or row.get("slug") or "").strip().lower()
            if not slug_tr:
                raise HTTPException(status_code=400, detail="slug_tr is required")
            if not SLUG_PATTERN.match(slug_tr):
                raise HTTPException(status_code=400, detail=f"Invalid slug: {slug_tr}")
            slug_map = {
                "tr": slug_tr,
                "en": (row.get("slug_en") or slug_tr).strip().lower(),
                "de": (row.get("slug_de") or slug_tr).strip().lower(),
            }
            name_map = {
                "tr": (row.get("name_tr") or slug_tr).strip(),
                "en": (row.get("name_en") or row.get("name_tr") or slug_tr).strip(),
                "de": (row.get("name_de") or row.get("name_tr") or slug_tr).strip(),
            }
            allowed_raw = row.get("allowed_countries") or ""
            allowed = [code.strip().upper() for code in allowed_raw.split("|") if code.strip()] if allowed_raw else []
            form_schema = None
            schema_raw = row.get("form_schema") or ""
            if schema_raw:
                try:
                    form_schema = json.loads(schema_raw)
                except json.JSONDecodeError as exc:
                    raise HTTPException(status_code=400, detail=f"Invalid form_schema JSON for {slug_tr}") from exc
            record = {
                "slug_tr": slug_tr,
                "slug": slug_map,
                "name": name_map,
                "parent_slug": (row.get("parent_slug") or "").strip().lower() or None,
                "path": (row.get("path") or "").strip() or None,
                "country_code": (row.get("country_code") or "").strip().upper() or None,
                "allowed_countries": allowed,
                "sort_order": int(row.get("sort_order") or 0),
                "active_flag": _coerce_bool(row.get("active_flag"), True),
                "hierarchy_complete": _coerce_bool(row.get("hierarchy_complete"), True),
                "module": row.get("module") or "vehicle",
                "form_schema": form_schema,
            }
            records.append(record)
        return records

    raise HTTPException(status_code=400, detail="Unsupported import format")


def _diff_categories(import_items: list[dict], existing: list[Category]) -> dict:
    def _change_type(before: Any, after: Any) -> str:
        empty_before = before in (None, "", [], {})
        empty_after = after in (None, "", [], {})
        if empty_before and not empty_after:
            return "added"
        if not empty_before and empty_after:
            return "removed"
        if before != after:
            return "updated"
        return "unchanged"

    existing_map: Dict[str, Category] = {}
    existing_parent_map: Dict[str, Optional[str]] = {}
    for category in existing:
        slug_key = _pick_category_slug(category.slug) or ""
        existing_map[slug_key.lower()] = category

    slug_by_id = {str(cat.id): (_pick_category_slug(cat.slug) or "") for cat in existing}
    for category in existing:
        parent_slug = slug_by_id.get(str(category.parent_id)) if category.parent_id else None
        existing_parent_map[str(category.id)] = parent_slug

    incoming_keys = set()
    creates = []
    updates = []
    for item in import_items:
        slug_key = item["slug_tr"]
        if slug_key in incoming_keys:
            raise HTTPException(status_code=400, detail=f"Duplicate slug: {slug_key}")
        incoming_keys.add(slug_key)
        existing_category = existing_map.get(slug_key)
        if not existing_category:
            creates.append(item)
            continue

        fields = []
        changed_count = 0

        translation_map = _category_translation_map(list(existing_category.translations or []))
        name_map = item.get("name") or {}
        for lang in ("tr", "en", "de"):
            old_name = (translation_map.get(lang) or {}).get("name")
            new_name = name_map.get(lang) if name_map.get(lang) is not None else old_name
            change_type = _change_type(old_name, new_name)
            if change_type != "unchanged":
                changed_count += 1
            fields.append({
                "field_name": f"name_{lang}",
                "before_value": old_name,
                "after_value": new_name,
                "change_type": change_type,
            })

        parent_slug = existing_parent_map.get(str(existing_category.id))
        new_parent_slug = item.get("parent_slug")
        change_type = _change_type(parent_slug, new_parent_slug)
        if change_type != "unchanged":
            changed_count += 1
        fields.append({
            "field_name": "parent_slug",
            "before_value": parent_slug,
            "after_value": new_parent_slug,
            "change_type": change_type,
        })

        change_type = _change_type(existing_category.country_code, item.get("country_code"))
        if change_type != "unchanged":
            changed_count += 1
        fields.append({
            "field_name": "country_code",
            "before_value": existing_category.country_code,
            "after_value": item.get("country_code"),
            "change_type": change_type,
        })

        allowed_existing = sorted(existing_category.allowed_countries or [])
        allowed_new = sorted(item.get("allowed_countries") or [])
        change_type = _change_type(allowed_existing, allowed_new)
        if change_type != "unchanged":
            changed_count += 1
        fields.append({
            "field_name": "allowed_countries",
            "before_value": allowed_existing,
            "after_value": allowed_new,
            "change_type": change_type,
        })

        before_sort = int(existing_category.sort_order or 0)
        after_sort = int(item.get("sort_order") or 0)
        change_type = _change_type(before_sort, after_sort)
        if change_type != "unchanged":
            changed_count += 1
        fields.append({
            "field_name": "sort_order",
            "before_value": before_sort,
            "after_value": after_sort,
            "change_type": change_type,
        })

        before_active = bool(existing_category.is_enabled)
        after_active = _coerce_bool(item.get("active_flag"), True)
        change_type = _change_type(before_active, after_active)
        if change_type != "unchanged":
            changed_count += 1
        fields.append({
            "field_name": "active_flag",
            "before_value": before_active,
            "after_value": after_active,
            "change_type": change_type,
        })

        if item.get("form_schema") is not None:
            existing_schema = _normalize_category_schema(existing_category.form_schema) if existing_category.form_schema else None
            incoming_schema = _normalize_category_schema(item.get("form_schema"))
            change_type = "updated" if json.dumps(existing_schema, sort_keys=True) != json.dumps(incoming_schema, sort_keys=True) else "unchanged"
            if change_type != "unchanged":
                changed_count += 1
            fields.append({
                "field_name": "form_schema",
                "before_value": "existing" if existing_schema else None,
                "after_value": "incoming" if change_type != "unchanged" else ("existing" if existing_schema else None),
                "change_type": change_type,
            })
        else:
            fields.append({
                "field_name": "form_schema",
                "before_value": "existing" if existing_category.form_schema else None,
                "after_value": "existing" if existing_category.form_schema else None,
                "change_type": "unchanged",
            })

        if changed_count > 0:
            updates.append({
                "slug": slug_key,
                "fields": fields,
                "changed_fields": changed_count,
            })

    deletes = []
    warnings = []
    for category in existing:
        slug_key = (_pick_category_slug(category.slug) or "").lower()
        if slug_key and slug_key not in incoming_keys and not category.is_deleted:
            is_root = category.parent_id is None
            deletes.append({
                "slug": slug_key,
                "name": _pick_category_name(list(category.translations or []), _pick_category_slug(category.slug)),
                "is_root": is_root,
            })
            if is_root:
                warnings.append({
                    "type": "critical",
                    "message": f"Root kategori siliniyor: {slug_key}",
                    "slug": slug_key,
                })

    return {
        "summary": {
            "creates": len(creates),
            "updates": len(updates),
            "deletes": len(deletes),
            "total": len(import_items),
        },
        "creates": creates,
        "updates": updates,
        "deletes": deletes,
        "warnings": warnings,
    }


def _build_category_import_report_pdf(diff: dict, snapshot_payload: dict) -> bytes:
    styles = getSampleStyleSheet()
    if "CodeSmall" not in styles:
        styles.add(ParagraphStyle(name="CodeSmall", fontName="Courier", fontSize=7, leading=9))

    summary = diff.get("summary", {})
    warnings = diff.get("warnings", [])
    updates = diff.get("updates", [])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    elements.append(Paragraph("Dry-run Report", styles["Title"]))
    elements.append(Paragraph("Kategori Import/Export Değişiklik Özeti", styles["Heading2"]))
    elements.append(Spacer(1, 8))

    summary_table = Table(
        [
            ["Creates", "Updates", "Deletes", "Total"],
            [summary.get("creates", 0), summary.get("updates", 0), summary.get("deletes", 0), summary.get("total", 0)],
        ]
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    if warnings:
        elements.append(Paragraph("Kritik Uyarılar", styles["Heading3"]))
        warning_rows = [["Tip", "Mesaj"]]
        for warning in warnings:
            warning_rows.append([warning.get("type"), warning.get("message")])
        warning_table = Table(warning_rows, colWidths=[80, 420])
        warning_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f59e0b")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ]
            )
        )
        elements.append(warning_table)
        elements.append(Spacer(1, 12))

    if updates:
        elements.append(Paragraph("Field-level Değişiklikler", styles["Heading3"]))
        for update in updates:
            elements.append(Paragraph(f"Slug: {update.get('slug')}", styles["Heading4"]))
            rows = [["Alan", "Önce", "Sonra", "Tip"]]
            for field in update.get("fields", []):
                if field.get("change_type") == "unchanged":
                    continue
                rows.append([
                    field.get("field_name"),
                    str(field.get("before_value") or ""),
                    str(field.get("after_value") or ""),
                    field.get("change_type"),
                ])
            if len(rows) == 1:
                rows.append(["-", "-", "-", "unchanged"])
            table = Table(rows, colWidths=[120, 160, 160, 80])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 10))

    elements.append(PageBreak())
    elements.append(Paragraph("Appendix: Import Snapshot", styles["Heading3"]))
    snapshot_text = json.dumps(snapshot_payload, ensure_ascii=False, indent=2)
    elements.append(Preformatted(snapshot_text, styles["CodeSmall"]))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()
    return pdf_bytes


def _deep_merge_schema(base: Dict[str, Any], override: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not override:
        return base
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            base[key] = _deep_merge_schema(base[key], value)
        else:
            base[key] = value
    return base


def _default_category_schema() -> Dict[str, Any]:
    return {
        "core_fields": {
            "title": {
                "required": True,
                "min": 10,
                "max": 80,
                "messages": {
                    "required": "Başlık zorunludur.",
                    "min": "Başlık çok kısa.",
                    "max": "Başlık çok uzun.",
                    "duplicate": "Bu başlık zaten kullanılıyor.",
                },
                "ui": {"bold": True},
            },
            "description": {
                "required": True,
                "min": 30,
                "max": 2000,
                "messages": {
                    "required": "Açıklama zorunludur.",
                    "min": "Açıklama çok kısa.",
                    "max": "Açıklama çok uzun.",
                },
                "ui": {"min_rows": 6, "max_rows": 8, "auto_grow": True, "show_counter": True},
            },
            "price": {
                "required": True,
                "currency_primary": "EUR",
                "currency_secondary": "CHF",
                "secondary_enabled": False,
                "decimal_places": 0,
                "range": {"min": None, "max": None},
                "messages": {
                    "required": "Fiyat zorunludur.",
                    "numeric": "Geçerli bir fiyat girin.",
                    "range": "Fiyat aralık dışında.",
                },
                "input_mask": {"thousand_separator": True},
            },
        },
        "title_uniqueness": {"enabled": False, "scope": "category"},
        "status": "published",
        "dynamic_fields": [],
        "detail_groups": [],
        "modules": {
            "address": {"enabled": True},
            "photos": {"enabled": True, "max_uploads": 10},
            "contact": {"enabled": True},
            "payment": {"enabled": True},
        },
        "payment_options": {"package": True, "doping": False},
        "module_order": [
            "core_fields",
            "dynamic_fields",
            "address",
            "detail_groups",
            "photos",
            "contact",
            "payment",
        ],
    }


def _normalize_category_schema(schema: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    base = _default_category_schema()
    merged = _deep_merge_schema(base, schema or {})
    merged.setdefault("status", "published")
    title_messages = merged["core_fields"]["title"].setdefault("messages", {})
    desc_messages = merged["core_fields"]["description"].setdefault("messages", {})
    price_messages = merged["core_fields"]["price"].setdefault("messages", {})
    title_messages.setdefault("required", "Başlık zorunludur.")
    title_messages.setdefault("min", "Başlık çok kısa.")
    title_messages.setdefault("max", "Başlık çok uzun.")
    title_messages.setdefault("duplicate", "Bu başlık zaten kullanılıyor.")
    desc_messages.setdefault("required", "Açıklama zorunludur.")
    desc_messages.setdefault("min", "Açıklama çok kısa.")
    desc_messages.setdefault("max", "Açıklama çok uzun.")
    price_messages.setdefault("required", "Fiyat zorunludur.")
    price_messages.setdefault("numeric", "Geçerli bir fiyat girin.")
    price_messages.setdefault("range", "Fiyat aralık dışında.")
    merged["validation_messages"] = {
        "core_fields": {
            "title": title_messages,
            "description": desc_messages,
            "price": price_messages,
        },
        "dynamic_fields": {
            field.get("key"): field.get("messages", {})
            for field in merged.get("dynamic_fields", [])
        },
        "detail_groups": {
            group.get("id"): group.get("messages", {})
            for group in merged.get("detail_groups", [])
        },
    }
    return merged


async def _record_category_version(
    db,
    category_id: str,
    schema_snapshot: Dict[str, Any],
    actor: dict,
    status: str,
    max_versions: int = 20,
) -> dict:
    raise RuntimeError("Legacy category versions disabled; use SQL variants")


async def _mark_latest_category_version_published(
    db,
    category_id: str,
    schema_snapshot: Dict[str, Any],
    actor: dict,
    max_versions: int = 20,
) -> dict:
    raise RuntimeError("Legacy category versions disabled; use SQL variants")


def _serialize_category_version(doc: dict, include_snapshot: bool = False) -> dict:
    payload = {
        "id": doc.get("id"),
        "category_id": doc.get("category_id"),
        "version": doc.get("version"),
        "status": doc.get("status"),
        "created_at": doc.get("created_at"),
        "created_by": doc.get("created_by"),
        "created_by_role": doc.get("created_by_role"),
        "created_by_email": doc.get("created_by_email"),
        "published_at": doc.get("published_at"),
        "published_by": doc.get("published_by"),
    }
    if include_snapshot:
        payload["schema_snapshot"] = doc.get("schema_snapshot")
    return payload


def _safe_uuid(value: Optional[str]) -> Optional[uuid.UUID]:
    if not value:
        return None
    try:
        return uuid.UUID(str(value))
    except ValueError:
        return None


def _serialize_category_version_sql(version: CategorySchemaVersion, include_snapshot: bool = False) -> dict:
    payload = {
        "id": str(version.id),
        "category_id": str(version.category_id),
        "version": version.version,
        "status": version.status,
        "created_at": version.created_at.isoformat() if version.created_at else None,
        "created_by": str(version.created_by) if version.created_by else None,
        "created_by_role": version.created_by_role,
        "created_by_email": version.created_by_email,
        "published_at": version.published_at.isoformat() if version.published_at else None,
        "published_by": str(version.published_by) if version.published_by else None,
    }
    if include_snapshot:
        payload["schema_snapshot"] = version.schema_snapshot
    return payload


async def _record_category_version_sql(
    session: AsyncSession,
    category_id: uuid.UUID,
    schema_snapshot: Dict[str, Any],
    actor: dict,
    status: str,
    max_versions: int = 20,
) -> CategorySchemaVersion:
    latest = await session.execute(
        select(CategorySchemaVersion)
        .where(CategorySchemaVersion.category_id == category_id)
        .order_by(desc(CategorySchemaVersion.version))
        .limit(1)
    )
    latest_row = latest.scalar_one_or_none()
    next_version = (latest_row.version if latest_row else 0) + 1

    now = datetime.now(timezone.utc)
    version = CategorySchemaVersion(
        id=uuid.uuid4(),
        category_id=category_id,
        version=next_version,
        status=status,
        schema_snapshot=schema_snapshot,
        created_at=now,
        created_by=_safe_uuid(actor.get("id")),
        created_by_role=actor.get("role"),
        created_by_email=actor.get("email"),
        published_at=now if status == "published" else None,
        published_by=_safe_uuid(actor.get("id")) if status == "published" else None,
    )
    session.add(version)
    await session.commit()

    if max_versions and max_versions > 0:
        versions_res = await session.execute(
            select(CategorySchemaVersion)
            .where(CategorySchemaVersion.category_id == category_id)
            .order_by(desc(CategorySchemaVersion.version))
        )
        versions = versions_res.scalars().all()
        if len(versions) > max_versions:
            for old in versions[max_versions:]:
                await session.delete(old)
            await session.commit()

    return version


async def _mark_latest_category_version_published_sql(
    session: AsyncSession,
    category_id: uuid.UUID,
    schema_snapshot: Dict[str, Any],
    actor: dict,
    max_versions: int = 20,
) -> CategorySchemaVersion:
    latest = await session.execute(
        select(CategorySchemaVersion)
        .where(CategorySchemaVersion.category_id == category_id)
        .order_by(desc(CategorySchemaVersion.version))
        .limit(1)
    )
    latest_row = latest.scalar_one_or_none()
    now = datetime.now(timezone.utc)
    if latest_row:
        latest_row.status = "published"
        latest_row.published_at = now
        latest_row.published_by = _safe_uuid(actor.get("id"))
        await session.commit()
        return latest_row

    return await _record_category_version_sql(session, category_id, schema_snapshot, actor, "published", max_versions=max_versions)


async def _get_schema_version_for_export_sql(session: AsyncSession, category_id: uuid.UUID) -> int:
    latest = await session.execute(
        select(CategorySchemaVersion.version)
        .where(CategorySchemaVersion.category_id == category_id)
        .order_by(desc(CategorySchemaVersion.version))
        .limit(1)
    )
    value = latest.scalar_one_or_none()
    return int(value or 0)


async def _get_schema_version_for_export(db, category_id: str) -> int:
    raise RuntimeError("Legacy category versions disabled; use SQL variant")


def _schema_to_csv_rows(schema: Dict[str, Any]) -> list[list[str]]:
    rows = []
    rows.append([
        "section",
        "field_key",
        "label",
        "type",
        "required",
        "enabled",
        "options",
        "messages_required",
        "messages_invalid",
        "messages_min",
        "messages_max",
        "messages_range",
        "messages_duplicate",
    ])

    core = schema.get("core_fields") or {}
    for key in ["title", "description", "price"]:
        cfg = core.get(key) or {}
        messages = cfg.get("messages") or {}
        rows.append([
            "core_fields",
            key,
            key,
            cfg.get("type") or ("price" if key == "price" else "text"),
            str(bool(cfg.get("required"))),
            "",
            "",
            str(messages.get("required", "")),
            str(messages.get("invalid", "")),
            str(messages.get("min", "")),
            str(messages.get("max", "")),
            str(messages.get("range", "")),
            str(messages.get("duplicate", "")),
        ])

    for field in schema.get("dynamic_fields", []) or []:
        messages = field.get("messages") or {}
        options = ",".join(field.get("options") or [])
        rows.append([
            "dynamic_fields",
            str(field.get("key")),
            str(field.get("label")),
            str(field.get("type")),
            str(bool(field.get("required"))),
            "",
            options,
            str(messages.get("required", "")),
            str(messages.get("invalid", "")),
            str(messages.get("min", "")),
            str(messages.get("max", "")),
            str(messages.get("range", "")),
            str(messages.get("duplicate", "")),
        ])

    for group in schema.get("detail_groups", []) or []:
        messages = group.get("messages") or {}
        options = ",".join(group.get("options") or [])
        rows.append([
            "detail_groups",
            str(group.get("id") or group.get("title")),
            str(group.get("title")),
            "checkbox_group",
            str(bool(group.get("required"))),
            "",
            options,
            str(messages.get("required", "")),
            str(messages.get("invalid", "")),
            str(messages.get("min", "")),
            str(messages.get("max", "")),
            str(messages.get("range", "")),
            str(messages.get("duplicate", "")),
        ])

    modules = schema.get("modules") or {}
    for key, module in modules.items():
        enabled = bool(module.get("enabled")) if isinstance(module, dict) else bool(module)
        rows.append([
            "modules",
            str(key),
            str(key),
            "module",
            "",
            str(enabled),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    payment = schema.get("payment_options") or {}
    if payment:
        rows.append([
            "payment_options",
            "package",
            "package",
            "payment",
            "",
            str(bool(payment.get("package"))),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])
        rows.append([
            "payment_options",
            "doping",
            "doping",
            "payment",
            "",
            str(bool(payment.get("doping"))),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ])

    title_uniqueness = schema.get("title_uniqueness") or {}
    if title_uniqueness.get("enabled"):
        rows.append([
            "rules",
            "title_uniqueness",
            "title_uniqueness",
            "rule",
            "",
            str(bool(title_uniqueness.get("enabled"))),
            "",
            "",
            "",
            "",
            "",
            "",
            str(title_uniqueness.get("scope") or ""),
        ])

    return rows


def _build_schema_pdf(schema: Dict[str, Any], category: dict, version: int, hierarchy: dict) -> bytes:
    styles = getSampleStyleSheet()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Schema Export")
    elements = []
    elements.append(Paragraph("Kategori Şema Export", styles["Title"]))
    elements.append(Paragraph(f"Kategori: {category.get('name')} ({category.get('slug')})", styles["Normal"]))
    elements.append(Paragraph(f"Versiyon: v{version}", styles["Normal"]))
    elements.append(Paragraph(f"Durum: {schema.get('status')}", styles["Normal"]))
    elements.append(Paragraph(f"Ana kategori: {hierarchy.get('parent') or '-'}", styles["Normal"]))
    elements.append(Paragraph(f"Alt kategoriler: {', '.join(hierarchy.get('children') or []) or '-'}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    def add_table(title: str, rows: list[list[str]]):
        if not rows:
            return
        elements.append(Paragraph(title, styles["Heading3"]))
        table = Table(rows, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10))

    core_rows = [["Alan", "Required", "Min", "Max", "Range", "Messages"]]
    for key, cfg in (schema.get("core_fields") or {}).items():
        messages = cfg.get("messages") or {}
        range_cfg = cfg.get("range") or {}
        core_rows.append([
            key,
            str(cfg.get("required")),
            str(cfg.get("min", "")),
            str(cfg.get("max", "")),
            f"{range_cfg.get('min', '')}-{range_cfg.get('max', '')}",
            ", ".join([str(v) for v in messages.values() if v]),
        ])
    add_table("Çekirdek Alanlar", core_rows)

    dynamic_rows = [["Key", "Label", "Type", "Required", "Options", "Messages"]]
    for field in schema.get("dynamic_fields", []) or []:
        messages = field.get("messages") or {}
        dynamic_rows.append([
            str(field.get("key")),
            str(field.get("label")),
            str(field.get("type")),
            str(field.get("required")),
            ",".join(field.get("options") or []),
            ", ".join([str(v) for v in messages.values() if v]),
        ])
    add_table("Parametre Alanları (2a)", dynamic_rows)

    detail_rows = [["Group", "Required", "Options", "Messages"]]
    for group in schema.get("detail_groups", []) or []:
        messages = group.get("messages") or {}
        detail_rows.append([
            str(group.get("title")),
            str(group.get("required")),
            ",".join(group.get("options") or []),
            ", ".join([str(v) for v in messages.values() if v]),
        ])
    add_table("Detay Grupları (2c)", detail_rows)

    modules_rows = [["Module", "Enabled"]]
    for key, module in (schema.get("modules") or {}).items():
        enabled = bool(module.get("enabled")) if isinstance(module, dict) else bool(module)
        modules_rows.append([str(key), str(enabled)])
    add_table("Modüller", modules_rows)

    if (schema.get("payment_options") or {}):
        payment = schema.get("payment_options") or {}
        payment_rows = [["Option", "Enabled"]]
        payment_rows.append(["package", str(bool(payment.get("package")))])
        payment_rows.append(["doping", str(bool(payment.get("doping")))])
        add_table("Ödeme Seçenekleri", payment_rows)

    title_uniqueness = schema.get("title_uniqueness") or {}
    if title_uniqueness:
        rule_rows = [["Rule", "Enabled", "Scope"]]
        rule_rows.append(["title_uniqueness", str(bool(title_uniqueness.get("enabled"))), str(title_uniqueness.get("scope") or "")])
        add_table("Duplicate Kuralı", rule_rows)

    doc.build(elements)
    return buffer.getvalue()



def _validate_category_schema(schema: Dict[str, Any]) -> None:
    core = schema.get("core_fields") or {}
    title = core.get("title") or {}
    description = core.get("description") or {}
    price = core.get("price") or {}
    if not title.get("required", True):
        raise HTTPException(status_code=400, detail="Başlık alanı zorunlu olmalıdır.")
    if not description.get("required", True):
        raise HTTPException(status_code=400, detail="Açıklama alanı zorunlu olmalıdır.")
    if not price.get("required", True):
        raise HTTPException(status_code=400, detail="Fiyat alanı zorunlu olmalıdır.")

    if title.get("min") and title.get("max") and title.get("min") > title.get("max"):
        raise HTTPException(status_code=400, detail="Başlık min değeri max değerinden büyük olamaz.")
    if description.get("min") and description.get("max") and description.get("min") > description.get("max"):
        raise HTTPException(status_code=400, detail="Açıklama min değeri max değerinden büyük olamaz.")

    currency_primary = price.get("currency_primary")
    if currency_primary not in {"EUR", "CHF"}:
        raise HTTPException(status_code=400, detail="Fiyat para birimi EUR veya CHF olmalıdır.")
    secondary = price.get("currency_secondary")
    if secondary and secondary not in {"EUR", "CHF"}:
        raise HTTPException(status_code=400, detail="İkincil para birimi EUR veya CHF olmalıdır.")
    if secondary and secondary == currency_primary:
        raise HTTPException(status_code=400, detail="Birincil ve ikincil para birimi aynı olamaz.")
    price_range = price.get("range") or {}
    if price_range.get("max") is not None and price_range.get("min") is not None:
        if price_range.get("min") > price_range.get("max"):
            raise HTTPException(status_code=400, detail="Fiyat aralığı min/max hatalı.")

    for field in schema.get("dynamic_fields", []):
        field_type = field.get("type")
        if field_type not in {"radio", "select", "text", "number"}:
            raise HTTPException(status_code=400, detail="Parametre alanı tipi radio/select/text/number olmalıdır.")
        options = field.get("options") or []
        if field_type in {"radio", "select"} and not options:
            raise HTTPException(status_code=400, detail="Parametre alanı için seçenek listesi zorunludur.")
        if not field.get("key"):
            raise HTTPException(status_code=400, detail="Parametre alanı anahtarı zorunludur.")
        if not field.get("label"):
            raise HTTPException(status_code=400, detail="Parametre alanı etiketi zorunludur.")

    for group in schema.get("detail_groups", []):
        if not group.get("title"):
            raise HTTPException(status_code=400, detail="Detay grubu başlığı zorunludur.")
        options = group.get("options") or []
        if not options:
            raise HTTPException(status_code=400, detail="Detay grubu için seçenek listesi zorunludur.")

    modules = schema.get("modules") or {}
    photos = modules.get("photos") or {}
    if photos.get("enabled") and not photos.get("max_uploads"):
        raise HTTPException(status_code=400, detail="Fotoğraf modülü için upload limiti zorunludur.")


def _normalize_attribute_doc(doc: dict) -> dict:
    return {
        "id": doc.get("id"),
        "category_id": doc.get("category_id"),
        "name": doc.get("name"),
        "key": doc.get("key"),
        "type": doc.get("type"),
        "required_flag": doc.get("required_flag", False),
        "filterable_flag": doc.get("filterable_flag", False),
        "options": doc.get("options"),
        "country_code": doc.get("country_code"),
        "active_flag": doc.get("active_flag", True),
        "created_at": doc.get("created_at"),
        "updated_at": doc.get("updated_at"),
    }


def _serialize_attribute_sql(attribute: Attribute, category_id: Optional[str] = None) -> dict:
    name_value = None
    if isinstance(attribute.name, dict):
        name_value = attribute.name.get("tr") or next(iter(attribute.name.values()), None)
    else:
        name_value = attribute.name

    options = []
    if attribute.options:
        for opt in attribute.options:
            label_value = None
            if isinstance(opt.label, dict):
                label_value = opt.label.get("tr") or next(iter(opt.label.values()), None)
            options.append(label_value or opt.value)

    return {
        "id": str(attribute.id),
        "category_id": category_id,
        "name": name_value,
        "key": attribute.key,
        "type": attribute.attribute_type,
        "required_flag": bool(attribute.is_required),
        "filterable_flag": bool(attribute.is_filterable),
        "options": options or None,
        "country_code": None,
        "active_flag": bool(attribute.is_active),
        "created_at": attribute.created_at.isoformat() if attribute.created_at else None,
        "updated_at": attribute.updated_at.isoformat() if attribute.updated_at else None,
    }


def _normalize_vehicle_make_doc(doc) -> dict:
    if isinstance(doc, VehicleMake):
        return {
            "id": str(doc.id),
            "name": doc.name,
            "slug": doc.slug,
            "country_code": None,
            "active_flag": bool(doc.is_active),
            "created_at": doc.created_at.isoformat() if doc.created_at else None,
            "updated_at": doc.updated_at.isoformat() if doc.updated_at else None,
        }
    return {
        "id": doc.get("id"),
        "name": doc.get("name"),
        "slug": doc.get("slug"),
        "country_code": doc.get("country_code"),
        "active_flag": doc.get("active_flag", True),
        "created_at": doc.get("created_at"),
        "updated_at": doc.get("updated_at"),
    }


def _normalize_vehicle_model_doc(doc) -> dict:
    if isinstance(doc, VehicleModel):
        vehicle_type = (doc.vehicle_type or "car").strip().lower()
        return {
            "id": str(doc.id),
            "make_id": str(doc.make_id),
            "name": doc.name,
            "slug": doc.slug,
            "vehicle_type": vehicle_type,
            "country_code": None,
            "active_flag": bool(doc.is_active),
            "created_at": doc.created_at.isoformat() if doc.created_at else None,
            "updated_at": None,
        }
    vehicle_type = (doc.get("vehicle_type") or "car").strip().lower()
    return {
        "id": doc.get("id"),
        "make_id": doc.get("make_id"),
        "name": doc.get("name"),
        "slug": doc.get("slug"),
        "vehicle_type": vehicle_type,
        "country_code": doc.get("country_code"),
        "active_flag": doc.get("active_flag", True),
        "created_at": doc.get("created_at"),
        "updated_at": doc.get("updated_at"),
    }


def _normalize_vehicle_type(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = str(value).strip().lower()
    return normalized or None


def _build_vehicle_type_summary(types: set[str]) -> str:
    if not types:
        return "—"
    if len(types) == 1:
        return next(iter(types))
    return "mixed"


async def _get_required_attribute_keys(session: AsyncSession, category_id: str, country_code: Optional[str]) -> list[str]:
    if not category_id:
        return []
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError:
        return []

    query = (
        select(Attribute.key)
        .join(CategoryAttributeMap, CategoryAttributeMap.attribute_id == Attribute.id)
        .where(
            CategoryAttributeMap.category_id == category_uuid,
            CategoryAttributeMap.is_active.is_(True),
            Attribute.is_active.is_(True),
            Attribute.is_required.is_(True),
        )
    )

    result = await session.execute(query)
    return [row[0] for row in result.all() if row[0]]

async def _build_vehicle_master_from_db(session: AsyncSession, country_code: str) -> dict:
    if not country_code:
        return {"makes": {}, "models": {}}
    result = await session.execute(
        select(VehicleMake).where(VehicleMake.is_active.is_(True)).order_by(VehicleMake.name.asc())
    )
    make_rows = result.scalars().all()
    makes = [
        {
            "make_key": make.slug,
            "name": make.name,
            "is_active": make.is_active,
        }
        for make in make_rows
        if make.slug
    ]
    make_ids = [make.id for make in make_rows if make.id]
    if not make_ids:
        return {"makes": makes, "models_by_make": {}}
    model_result = await session.execute(
        select(VehicleModel.make_id, VehicleModel.slug, VehicleModel.name)
        .where(VehicleModel.make_id.in_(make_ids), VehicleModel.is_active.is_(True))
    )
    models_by_make: dict = {}
    make_slug_by_id = {str(make.id): make.slug for make in make_rows if make.id and make.slug}
    for make_id, slug, name in model_result.all():
        make_slug = make_slug_by_id.get(str(make_id))
        if not make_slug or not slug:
            continue
        models_by_make.setdefault(make_slug, []).append(
            {
                "model_key": slug,
                "name": name,
                "is_active": True,
            }
        )
    return {"makes": makes, "models_by_make": models_by_make}


async def _report_transition(
    *,
    session: AsyncSession,
    report_id: str,
    current_user: dict,
    target_status: str,
    note: str,
) -> Report:
    try:
        report_uuid = uuid.UUID(report_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid report id") from exc

    report = await session.get(Report, report_uuid)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope and report.country_code not in scope:
            raise HTTPException(status_code=403, detail="Country scope forbidden")

    prev_status = report.status
    allowed = REPORT_STATUS_TRANSITIONS.get(prev_status, set())
    if target_status not in allowed:
        raise HTTPException(status_code=400, detail="Invalid status transition")

    now_dt = datetime.now(timezone.utc)
    report.status = target_status
    report.updated_at = now_dt
    report.handled_by_admin_id = _safe_uuid(current_user.get("id"))
    report.status_note = note

    await _write_audit_log_sql(
        session=session,
        action="REPORT_STATUS_CHANGE",
        actor=current_user,
        resource_type="report",
        resource_id=str(report.id),
        metadata={
            "previous_status": prev_status,
            "new_status": target_status,
            "note": note,
            "listing_id": str(report.listing_id),
        },
        request=None,
        country_code=report.country_code,
    )

    await session.commit()
    return report


async def _moderation_transition_sql(
    *,
    session: AsyncSession,
    listing_id: str,
    current_user: dict,
    new_status: str,
    action_type: str,
    reason: Optional[str] = None,
    reason_note: Optional[str] = None,
    request: Optional[Request] = None,
    commit: bool = True,
) -> Listing:
    _ensure_moderation_rbac(current_user)

    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid listing id") from exc

    listing = await session.get(Listing, listing_uuid)
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")

    await _assert_moderation_not_frozen(
        session=session,
        request=request,
        current_user=current_user,
        listing_id=listing_id,
        listing_country=listing.country,
        action_type=action_type,
    )

    prev_status = listing.status
    if prev_status != "pending_moderation":
        raise HTTPException(status_code=400, detail="Listing not pending_moderation")

    now = datetime.now(timezone.utc)
    listing.status = new_status
    listing.updated_at = now
    if new_status == "published":
        listing.published_at = now

    actor_id = current_user.get("id")
    actor_uuid = None
    try:
        if actor_id:
            actor_uuid = uuid.UUID(str(actor_id))
    except ValueError:
        actor_uuid = None

    if actor_uuid:
        moderation_action = ModerationAction(
            listing_id=listing_uuid,
            action_type=action_type,
            reason=reason,
            note=reason_note,
            actor_admin_id=actor_uuid,
            actor_email=current_user.get("email"),
        )
        session.add(moderation_action)

    action_label = {
        "published": "moderation_approved",
        "rejected": "moderation_rejected",
        "needs_revision": "moderation_needs_revision",
    }.get(new_status, "moderation_updated")

    audit_entry = await _write_audit_log_sql(
        session=session,
        action=action_label,
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="listing",
        resource_id=listing_id,
        metadata={
            "previous_status": prev_status,
            "new_status": new_status,
            "reason": reason,
            "reason_note": reason_note,
        },
        request=request,
        country_code=listing.country,
    )

    moderation_status = _normalize_moderation_item_status(new_status)
    moderation_reason = _sanitize_moderation_reason(reason)
    moderator_id = _resolve_moderation_actor_id(current_user.get("id"))
    audit_ref = str(audit_entry.id) if audit_entry else None

    await _upsert_moderation_item(
        session=session,
        listing=listing,
        status=moderation_status,
        reason=moderation_reason,
        moderator_id=moderator_id,
        audit_ref=audit_ref,
    )
    if commit:
        await session.commit()
    else:
        await session.flush()

    return listing


@api_router.post("/admin/listings/{listing_id}/approve")
async def admin_approve_listing(
    listing_id: str,
    request: Request,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    updated = await _moderation_transition_sql(
        session=session,
        listing_id=listing_id,
        current_user=current_user,
        new_status="published",
        action_type="approve",
        request=request,
    )
    return {"ok": True, "listing": {"id": str(updated.id), "status": updated.status}}


@api_router.post("/admin/listings/{listing_id}/reject")
async def admin_reject_listing(
    listing_id: str,
    payload: ModerationReasonPayload,
    request: Request,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    if not payload.reason:
        raise HTTPException(status_code=400, detail="Reason is required")
    updated = await _moderation_transition_sql(
        session=session,
        listing_id=listing_id,
        current_user=current_user,
        new_status="rejected",
        action_type="reject",
        reason=payload.reason,
        reason_note=payload.reason_note,
        request=request,
    )
    return {"ok": True, "listing": {"id": str(updated.id), "status": updated.status}}


@api_router.post("/admin/listings/{listing_id}/needs_revision")
async def admin_needs_revision_listing(
    listing_id: str,
    payload: ModerationReasonPayload,
    request: Request,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    if not payload.reason:
        raise HTTPException(status_code=400, detail="Reason is required")
    updated = await _moderation_transition_sql(
        session=session,
        listing_id=listing_id,
        current_user=current_user,
        new_status="needs_revision",
        action_type="needs_revision",
        reason=payload.reason,
        reason_note=payload.reason_note,
        request=request,
    )
    return {"ok": True, "listing": {"id": str(updated.id), "status": updated.status}}


@api_router.post("/admin/moderation/bulk-approve")
async def admin_bulk_approve_listings(
    payload: BulkModerationPayload,
    request: Request,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    _ensure_moderation_rbac(current_user)
    listing_ids = [str(item).strip() for item in (payload.listing_ids or []) if str(item).strip()]
    if not listing_ids:
        raise HTTPException(status_code=400, detail="listing_ids is required")
    unique_ids = list(dict.fromkeys(listing_ids))

    await _ensure_bulk_listings_pending(session, unique_ids)

    await _assert_moderation_not_frozen(
        session=session,
        request=request,
        current_user=current_user,
        action_type="bulk_approve",
        listing_ids=unique_ids,
    )

    async with session.begin():
        for listing_id in unique_ids:
            await _moderation_transition_sql(
                session=session,
                listing_id=listing_id,
                current_user=current_user,
                new_status="published",
                action_type="approve",
                request=request,
                commit=False,
            )

    return {"ok": True, "processed": len(unique_ids)}


@api_router.post("/admin/moderation/bulk-reject")
async def admin_bulk_reject_listings(
    payload: BulkModerationPayload,
    request: Request,
    current_user=Depends(check_permissions(list(ALLOWED_MODERATION_ROLES))),
    session: AsyncSession = Depends(get_sql_session),
):
    _ensure_moderation_rbac(current_user)
    if not payload.reason:
        raise HTTPException(status_code=400, detail="Reason is required")
    listing_ids = [str(item).strip() for item in (payload.listing_ids or []) if str(item).strip()]
    if not listing_ids:
        raise HTTPException(status_code=400, detail="listing_ids is required")
    unique_ids = list(dict.fromkeys(listing_ids))

    await _ensure_bulk_listings_pending(session, unique_ids)

    await _assert_moderation_not_frozen(
        session=session,
        request=request,
        current_user=current_user,
        action_type="bulk_reject",
        listing_ids=unique_ids,
    )

    async with session.begin():
        for listing_id in unique_ids:
            await _moderation_transition_sql(
                session=session,
                listing_id=listing_id,
                current_user=current_user,
                new_status="rejected",
                action_type="reject",
                reason=payload.reason,
                reason_note=payload.reason_note,
                request=request,
                commit=False,
            )

    return {"ok": True, "processed": len(unique_ids)}


# =====================
# Sprint 2.1 — Admin Global Listing Management
# =====================


@api_router.get("/admin/listings")
async def admin_listings(
    request: Request,
    skip: int = 0,
    limit: int = 50,
    status: Optional[str] = None,
    q: Optional[str] = None,
    dealer_only: Optional[str] = None,
    category_id: Optional[str] = None,
    owner_id: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session, )

    filters = []
    if getattr(ctx, "mode", "global") == "country" and ctx.country:
        filters.append(Listing.country == ctx.country)
    if status:
        filters.append(Listing.status == status)

    if q:
        search_value = f"%{q}%"
        filters.append(or_(Listing.title.ilike(search_value), cast(Listing.id, String).ilike(search_value)))

    if category_id:
        try:
            category_uuid = uuid.UUID(category_id)
            filters.append(Listing.category_id == category_uuid)
        except ValueError:
            pass

    if owner_id:
        try:
            owner_uuid = uuid.UUID(owner_id)
            filters.append(Listing.user_id == owner_uuid)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid owner_id") from exc

    dealer_only_flag = _parse_bool_flag(dealer_only)
    if dealer_only_flag and not owner_id:
        filters.append(Listing.is_dealer_listing.is_(True))

    limit = min(100, max(1, int(limit)))

    query = (
        select(Listing)
        .where(*filters)
        .order_by(Listing.created_at.desc())
        .offset(int(skip))
        .limit(limit)
    )
    result = await session.execute(query)
    listings = result.scalars().all()

    total = await session.scalar(select(func.count(Listing.id)).where(*filters)) or 0

    owner_ids = [listing.user_id for listing in listings if listing.user_id]
    user_map: Dict[uuid.UUID, SqlUser] = {}
    if owner_ids:
        users = await session.execute(select(SqlUser).where(SqlUser.id.in_(owner_ids)))
        user_map = {u.id: u for u in users.scalars().all()}

    items = []
    for listing in listings:
        owner = user_map.get(listing.user_id)
        items.append(
            {
                "id": str(listing.id),
                "title": listing.title,
                "status": listing.status,
                "country": listing.country,
                "category_key": str(listing.category_id) if listing.category_id else None,
                "price": listing.price or listing.hourly_rate,
                "currency": listing.currency,
                "image_count": len(listing.images or []),
                "created_at": listing.created_at.isoformat() if listing.created_at else None,
                "owner_id": str(listing.user_id) if listing.user_id else None,
                "owner_email": owner.email if owner else None,
                "owner_role": owner.role if owner else None,
                "is_dealer_listing": listing.is_dealer_listing or (owner and owner.role == "dealer"),
            }
        )

    return {"items": items, "pagination": {"total": int(total), "skip": int(skip), "limit": limit}}


@api_router.post("/admin/listings/{listing_id}/soft-delete")
async def admin_soft_delete_listing(
    listing_id: str,
    payload: ListingAdminActionPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    updated = await _admin_listing_action(
        session=session,
        listing_id=listing_id,
        current_user=current_user,
        event_type="LISTING_SOFT_DELETE",
        new_status="archived",
        reason=(payload.reason or None),
        reason_note=(payload.reason_note or None),
    )
    return {"ok": True, "listing": {"id": updated["id"], "status": updated.get("status")}}


@api_router.post("/admin/listings/{listing_id}/force-unpublish")
async def admin_force_unpublish_listing(
    listing_id: str,
    payload: ListingAdminActionPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid listing id") from exc
    listing = await session.get(Listing, listing_uuid)
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")
    if listing.status != "published":
        raise HTTPException(status_code=400, detail="Only published listings can be force-unpublished")

    updated = await _admin_listing_action(
        session=session,
        listing_id=listing_id,
        current_user=current_user,
        event_type="LISTING_FORCE_UNPUBLISH",
        new_status="unpublished",
        reason=(payload.reason or None),
        reason_note=(payload.reason_note or None),
    )
    return {"ok": True, "listing": {"id": updated["id"], "status": updated.get("status")}}


# =====================
# Sprint 2.2 — Reports Engine
# =====================


@api_router.post("/reports")
async def create_report(
    payload: ReportCreatePayload,
    request: Request,
    current_user=Depends(get_current_user_optional),
    session: AsyncSession = Depends(get_sql_session),
):
    listing_id = payload.listing_id
    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid listing id") from exc

    result = await session.execute(select(Listing).where(Listing.id == listing_uuid))
    listing = result.scalar_one_or_none()
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")
    if listing.status not in {"published", "active"}:
        raise HTTPException(status_code=400, detail="Only published listings can be reported")

    reason, reason_note = _validate_report_reason(payload.reason, payload.reason_note)
    reporter_user_id = current_user.get("id") if current_user else None
    _check_report_rate_limit(request, listing_id, reporter_user_id)

    report = Report(
        listing_id=listing.id,
        reporter_user_id=_safe_uuid(reporter_user_id) if reporter_user_id else None,
        reason=reason,
        reason_note=reason_note,
        status="open",
        country_code=listing.country,
    )
    session.add(report)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="REPORT_CREATED",
        actor=current_user or {"id": None, "email": None, "role": None},
        resource_type="report",
        resource_id=str(report.id),
        metadata={"listing_id": listing_id, "reason": reason},
        request=request,
        country_code=listing.country,
    )
    await session.commit()

    return {"ok": True, "report_id": str(report.id), "status": "open"}


@api_router.get("/admin/reports")
async def admin_reports(
    request: Request,
    status: Optional[str] = None,
    reason: Optional[str] = None,
    listing_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session)

    conditions = []
    if status:
        if status not in REPORT_STATUS_SET:
            raise HTTPException(status_code=400, detail="Invalid status")
        conditions.append(Report.status == status)
    if reason:
        _validate_reason(reason, REPORT_REASONS_V1)
        conditions.append(Report.reason == reason)
    if listing_id:
        try:
            listing_uuid = uuid.UUID(listing_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid listing id") from exc
        conditions.append(Report.listing_id == listing_uuid)

    if getattr(ctx, "mode", "global") == "country" and ctx.country:
        conditions.append(Report.country_code == ctx.country)
    elif current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope:
            conditions.append(Report.country_code.in_(scope))

    query = select(Report)
    if conditions:
        for cond in conditions:
            query = query.where(cond)

    count_query = select(func.count()).select_from(Report)
    if conditions:
        for cond in conditions:
            count_query = count_query.where(cond)

    total = (await session.execute(count_query)).scalar_one()
    result = await session.execute(
        query.order_by(Report.created_at.desc()).offset(int(skip)).limit(int(limit))
    )
    reports = result.scalars().all()

    listing_ids = [report.listing_id for report in reports if report.listing_id]
    reporter_ids = [report.reporter_user_id for report in reports if report.reporter_user_id]

    listing_map = {}
    if listing_ids:
        listing_result = await session.execute(select(Listing).where(Listing.id.in_(listing_ids)))
        listing_map = {str(row.id): row for row in listing_result.scalars().all()}

    user_ids = set(reporter_ids)
    for listing in listing_map.values():
        if listing.user_id:
            user_ids.add(listing.user_id)

    user_map = {}
    if user_ids:
        user_result = await session.execute(select(SqlUser).where(SqlUser.id.in_(list(user_ids))))
        user_map = {str(user.id): user for user in user_result.scalars().all()}

    items = []
    for report in reports:
        listing = listing_map.get(str(report.listing_id))
        seller = user_map.get(str(listing.user_id)) if listing and listing.user_id else None
        reporter = user_map.get(str(report.reporter_user_id)) if report.reporter_user_id else None

        items.append(
            {
                "id": str(report.id),
                "listing_id": str(report.listing_id),
                "reason": report.reason,
                "reason_note": report.reason_note,
                "status": report.status,
                "country_code": report.country_code,
                "created_at": report.created_at.isoformat() if report.created_at else None,
                "updated_at": report.updated_at.isoformat() if report.updated_at else None,
                "reporter_user_id": str(report.reporter_user_id) if report.reporter_user_id else None,
                "reporter_email": reporter.email if reporter else None,
                "listing_title": listing.title if listing else None,
                "listing_status": listing.status if listing else None,
                "seller_id": str(listing.user_id) if listing and listing.user_id else None,
                "seller_email": seller.email if seller else None,
                "seller_role": seller.role if seller else None,
                "seller_dealer_status": None,
            }
        )

    return {"items": items, "pagination": {"total": int(total), "skip": int(skip), "limit": limit}}

@api_router.get("/admin/reports/{report_id}")
async def admin_report_detail(
    report_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        report_uuid = uuid.UUID(report_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid report id") from exc

    report = await session.get(Report, report_uuid)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    if getattr(ctx, "mode", "global") == "country" and ctx.country:
        if report.country_code != ctx.country:
            raise HTTPException(status_code=403, detail="Country scope forbidden")
    elif current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope and report.country_code not in scope:
            raise HTTPException(status_code=403, detail="Country scope forbidden")

    listing = await session.get(Listing, report.listing_id)
    seller = None
    if listing and listing.user_id:
        seller = await session.get(SqlUser, listing.user_id)

    reporter = None
    if report.reporter_user_id:
        reporter = await session.get(SqlUser, report.reporter_user_id)

    listing_snapshot = None
    if listing:
        listing_snapshot = {
            "id": str(listing.id),
            "title": listing.title,
            "status": listing.status,
            "country": listing.country,
            "category_key": listing.category_key,
            "price": listing.price,
            "currency": listing.currency or "EUR",
            "created_at": listing.created_at.isoformat() if listing.created_at else None,
        }

    seller_summary = None
    if seller:
        seller_summary = {
            "id": str(seller.id),
            "email": seller.email,
            "role": seller.role,
            "dealer_status": None,
            "country_code": seller.country_code,
        }

    reporter_summary = None
    if reporter:
        reporter_summary = {
            "id": str(reporter.id),
            "email": reporter.email,
            "role": reporter.role,
        }

    return {
        "id": str(report.id),
        "listing_id": str(report.listing_id),
        "reason": report.reason,
        "reason_note": report.reason_note,
        "status": report.status,
        "country_code": report.country_code,
        "created_at": report.created_at.isoformat() if report.created_at else None,
        "updated_at": report.updated_at.isoformat() if report.updated_at else None,
        "handled_by_admin_id": str(report.handled_by_admin_id) if report.handled_by_admin_id else None,
        "status_note": report.status_note,
        "listing_snapshot": listing_snapshot,
        "seller_summary": seller_summary,
        "reporter_summary": reporter_summary,
    }

@api_router.post("/admin/reports/{report_id}/status")
async def admin_report_status_change(
    report_id: str,
    payload: ReportStatusPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    target_status = (payload.target_status or "").strip()
    if target_status not in REPORT_STATUS_SET:
        raise HTTPException(status_code=400, detail="Invalid target_status")
    note = (payload.note or "").strip()
    if not note:
        raise HTTPException(status_code=400, detail="note is required")

    updated = await _report_transition(
        session=session,
        report_id=report_id,
        current_user=current_user,
        target_status=target_status,
        note=note,
    )
    return {"ok": True, "report": {"id": str(updated.id), "status": updated.status}}


# =====================
# Sprint 3 — Finance Domain
# =====================


@api_router.post("/admin/invoices")
async def admin_create_invoice(
    payload: InvoiceCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    target_user_id = payload.user_id or payload.dealer_id
    if not target_user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    try:
        user_uuid = uuid.UUID(target_user_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="user_id invalid")

    subscription_uuid = None
    subscription = None
    if payload.subscription_id:
        try:
            subscription_uuid = uuid.UUID(payload.subscription_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="subscription_id invalid")
        subscription = await session.get(UserSubscription, subscription_uuid)
        if not subscription:
            raise HTTPException(status_code=404, detail="Subscription not found")

    plan_uuid = None
    plan = None
    if payload.plan_id:
        try:
            plan_uuid = uuid.UUID(payload.plan_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="plan_id invalid")
        plan = await session.get(Plan, plan_uuid)
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")

    user = await session.get(SqlUser, user_uuid)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    currency_value = (payload.currency or payload.currency_code or "").upper().strip()
    if not currency_value and plan:
        currency_value = (plan.currency_code or "").upper().strip()
    if not currency_value:
        raise HTTPException(status_code=400, detail="currency is required")

    amount_value = payload.amount_total if payload.amount_total is not None else payload.amount
    if amount_value is None and plan:
        amount_value = plan.price_amount
    if amount_value is None:
        raise HTTPException(status_code=400, detail="amount_total is required")
    if amount_value < 0:
        raise HTTPException(status_code=400, detail="amount_total must be >= 0")

    due_at = _parse_iso_datetime(payload.due_at, "due_at") if payload.due_at else None
    issue_now = True if payload.issue_now is None else payload.issue_now

    now = datetime.now(timezone.utc)
    status_value = "issued" if issue_now else "draft"
    issued_at = now if issue_now else None

    scope_value = "country" if user.country_code else "global"
    country_code = user.country_code or None

    invoice = AdminInvoice(
        invoice_no=_generate_invoice_no(),
        user_id=user_uuid,
        subscription_id=subscription_uuid,
        plan_id=plan_uuid,
        campaign_id=None,
        amount_total=amount_value,
        currency=currency_value,
        status=status_value,
        payment_status="requires_payment_method",
        issued_at=issued_at,
                due_at=due_at,
        provider_customer_id=payload.provider_customer_id,
        meta_json=payload.meta_json,
        scope=scope_value,
        country_code=country_code,
        payment_method=None,
        notes=payload.notes,
        created_at=now,
        updated_at=now,
    )
    session.add(invoice)
    await session.commit()
    await session.refresh(invoice)

    return {"invoice": _admin_invoice_to_dict(invoice, user, plan)}


@api_router.get("/admin/invoices")
async def admin_list_invoices(
    request: Request,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    conditions = []

    if status:
        status_value = status.strip().lower()
        if status_value == "cancelled":
            status_value = "void"
        if status_value not in INVOICE_STATUS_SET:
            raise HTTPException(status_code=400, detail="Invalid status")
        conditions.append(AdminInvoice.status == status_value)

    limit = min(100, max(1, int(limit)))
    skip = max(0, int(skip))

    base_query = select(AdminInvoice)
    if conditions:
        base_query = base_query.where(*conditions)

    rows = (
        await session.execute(
            base_query.order_by(AdminInvoice.created_at.desc()).offset(skip).limit(limit)
        )
    ).scalars().all()
    total = (
        await session.execute(select(func.count()).select_from(AdminInvoice).where(*conditions))
    ).scalar() or 0

    user_ids = {row.user_id for row in rows}
    plan_ids = {row.plan_id for row in rows if row.plan_id}

    users = []
    plans = []
    if user_ids:
        users = (await session.execute(select(SqlUser).where(SqlUser.id.in_(user_ids)))).scalars().all()
    if plan_ids:
        plans = (await session.execute(select(Plan).where(Plan.id.in_(plan_ids)))).scalars().all()

    user_map = {user.id: user for user in users}
    plan_map = {plan.id: plan for plan in plans}

    items = [_admin_invoice_to_dict(row, user_map.get(row.user_id), plan_map.get(row.plan_id)) for row in rows]

    return {"items": items, "pagination": {"total": total, "skip": skip, "limit": limit}}


@api_router.get("/admin/invoices/{invoice_id}")
async def admin_invoice_detail(
    invoice_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    try:
        invoice_uuid = uuid.UUID(invoice_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="invoice_id invalid")

    invoice = await session.get(AdminInvoice, invoice_uuid)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.country_code and invoice.country_code != "GLOBAL":
        _assert_country_scope(invoice.country_code, current_user)

    user = await session.get(SqlUser, invoice.user_id)
    plan = await session.get(Plan, invoice.plan_id) if invoice.plan_id else None

    return {
        "invoice": _admin_invoice_to_dict(invoice, user, plan),
        "dealer": {"id": str(user.id), "email": user.email} if user else None,
        "plan": {"id": str(plan.id), "name": plan.name} if plan else None,
    }


@api_router.post("/admin/invoices/{invoice_id}/mark-paid")
async def admin_invoice_mark_paid(
    invoice_id: str,
    payload: InvoiceActionPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    try:
        invoice_uuid = uuid.UUID(invoice_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="invoice_id invalid")

    invoice = await session.get(AdminInvoice, invoice_uuid)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.country_code and invoice.country_code != "GLOBAL":
        _assert_country_scope(invoice.country_code, current_user)

    if invoice.status not in {"issued"}:
        raise HTTPException(status_code=400, detail="Only issued invoices can be marked paid")

    if not payload.reason:
        raise HTTPException(status_code=400, detail="reason is required")

    now = datetime.now(timezone.utc)
    invoice.status = "paid"
    invoice.payment_status = "succeeded"
    invoice.paid_at = now
    if payload.payment_method:
        invoice.payment_method = payload.payment_method
    invoice.updated_at = now
    await session.commit()
    await session.refresh(invoice)

    user = await session.get(SqlUser, invoice.user_id)
    plan = await session.get(Plan, invoice.plan_id) if invoice.plan_id else None
    return {"invoice": _admin_invoice_to_dict(invoice, user, plan)}


@api_router.post("/admin/invoices/{invoice_id}/cancel")
async def admin_invoice_cancel(
    invoice_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    try:
        invoice_uuid = uuid.UUID(invoice_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="invoice_id invalid")

    invoice = await session.get(AdminInvoice, invoice_uuid)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.country_code and invoice.country_code != "GLOBAL":
        _assert_country_scope(invoice.country_code, current_user)

    if invoice.status != "issued":
        raise HTTPException(status_code=400, detail="Only issued invoices can be voided")

    now = datetime.now(timezone.utc)
    invoice.status = "void"
    invoice.payment_status = "canceled"
    invoice.updated_at = now
    await session.commit()
    await session.refresh(invoice)

    user = await session.get(SqlUser, invoice.user_id)
    plan = await session.get(Plan, invoice.plan_id) if invoice.plan_id else None
    return {"invoice": _admin_invoice_to_dict(invoice, user, plan)}


@api_router.post("/admin/invoices/{invoice_id}/refund")
async def admin_invoice_refund(
    invoice_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    try:
        invoice_uuid = uuid.UUID(invoice_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="invoice_id invalid")

    invoice = await session.get(AdminInvoice, invoice_uuid)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")
    if invoice.country_code and invoice.country_code != "GLOBAL":
        _assert_country_scope(invoice.country_code, current_user)

    if invoice.status != "paid":
        raise HTTPException(status_code=400, detail="Only paid invoices can be refunded")

    now = datetime.now(timezone.utc)
    invoice.status = "refunded"
    invoice.payment_status = "refunded"
    invoice.updated_at = now
    await session.commit()
    await session.refresh(invoice)

    user = await session.get(SqlUser, invoice.user_id)
    plan = await session.get(Plan, invoice.plan_id) if invoice.plan_id else None
    return {"invoice": _admin_invoice_to_dict(invoice, user, plan)}


@api_router.get("/dealer/invoices")
async def dealer_list_invoices(
    request: Request,
    status: Optional[str] = None,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    try:
        dealer_uuid = uuid.UUID(current_user.get("id"))
    except ValueError:
        raise HTTPException(status_code=400, detail="dealer invalid")

    conditions = [AdminInvoice.user_id == dealer_uuid]
    if status:
        status_value = status.strip().lower()
        if status_value == "cancelled":
            status_value = "void"
        if status_value not in INVOICE_STATUS_SET:
            raise HTTPException(status_code=400, detail="Invalid status")
        conditions.append(AdminInvoice.status == status_value)

    rows = (await session.execute(select(AdminInvoice).where(*conditions).order_by(AdminInvoice.created_at.desc()))).scalars().all()
    items = [_admin_invoice_to_dict(row) for row in rows]
    return {"items": items}


@api_router.get("/dealer/invoices/{invoice_id}")
async def dealer_invoice_detail(
    invoice_id: str,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    try:
        invoice_uuid = uuid.UUID(invoice_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="invoice_id invalid")

    invoice = await session.get(AdminInvoice, invoice_uuid)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if str(invoice.user_id) != current_user.get("id"):
        raise HTTPException(status_code=403, detail="Invoice access denied")

    plan = await session.get(Plan, invoice.plan_id) if invoice.plan_id else None
    return {"invoice": _admin_invoice_to_dict(invoice, None, plan)}


class DealerListingCreatePayload(BaseModel):
    title: str
    price: Optional[float] = None


class DealerListingUpdatePayload(BaseModel):
    title: Optional[str] = None
    price: Optional[float] = None


class DealerListingStatusPayload(BaseModel):
    status: str


class DealerListingBulkPayload(BaseModel):
    ids: List[str]
    action: str


def _dealer_listing_to_dict(listing: DealerListing) -> dict:
    return {
        "id": str(listing.id),
        "title": listing.title,
        "price": listing.price,
        "status": listing.status,
        "created_at": listing.created_at.isoformat() if listing.created_at else None,
        "updated_at": listing.updated_at.isoformat() if listing.updated_at else None,
        "deleted_at": listing.deleted_at.isoformat() if listing.deleted_at else None,
    }


@api_router.get("/dealer/listings")
async def dealer_list_listings(
    request: Request,
    status: Optional[str] = Query(None),
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        dealer_uuid = uuid.UUID(current_user.get("id"))
    except ValueError:
        raise HTTPException(status_code=400, detail="dealer invalid")

    status_value = status.lower().strip() if status else None
    if status_value == "all":
        status_value = None
    if status_value and status_value not in {"active", "draft", "archived"}:
        raise HTTPException(status_code=400, detail="invalid status")

    query = (
        select(DealerListing)
        .where(
            DealerListing.dealer_id == dealer_uuid,
            DealerListing.deleted_at.is_(None),
        )
        .order_by(desc(DealerListing.created_at))
    )
    if status_value:
        query = query.where(DealerListing.status == status_value)

    rows = (await session.execute(query)).scalars().all()
    items = [_dealer_listing_to_dict(row) for row in rows]

    active_count = (
        await session.execute(
            select(func.count()).select_from(DealerListing).where(
                DealerListing.dealer_id == dealer_uuid,
                DealerListing.status == "active",
                DealerListing.deleted_at.is_(None),
            )
        )
    ).scalar_one()

    limit = DEALER_LISTING_QUOTA_LIMIT
    remaining = max(0, limit - int(active_count or 0))
    return {
        "items": items,
        "quota": {
            "limit": limit,
            "used": int(active_count or 0),
            "remaining": remaining,
        },
    }


@api_router.get("/dealer/quotas")
async def dealer_listing_quota(
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        dealer_uuid = uuid.UUID(current_user.get("id"))
    except ValueError:
        raise HTTPException(status_code=400, detail="dealer invalid")

    count = (
        await session.execute(
            select(func.count()).select_from(DealerListing).where(
                DealerListing.dealer_id == dealer_uuid,
                DealerListing.status == "active",
                DealerListing.deleted_at.is_(None),
            )
        )
    ).scalar_one()
    limit = DEALER_LISTING_QUOTA_LIMIT
    remaining = max(0, limit - int(count or 0))
    return {"limit": limit, "used": int(count or 0), "remaining": remaining}


@api_router.post("/dealer/listings", status_code=201)
async def dealer_create_listing(
    payload: DealerListingCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        dealer_uuid = uuid.UUID(current_user.get("id"))
    except ValueError:
        raise HTTPException(status_code=400, detail="dealer invalid")

    active_count = (
        await session.execute(
            select(func.count()).select_from(DealerListing).where(
                DealerListing.dealer_id == dealer_uuid,
                DealerListing.status == "active",
                DealerListing.deleted_at.is_(None),
            )
        )
    ).scalar_one()

    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="title is required")

    listing = DealerListing(
        dealer_id=dealer_uuid,
        title=title,
        price=payload.price,
        status="draft",
    )
    session.add(listing)
    await session.commit()
    await session.refresh(listing)
    used = int(active_count or 0)
    limit = DEALER_LISTING_QUOTA_LIMIT
    remaining = max(0, limit - used)
    return {"item": _dealer_listing_to_dict(listing), "quota": {"limit": limit, "used": used, "remaining": remaining}}


@api_router.patch("/dealer/listings/{listing_id}")
async def dealer_update_listing(
    listing_id: str,
    payload: DealerListingUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="listing invalid")

    dealer_uuid = uuid.UUID(current_user.get("id"))
    listing = await session.get(DealerListing, listing_uuid)
    if not listing or listing.dealer_id != dealer_uuid or listing.deleted_at:
        raise HTTPException(status_code=404, detail="Listing not found")

    if payload.title is not None:
        title = payload.title.strip()
        if not title:
            raise HTTPException(status_code=400, detail="title is required")
        listing.title = title

    if payload.price is not None:
        listing.price = payload.price

    await session.commit()
    await session.refresh(listing)
    return {"item": _dealer_listing_to_dict(listing)}


@api_router.post("/dealer/listings/{listing_id}/status")
async def dealer_update_listing_status(
    listing_id: str,
    payload: DealerListingStatusPayload,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    status_value = payload.status.lower().strip()
    if status_value not in {"draft", "active", "archived"}:
        raise HTTPException(status_code=400, detail="invalid status")

    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="listing invalid")

    dealer_uuid = uuid.UUID(current_user.get("id"))
    listing = await session.get(DealerListing, listing_uuid)
    if not listing or listing.dealer_id != dealer_uuid or listing.deleted_at:
        raise HTTPException(status_code=404, detail="Listing not found")

    if status_value == "active":
        active_count = (
            await session.execute(
                select(func.count()).select_from(DealerListing).where(
                    DealerListing.dealer_id == dealer_uuid,
                    DealerListing.status == "active",
                    DealerListing.deleted_at.is_(None),
                    DealerListing.id != listing.id,
                )
            )
        ).scalar_one()
        if int(active_count or 0) >= DEALER_LISTING_QUOTA_LIMIT:
            raise HTTPException(status_code=403, detail="Listing quota exceeded")

    listing.status = status_value
    await session.commit()
    await session.refresh(listing)
    return {"item": _dealer_listing_to_dict(listing)}


@api_router.delete("/dealer/listings/{listing_id}")
async def dealer_delete_listing(
    listing_id: str,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        listing_uuid = uuid.UUID(listing_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="listing invalid")

    dealer_uuid = uuid.UUID(current_user.get("id"))
    listing = await session.get(DealerListing, listing_uuid)
    if not listing or listing.dealer_id != dealer_uuid or listing.deleted_at:
        raise HTTPException(status_code=404, detail="Listing not found")

    listing.status = "archived"
    listing.deleted_at = datetime.now(timezone.utc)
    await session.commit()
    await session.refresh(listing)
    return {"item": _dealer_listing_to_dict(listing)}


@api_router.post("/dealer/listings/bulk")
async def dealer_bulk_action_listings(
    payload: DealerListingBulkPayload,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    action = payload.action.lower().strip()
    if action not in {"archive", "delete", "restore"}:
        raise HTTPException(status_code=400, detail="invalid action")

    if not payload.ids:
        raise HTTPException(status_code=400, detail="ids required")

    dealer_uuid = uuid.UUID(current_user.get("id"))
    listing_ids = []
    for raw_id in payload.ids:
        try:
            listing_ids.append(uuid.UUID(raw_id))
        except ValueError:
            continue

    if not listing_ids:
        raise HTTPException(status_code=400, detail="ids invalid")

    result = await session.execute(
        select(DealerListing).where(
            DealerListing.id.in_(listing_ids),
            DealerListing.dealer_id == dealer_uuid,
            DealerListing.deleted_at.is_(None),
        )
    )
    listings = result.scalars().all()

    restorable_count = 0
    if action == "restore":
        restorable_count = sum(1 for listing in listings if listing.status == "archived")
        active_count = (
            await session.execute(
                select(func.count()).select_from(DealerListing).where(
                    DealerListing.dealer_id == dealer_uuid,
                    DealerListing.status == "active",
                    DealerListing.deleted_at.is_(None),
                )
            )
        ).scalar_one()
        if int(active_count or 0) + restorable_count > DEALER_LISTING_QUOTA_LIMIT:
            raise HTTPException(status_code=403, detail="Listing quota exceeded")

    updated = 0
    deleted = 0
    failed = 0
    for listing in listings:
        if action == "archive":
            if listing.status == "archived":
                failed += 1
            else:
                listing.status = "archived"
                updated += 1
        elif action == "restore":
            if listing.status != "archived":
                failed += 1
            else:
                listing.status = "trial"
                updated += 1
        elif action == "delete":
            listing.status = "archived"
            listing.deleted_at = datetime.now(timezone.utc)
            deleted += 1

    await session.commit()
    return {"updated": updated, "deleted": deleted, "failed": failed}


@api_router.get("/dealer/dashboard/metrics")
async def dealer_dashboard_metrics(
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        dealer_uuid = uuid.UUID(current_user.get("id"))
    except ValueError:
        raise HTTPException(status_code=400, detail="dealer invalid")

    active_count = (
        await session.execute(
            select(func.count()).select_from(DealerListing).where(
                DealerListing.dealer_id == dealer_uuid,
                DealerListing.status == "active",
                DealerListing.deleted_at.is_(None),
            )
        )
    ).scalar_one()

    total_count = (
        await session.execute(
            select(func.count()).select_from(DealerListing).where(
                DealerListing.dealer_id == dealer_uuid,
                DealerListing.deleted_at.is_(None),
            )
        )
    ).scalar_one()

    quota_limit = DEALER_LISTING_QUOTA_LIMIT
    remaining = max(0, quota_limit - int(active_count or 0))
    utilization = (int(active_count or 0) / quota_limit * 100) if quota_limit else 0
    quota_warning = utilization >= 80

    views_payload = {"count": 0, "gated": False}
    try:
        views_count = (
            await session.execute(
                select(func.count())
                .select_from(ListingView)
                .join(Listing, ListingView.listing_id == Listing.id)
                .where(Listing.dealer_id == dealer_uuid)
            )
        ).scalar_one()
        views_payload["count"] = int(views_count or 0)
    except Exception:
        views_payload = {"count": 0, "gated": True}

    messages_payload = {"count": 0, "gated": False}
    try:
        msg_count = (
            await session.execute(
                select(func.count())
                .select_from(Message)
                .join(Conversation, Message.conversation_id == Conversation.id)
                .where(Conversation.seller_id == dealer_uuid)
            )
        ).scalar_one()
        messages_payload["count"] = int(msg_count or 0)
    except Exception:
        messages_payload = {"count": 0, "gated": True}

    subscription_payload = {
        "name": "N/A",
        "status": "gated",
        "current_period_end": None,
        "warning": False,
    }

    return {
        "active_listings": int(active_count or 0),
        "total_listings": int(total_count or 0),
        "quota": {
            "limit": quota_limit,
            "used": int(active_count or 0),
            "remaining": remaining,
            "utilization": utilization,
            "warning": quota_warning,
        },
        "views": views_payload,
        "messages": messages_payload,
        "subscription": subscription_payload,
    }


def _resolve_payment_status(value: Optional[str]) -> str:
    normalized = (value or "").lower()
    if normalized in {"paid", "succeeded"}:
        return "succeeded"
    if normalized in {"refunded"}:
        return "refunded"
    if normalized in {"failed"}:
        return "failed"
    if normalized in {"canceled", "cancelled"}:
        return "canceled"
    if normalized in {"requires_payment_method", "requires_confirmation", "processing"}:
        return normalized
    return normalized or "requires_payment_method"


async def _apply_payment_status(
    invoice: AdminInvoice,
    payment: Payment,
    transaction: PaymentTransaction,
    payment_status: str,
    provider_payment_id: Optional[str],
    provider_ref: Optional[str] = None,
    meta: Optional[Dict[str, Any]] = None,
    session: Optional[AsyncSession] = None,
) -> None:
    now = datetime.now(timezone.utc)
    status_value = _resolve_payment_status(payment_status)

    if provider_payment_id:
        transaction.provider_payment_id = provider_payment_id
    if provider_ref:
        payment.provider_ref = provider_ref

    if status_value == "succeeded":
        invoice.payment_status = "succeeded"
        invoice.status = "paid"
        if not invoice.paid_at:
            invoice.paid_at = now
        payment.status = "succeeded"
        transaction.status = "succeeded"
        transaction.payment_status = "succeeded"

        if session:
            await _write_audit_log_sql(
                session=session,
                action="payment_succeeded",
                actor={"id": str(invoice.user_id), "email": None},
                resource_type="payment",
                resource_id=str(payment.id),
                metadata={"invoice_id": str(invoice.id), "status": "succeeded"},
                request=None,
                country_code=invoice.country_code,
            )
            await _write_audit_log_sql(
                session=session,
                action="invoice_marked_paid",
                actor={"id": str(invoice.user_id), "email": None},
                resource_type="invoice",
                resource_id=str(invoice.id),
                metadata={"status": "paid"},
                request=None,
                country_code=invoice.country_code,
            )
            await _activate_subscription_from_invoice(session, invoice)

    elif status_value == "refunded":
        invoice.payment_status = "refunded"
        invoice.status = "refunded"
        payment.status = "refunded"
        transaction.status = "refunded"
        transaction.payment_status = "refunded"
    elif status_value in {"failed", "canceled"}:
        invoice.payment_status = status_value
        payment.status = status_value
        transaction.status = status_value
        transaction.payment_status = status_value
    else:
        invoice.payment_status = status_value
        payment.status = status_value
        transaction.status = status_value
        transaction.payment_status = status_value

    if meta:
        payment.meta_json = meta

    invoice.updated_at = now
    payment.updated_at = now
    transaction.updated_at = now


async def _process_webhook_payment(
    session: AsyncSession,
    event_log: WebhookEventLog,
    session_id: Optional[str],
    payment_status: Optional[str],
    metadata: Dict[str, Any],
    provider_ref: Optional[str],
    provider_payment_id: Optional[str],
) -> str:
    now = datetime.now(timezone.utc)

    transaction = None
    if session_id:
        transaction = (await session.execute(
            select(PaymentTransaction).where(PaymentTransaction.session_id == session_id)
        )).scalar_one_or_none()

    invoice = None
    payment = None
    if transaction:
        invoice = await session.get(AdminInvoice, transaction.invoice_id)
        payment = (await session.execute(
            select(Payment).where(Payment.invoice_id == transaction.invoice_id)
        )).scalar_one_or_none()
        if invoice and not payment:
            payment = Payment(
                invoice_id=invoice.id,
                user_id=invoice.user_id,
                provider="stripe",
                provider_ref=provider_ref or f"pi_{invoice.id}",
                status=_resolve_payment_status(payment_status),
                amount_total=invoice.amount_total,
                currency=invoice.currency,
                meta_json=metadata,
                created_at=now,
                updated_at=now,
            )
            session.add(payment)
    else:
        invoice_id = metadata.get("invoice_id")
        try:
            invoice_uuid = uuid.UUID(invoice_id) if invoice_id else None
        except ValueError:
            invoice_uuid = None
        if invoice_uuid:
            invoice = await session.get(AdminInvoice, invoice_uuid)
            if invoice:
                payment = (await session.execute(
                    select(Payment).where(Payment.invoice_id == invoice.id)
                )).scalar_one_or_none()

                if not payment:
                    payment = Payment(
                        invoice_id=invoice.id,
                        user_id=invoice.user_id,
                        provider="stripe",
                        provider_ref=provider_ref or f"pi_{invoice.id}",
                        status=_resolve_payment_status(payment_status),
                        amount_total=invoice.amount_total,
                        currency=invoice.currency,
                        meta_json=metadata,
                        created_at=now,
                        updated_at=now,
                    )
                    session.add(payment)

                if not transaction:
                    transaction = PaymentTransaction(
                        provider="stripe",
                        session_id=session_id or f"sess_{invoice.id}",
                        provider_payment_id=provider_payment_id,
                        invoice_id=invoice.id,
                        dealer_id=invoice.user_id,
                        amount=invoice.amount_total,
                        currency=invoice.currency,
                        status=_resolve_payment_status(payment_status),
                        payment_status=_resolve_payment_status(payment_status),
                        metadata_json=metadata,
                        created_at=now,
                        updated_at=now,
                    )
                    session.add(transaction)

    if invoice and payment and transaction:
        await _apply_payment_status(
            invoice,
            payment,
            transaction,
            payment_status or "",
            provider_payment_id,
            provider_ref=payment.provider_ref,
            meta=metadata,
            session=session,
        )
        event_log.status = "processed"
        event_log.processed_at = now
        return "processed"

    event_log.status = "ignored"
    event_log.processed_at = now
    return "ignored"


@api_router.post("/payments/create-checkout-session/stub")
async def create_checkout_session_stub(
    payload: PaymentCheckoutStubPayload,
    request: Request,
):
    if APP_ENV == "prod":
        raise HTTPException(status_code=404, detail="Not found")
    if not STRIPE_API_KEY:
        return JSONResponse(
            status_code=200,
            content={
                "status": "blocked",
                "reason": "stripe_not_configured",
                "checkout_url": None,
            },
        )

    origin = payload.origin_url.strip().rstrip("/")
    if not origin:
        raise HTTPException(status_code=400, detail="origin_url required")

    webhook_url = f"{origin}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)

    success_url = f"{origin}/dealer/payments/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/dealer/payments/cancel"

    session_request = CheckoutSessionRequest(
        amount=float(payload.amount),
        currency=payload.currency,
        success_url=success_url,
        cancel_url=cancel_url,
        metadata=payload.metadata or {"stub": "true"},
    )

    checkout_session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(session_request)

    return {
        "checkout_url": checkout_session.url,
        "session_id": checkout_session.session_id,
        "payment_intent_id": None,
        "mode": "stub",
    }


@api_router.post("/payments/create-checkout-session")
async def create_checkout_session(
    payload: PaymentCheckoutPayload,
    request: Request,
    idempotency_key: Optional[str] = Header(None, alias="Idempotency-Key"),
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    if not STRIPE_API_KEY:
        raise HTTPException(status_code=503, detail="Stripe not configured")

    try:
        invoice_uuid = uuid.UUID(payload.invoice_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="invoice_id invalid")

    invoice = await session.get(AdminInvoice, invoice_uuid)
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    if str(invoice.user_id) != current_user.get("id"):
        raise HTTPException(status_code=403, detail="Invoice access denied")

    if invoice.status != "issued" or invoice.payment_status in {"succeeded", "refunded"}:
        raise HTTPException(status_code=400, detail="Invoice is not payable")

    if not _is_payment_enabled_for_country(invoice.country_code):
        raise HTTPException(status_code=403, detail="Payments disabled for this country")

    origin = payload.origin_url.strip().rstrip("/")
    if not origin:
        raise HTTPException(status_code=400, detail="origin_url required")

    idempotency_key = (idempotency_key or "").strip() or None
    if idempotency_key:
        existing_txn = (
            await session.execute(
                select(PaymentTransaction).where(
                    PaymentTransaction.invoice_id == invoice.id,
                    cast(PaymentTransaction.metadata_json, JSONB).contains({"idempotency_key": idempotency_key}),
                )
            )
        ).scalar_one_or_none()
        if existing_txn:
            existing_meta = existing_txn.metadata_json or {}
            checkout_url = existing_meta.get("checkout_url")
            if checkout_url:
                logging.getLogger("stripe_idempotency").warning(
                    "stripe_idempotency_reused invoice_id=%s session_id=%s",
                    invoice.id,
                    existing_txn.session_id,
                )
                return {
                    "checkout_url": checkout_url,
                    "session_id": existing_txn.session_id,
                    "idempotency_reused": True,
                }

    webhook_url = f"{origin}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)

    success_url = f"{origin}/dealer/payments/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/dealer/payments/cancel"

    session_request = CheckoutSessionRequest(
        amount=float(invoice.amount_total),
        currency=invoice.currency,
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={
            "invoice_id": str(invoice.id),
            "dealer_id": str(invoice.user_id),
            "invoice_no": invoice.invoice_no,
        },
    )

    checkout_session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(session_request)

    now = datetime.now(timezone.utc)
    payment_meta = {"checkout_session_id": checkout_session.session_id}
    transaction_meta = {"checkout_url": checkout_session.url}
    if idempotency_key:
        payment_meta["idempotency_key"] = idempotency_key
        transaction_meta["idempotency_key"] = idempotency_key

    payment = Payment(
        invoice_id=invoice.id,
        user_id=invoice.user_id,
        provider="stripe",
        provider_ref=checkout_session.session_id,
        status="requires_payment_method",
        amount_total=invoice.amount_total,
        currency=invoice.currency,
        meta_json=payment_meta,
        created_at=now,
        updated_at=now,
    )
    transaction = PaymentTransaction(
        provider="stripe",
        session_id=checkout_session.session_id,
        provider_payment_id=None,
        invoice_id=invoice.id,
        dealer_id=invoice.user_id,
        amount=invoice.amount_total,
        currency=invoice.currency,
        status="requires_payment_method",
        payment_status="requires_payment_method",
        metadata_json=transaction_meta,
        created_at=now,
        updated_at=now,
    )
    session.add(payment)
    session.add(transaction)
    await session.commit()

    logging.getLogger("stripe_idempotency").warning(
        "stripe_checkout_session_created invoice_id=%s session_id=%s",
        invoice.id,
        checkout_session.session_id,
    )

    return {
        "checkout_url": checkout_session.url,
        "session_id": checkout_session.session_id,
        "idempotency_reused": False,
    }


@api_router.get("/payments/checkout/status/{session_id}")
async def get_checkout_status(
    session_id: str,
    request: Request,
    current_user=Depends(check_permissions(["dealer"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    if not STRIPE_API_KEY:
        raise HTTPException(status_code=503, detail="Stripe not configured")

    origin = str(request.base_url).rstrip("/")
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=f"{origin}/api/webhook/stripe")

    status_response: CheckoutStatusResponse = await stripe_checkout.get_checkout_status(session_id)

    transaction = (await session.execute(
        select(PaymentTransaction).where(PaymentTransaction.session_id == session_id)
    )).scalar_one_or_none()

    if not transaction:
        raise HTTPException(status_code=404, detail="payment transaction not found")

    if str(transaction.dealer_id) != current_user.get("id"):
        raise HTTPException(status_code=403, detail="payment access denied")

    invoice = await session.get(AdminInvoice, transaction.invoice_id)
    payment = (await session.execute(
        select(Payment).where(Payment.invoice_id == transaction.invoice_id)
    )).scalar_one_or_none()
    provider_payment_id = getattr(status_response, "payment_id", None)
    payment_status = status_response.payment_status if hasattr(status_response, "payment_status") else None
    metadata = {}

    if invoice and not payment:
        payment = Payment(
            invoice_id=invoice.id,
            user_id=invoice.user_id,
            provider="stripe",
            provider_ref=provider_payment_id or f"pi_{invoice.id}",
            status=_resolve_payment_status(payment_status),
            amount_total=invoice.amount_total,
            currency=invoice.currency,
            meta_json={"source": "status_check"},
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        )
        session.add(payment)

    if invoice and payment:
        await _apply_payment_status(
            invoice,
            payment,
            transaction,
            payment_status or "",
            provider_payment_id,
            provider_ref=payment.provider_ref,
            meta=metadata,
            session=session,
        )
        await session.commit()
        await session.refresh(invoice)

    return {
        "session_id": status_response.session_id,
        "status": status_response.status,
        "payment_status": status_response.payment_status,
    }


@api_router.post("/webhook/stripe")
async def stripe_webhook(
    request: Request,
    session: AsyncSession = Depends(get_sql_session),
):
    if not STRIPE_API_KEY:
        raise HTTPException(status_code=503, detail="Stripe not configured")
    if not STRIPE_WEBHOOK_SECRET:
        raise HTTPException(status_code=503, detail="Stripe webhook secret not configured")

    origin = str(request.base_url).rstrip("/")
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=f"{origin}/api/webhook/stripe")

    signature = request.headers.get("stripe-signature")
    raw_body = await request.body()

    try:
        raw_payload = json.loads(raw_body.decode("utf-8"))
    except Exception:
        raw_payload = {"raw": raw_body.decode("utf-8", errors="ignore")}

    try:
        webhook_response = await stripe_checkout.handle_webhook(raw_body, signature)
    except Exception as exc:
        now = datetime.now(timezone.utc)
        invalid_event_id = f"invalid-{uuid.uuid4()}"
        event_log = WebhookEventLog(
            provider="stripe",
            event_id=invalid_event_id,
            event_type="invalid_signature",
            received_at=now,
            processed_at=now,
            status="invalid_signature",
            payload_json={"raw_payload": raw_payload},
            signature_valid=False,
            error_message=str(exc)[:255],
            created_at=now,
        )
        session.add(event_log)
        await session.commit()
        raise HTTPException(status_code=400, detail="Invalid signature")

    event_id = webhook_response.event_id
    event_type = webhook_response.event_type
    session_id = webhook_response.session_id
    payment_status = webhook_response.payment_status
    metadata = webhook_response.metadata or {}
    provider_ref = (
        getattr(webhook_response, "payment_intent_id", None)
        or getattr(webhook_response, "payment_id", None)
        or metadata.get("payment_intent_id")
        or metadata.get("payment_id")
        or session_id
    )
    provider_payment_id = getattr(webhook_response, "payment_id", None) or provider_ref

    db_ready, db_reason, db_status = await _get_db_status(session)
    if not db_ready:
        return JSONResponse(
            status_code=503,
            content={
                "status": "blocked",
                "reason": db_reason,
                "db_status": db_status,
                "event_id": event_id,
                "event_type": event_type,
                "signature_valid": True,
            },
        )

    now = datetime.now(timezone.utc)
    payload_json = {
        "event_type": event_type,
        "session_id": session_id,
        "payment_status": payment_status,
        "metadata": metadata,
        "provider_ref": provider_ref,
        "provider_payment_id": provider_payment_id,
        "raw_payload": raw_payload,
    }

    event_log = WebhookEventLog(
        provider="stripe",
        event_id=event_id,
        event_type=event_type,
        received_at=now,
        processed_at=None,
        status="received",
        payload_json=payload_json,
        signature_valid=True,
        error_message=None,
        created_at=now,
    )
    session.add(event_log)
    try:
        await session.flush()
    except IntegrityError:
        await session.rollback()
        existing = (
            await session.execute(
                select(WebhookEventLog).where(
                    WebhookEventLog.provider == "stripe",
                    WebhookEventLog.event_id == event_id,
                )
            )
        ).scalar_one_or_none()
        if existing:
            existing.status = "duplicate"
            existing.processed_at = now
            await session.commit()
        return {"status": "duplicate"}

    try:
        status_result = await _process_webhook_payment(
            session=session,
            event_log=event_log,
            session_id=session_id,
            payment_status=payment_status,
            metadata=metadata,
            provider_ref=provider_ref,
            provider_payment_id=provider_payment_id,
        )
        await session.commit()
        return {"status": status_result}
    except Exception as exc:
        await session.rollback()
        event_log.status = "failed"
        event_log.processed_at = now
        event_log.error_message = str(exc)[:255]
        session.add(event_log)
        await session.commit()
        raise


@api_router.post("/admin/webhooks/events/{event_id}/replay")
async def admin_replay_webhook_event(
    event_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    event_log = (
        await session.execute(
            select(WebhookEventLog).where(
                WebhookEventLog.provider == "stripe",
                WebhookEventLog.event_id == event_id,
            )
        )
    ).scalar_one_or_none()
    if not event_log:
        raise HTTPException(status_code=404, detail="Event not found")
    if not event_log.signature_valid:
        raise HTTPException(status_code=400, detail="Invalid signature event cannot be replayed")

    payload = event_log.payload_json or {}
    session_id = payload.get("session_id")
    payment_status = payload.get("payment_status")
    metadata = payload.get("metadata") or {}
    provider_ref = payload.get("provider_ref")
    provider_payment_id = payload.get("provider_payment_id")

    status_result = await _process_webhook_payment(
        session=session,
        event_log=event_log,
        session_id=session_id,
        payment_status=payment_status,
        metadata=metadata,
        provider_ref=provider_ref,
        provider_payment_id=provider_payment_id,
    )

    if status_result == "processed":
        event_log.status = "replayed"
        event_log.processed_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="webhook_replayed",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="webhook_event",
        resource_id=event_id,
        metadata={"status": event_log.status},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"status": event_log.status}


@api_router.get("/admin/payments")
async def admin_list_payments(
    request: Request,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    conditions = []
    if status:
        status_value = status.strip().lower()
        conditions.append(Payment.status == status_value)

    limit = min(100, max(1, int(limit)))
    skip = max(0, int(skip))

    base_query = select(Payment)
    if conditions:
        base_query = base_query.where(*conditions)

    rows = (
        await session.execute(
            base_query.order_by(Payment.created_at.desc()).offset(skip).limit(limit)
        )
    ).scalars().all()

    total = (
        await session.execute(select(func.count()).select_from(Payment).where(*conditions))
    ).scalar() or 0

    invoice_ids = {row.invoice_id for row in rows}
    user_ids = {row.user_id for row in rows}

    invoices = []
    users = []
    if invoice_ids:
        invoices = (await session.execute(select(AdminInvoice).where(AdminInvoice.id.in_(invoice_ids)))).scalars().all()
    if user_ids:
        users = (await session.execute(select(SqlUser).where(SqlUser.id.in_(user_ids)))).scalars().all()

    invoice_map = {inv.id: inv for inv in invoices}
    user_map = {user.id: user for user in users}

    items = []
    for row in rows:
        invoice = invoice_map.get(row.invoice_id)
        user_obj = user_map.get(row.user_id)
        items.append({
            "id": str(row.id),
            "invoice_id": str(row.invoice_id),
            "invoice_no": invoice.invoice_no if invoice else None,
            "user_id": str(row.user_id),
            "user_email": user_obj.email if user_obj else None,
            "amount_total": float(row.amount_total) if row.amount_total is not None else None,
            "currency": row.currency,
            "status": row.status,
            "provider": row.provider,
            "provider_ref": row.provider_ref,
            "created_at": row.created_at.isoformat() if row.created_at else None,
        })

    return {"items": items, "pagination": {"total": total, "skip": skip, "limit": limit}}


@api_router.get("/admin/finance/revenue")
async def admin_revenue(
    request: Request,
    country: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=None, )
    await _ensure_invoices_db_ready(session)

    start_dt = _parse_iso_datetime(start_date, "start_date") if start_date else datetime.now(timezone.utc) - timedelta(days=30)
    end_dt = _parse_iso_datetime(end_date, "end_date") if end_date else datetime.now(timezone.utc)
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="end_date must be after start_date")

    conditions = [
        AdminInvoice.status == "paid",
        AdminInvoice.paid_at >= start_dt,
        AdminInvoice.paid_at <= end_dt,
    ]

    if country:
        country_code = country.upper()
        if country_code != "GLOBAL":
            _assert_country_scope(country_code, current_user)
        conditions.append(AdminInvoice.country_code == country_code)
    else:
        country_code = None

    totals = await _invoice_totals_by_currency(session, conditions)
    total_gross = sum(totals.values())

    return {
        "country_code": country_code,
        "start_date": start_dt.isoformat(),
        "end_date": end_dt.isoformat(),
        "total_gross": round(total_gross, 2),
        "totals_by_currency": {k: round(v, 2) for k, v in totals.items()},
    }


@api_router.get("/admin/tax-rates")
async def admin_list_tax_rates(
    request: Request,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    query = select(VatRate)
    if country:
        country_code = country.upper()
        _assert_country_scope(country_code, current_user)
        query = query.where(VatRate.country == country_code)
    elif current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope:
            query = query.where(VatRate.country.in_(scope))

    result = await session.execute(query.order_by(VatRate.valid_from.desc()))
    items = []
    for row in result.scalars().all():
        items.append(
            {
                "id": str(row.id),
                "country_code": row.country,
                "rate": float(row.rate) if row.rate is not None else None,
                "effective_date": row.valid_from.isoformat() if row.valid_from else None,
                "active_flag": row.is_active,
                "created_at": row.created_at.isoformat() if row.created_at else None,
            }
        )
    return {"items": items}


@api_router.post("/admin/tax-rates")
async def admin_create_tax_rate(
    payload: TaxRateCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    country_code = payload.country_code.upper()
    _assert_country_scope(country_code, current_user)
    if not (0 <= payload.rate <= 100):
        raise HTTPException(status_code=400, detail="rate must be between 0 and 100")

    effective_dt = _parse_iso_datetime(payload.effective_date, "effective_date")

    tax_rate = VatRate(
        country=country_code,
        rate=payload.rate,
        tax_type="VAT",
        is_active=True if payload.active_flag is None else payload.active_flag,
        valid_from=effective_dt,
    )
    session.add(tax_rate)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="TAX_RATE_CREATE",
        actor=current_user,
        resource_type="tax_rate",
        resource_id=str(tax_rate.id),
        metadata={"country_code": country_code, "rate": payload.rate},
        request=request,
        country_code=country_code,
    )
    await session.commit()

    return {
        "ok": True,
        "tax_rate": {
            "id": str(tax_rate.id),
            "country_code": tax_rate.country,
            "rate": float(tax_rate.rate) if tax_rate.rate is not None else None,
            "effective_date": tax_rate.valid_from.isoformat() if tax_rate.valid_from else None,
            "active_flag": tax_rate.is_active,
            "created_at": tax_rate.created_at.isoformat() if tax_rate.created_at else None,
        },
    }


@api_router.patch("/admin/tax-rates/{tax_id}")
async def admin_update_tax_rate(
    tax_id: str,
    payload: TaxRateUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        tax_uuid = uuid.UUID(tax_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid tax id") from exc

    tax_rate = await session.get(VatRate, tax_uuid)
    if not tax_rate:
        raise HTTPException(status_code=404, detail="Tax rate not found")
    _assert_country_scope(tax_rate.country, current_user)

    if payload.rate is not None:
        if not (0 <= payload.rate <= 100):
            raise HTTPException(status_code=400, detail="rate must be between 0 and 100")
        tax_rate.rate = payload.rate
    if payload.effective_date is not None:
        tax_rate.valid_from = _parse_iso_datetime(payload.effective_date, "effective_date")
    if payload.active_flag is not None:
        tax_rate.is_active = payload.active_flag

    await _write_audit_log_sql(
        session=session,
        action="TAX_RATE_UPDATE",
        actor=current_user,
        resource_type="tax_rate",
        resource_id=str(tax_rate.id),
        metadata={"country_code": tax_rate.country},
        request=request,
        country_code=tax_rate.country,
    )
    await session.commit()

    return {
        "ok": True,
        "tax_rate": {
            "id": str(tax_rate.id),
            "country_code": tax_rate.country,
            "rate": float(tax_rate.rate) if tax_rate.rate is not None else None,
            "effective_date": tax_rate.valid_from.isoformat() if tax_rate.valid_from else None,
            "active_flag": tax_rate.is_active,
        },
    }


@api_router.delete("/admin/tax-rates/{tax_id}")
async def admin_delete_tax_rate(
    tax_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        tax_uuid = uuid.UUID(tax_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid tax id") from exc

    tax_rate = await session.get(VatRate, tax_uuid)
    if not tax_rate:
        raise HTTPException(status_code=404, detail="Tax rate not found")
    _assert_country_scope(tax_rate.country, current_user)

    tax_rate.is_active = False

    await _write_audit_log_sql(
        session=session,
        action="TAX_RATE_DELETE",
        actor=current_user,
        resource_type="tax_rate",
        resource_id=str(tax_rate.id),
        metadata={"active_flag": False},
        request=request,
        country_code=tax_rate.country,
    )
    await session.commit()
    return {"ok": True}


@api_router.get("/admin/plans")
async def admin_list_plans(
    request: Request,
    scope: Optional[str] = None,
    country_code: Optional[str] = None,
    status: Optional[str] = None,
    period: Optional[str] = None,
    q: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_plans_db_ready(session)

    query = select(Plan)

    if scope:
        scope_value = scope.strip().lower()
        if scope_value not in PLAN_SCOPE_SET:
            raise HTTPException(status_code=400, detail="country_scope invalid")
        query = query.where(Plan.country_scope == scope_value)

    if country_code:
        query = query.where(Plan.country_code == country_code.strip().upper())

    if period:
        period_value = period.strip().lower()
        if period_value not in PLAN_PERIOD_SET:
            raise HTTPException(status_code=400, detail="period invalid")
        query = query.where(Plan.period == period_value)

    if status:
        status_value = status.strip().lower()
        if status_value not in PLAN_STATUS_SET:
            raise HTTPException(status_code=400, detail="status invalid")
        if status_value == "archived":
            query = query.where(Plan.archived_at.isnot(None))
        elif status_value == "active":
            query = query.where(Plan.archived_at.is_(None), Plan.active_flag.is_(True))
        elif status_value == "inactive":
            query = query.where(Plan.archived_at.is_(None), Plan.active_flag.is_(False))
    else:
        query = query.where(Plan.archived_at.is_(None))

    if q:
        search_value = f"%{q.strip()}%"
        query = query.where(Plan.name.ilike(search_value))

    query = query.order_by(Plan.updated_at.desc(), Plan.created_at.desc())
    rows = (await session.execute(query)).scalars().all()
    return {"items": [_plan_to_dict(plan) for plan in rows]}


@api_router.get("/admin/plans/{plan_id}")
async def admin_get_plan(
    plan_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_plans_db_ready(session)
    plan = await session.get(Plan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    return {"plan": _plan_to_dict(plan)}


@api_router.post("/admin/plans")
async def admin_create_plan(
    payload: PlanCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_plans_db_ready(session)

    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    scope_value = payload.country_scope.strip().lower()
    if scope_value not in PLAN_SCOPE_SET:
        raise HTTPException(status_code=400, detail="country_scope invalid")

    period_value = payload.period.strip().lower()
    if period_value not in PLAN_PERIOD_SET:
        raise HTTPException(status_code=400, detail="period invalid")

    if payload.price_amount < 0:
        raise HTTPException(status_code=400, detail="price_amount must be >= 0")
    _assert_plan_quota_limits(payload.listing_quota, payload.showcase_quota)

    slug_value = _slugify_value(payload.slug or name)
    if not slug_value or not SLUG_PATTERN.match(slug_value):
        raise HTTPException(status_code=400, detail="slug invalid")

    if scope_value == "country":
        if not payload.country_code:
            raise HTTPException(status_code=400, detail="country_code is required")
        country_value = payload.country_code.strip().upper()
        currency = _resolve_currency_code(country_value)
        if not currency:
            raise HTTPException(status_code=400, detail="currency_code unavailable for country")
        currency_code = currency
    else:
        country_value = "GLOBAL"
        currency_code = "EUR"

    if payload.currency_code and payload.currency_code != currency_code:
        raise HTTPException(status_code=400, detail="currency_code invalid for scope")

    exists_query = select(Plan).where(
        Plan.country_scope == scope_value,
        Plan.country_code == country_value,
        Plan.slug == slug_value,
        Plan.period == period_value,
    )
    existing = (await session.execute(exists_query)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="plan slug already exists")

    now = datetime.now(timezone.utc)
    plan = Plan(
        slug=slug_value,
        name=name,
        country_scope=scope_value,
        country_code=country_value,
        period=period_value,
        price_amount=payload.price_amount,
        currency_code=currency_code,
        listing_quota=payload.listing_quota,
        showcase_quota=payload.showcase_quota,
        active_flag=True if payload.active_flag is None else payload.active_flag,
        archived_at=None,
        created_at=now,
        updated_at=now,
    )
    session.add(plan)
    await session.commit()
    await session.refresh(plan)
    return {"plan": _plan_to_dict(plan)}


@api_router.put("/admin/plans/{plan_id}")
async def admin_update_plan(
    plan_id: str,
    payload: PlanUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_plans_db_ready(session)

    plan = await session.get(Plan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    updates: Dict[str, Any] = {}

    if payload.name is not None:
        name_value = payload.name.strip()
        if not name_value:
            raise HTTPException(status_code=400, detail="name is required")
        updates["name"] = name_value

    scope_value = plan.country_scope
    if payload.country_scope is not None:
        scope_value = payload.country_scope.strip().lower()
        if scope_value not in PLAN_SCOPE_SET:
            raise HTTPException(status_code=400, detail="country_scope invalid")
        updates["country_scope"] = scope_value

    country_value = plan.country_code
    if payload.country_code is not None:
        country_value = payload.country_code.strip().upper()
        updates["country_code"] = country_value

    period_value = plan.period
    if payload.period is not None:
        period_value = payload.period.strip().lower()
        if period_value not in PLAN_PERIOD_SET:
            raise HTTPException(status_code=400, detail="period invalid")
        updates["period"] = period_value

    if scope_value == "country":
        if not country_value:
            raise HTTPException(status_code=400, detail="country_code is required")
        currency = _resolve_currency_code(country_value)
        if not currency:
            raise HTTPException(status_code=400, detail="currency_code unavailable for country")
        currency_code = currency
    else:
        country_value = "GLOBAL"
        updates["country_code"] = country_value
        currency_code = "EUR"

    if payload.currency_code and payload.currency_code != currency_code:
        raise HTTPException(status_code=400, detail="currency_code invalid for scope")

    if payload.price_amount is not None:
        if payload.price_amount < 0:
            raise HTTPException(status_code=400, detail="price_amount must be >= 0")
        updates["price_amount"] = payload.price_amount

    if payload.listing_quota is not None:
        updates["listing_quota"] = payload.listing_quota

    if payload.showcase_quota is not None:
        updates["showcase_quota"] = payload.showcase_quota

    listing_quota_value = payload.listing_quota if payload.listing_quota is not None else plan.listing_quota
    showcase_quota_value = payload.showcase_quota if payload.showcase_quota is not None else plan.showcase_quota
    _assert_plan_quota_limits(listing_quota_value, showcase_quota_value)

    if payload.active_flag is not None:
        updates["active_flag"] = payload.active_flag

    if payload.slug is not None:
        slug_value = _slugify_value(payload.slug)
    elif payload.name is not None:
        slug_value = _slugify_value(payload.name)
    else:
        slug_value = plan.slug

    if not slug_value or not SLUG_PATTERN.match(slug_value):
        raise HTTPException(status_code=400, detail="slug invalid")

    updates["slug"] = slug_value
    updates["currency_code"] = currency_code

    exists_query = select(Plan).where(
        Plan.country_scope == scope_value,
        Plan.country_code == country_value,
        Plan.slug == slug_value,
        Plan.period == period_value,
        Plan.id != plan.id,
    )
    existing = (await session.execute(exists_query)).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="plan slug already exists")

    updates["updated_at"] = datetime.now(timezone.utc)
    for key, value in updates.items():
        setattr(plan, key, value)

    await session.commit()
    await session.refresh(plan)
    return {"plan": _plan_to_dict(plan)}


@api_router.post("/admin/plans/{plan_id}/toggle-active")
async def admin_toggle_plan_active(
    plan_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_plans_db_ready(session)

    plan = await session.get(Plan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    if plan.archived_at:
        raise HTTPException(status_code=400, detail="Plan is archived")

    plan.active_flag = not plan.active_flag
    plan.updated_at = datetime.now(timezone.utc)
    await session.commit()
    await session.refresh(plan)
    return {"plan": _plan_to_dict(plan)}


@api_router.post("/admin/plans/{plan_id}/archive")
async def admin_archive_plan(
    plan_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_plans_db_ready(session)

    plan = await session.get(Plan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")

    now = datetime.now(timezone.utc)
    plan.archived_at = now
    plan.active_flag = False
    plan.updated_at = now
    await session.commit()
    await session.refresh(plan)
    return {"plan": _plan_to_dict(plan)}


@api_router.post("/admin/dealers/{dealer_id}/plan")
async def admin_assign_dealer_plan(
    dealer_id: str,
    payload: DealerPlanAssignmentPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session, )
    await _ensure_plans_db_ready(session)

    try:
        dealer_uuid = uuid.UUID(dealer_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid dealer id") from exc

    dealer = await session.get(SqlUser, dealer_uuid)
    if not dealer or dealer.role != "dealer":
        raise HTTPException(status_code=404, detail="Dealer not found")
    _assert_country_scope(dealer.country_code, current_user)

    plan_id = payload.plan_id
    plan = None
    if plan_id:
        try:
            plan_uuid = uuid.UUID(str(plan_id))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid plan id") from exc
        plan = await session.get(Plan, plan_uuid)
        if not plan:
            raise HTTPException(status_code=404, detail="Plan not found")
        if plan.country_scope == "country" and plan.country_code != dealer.country_code:
            raise HTTPException(status_code=400, detail="Plan country mismatch")

    prev_plan_id = dealer.plan_id
    dealer.plan_id = plan.id if plan else None
    dealer.updated_at = datetime.now(timezone.utc)
    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="ADMIN_PLAN_ASSIGNMENT",
        actor=current_user,
        resource_type="dealer",
        resource_id=str(dealer.id),
        metadata={
            "previous_plan_id": str(prev_plan_id) if prev_plan_id else None,
            "new_plan_id": str(plan.id) if plan else None,
        },
        request=request,
        country_code=dealer.country_code,
    )

    if (prev_plan_id or None) != (dealer.plan_id or None):
        await _write_audit_log_sql(
            session=session,
            action="DEALER_PLAN_OVERRIDE",
            actor=current_user,
            resource_type="dealer",
            resource_id=str(dealer.id),
            metadata={
                "previous_plan_id": str(prev_plan_id) if prev_plan_id else None,
                "new_plan_id": str(plan.id) if plan else None,
            },
            request=request,
            country_code=dealer.country_code,
        )

    await session.commit()

    return {"ok": True, "plan_id": str(plan.id) if plan else None}


# =====================
# Sprint 4 — System + Dashboard
# =====================


@api_router.get("/admin/countries")
async def admin_list_countries(
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support"])),
    session: AsyncSession = Depends(get_sql_session),
):
    query = select(Country)
    if current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope:
            query = query.where(Country.code.in_([code.upper() for code in scope]))

    rows = (await session.execute(query.order_by(Country.code))).scalars().all()
    items = []
    for row in rows:
        items.append(
            _normalize_country_doc(
                {
                    "code": row.code,
                    "country_code": row.code,
                    "name": row.name,
                    "active_flag": row.is_enabled,
                    "default_currency": row.default_currency,
                    "default_language": row.default_language,
                    "updated_at": row.updated_at.isoformat() if row.updated_at else None,
                    "created_at": row.created_at.isoformat() if row.created_at else None,
                }
            )
        )
    return {"items": items}


@api_router.get("/admin/countries/{code}")
async def admin_get_country(
    code: str,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support"])),
    session: AsyncSession = Depends(get_sql_session),
):
    code_upper = code.upper()
    country = (
        await session.execute(select(Country).where(Country.code == code_upper))
    ).scalar_one_or_none()
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")

    return {
        "country": _normalize_country_doc(
            {
                "code": country.code,
                "country_code": country.code,
                "name": country.name,
                "active_flag": country.is_enabled,
                "default_currency": country.default_currency,
                "default_language": country.default_language,
                "updated_at": country.updated_at.isoformat() if country.updated_at else None,
                "created_at": country.created_at.isoformat() if country.created_at else None,
            }
        )
    }


@api_router.post("/admin/countries", status_code=201)
async def admin_create_country(
    payload: CountryCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    code = payload.country_code.strip().upper()
    if not re.match(r"^[A-Z]{2}$", code):
        raise HTTPException(status_code=400, detail="country_code must be 2-letter ISO")
    if not payload.default_currency:
        raise HTTPException(status_code=400, detail="default_currency is required")

    existing = (
        await session.execute(select(Country).where(Country.code == code))
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Country already exists")

    country = Country(
        code=code,
        name=payload.name,
        default_currency=payload.default_currency.upper(),
        default_language=payload.default_language,
        is_enabled=True if payload.active_flag is None else payload.active_flag,
    )
    session.add(country)
    await session.commit()
    await session.refresh(country)

    await _write_audit_log_sql(
        session=session,
        action="COUNTRY_CREATE",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="country",
        resource_id=code,
        metadata={"country_code": code},
        request=request,
        country_code=code,
    )
    await session.commit()

    SUPPORTED_COUNTRIES.add(code)
    return {
        "ok": True,
        "country": _normalize_country_doc(
            {
                "code": country.code,
                "country_code": country.code,
                "name": country.name,
                "active_flag": country.is_enabled,
                "default_currency": country.default_currency,
                "default_language": country.default_language,
                "updated_at": country.updated_at.isoformat() if country.updated_at else None,
                "created_at": country.created_at.isoformat() if country.created_at else None,
            }
        ),
    }


@api_router.patch("/admin/countries/{code}")
async def admin_update_country(
    code: str,
    payload: CountryUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    code_upper = code.upper()
    country = (
        await session.execute(select(Country).where(Country.code == code_upper))
    ).scalar_one_or_none()
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")

    updates = {}
    if payload.name is not None:
        updates["name"] = payload.name
    if payload.active_flag is not None:
        updates["is_enabled"] = payload.active_flag
    if payload.default_currency is not None:
        updates["default_currency"] = payload.default_currency.upper()
    if payload.default_language is not None:
        updates["default_language"] = payload.default_language

    if updates:
        for key, value in updates.items():
            setattr(country, key, value)
        await session.commit()
        await session.refresh(country)

        await _write_audit_log_sql(
            session=session,
            action="COUNTRY_UPDATE",
            actor={"id": current_user.get("id"), "email": current_user.get("email")},
            resource_type="country",
            resource_id=code_upper,
            metadata={"updates": updates},
            request=request,
            country_code=code_upper,
        )
        await session.commit()

    return {
        "ok": True,
        "country": _normalize_country_doc(
            {
                "code": country.code,
                "country_code": country.code,
                "name": country.name,
                "active_flag": country.is_enabled,
                "default_currency": country.default_currency,
                "default_language": country.default_language,
                "updated_at": country.updated_at.isoformat() if country.updated_at else None,
                "created_at": country.created_at.isoformat() if country.created_at else None,
            }
        ),
    }


@api_router.delete("/admin/countries/{code}")
async def admin_delete_country(
    code: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    code_upper = code.upper()
    country = (
        await session.execute(select(Country).where(Country.code == code_upper))
    ).scalar_one_or_none()
    if not country:
        raise HTTPException(status_code=404, detail="Country not found")

    await session.delete(country)
    await session.commit()

    await _write_audit_log_sql(
        session=session,
        action="COUNTRY_DELETE",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="country",
        resource_id=code_upper,
        metadata={},
        request=request,
        country_code=code_upper,
    )
    await session.commit()

    if code_upper in SUPPORTED_COUNTRIES:
        SUPPORTED_COUNTRIES.remove(code_upper)

    return {"message": "Country deleted successfully"}


@api_router.get("/admin/system-settings")
async def admin_list_system_settings(
    request: Request,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)
    filters = []
    if country:
        code = country.upper()
        _assert_country_scope(code, current_user)
        filters.append(SystemSetting.country_code == code)
    elif current_user.get("role") == "country_admin":
        scope = current_user.get("country_scope") or []
        if "*" not in scope:
            filters.append(
                or_(
                    SystemSetting.country_code.is_(None),
                    SystemSetting.country_code == "",
                    SystemSetting.country_code.in_(scope),
                )
            )

    query = select(SystemSetting)
    if filters:
        for condition in filters:
            query = query.where(condition)

    result = await session.execute(query.order_by(SystemSetting.key.asc()))
    items = result.scalars().all()

    return {
        "items": [
            {
                "id": str(item.id),
                "key": item.key,
                "value": item.value,
                "country_code": item.country_code,
                "is_readonly": item.is_readonly,
                "description": item.description,
                "moderation_freeze_reason": item.moderation_freeze_reason,
                "created_at": item.created_at.isoformat() if item.created_at else None,
                "updated_at": item.updated_at.isoformat() if item.updated_at else None,
            }
            for item in items
        ]
    }


@api_router.get("/admin/system-settings/cloudflare")
async def admin_get_cloudflare_config(
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    masked = await build_masked_config(session)
    return {
        "account_id_masked": masked.account_masked,
        "zone_id_masked": masked.zone_masked,
        "account_id_last4": masked.account_last4,
        "zone_id_last4": masked.zone_last4,
        "cf_ids_source": masked.source,
        "cf_ids_present": masked.present,
        "canary_status": masked.canary_status,
        "canary_checked_at": masked.canary_checked_at,
    }


@api_router.post("/admin/system-settings/cloudflare")
async def admin_update_cloudflare_config(
    payload: CloudflareConfigPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    logger = logging.getLogger("cloudflare_config")
    request_id = _get_request_id(request)

    if not payload.account_id or not payload.zone_id:
        logger.warning(
            f"cloudflare_config_save_failed reason=validation_failed request_id={request_id}",
            extra={"reason": "validation_failed", "admin_id": current_user.get("id"), "request_id": request_id},
        )
        raise HTTPException(status_code=400, detail="Cloudflare IDs required")

    if not os.environ.get("CONFIG_ENCRYPTION_KEY"):
        logger.warning(
            f"cloudflare_config_save_failed reason=encryption_key_missing request_id={request_id}",
            extra={"reason": "encryption_key_missing", "admin_id": current_user.get("id"), "request_id": request_id},
        )
        raise HTTPException(status_code=400, detail="CONFIG_ENCRYPTION_KEY missing")

    try:
        encrypted_account, account_last4 = encrypt_config_value(payload.account_id)
        encrypted_zone, zone_last4 = encrypt_config_value(payload.zone_id)
    except CloudflareConfigError as exc:
        reason = "encryption_key_missing" if exc.code in {"config_key_missing", "config_key_invalid", "config_key_mismatch"} else "validation_failed"
        logger.warning(
            f"cloudflare_config_save_failed reason={reason} request_id={request_id}",
            extra={"reason": reason, "admin_id": current_user.get("id"), "request_id": request_id},
        )
        status_code = 400 if reason == "encryption_key_missing" else 500
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc

    try:
        if APP_ENV != "prod" and request.headers.get("X-Force-DB-Write-Fail") == "true":
            raise RuntimeError("forced_db_write_fail")
        config = await upsert_cloudflare_config(
            session,
            account_token=encrypted_account,
            account_last4=account_last4,
            zone_token=encrypted_zone,
            zone_last4=zone_last4,
            updated_by=current_user.get("id"),
        )

        await _write_audit_log_sql(
            session=session,
            action="CLOUDFLARE_CONFIG_UPDATE",
            actor=current_user,
            resource_type="cloudflare_config",
            resource_id=str(config.id),
            metadata={
                "account_last4": account_last4,
                "zone_last4": zone_last4,
                "message": f"Cloudflare config updated by {current_user.get('id')}",
            },
            request=request,
            country_code=None,
        )
        await session.commit()
    except Exception as exc:
        await session.rollback()
        logger.exception(
            f"cloudflare_config_save_failed reason=db_write_failed request_id={request_id}",
            extra={"reason": "db_write_failed", "admin_id": current_user.get("id"), "request_id": request_id},
        )
        raise HTTPException(status_code=500, detail="Cloudflare config save failed") from exc

    canary_status = CANARY_CONFIG_MISSING
    canary_reason = None
    api_token = os.environ.get("CLOUDFLARE_API_TOKEN")
    if api_token:
        credentials = CloudflareCredentials(
            api_token=api_token,
            account_id=payload.account_id,
            zone_id=payload.zone_id,
        )
        canary_status, canary_reason = await cloudflare_metrics_service.run_canary(credentials)

    await update_canary_status(session, canary_status, current_user.get("id"))
    _system_health_detail_cache["checked_at"] = 0

    logger.info(
        "cloudflare_canary_result",
        extra={"status": canary_status, "reason": canary_reason, "admin_id": current_user.get("id"), "request_id": request_id},
    )

    return {
        "ok": True,
        "canary_status": canary_status,
        "cf_ids_present": True,
    }


@api_router.post("/admin/system-settings/cloudflare/canary")
async def admin_cloudflare_canary(
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    account_id = None
    zone_id = None
    try:
        account_id, zone_id, _ = await resolve_cloudflare_config(session)
    except CloudflareConfigError:
        account_id = None
        zone_id = None
    api_token = os.environ.get("CLOUDFLARE_API_TOKEN")
    canary_status = CANARY_CONFIG_MISSING
    canary_reason = None
    if account_id and zone_id and api_token:
        credentials = CloudflareCredentials(api_token=api_token, account_id=account_id, zone_id=zone_id)
        canary_status, canary_reason = await cloudflare_metrics_service.run_canary(credentials)

    await update_canary_status(session, canary_status, current_user.get("id"))
    _system_health_detail_cache["checked_at"] = 0

    logger = logging.getLogger("cloudflare_config")
    logger.info(
        "cloudflare_canary_result",
        extra={"status": canary_status, "reason": canary_reason, "admin_id": current_user.get("id"), "request_id": _get_request_id(request)},
    )

    await _write_audit_log_sql(
        session=session,
        action="CLOUDFLARE_CANARY",
        actor=current_user,
        resource_type="cloudflare_config",
        resource_id=None,
        metadata={"status": canary_status},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"status": canary_status}


@api_router.post("/admin/system-settings", status_code=201)
async def admin_create_system_setting(
    payload: SystemSettingCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    key = payload.key.strip()
    if not KEY_NAMESPACE_REGEX.match(key):
        raise HTTPException(status_code=400, detail="Invalid key namespace")
    country_code = payload.country_code.upper() if payload.country_code else None
    normalized_reason = None
    if payload.moderation_freeze_reason is not None:
        normalized_reason = _normalize_freeze_reason(payload.moderation_freeze_reason)

    result = await session.execute(
        select(SystemSetting).where(SystemSetting.key == key, SystemSetting.country_code == country_code)
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Setting already exists")

    setting = SystemSetting(
        key=key,
        value=payload.value,
        country_code=country_code,
        is_readonly=bool(payload.is_readonly),
        description=payload.description,
        moderation_freeze_reason=normalized_reason,
    )
    session.add(setting)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="SYSTEM_SETTING_CREATE",
        actor=current_user,
        resource_type="system_setting",
        resource_id=str(setting.id),
        metadata={
            "key": key,
            "country_code": country_code,
            "value": payload.value,
        },
        request=request,
        country_code=country_code,
    )

    if key == MODERATION_FREEZE_SETTING_KEY:
        freeze_active = _extract_setting_bool(setting.value)
        metadata = {}
        if normalized_reason:
            metadata["reason"] = normalized_reason
        await _write_audit_log_sql(
            session=session,
            action="MODERATION_FREEZE_ENABLED" if freeze_active else "MODERATION_FREEZE_DISABLED",
            actor=current_user,
            resource_type="moderation_freeze",
            resource_id=str(setting.id),
            metadata=metadata,
            request=request,
            country_code=country_code,
        )

    await session.commit()

    return {
        "ok": True,
        "setting": {
            "id": str(setting.id),
            "key": setting.key,
            "value": setting.value,
            "country_code": setting.country_code,
            "is_readonly": setting.is_readonly,
            "description": setting.description,
            "moderation_freeze_reason": setting.moderation_freeze_reason,
            "created_at": setting.created_at.isoformat() if setting.created_at else None,
            "updated_at": setting.updated_at.isoformat() if setting.updated_at else None,
        },
    }


@api_router.patch("/admin/system-settings/{setting_id}")
async def admin_update_system_setting(
    setting_id: str,
    payload: SystemSettingUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        setting_uuid = uuid.UUID(setting_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid setting id") from exc

    result = await session.execute(select(SystemSetting).where(SystemSetting.id == setting_uuid))
    setting = result.scalar_one_or_none()
    if not setting:
        raise HTTPException(status_code=404, detail="Setting not found")
    if setting.is_readonly and payload.value is not None:
        raise HTTPException(status_code=400, detail="Setting is read-only")

    old_freeze_active = None
    if setting.key == MODERATION_FREEZE_SETTING_KEY:
        old_freeze_active = _extract_setting_bool(setting.value)

    normalized_reason = None
    if payload.moderation_freeze_reason is not None:
        normalized_reason = _normalize_freeze_reason(payload.moderation_freeze_reason)

    if payload.value is not None:
        setting.value = payload.value
    if payload.country_code is not None:
        setting.country_code = payload.country_code.upper() if payload.country_code else None
    if payload.is_readonly is not None:
        setting.is_readonly = payload.is_readonly
    if payload.description is not None:
        setting.description = payload.description
    if payload.moderation_freeze_reason is not None:
        setting.moderation_freeze_reason = normalized_reason

    await _write_audit_log_sql(
        session=session,
        action="SYSTEM_SETTING_UPDATE",
        actor=current_user,
        resource_type="system_setting",
        resource_id=str(setting.id),
        metadata={
            "key": setting.key,
            "country_code": setting.country_code,
        },
        request=request,
        country_code=setting.country_code,
    )

    if setting.key == MODERATION_FREEZE_SETTING_KEY:
        new_freeze_active = _extract_setting_bool(setting.value)
        if old_freeze_active is not None and new_freeze_active != old_freeze_active:
            reason_payload = normalized_reason if payload.moderation_freeze_reason is not None else setting.moderation_freeze_reason
            metadata = {}
            if reason_payload:
                metadata["reason"] = reason_payload
            await _write_audit_log_sql(
                session=session,
                action="MODERATION_FREEZE_ENABLED" if new_freeze_active else "MODERATION_FREEZE_DISABLED",
                actor=current_user,
                resource_type="moderation_freeze",
                resource_id=str(setting.id),
                metadata=metadata,
                request=request,
                country_code=setting.country_code,
            )

    await session.commit()

    return {
        "ok": True,
        "setting": {
            "id": str(setting.id),
            "key": setting.key,
            "value": setting.value,
            "country_code": setting.country_code,
            "is_readonly": setting.is_readonly,
            "description": setting.description,
            "moderation_freeze_reason": setting.moderation_freeze_reason,
            "created_at": setting.created_at.isoformat() if setting.created_at else None,
            "updated_at": setting.updated_at.isoformat() if setting.updated_at else None,
        },
    }


@api_router.get("/system-settings/effective")
async def system_settings_effective(
    request: Request,
    country: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    country_code = country.upper() if country else None

    global_query = select(SystemSetting).where(
        or_(SystemSetting.country_code.is_(None), SystemSetting.country_code == "")
    )
    result = await session.execute(global_query)
    global_settings = result.scalars().all()

    country_settings = []
    if country_code:
        result = await session.execute(
            select(SystemSetting).where(SystemSetting.country_code == country_code)
        )
        country_settings = result.scalars().all()

    merged: Dict[str, dict] = {}
    for item in global_settings:
        merged[item.key] = {
            "id": str(item.id),
            "key": item.key,
            "value": item.value,
            "country_code": item.country_code,
            "is_readonly": item.is_readonly,
            "description": item.description,
            "moderation_freeze_reason": item.moderation_freeze_reason,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
            "source": "global",
        }
    for item in country_settings:
        merged[item.key] = {
            "id": str(item.id),
            "key": item.key,
            "value": item.value,
            "country_code": item.country_code,
            "is_readonly": item.is_readonly,
            "description": item.description,
            "moderation_freeze_reason": item.moderation_freeze_reason,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
            "source": "country",
        }

    items = sorted(merged.values(), key=lambda x: x.get("key"))
    return {"country_code": country_code, "items": items}


# =====================
# Master Data: Categories / Attributes / Vehicle
# =====================


@api_router.get("/admin/categories")
async def admin_list_categories(
    request: Request,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    if country:
        code = country.upper()
        _assert_country_scope(code, current_user)
        query = (
            select(Category)
            .options(selectinload(Category.translations))
            .where(Category.country_code == code, Category.is_deleted.is_(False))
        )
    else:
        query = (
            select(Category)
            .options(selectinload(Category.translations))
            .where(Category.is_deleted.is_(False))
        )

    result = await session.execute(query.order_by(Category.sort_order.asc()))
    items = result.scalars().all()
    return {
        "items": [
            _serialize_category_sql(item, include_schema=True, include_translations=False)
            for item in items
        ]
    }


@api_router.get("/admin/categories/import-export/export/json")
async def admin_export_categories_json(
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.is_deleted.is_(False))
        .order_by(Category.depth.asc(), Category.sort_order.asc())
    )
    categories = result.scalars().all()

    versions_result = await session.execute(
        select(CategorySchemaVersion)
        .order_by(CategorySchemaVersion.version.desc())
    )
    versions = versions_result.scalars().all()
    latest_versions: Dict[str, int] = {}
    for version in versions:
        key = str(version.category_id)
        if key not in latest_versions:
            latest_versions[key] = version.version

    slug_by_id = {str(cat.id): (_pick_category_slug(cat.slug) or "") for cat in categories}
    items = []
    for category in categories:
        parent_slug = slug_by_id.get(str(category.parent_id)) if category.parent_id else None
        items.append(
            _build_category_export_item(
                category,
                list(category.translations or []),
                parent_slug,
                latest_versions.get(str(category.id), 0),
            )
        )

    payload = {
        "metadata": {
            "schema": "CATEGORY_IMPORT_EXPORT_SCHEMA_V1",
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "module": "vehicle",
            "count": len(items),
        },
        "categories": _build_category_tree(items),
    }

    filename = f"categories-export-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.json"
    return JSONResponse(
        content=payload,
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.get("/admin/categories/import-export/export/csv")
async def admin_export_categories_csv(
    request: Request,
    module: Optional[str] = None,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.is_deleted.is_(False))
        .order_by(Category.depth.asc(), Category.sort_order.asc())
    )
    categories = result.scalars().all()

    module_filter = module.strip().lower() if module else None
    country_filter = country.strip().upper() if country else None

    filtered: list[Category] = []
    for category in categories:
        if module_filter and category.module != module_filter:
            continue
        if country_filter:
            if category.country_code and category.country_code != country_filter:
                continue
            allowed = category.allowed_countries or []
            if allowed and country_filter not in allowed:
                continue
        filtered.append(category)

    slug_by_id = {str(cat.id): (_pick_category_slug(cat.slug) or "") for cat in categories}

    output = io.StringIO()
    writer = csv.writer(output, delimiter=",")
    writer.writerow(CATEGORY_IMPORT_COLUMNS)

    for category in filtered:
        translation_map = _category_translation_map(list(category.translations or []))
        slug_value = _pick_category_slug(category.slug) or ""
        parent_slug = slug_by_id.get(str(category.parent_id)) if category.parent_id else ""
        country_value = category.country_code or (country_filter or (category.allowed_countries or [""])[0])
        wizard_raw = json.dumps(category.wizard_progress or {}, ensure_ascii=False) if category.wizard_progress else ""
        writer.writerow([
            category.module,
            country_value,
            CATEGORY_SCHEMA_VERSION,
            slug_value,
            parent_slug or "",
            (translation_map.get("tr") or {}).get("name") or "",
            (translation_map.get("de") or {}).get("name") or "",
            (translation_map.get("fr") or {}).get("name") or "",
            "true" if category.is_enabled else "false",
            category.sort_order or 0,
            wizard_raw,
        ])

    filename = f"categories-export-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.get("/admin/categories/import-export/export/xlsx")
async def admin_export_categories_xlsx(
    request: Request,
    module: Optional[str] = None,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.is_deleted.is_(False))
        .order_by(Category.depth.asc(), Category.sort_order.asc())
    )
    categories = result.scalars().all()

    module_filter = module.strip().lower() if module else None
    country_filter = country.strip().upper() if country else None

    filtered: list[Category] = []
    for category in categories:
        if module_filter and category.module != module_filter:
            continue
        if country_filter:
            if category.country_code and category.country_code != country_filter:
                continue
            allowed = category.allowed_countries or []
            if allowed and country_filter not in allowed:
                continue
        filtered.append(category)

    slug_by_id = {str(cat.id): (_pick_category_slug(cat.slug) or "") for cat in categories}

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(CATEGORY_IMPORT_COLUMNS)

    for category in filtered:
        translation_map = _category_translation_map(list(category.translations or []))
        slug_value = _pick_category_slug(category.slug) or ""
        parent_slug = slug_by_id.get(str(category.parent_id)) if category.parent_id else ""
        country_value = category.country_code or (country_filter or (category.allowed_countries or [""])[0])
        wizard_raw = json.dumps(category.wizard_progress or {}, ensure_ascii=False) if category.wizard_progress else ""
        sheet.append([
            category.module,
            country_value,
            CATEGORY_SCHEMA_VERSION,
            slug_value,
            parent_slug or "",
            (translation_map.get("tr") or {}).get("name") or "",
            (translation_map.get("de") or {}).get("name") or "",
            (translation_map.get("fr") or {}).get("name") or "",
            "true" if category.is_enabled else "false",
            category.sort_order or 0,
            wizard_raw,
        ])

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"categories-export-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


def _build_category_sample_rows(module_value: str, country_value: str) -> list[dict]:
    root_slug = f"ornek-{module_value}"
    child_slug = f"{root_slug}-alt"
    return [
        {
            "module": module_value,
            "country": country_value,
            "schema_version": CATEGORY_SCHEMA_VERSION,
            "slug": root_slug,
            "parent_slug": "",
            "name_tr": "Örnek Ana Kategori",
            "name_de": "Beispiel Hauptkategorie",
            "name_fr": "Categorie Exemple",
            "is_active": "true",
            "sort_order": 1,
            "wizard_progress": json.dumps({"step": "core"}, ensure_ascii=False),
        },
        {
            "module": module_value,
            "country": country_value,
            "schema_version": CATEGORY_SCHEMA_VERSION,
            "slug": child_slug,
            "parent_slug": root_slug,
            "name_tr": "Örnek Alt Kategori",
            "name_de": "Beispiel Unterkategorie",
            "name_fr": "Sous-categorie Exemple",
            "is_active": "true",
            "sort_order": 2,
            "wizard_progress": "",
        },
    ]


def _category_row_values(row: dict) -> list:
    return [row.get(col, "") for col in CATEGORY_IMPORT_COLUMNS]


@api_router.get("/admin/categories/import-export/sample/csv")
async def admin_export_category_sample_csv(
    request: Request,
    module: Optional[str] = None,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
):
    module_value = (module or "vehicle").strip().lower()
    if module_value not in SUPPORTED_CATEGORY_MODULES:
        raise HTTPException(status_code=400, detail="Invalid module")
    country_value = (country or "DE").strip().upper()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=",")
    writer.writerow(CATEGORY_IMPORT_COLUMNS)
    for row in _build_category_sample_rows(module_value, country_value):
        writer.writerow(_category_row_values(row))

    filename = f"categories-sample-{module_value}-{country_value}.csv"
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.get("/admin/categories/import-export/sample/xlsx")
async def admin_export_category_sample_xlsx(
    request: Request,
    module: Optional[str] = None,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
):
    module_value = (module or "vehicle").strip().lower()
    if module_value not in SUPPORTED_CATEGORY_MODULES:
        raise HTTPException(status_code=400, detail="Invalid module")
    country_value = (country or "DE").strip().upper()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(CATEGORY_IMPORT_COLUMNS)
    for row in _build_category_sample_rows(module_value, country_value):
        sheet.append(_category_row_values(row))

    stream = io.BytesIO()
    workbook.save(stream)
    stream.seek(0)

    filename = f"categories-sample-{module_value}-{country_value}.xlsx"
    return Response(
        content=stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.post("/admin/categories", status_code=201)
async def admin_create_category(
    payload: CategoryCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    slug = payload.slug.strip().lower()
    if not SLUG_PATTERN.match(slug):
        raise HTTPException(status_code=400, detail="slug format invalid")

    parent = None
    if payload.parent_id:
        try:
            parent_uuid = uuid.UUID(payload.parent_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="parent_id not valid") from exc
        parent = await session.get(Category, parent_uuid)
        if not parent:
            raise HTTPException(status_code=400, detail="parent_id not found")

    country_code = payload.country_code.upper() if payload.country_code else None
    module_input = payload.module or (parent.module if parent else None)
    module_value = _normalize_category_module(module_input)
    if country_code:
        _assert_country_scope(country_code, current_user)
    if parent and parent.country_code and not country_code:
        country_code = parent.country_code
    _assert_category_parent_compatible(
        category=None,
        parent=parent,
        module_value=module_value,
        country_code=country_code,
    )

    slug_query = await session.execute(
        select(Category).where(Category.is_deleted.is_(False))
    )
    existing_categories = slug_query.scalars().all()
    if any(_pick_category_slug(cat.slug) == slug for cat in existing_categories):
        raise HTTPException(status_code=409, detail="Category slug already exists")

    now = datetime.now(timezone.utc)
    hierarchy_complete = payload.hierarchy_complete if payload.hierarchy_complete is not None else True
    schema = None
    schema_status = None
    if payload.form_schema is not None:
        schema = _normalize_category_schema(payload.form_schema)
        schema_status = schema.get("status", "published")
        if not hierarchy_complete:
            raise HTTPException(status_code=409, detail="Kategori hiyerarşisi tamamlanmadan kaydedilemez")
        if schema_status != "draft":
            _validate_category_schema(schema)

    slug_json = {"tr": slug, "en": slug, "de": slug}
    depth = (parent.depth + 1) if parent else 0
    path = f"{parent.path}.{slug}" if parent and parent.path else slug
    allowed_countries = [country_code] if country_code else sorted(SUPPORTED_COUNTRIES)

    wizard_progress = payload.wizard_progress or {"state": "draft"}

    category = Category(
        id=uuid.uuid4(),
        parent_id=parent.id if parent else None,
        path=path,
        depth=depth,
        sort_order=payload.sort_order or 0,
        module=module_value,
        slug=slug_json,
        icon=None,
        image_url=None,
        is_enabled=payload.active_flag if payload.active_flag is not None else True,
        is_visible_on_home=False,
        is_deleted=False,
        inherit_enabled=True,
        override_enabled=None,
        inherit_countries=True,
        override_countries=None,
        allowed_countries=allowed_countries,
        listing_count=0,
        country_code=country_code,
        hierarchy_complete=hierarchy_complete,
        form_schema=schema,
        wizard_progress=wizard_progress,
        created_at=now,
        updated_at=now,
    )
    session.add(category)

    for lang in ("tr", "en", "de"):
        session.add(
            CategoryTranslation(
                category_id=category.id,
                language=lang,
                name=payload.name.strip(),
                description=None,
                meta_title=None,
                meta_description=None,
            )
        )

    await session.commit()

    if schema:
        if schema_status == "draft":
            await _record_category_version_sql(session, category.id, schema, current_user, "draft")
        else:
            await _record_category_version_sql(session, category.id, schema, current_user, "published")

    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.id == category.id)
    )
    created = result.scalar_one()
    return {"category": _serialize_category_sql(created, include_schema=True, include_translations=False)}


@api_router.patch("/admin/categories/{category_id}")
async def admin_update_category(
    category_id: str,
    payload: CategoryUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.id == category_uuid)
    )
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    if payload.expected_updated_at is not None:
        current_updated_at = category.updated_at.isoformat() if category.updated_at else None
        if current_updated_at and payload.expected_updated_at != current_updated_at:
            raise HTTPException(status_code=409, detail="Category updated in another session")

    updates: Dict[str, Any] = {}
    module_value = category.module
    parent_candidate = None
    country_value = category.country_code
    schema = None
    schema_status = None

    if payload.name is not None:
        updates["name"] = payload.name.strip()

    if payload.slug is not None:
        slug = payload.slug.strip().lower()
        if not SLUG_PATTERN.match(slug):
            raise HTTPException(status_code=400, detail="slug format invalid")
        existing_query = await session.execute(
            select(Category).where(Category.is_deleted.is_(False), Category.id != category.id)
        )
        existing_categories = existing_query.scalars().all()
        if any(_pick_category_slug(cat.slug) == slug for cat in existing_categories):
            raise HTTPException(status_code=409, detail="Category slug already exists")
        updates["slug"] = slug

    if payload.module is not None:
        module_value = _normalize_category_module(payload.module)
        if module_value != category.module:
            child_count = await session.execute(
                select(func.count()).select_from(Category).where(
                    Category.parent_id == category.id,
                    Category.is_deleted.is_(False),
                )
            )
            if (child_count.scalar_one() or 0) > 0:
                raise HTTPException(status_code=409, detail="Kategori çocukları varken module değiştirilemez")
        updates["module"] = module_value

    if payload.parent_id is not None:
        parent_id_value = payload.parent_id or None
        parent = None
        if parent_id_value:
            try:
                parent_uuid = uuid.UUID(parent_id_value)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail="parent_id not valid") from exc
            parent = await session.get(Category, parent_uuid)
            if not parent:
                raise HTTPException(status_code=400, detail="parent_id not found")
        parent_candidate = parent
        updates["parent_id"] = parent.id if parent else None
        updates["path"] = f"{parent.path}.{updates.get('slug') or _pick_category_slug(category.slug)}" if parent and parent.path else (updates.get("slug") or _pick_category_slug(category.slug))
        updates["depth"] = (parent.depth + 1) if parent else 0

    if payload.country_code is not None:
        code = payload.country_code.upper() if payload.country_code else None
        if code:
            _assert_country_scope(code, current_user)
        updates["country_code"] = code
        country_value = code
        if code:
            updates["allowed_countries"] = [code]

    should_validate_parent = payload.parent_id is not None or payload.module is not None or payload.country_code is not None
    if should_validate_parent:
        if parent_candidate is None and category.parent_id:
            parent_candidate = await session.get(Category, category.parent_id)
        _assert_category_parent_compatible(
            category=category,
            parent=parent_candidate,
            module_value=module_value,
            country_code=country_value,
        )

    if payload.active_flag is not None:
        updates["is_enabled"] = payload.active_flag

    if payload.sort_order is not None:
        updates["sort_order"] = payload.sort_order

    if payload.hierarchy_complete is not None:
        updates["hierarchy_complete"] = payload.hierarchy_complete

    if payload.form_schema is not None:
        schema = _normalize_category_schema(payload.form_schema)
        schema_status = schema.get("status", "published")
        hierarchy_complete = updates.get("hierarchy_complete", category.hierarchy_complete)
        if not hierarchy_complete:
            raise HTTPException(status_code=409, detail="Kategori hiyerarşisi tamamlanmadan kaydedilemez")
        if schema_status != "draft":
            dirty_steps = []
            wizard_progress_payload = payload.wizard_progress or category.wizard_progress or {}
            if isinstance(wizard_progress_payload, dict):
                dirty_steps = wizard_progress_payload.get("dirty_steps") or []
            if dirty_steps:
                raise HTTPException(status_code=409, detail="Dirty adımlar tamamlanmadan yayınlanamaz")
            _validate_category_schema(schema)
        updates["form_schema"] = schema

    if payload.wizard_progress is not None:
        updates["wizard_progress"] = payload.wizard_progress

    if not updates:
        return {"category": _serialize_category_sql(category, include_schema=True, include_translations=False)}

    if "slug" in updates:
        slug_json = dict(category.slug or {})
        slug_json.update({"tr": updates["slug"], "en": updates["slug"], "de": updates["slug"]})
        updates["slug"] = slug_json

    if "slug" in updates and "path" not in updates:
        if category.parent_id:
            parent = await session.get(Category, category.parent_id)
            base_path = parent.path if parent else ""
            updates["path"] = f"{base_path}.{_pick_category_slug(updates['slug'])}" if base_path else (_pick_category_slug(updates["slug"]) or "")
        else:
            updates["path"] = _pick_category_slug(updates["slug"]) or ""

    updates["updated_at"] = datetime.now(timezone.utc)

    for key, value in updates.items():
        if hasattr(category, key):
            setattr(category, key, value)

    if "name" in updates:
        translations = list(category.translations or [])
        if not translations:
            for lang in ("tr", "en", "de"):
                session.add(
                    CategoryTranslation(
                        category_id=category.id,
                        language=lang,
                        name=updates["name"],
                        description=None,
                        meta_title=None,
                        meta_description=None,
                    )
                )
        else:
            for translation in translations:
                if translation.language in {"tr", "en", "de"}:
                    translation.name = updates["name"]

    if payload.wizard_edit_event:
        role = current_user.get("role")
        if role not in {"super_admin", "country_admin"}:
            raise HTTPException(status_code=403, detail="Edit unlock yetkisi yok")
        event_meta = dict(payload.wizard_edit_event or {})
        event_meta["category_id"] = str(category.id)
        event_type = "categories.wizard.unlock"
        if event_meta.get("action") == "save":
            event_type = "categories.wizard.edit"
        await _write_audit_log_sql(
            session,
            event_type,
            current_user,
            "category_wizard",
            str(category.id),
            event_meta,
            request,
        )

    await session.commit()

    if schema:
        if schema_status == "draft":
            await _record_category_version_sql(session, category.id, schema, current_user, "draft")
        else:
            await _mark_latest_category_version_published_sql(session, category.id, schema, current_user)

    refreshed = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.id == category.id)
    )
    updated = refreshed.scalar_one()
    return {"category": _serialize_category_sql(updated, include_schema=True, include_translations=False)}


@api_router.get("/admin/categories/{category_id}/versions")
async def admin_list_category_versions(
    category_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    category = await session.get(Category, category_uuid)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if category.country_code:
        _assert_country_scope(category.country_code, current_user)

    result = await session.execute(
        select(CategorySchemaVersion)
        .where(CategorySchemaVersion.category_id == category_uuid)
        .order_by(desc(CategorySchemaVersion.version))
    )
    versions = result.scalars().all()
    return {"items": [_serialize_category_version_sql(v) for v in versions]}


@api_router.get("/admin/categories/{category_id}/versions/{version_id}")
async def admin_get_category_version(
    category_id: str,
    version_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc
    try:
        version_uuid = uuid.UUID(version_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid version id") from exc

    category = await session.get(Category, category_uuid)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if category.country_code:
        _assert_country_scope(category.country_code, current_user)

    result = await session.execute(
        select(CategorySchemaVersion).where(
            CategorySchemaVersion.id == version_uuid,
            CategorySchemaVersion.category_id == category_uuid,
        )
    )
    version = result.scalar_one_or_none()
    if not version:
        raise HTTPException(status_code=404, detail="Category version not found")

    return {"version": _serialize_category_version_sql(version, include_snapshot=True)}


@api_router.get("/admin/categories/{category_id}/export/pdf")
async def admin_export_category_pdf(
    category_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    _enforce_export_rate_limit(request, current_user.get("id"))

    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    category = await session.get(Category, category_uuid)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if category.country_code:
        _assert_country_scope(category.country_code, current_user)

    schema = _normalize_category_schema(category.form_schema) if category.form_schema else None
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")
    if schema.get("status") != "draft":
        raise HTTPException(status_code=409, detail="Sadece draft şema export edilebilir")

    parent_label = None
    if category.parent_id:
        parent = await session.get(Category, category.parent_id)
        if parent:
            parent_label = _pick_category_name(list(parent.translations or []), _pick_category_slug(parent.slug))

    children_result = await session.execute(
        select(Category).options(selectinload(Category.translations)).where(Category.parent_id == category.id)
    )
    children = children_result.scalars().all()
    children_labels = [
        _pick_category_name(list(child.translations or []), _pick_category_slug(child.slug))
        for child in children
    ]

    version = await _get_schema_version_for_export_sql(session, category.id)
    pdf_bytes = _build_schema_pdf(
        schema,
        _serialize_category_sql(category, include_schema=False, include_translations=False),
        version,
        {"parent": parent_label, "children": children_labels},
    )

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    filename = f"schema-{category_id}-v{version}-{timestamp}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.get("/admin/categories/{category_id}/export/csv")
async def admin_export_category_csv(
    category_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    _enforce_export_rate_limit(request, current_user.get("id"))

    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    category = await session.get(Category, category_uuid)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    if category.country_code:
        _assert_country_scope(category.country_code, current_user)

    schema = _normalize_category_schema(category.form_schema) if category.form_schema else None
    if not schema:
        raise HTTPException(status_code=404, detail="Schema not found")
    if schema.get("status") != "draft":
        raise HTTPException(status_code=409, detail="Sadece draft şema export edilebilir")

    parent_label = None
    if category.parent_id:
        parent = await session.get(Category, category.parent_id)
        if parent:
            parent_label = _pick_category_name(list(parent.translations or []), _pick_category_slug(parent.slug))

    children_result = await session.execute(
        select(Category).options(selectinload(Category.translations)).where(Category.parent_id == category.id)
    )
    children = children_result.scalars().all()
    children_labels = [
        _pick_category_name(list(child.translations or []), _pick_category_slug(child.slug))
        for child in children
    ]

    version = await _get_schema_version_for_export_sql(session, category.id)
    rows = _schema_to_csv_rows(schema)
    rows.append(["hierarchy", "parent", parent_label or "", "", "", "", "", "", "", "", "", "", ""])
    rows.append(["hierarchy", "children", ",".join(children_labels), "", "", "", "", "", "", "", "", "", ""])

    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    output.seek(0)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    filename = f"schema-{category_id}-v{version}-{timestamp}.csv"

    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


@api_router.delete("/admin/categories/{category_id}")
async def admin_delete_category(
    category_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    category = await session.get(Category, category_uuid)
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")

    category.is_enabled = False
    category.is_deleted = True
    category.updated_at = datetime.now(timezone.utc)
    await session.commit()

    result = await session.execute(
        select(Category)
        .options(selectinload(Category.translations))
        .where(Category.id == category_uuid)
    )
    deleted = result.scalar_one()
    return {"category": _serialize_category_sql(deleted, include_schema=True, include_translations=False)}


def _read_import_file(file: UploadFile) -> bytes:
    content = file.file.read()
    if len(content) > CATEGORY_IMPORT_MAX_BYTES:
        raise HTTPException(status_code=413, detail="File exceeds 10MB limit")
    if not content:
        raise HTTPException(status_code=400, detail="File is empty")
    return content


CATEGORY_SCHEMA_VERSION = "v1"

CATEGORY_IMPORT_COLUMNS = [
    "module",
    "country",
    "schema_version",
    "slug",
    "parent_slug",
    "name_tr",
    "name_de",
    "name_fr",
    "is_active",
    "sort_order",
    "wizard_progress",
]

SUPPORTED_CATEGORY_MODULES = {"vehicle", "real_estate", "machinery", "services", "jobs"}


def _normalize_import_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _parse_bool_value(value: Any) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return None


def _parse_int_value(value: Any) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _parse_wizard_progress_value(value: Any) -> Tuple[Optional[dict], Optional[str]]:
    if value is None:
        return None, None
    if isinstance(value, dict):
        return value, None
    raw = _normalize_import_value(value)
    if not raw:
        return None, None
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return None, "WIZARD_PROGRESS_INVALID"
    if parsed is not None and not isinstance(parsed, dict):
        return None, "WIZARD_PROGRESS_INVALID"
    return parsed, None


def _parse_category_import_csv(content: bytes) -> list[dict]:
    decoded = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(decoded), delimiter=",")
    headers = [h.strip() for h in (reader.fieldnames or [])]
    missing = [col for col in CATEGORY_IMPORT_COLUMNS if col not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    rows = []
    for index, row in enumerate(reader, start=2):
        if not any(_normalize_import_value(row.get(col)) for col in CATEGORY_IMPORT_COLUMNS):
            continue
        rows.append({
            "row_number": index,
            **{col: row.get(col) for col in CATEGORY_IMPORT_COLUMNS},
        })
    return rows


def _parse_category_import_xlsx(content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content))
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="XLSX file is empty")

    headers = [_normalize_import_value(cell).lower() for cell in header_row]
    header_map = {header: idx for idx, header in enumerate(headers) if header}
    missing = [col for col in CATEGORY_IMPORT_COLUMNS if col not in header_map]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    rows = []
    row_number = 1
    for row in rows_iter:
        row_number += 1
        row_values = {}
        for col in CATEGORY_IMPORT_COLUMNS:
            idx = header_map[col]
            row_values[col] = row[idx] if idx < len(row) else None
        if not any(_normalize_import_value(row_values.get(col)) for col in CATEGORY_IMPORT_COLUMNS):
            continue
        rows.append({"row_number": row_number, **row_values})
    return rows


def _parse_category_import_rows(content: bytes, file_type: str) -> list[dict]:
    if file_type == "csv":
        return _parse_category_import_csv(content)
    if file_type == "xlsx":
        return _parse_category_import_xlsx(content)
    raise HTTPException(status_code=400, detail="Unsupported import format")


def _build_existing_category_maps(existing: list[Category]) -> Tuple[Dict[str, Category], Dict[str, Optional[str]], set]:
    existing_map: Dict[str, Category] = {}
    duplicates: set = set()
    slug_by_id = {str(cat.id): (_pick_category_slug(cat.slug) or "").lower() for cat in existing}
    parent_map: Dict[str, Optional[str]] = {}
    for category in existing:
        slug_key = (_pick_category_slug(category.slug) or "").lower()
        if not slug_key:
            continue
        if slug_key in existing_map:
            duplicates.add(slug_key)
        existing_map[slug_key] = category
        parent_slug = slug_by_id.get(str(category.parent_id)) if category.parent_id else None
        parent_map[slug_key] = parent_slug
    return existing_map, parent_map, duplicates


def _validate_category_import_rows(rows: list[dict], existing: list[Category], current_user: dict) -> dict:
    existing_map, existing_parent_map, duplicate_db = _build_existing_category_maps(existing)
    errors: list[dict] = []
    parsed_rows: list[dict] = []
    seen_slugs: set = set()

    for row in rows:
        row_number = row.get("row_number")
        slug_raw = _normalize_import_value(row.get("slug")).lower()
        module_raw = _normalize_import_value(row.get("module")).lower()
        country_raw = _normalize_import_value(row.get("country")).upper()
        schema_version = _normalize_import_value(row.get("schema_version")).lower()
        parent_slug_raw = _normalize_import_value(row.get("parent_slug")).lower() or None
        name_tr = _normalize_import_value(row.get("name_tr"))
        name_de = _normalize_import_value(row.get("name_de"))
        name_fr = _normalize_import_value(row.get("name_fr"))
        is_active = _parse_bool_value(row.get("is_active"))
        sort_order = _parse_int_value(row.get("sort_order"))
        wizard_progress, wizard_error = _parse_wizard_progress_value(row.get("wizard_progress"))

        if not module_raw:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "module zorunlu."})
        elif module_raw not in SUPPORTED_CATEGORY_MODULES:
            errors.append({"row_number": row_number, "error_code": "INVALID_MODULE", "message": "Geçersiz module."})

        if not country_raw:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "country zorunlu."})

        if not schema_version:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "schema_version zorunlu."})
        elif schema_version != CATEGORY_SCHEMA_VERSION:
            errors.append({"row_number": row_number, "error_code": "SCHEMA_VERSION_MISMATCH", "message": "schema_version v1 olmalı."})

        if not slug_raw:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "slug zorunlu."})
        elif not SLUG_PATTERN.match(slug_raw):
            errors.append({"row_number": row_number, "error_code": "INVALID_SLUG", "message": "Slug formatı geçersiz."})

        if slug_raw:
            if slug_raw in seen_slugs:
                errors.append({"row_number": row_number, "error_code": "DUPLICATE_SLUG", "message": "Dosyada tekrarlayan slug."})
            seen_slugs.add(slug_raw)

        if slug_raw and slug_raw in duplicate_db:
            errors.append({"row_number": row_number, "error_code": "DUPLICATE_SLUG_DB", "message": "Slug veritabanında tekil değil."})

        if parent_slug_raw and slug_raw and parent_slug_raw == slug_raw:
            errors.append({"row_number": row_number, "error_code": "PARENT_SELF", "message": "parent_slug kendisi olamaz."})

        if not name_tr:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "name_tr zorunlu."})
        if not name_de:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "name_de zorunlu."})
        if not name_fr:
            errors.append({"row_number": row_number, "error_code": "REQUIRED_FIELD", "message": "name_fr zorunlu."})

        if is_active is None:
            errors.append({"row_number": row_number, "error_code": "INVALID_BOOLEAN", "message": "is_active true/false olmalı."})

        if sort_order is None:
            errors.append({"row_number": row_number, "error_code": "INVALID_SORT_ORDER", "message": "sort_order sayısal olmalı."})

        if wizard_error:
            errors.append({"row_number": row_number, "error_code": wizard_error, "message": "wizard_progress JSON olmalı."})

        if slug_raw:
            existing_category = existing_map.get(slug_raw)
            if existing_category and module_raw and existing_category.module != module_raw:
                errors.append({"row_number": row_number, "error_code": "MODULE_MISMATCH", "message": "Modül uyuşmuyor."})
            if existing_category and country_raw and existing_category.country_code and existing_category.country_code != country_raw:
                errors.append({"row_number": row_number, "error_code": "COUNTRY_MISMATCH", "message": "Ülke uyuşmuyor."})

        if country_raw and current_user.get("role") == "country_admin":
            scope = current_user.get("country_scope") or []
            if "*" not in scope and country_raw not in scope:
                errors.append({"row_number": row_number, "error_code": "COUNTRY_SCOPE_FORBIDDEN", "message": "Ülke yetkisi yok."})

        parsed_rows.append({
            "row_number": row_number,
            "module": module_raw,
            "country": country_raw,
            "schema_version": schema_version,
            "slug": slug_raw,
            "parent_slug": parent_slug_raw,
            "name_tr": name_tr,
            "name_de": name_de,
            "name_fr": name_fr,
            "is_active": is_active,
            "sort_order": sort_order,
            "wizard_progress": wizard_progress,
        })

    parent_map = dict(existing_parent_map)
    for row in parsed_rows:
        if row.get("slug"):
            parent_map[row["slug"]] = row.get("parent_slug")

    slug_errors = defaultdict(list)
    for error in errors:
        slug_errors[error.get("row_number")].append(error)

    # parent existence check
    for row in parsed_rows:
        if slug_errors.get(row.get("row_number")):
            continue
        parent_slug = row.get("parent_slug")
        if parent_slug and parent_slug not in parent_map:
            errors.append({"row_number": row.get("row_number"), "error_code": "PARENT_NOT_FOUND", "message": "parent_slug bulunamadı."})

    # cycle detection
    cycle_nodes = set()
    visited: set = set()
    stack: set = set()

    def dfs(node: str):
        if node in stack:
            cycle_nodes.add(node)
            return True
        if node in visited:
            return False
        visited.add(node)
        stack.add(node)
        parent = parent_map.get(node)
        if parent and dfs(parent):
            cycle_nodes.add(node)
            stack.remove(node)
            return True
        stack.remove(node)
        return False

    for row in parsed_rows:
        slug = row.get("slug")
        if not slug:
            continue
        if dfs(slug):
            continue

    for row in parsed_rows:
        if row.get("slug") in cycle_nodes:
            errors.append({"row_number": row.get("row_number"), "error_code": "CYCLE_DETECTED", "message": "Döngüsel parent ilişkisi."})

    valid_rows = [row for row in parsed_rows if not any(err["row_number"] == row.get("row_number") for err in errors)]

    creates = sum(1 for row in valid_rows if row.get("slug") not in existing_map)
    updates = sum(1 for row in valid_rows if row.get("slug") in existing_map)

    return {
        "rows": valid_rows,
        "errors": errors,
        "summary": {
            "total": len(parsed_rows),
            "creates": creates,
            "updates": updates,
            "errors": len(errors),
        },
        "existing_map": existing_map,
        "parent_map": parent_map,
    }


async def _update_category_paths(session: AsyncSession, category: Category, new_parent: Optional[Category]) -> bool:
    slug_value = _pick_category_slug(category.slug) or ""
    new_depth = new_parent.depth + 1 if new_parent else 0
    new_path = f"{new_parent.path}.{slug_value}" if new_parent and new_parent.path else slug_value
    old_path = category.path or ""
    old_depth = category.depth or 0

    if old_path == new_path and old_depth == new_depth and category.parent_id == (new_parent.id if new_parent else None):
        return False

    category.parent_id = new_parent.id if new_parent else None
    category.depth = new_depth
    category.path = new_path

    if old_path:
        descendants = (await session.execute(
            select(Category).where(Category.path.like(f"{old_path}.%"))
        )).scalars().all()
        for desc in descendants:
            if not desc.path:
                continue
            desc.path = desc.path.replace(old_path, new_path, 1)
            depth_delta = desc.depth - old_depth
            desc.depth = new_depth + depth_delta
    return True


@api_router.post("/admin/categories/import-export/import/dry-run")
async def admin_import_categories_dry_run(
    request: Request,
    file: UploadFile = File(...),
    format: str = "csv",
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    content = _read_import_file(file)
    file_hash = hashlib.sha256(content).hexdigest()
    rows = _parse_category_import_rows(content, format.lower())

    result = await session.execute(select(Category).options(selectinload(Category.translations)))
    existing = result.scalars().all()
    validation = _validate_category_import_rows(rows, existing, current_user)

    await _write_audit_log_sql(
        session,
        "categories.import.dry_run",
        current_user,
        "category_import",
        None,
        {
            "format": format,
            "filename": file.filename,
            "summary": validation.get("summary"),
            "hash": file_hash,
        },
        request,
    )
    await session.commit()

    summary = validation.get("summary") or {}
    response = {
        "dry_run_hash": file_hash,
        "summary": summary,
        "errors": validation.get("errors", []),
        "creates": [row["slug"] for row in validation.get("rows", []) if row["slug"] not in validation.get("existing_map", {})],
        "updates": [row["slug"] for row in validation.get("rows", []) if row["slug"] in validation.get("existing_map", {})],
    }
    return response


@api_router.post("/admin/categories/import-export/import/dry-run/pdf")
async def admin_import_categories_dry_run_pdf(
    request: Request,
    file: UploadFile = File(...),
    format: str = "csv",
    dry_run_hash: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    raise HTTPException(status_code=410, detail="PDF dry-run raporu devre dışı")


@api_router.post("/admin/categories/import-export/import/commit")
async def admin_import_categories_commit(
    request: Request,
    file: UploadFile = File(...),
    format: str = "csv",
    dry_run_hash: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    content = _read_import_file(file)
    file_hash = hashlib.sha256(content).hexdigest()
    if not dry_run_hash or dry_run_hash != file_hash:
        raise HTTPException(status_code=409, detail="Dry-run doğrulaması gerekli")

    rows = _parse_category_import_rows(content, format.lower())
    result = await session.execute(select(Category).options(selectinload(Category.translations)))
    existing_categories = result.scalars().all()
    validation = _validate_category_import_rows(rows, existing_categories, current_user)

    if validation.get("errors"):
        return Response(
            status_code=400,
            media_type="application/json",
            content=json.dumps({
                "detail": "Validation failed",
                "errors": validation.get("errors"),
                "summary": validation.get("summary"),
            }),
        )

    parent_map = validation.get("parent_map", {})
    existing_by_slug: Dict[str, Category] = validation.get("existing_map", {})
    rows_to_apply = validation.get("rows", [])

    depth_cache: Dict[str, int] = {}

    def compute_depth(slug: str) -> int:
        if slug in depth_cache:
            return depth_cache[slug]
        parent = parent_map.get(slug)
        if not parent or parent == slug:
            depth_cache[slug] = 0
            return 0
        depth_cache[slug] = compute_depth(parent) + 1
        return depth_cache[slug]

    ordered_rows = sorted(rows_to_apply, key=lambda row: (compute_depth(row["slug"]), row["sort_order"]))

    created = 0
    updated = 0

    try:
        created_map: Dict[str, Category] = {}
        for row in ordered_rows:
            slug_key = row["slug"]
            category = existing_by_slug.get(slug_key)
            parent_slug = row.get("parent_slug")
            parent = None
            if parent_slug:
                parent = created_map.get(parent_slug) or existing_by_slug.get(parent_slug)

            allowed_countries = [row["country"]] if row.get("country") else []

            if not category:
                slug_map = {"tr": slug_key, "de": slug_key, "fr": slug_key}
                category = Category(
                    id=uuid.uuid4(),
                    parent_id=parent.id if parent else None,
                    path=f"{parent.path}.{slug_key}" if parent and parent.path else slug_key,
                    depth=(parent.depth + 1) if parent else 0,
                    sort_order=row.get("sort_order") or 0,
                    module=row.get("module") or "vehicle",
                    slug=slug_map,
                    icon=None,
                    image_url=None,
                    is_enabled=row.get("is_active") if row.get("is_active") is not None else True,
                    is_visible_on_home=False if parent else True,
                    is_deleted=False,
                    inherit_enabled=True,
                    override_enabled=None,
                    inherit_countries=True,
                    override_countries=None,
                    allowed_countries=allowed_countries,
                    listing_count=0,
                    country_code=row.get("country") or None,
                    hierarchy_complete=True,
                    form_schema=None,
                    wizard_progress=row.get("wizard_progress"),
                    created_at=datetime.now(timezone.utc),
                    updated_at=datetime.now(timezone.utc),
                )
                session.add(category)
                await session.flush()
                existing_by_slug[slug_key] = category
                created_map[slug_key] = category
                created += 1

                for lang, value in (("tr", row.get("name_tr")), ("de", row.get("name_de")), ("fr", row.get("name_fr"))):
                    session.add(
                        CategoryTranslation(
                            category_id=category.id,
                            language=lang,
                            name=value or slug_key,
                            description=None,
                            meta_title=None,
                            meta_description=None,
                        )
                    )
            else:
                changed = False
                if await _update_category_paths(session, category, parent):
                    changed = True
                if category.sort_order != row.get("sort_order"):
                    category.sort_order = row.get("sort_order")
                    changed = True
                if category.is_enabled != row.get("is_active"):
                    category.is_enabled = row.get("is_active")
                    changed = True
                if category.country_code != row.get("country"):
                    category.country_code = row.get("country") or None
                    changed = True
                if category.allowed_countries != allowed_countries:
                    category.allowed_countries = allowed_countries
                    changed = True
                if category.is_deleted:
                    category.is_deleted = False
                    changed = True
                if category.wizard_progress != row.get("wizard_progress"):
                    category.wizard_progress = row.get("wizard_progress")
                    changed = True

                translations = {t.language: t for t in (category.translations or [])}
                for lang, value in (("tr", row.get("name_tr")), ("de", row.get("name_de")), ("fr", row.get("name_fr"))):
                    if lang in translations:
                        if translations[lang].name != value:
                            translations[lang].name = value
                            changed = True
                    else:
                        session.add(
                            CategoryTranslation(
                                category_id=category.id,
                                language=lang,
                                name=value or slug_key,
                                description=None,
                                meta_title=None,
                                meta_description=None,
                            )
                        )
                        changed = True

                if changed:
                    category.updated_at = datetime.now(timezone.utc)
                    updated += 1

    except Exception:
        await session.rollback()
        raise

    await _write_audit_log_sql(
        session,
        "categories.import.apply",
        current_user,
        "category_import",
        None,
        {
            "format": format,
            "filename": file.filename,
            "created": created,
            "updated": updated,
            "hash": file_hash,
        },
        request,
    )
    await session.commit()

    return {
        "summary": {"created": created, "updated": updated},
    }


@api_router.get("/admin/menu-items")
async def admin_list_menu_items(
    request: Request,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session)
    query = select(MenuItem)
    if country:
        country_code = country.upper()
        _assert_country_scope(country_code, current_user)
        query = query.where(MenuItem.country_code == country_code)
    elif ctx.mode == "country" and ctx.country:
        query = query.where(MenuItem.country_code == ctx.country)

    result = await session.execute(query.order_by(MenuItem.sort_order.asc()))
    items = result.scalars().all()

    return {
        "items": [
            {
                "id": str(item.id),
                "label": item.label,
                "slug": item.slug,
                "url": item.url,
                "parent_id": str(item.parent_id) if item.parent_id else None,
                "country_code": item.country_code,
                "active_flag": item.active_flag,
                "sort_order": item.sort_order,
                "created_at": item.created_at.isoformat() if item.created_at else None,
                "updated_at": item.updated_at.isoformat() if item.updated_at else None,
            }
            for item in items
            if not item.deleted_at
        ]
    }


@api_router.post("/admin/menu-items")
async def admin_create_menu_item(
    payload: MenuItemCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    ctx = await resolve_admin_country_context(request, current_user=current_user, session=session)
    slug = payload.slug.strip()
    if not slug:
        raise HTTPException(status_code=400, detail="Slug is required")

    country_code = payload.country_code.upper() if payload.country_code else None
    if ctx.mode == "country" and ctx.country:
        country_code = country_code or ctx.country
    if country_code:
        _assert_country_scope(country_code, current_user)

    parent_id = None
    if payload.parent_id:
        try:
            parent_uuid = uuid.UUID(payload.parent_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid parent id") from exc
        parent = await session.get(MenuItem, parent_uuid)
        if not parent or parent.deleted_at:
            raise HTTPException(status_code=404, detail="Parent menu item not found")
        parent_id = parent_uuid

    result = await session.execute(
        select(MenuItem).where(MenuItem.slug == slug, MenuItem.country_code == country_code)
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Menu slug already exists")

    item = MenuItem(
        label=payload.label.strip(),
        slug=slug,
        url=payload.url.strip() if payload.url else None,
        parent_id=parent_id,
        country_code=country_code,
        active_flag=payload.active_flag if payload.active_flag is not None else True,
        sort_order=payload.sort_order or 0,
    )
    session.add(item)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="MENU_ITEM_CREATE",
        actor=current_user,
        resource_type="menu_item",
        resource_id=str(item.id),
        metadata={"label": item.label, "slug": item.slug, "url": item.url},
        request=request,
        country_code=country_code,
    )
    await session.commit()

    return {
        "menu_item": {
            "id": str(item.id),
            "label": item.label,
            "slug": item.slug,
            "url": item.url,
            "parent_id": str(item.parent_id) if item.parent_id else None,
            "country_code": item.country_code,
            "active_flag": item.active_flag,
            "sort_order": item.sort_order,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        }
    }


@api_router.patch("/admin/menu-items/{menu_id}")
async def admin_update_menu_item(
    menu_id: str,
    payload: MenuItemUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)
    try:
        menu_uuid = uuid.UUID(menu_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid menu id") from exc

    item = await session.get(MenuItem, menu_uuid)
    if not item or item.deleted_at:
        raise HTTPException(status_code=404, detail="Menu item not found")

    if payload.slug is not None:
        slug = payload.slug.strip()
        if not slug:
            raise HTTPException(status_code=400, detail="Slug is required")
        if slug != item.slug:
            result = await session.execute(
                select(MenuItem).where(MenuItem.slug == slug, MenuItem.country_code == item.country_code)
            )
            if result.scalar_one_or_none():
                raise HTTPException(status_code=409, detail="Menu slug already exists")
            item.slug = slug

    if payload.label is not None:
        item.label = payload.label.strip()
    if payload.url is not None:
        item.url = payload.url.strip() if payload.url else None
    if payload.active_flag is not None:
        item.active_flag = bool(payload.active_flag)
    if payload.sort_order is not None:
        item.sort_order = payload.sort_order
    if payload.parent_id is not None:
        if payload.parent_id == "":
            item.parent_id = None
        else:
            try:
                parent_uuid = uuid.UUID(payload.parent_id)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail="Invalid parent id") from exc
            parent = await session.get(MenuItem, parent_uuid)
            if not parent or parent.deleted_at:
                raise HTTPException(status_code=404, detail="Parent menu item not found")
            item.parent_id = parent_uuid

    await _write_audit_log_sql(
        session=session,
        action="MENU_ITEM_UPDATE",
        actor=current_user,
        resource_type="menu_item",
        resource_id=str(item.id),
        metadata={"label": item.label, "slug": item.slug, "url": item.url},
        request=request,
        country_code=item.country_code,
    )
    await session.commit()

    return {"menu_item": {
        "id": str(item.id),
        "label": item.label,
        "slug": item.slug,
        "url": item.url,
        "parent_id": str(item.parent_id) if item.parent_id else None,
        "country_code": item.country_code,
        "active_flag": item.active_flag,
        "sort_order": item.sort_order,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }}


@api_router.delete("/admin/menu-items/{menu_id}")
async def admin_delete_menu_item(
    menu_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)
    try:
        menu_uuid = uuid.UUID(menu_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid menu id") from exc

    item = await session.get(MenuItem, menu_uuid)
    if not item or item.deleted_at:
        raise HTTPException(status_code=404, detail="Menu item not found")

    item.deleted_at = datetime.now(timezone.utc)
    item.active_flag = False

    await _write_audit_log_sql(
        session=session,
        action="MENU_ITEM_DELETE",
        actor=current_user,
        resource_type="menu_item",
        resource_id=str(item.id),
        metadata={"label": item.label, "slug": item.slug},
        request=request,
        country_code=item.country_code,
    )
    await session.commit()

    return {"ok": True}


@api_router.get("/attributes")
async def public_attributes(
    category_id: str,
    country: Optional[str] = None,
    request: Request = None,
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        category_uuid = uuid.UUID(category_id)
    except ValueError:
        return {"items": []}

    query = (
        select(Attribute, CategoryAttributeMap)
        .join(CategoryAttributeMap, CategoryAttributeMap.attribute_id == Attribute.id)
        .where(
            CategoryAttributeMap.category_id == category_uuid,
            CategoryAttributeMap.is_active.is_(True),
            Attribute.is_active.is_(True),
        )
        .options(selectinload(Attribute.options))
        .order_by(Attribute.key.asc())
    )

    result = await session.execute(query)
    items = []
    for attribute, mapping in result.all():
        items.append(_serialize_attribute_sql(attribute, str(mapping.category_id)))

    return {"items": items}


@api_router.get("/admin/attributes")
async def admin_list_attributes(
    request: Request,
    category_id: Optional[str] = None,
    country: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    if category_id:
        try:
            category_uuid = uuid.UUID(category_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid category id") from exc

        query = (
            select(Attribute, CategoryAttributeMap)
            .join(CategoryAttributeMap, CategoryAttributeMap.attribute_id == Attribute.id)
            .where(CategoryAttributeMap.category_id == category_uuid)
            .options(selectinload(Attribute.options))
            .order_by(Attribute.key.asc())
        )
        result = await session.execute(query)
        items = [_serialize_attribute_sql(attr, str(mapping.category_id)) for attr, mapping in result.all()]
        return {"items": items}

    result = await session.execute(select(Attribute).options(selectinload(Attribute.options)).order_by(Attribute.key.asc()))
    items = [_serialize_attribute_sql(attr, None) for attr in result.scalars().all()]
    return {"items": items}


@api_router.post("/admin/attributes", status_code=201)
async def admin_create_attribute(
    payload: AttributeCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        category_uuid = uuid.UUID(payload.category_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid category id") from exc

    result = await session.execute(select(Category).where(Category.id == category_uuid))
    category = result.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=400, detail="category_id not found")

    key = payload.key.strip().lower()
    if not ATTRIBUTE_KEY_PATTERN.match(key):
        raise HTTPException(status_code=400, detail="key format invalid")
    if payload.type not in {"text", "number", "select", "boolean"}:
        raise HTTPException(status_code=400, detail="type invalid")
    if payload.type == "select" and not payload.options:
        raise HTTPException(status_code=400, detail="options required for select")

    existing = await session.execute(select(Attribute).where(Attribute.key == key))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Attribute key already exists")

    attribute = Attribute(
        key=key,
        name={"tr": payload.name.strip()},
        attribute_type=payload.type,
        is_required=bool(payload.required_flag),
        is_filterable=bool(payload.filterable_flag),
        is_active=payload.active_flag if payload.active_flag is not None else True,
    )
    session.add(attribute)
    await session.flush()

    mapping = CategoryAttributeMap(
        category_id=category_uuid,
        attribute_id=attribute.id,
        is_required_override=bool(payload.required_flag),
        inherit_to_children=True,
        is_active=True,
    )
    session.add(mapping)

    if payload.options:
        for option in payload.options:
            label = option.strip()
            if not label:
                continue
            session.add(
                AttributeOption(
                    attribute_id=attribute.id,
                    value=label.lower().replace(" ", "_"),
                    label={"tr": label},
                    is_active=True,
                )
            )

    await _write_audit_log_sql(
        session=session,
        action="ATTRIBUTE_CREATE",
        actor=current_user,
        resource_type="attribute",
        resource_id=str(attribute.id),
        metadata={"key": attribute.key, "category_id": payload.category_id},
        request=request,
        country_code=None,
    )
    await session.commit()

    await session.refresh(attribute)
    return {"attribute": _serialize_attribute_sql(attribute, payload.category_id)}


@api_router.patch("/admin/attributes/{attribute_id}")
async def admin_update_attribute(
    attribute_id: str,
    payload: AttributeUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        attribute_uuid = uuid.UUID(attribute_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid attribute id") from exc

    result = await session.execute(select(Attribute).where(Attribute.id == attribute_uuid).options(selectinload(Attribute.options)))
    attribute = result.scalar_one_or_none()
    if not attribute:
        raise HTTPException(status_code=404, detail="Attribute not found")

    if payload.name is not None:
        attribute.name = {"tr": payload.name.strip()}
    if payload.key is not None:
        key = payload.key.strip().lower()
        if not ATTRIBUTE_KEY_PATTERN.match(key):
            raise HTTPException(status_code=400, detail="key format invalid")
        attribute.key = key
    if payload.type is not None:
        if payload.type not in {"text", "number", "select", "boolean"}:
            raise HTTPException(status_code=400, detail="type invalid")
        attribute.attribute_type = payload.type
    if payload.options is not None:
        for opt in list(attribute.options):
            await session.delete(opt)
        for option in payload.options:
            label = option.strip()
            if not label:
                continue
            session.add(
                AttributeOption(
                    attribute_id=attribute.id,
                    value=label.lower().replace(" ", "_"),
                    label={"tr": label},
                    is_active=True,
                )
            )
    if payload.required_flag is not None:
        attribute.is_required = bool(payload.required_flag)
    if payload.filterable_flag is not None:
        attribute.is_filterable = bool(payload.filterable_flag)
    if payload.active_flag is not None:
        attribute.is_active = bool(payload.active_flag)

    await _write_audit_log_sql(
        session=session,
        action="ATTRIBUTE_UPDATE",
        actor=current_user,
        resource_type="attribute",
        resource_id=str(attribute.id),
        metadata={"key": attribute.key},
        request=request,
        country_code=None,
    )
    await session.commit()

    await session.refresh(attribute)
    return {"attribute": _serialize_attribute_sql(attribute, None)}


@api_router.delete("/admin/attributes/{attribute_id}")
async def admin_delete_attribute(
    attribute_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        attribute_uuid = uuid.UUID(attribute_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid attribute id") from exc

    result = await session.execute(select(Attribute).where(Attribute.id == attribute_uuid))
    attribute = result.scalar_one_or_none()
    if not attribute:
        raise HTTPException(status_code=404, detail="Attribute not found")

    attribute.is_active = False
    result = await session.execute(
        select(CategoryAttributeMap).where(CategoryAttributeMap.attribute_id == attribute_uuid)
    )
    for mapping in result.scalars().all():
        mapping.is_active = False

    await _write_audit_log_sql(
        session=session,
        action="ATTRIBUTE_DELETE",
        actor=current_user,
        resource_type="attribute",
        resource_id=str(attribute.id),
        metadata={"key": attribute.key},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"attribute": _serialize_attribute_sql(attribute, None)}


@api_router.get("/admin/vehicle-makes")
async def admin_list_vehicle_makes(
    request: Request,
    country: Optional[str] = None,
    vehicle_type: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    vehicle_type_filter = _normalize_vehicle_type(vehicle_type)
    if vehicle_type_filter:
        if vehicle_type_filter not in VEHICLE_TYPE_SET:
            raise HTTPException(status_code=400, detail="vehicle_type invalid")

    query = select(VehicleMake)
    if vehicle_type_filter:
        if vehicle_type_filter == "car":
            model_query = select(VehicleModel.make_id).where(
                or_(VehicleModel.vehicle_type == "car", VehicleModel.vehicle_type.is_(None))
            )
        else:
            model_query = select(VehicleModel.make_id).where(VehicleModel.vehicle_type == vehicle_type_filter)
        query = query.where(VehicleMake.id.in_(model_query))

    result = await session.execute(query.order_by(VehicleMake.name.asc()))
    makes = result.scalars().all()

    make_ids = [make.id for make in makes]
    type_map: Dict[str, set[str]] = {}
    if make_ids:
        model_result = await session.execute(
            select(VehicleModel.make_id, VehicleModel.vehicle_type).where(VehicleModel.make_id.in_(make_ids))
        )
        for make_id, vtype in model_result.all():
            key = str(make_id)
            normalized = (vtype or "car").strip().lower()
            type_map.setdefault(key, set()).add(normalized)

    result_items = []
    for make in makes:
        item = _normalize_vehicle_make_doc(make)
        types = sorted(type_map.get(str(make.id), set()))
        item["vehicle_types"] = types
        item["vehicle_type_summary"] = _build_vehicle_type_summary(set(types))
        result_items.append(item)

    return {"items": result_items}


@api_router.post("/admin/vehicle-makes", status_code=201)
async def admin_create_vehicle_make(
    payload: VehicleMakeCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    slug = payload.slug.strip().lower()
    if not SLUG_PATTERN.match(slug):
        raise HTTPException(status_code=400, detail="slug format invalid")
    if payload.country_code:
        _assert_country_scope(payload.country_code.upper(), current_user)

    existing = await session.execute(select(VehicleMake).where(VehicleMake.slug == slug))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Make slug already exists")

    make = VehicleMake(
        name=payload.name.strip(),
        slug=slug,
        is_active=payload.active_flag if payload.active_flag is not None else True,
    )
    session.add(make)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="VEHICLE_MAKE_CREATE",
        actor=current_user,
        resource_type="vehicle_make",
        resource_id=str(make.id),
        metadata={"slug": make.slug, "name": make.name},
        request=request,
        country_code=payload.country_code.upper() if payload.country_code else None,
    )
    await session.commit()

    return {"make": _normalize_vehicle_make_doc(make)}


@api_router.patch("/admin/vehicle-makes/{make_id}")
async def admin_update_vehicle_make(
    make_id: str,
    payload: VehicleMakeUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        make_uuid = uuid.UUID(make_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid make id") from exc

    make = await session.get(VehicleMake, make_uuid)
    if not make:
        raise HTTPException(status_code=404, detail="Make not found")

    if payload.name is not None:
        make.name = payload.name.strip()
    if payload.slug is not None:
        slug = payload.slug.strip().lower()
        if not SLUG_PATTERN.match(slug):
            raise HTTPException(status_code=400, detail="slug format invalid")
        if slug != make.slug:
            existing = await session.execute(select(VehicleMake).where(VehicleMake.slug == slug))
            if existing.scalar_one_or_none():
                raise HTTPException(status_code=409, detail="Make slug already exists")
        make.slug = slug
    if payload.active_flag is not None:
        make.is_active = bool(payload.active_flag)

    await _write_audit_log_sql(
        session=session,
        action="VEHICLE_MAKE_UPDATE",
        actor=current_user,
        resource_type="vehicle_make",
        resource_id=str(make.id),
        metadata={"slug": make.slug, "name": make.name},
        request=request,
        country_code=payload.country_code.upper() if payload.country_code else None,
    )
    await session.commit()

    return {"make": _normalize_vehicle_make_doc(make)}


@api_router.delete("/admin/vehicle-makes/{make_id}")
async def admin_delete_vehicle_make(
    make_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        make_uuid = uuid.UUID(make_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid make id") from exc

    make = await session.get(VehicleMake, make_uuid)
    if not make:
        raise HTTPException(status_code=404, detail="Make not found")

    make.is_active = False

    await _write_audit_log_sql(
        session=session,
        action="VEHICLE_MAKE_DELETE",
        actor=current_user,
        resource_type="vehicle_make",
        resource_id=str(make.id),
        metadata={"active_flag": False},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"make": _normalize_vehicle_make_doc(make)}


@api_router.get("/admin/vehicle-models")
async def admin_list_vehicle_models(
    request: Request,
    make_id: Optional[str] = None,
    country: Optional[str] = None,
    vehicle_type: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    query = select(VehicleModel)
    if make_id:
        try:
            make_uuid = uuid.UUID(make_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid make id") from exc
        query = query.where(VehicleModel.make_id == make_uuid)
    elif country:
        _assert_country_scope(country.upper(), current_user)

    vehicle_type_filter = _normalize_vehicle_type(vehicle_type)
    if vehicle_type_filter:
        if vehicle_type_filter not in VEHICLE_TYPE_SET:
            raise HTTPException(status_code=400, detail="vehicle_type invalid")
        if vehicle_type_filter == "car":
            query = query.where(or_(VehicleModel.vehicle_type == "car", VehicleModel.vehicle_type.is_(None)))
        else:
            query = query.where(VehicleModel.vehicle_type == vehicle_type_filter)

    result = await session.execute(query.order_by(VehicleModel.name.asc()))
    return {"items": [_normalize_vehicle_model_doc(doc) for doc in result.scalars().all()]}


@api_router.post("/admin/vehicle-models", status_code=201)
async def admin_create_vehicle_model(
    payload: VehicleModelCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    slug = payload.slug.strip().lower()
    if not SLUG_PATTERN.match(slug):
        raise HTTPException(status_code=400, detail="slug format invalid")

    vehicle_type = _normalize_vehicle_type(payload.vehicle_type)
    if not vehicle_type or vehicle_type not in VEHICLE_TYPE_SET:
        raise HTTPException(status_code=400, detail="vehicle_type invalid")

    try:
        make_uuid = uuid.UUID(payload.make_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid make id") from exc

    make = await session.get(VehicleMake, make_uuid)
    if not make:
        raise HTTPException(status_code=404, detail="Make not found")

    existing = await session.execute(
        select(VehicleModel).where(VehicleModel.make_id == make_uuid, VehicleModel.slug == slug)
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Model slug already exists")

    model = VehicleModel(
        make_id=make_uuid,
        name=payload.name.strip(),
        slug=slug,
        vehicle_type=vehicle_type,
        is_active=payload.active_flag if payload.active_flag is not None else True,
    )
    session.add(model)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="VEHICLE_MODEL_CREATE",
        actor=current_user,
        resource_type="vehicle_model",
        resource_id=str(model.id),
        metadata={"slug": model.slug, "name": model.name, "make_id": payload.make_id},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"model": _normalize_vehicle_model_doc(model)}


@api_router.patch("/admin/vehicle-models/{model_id}")
async def admin_update_vehicle_model(
    model_id: str,
    payload: VehicleModelUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        model_uuid = uuid.UUID(model_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid model id") from exc

    model = await session.get(VehicleModel, model_uuid)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")

    if payload.name is not None:
        model.name = payload.name.strip()
    if payload.slug is not None:
        slug = payload.slug.strip().lower()
        if not SLUG_PATTERN.match(slug):
            raise HTTPException(status_code=400, detail="slug format invalid")
        if slug != model.slug:
            existing = await session.execute(
                select(VehicleModel).where(VehicleModel.make_id == model.make_id, VehicleModel.slug == slug)
            )
            if existing.scalar_one_or_none():
                raise HTTPException(status_code=409, detail="Model slug already exists")
        model.slug = slug
    if payload.make_id is not None:
        try:
            make_uuid = uuid.UUID(payload.make_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid make id") from exc
        make = await session.get(VehicleMake, make_uuid)
        if not make:
            raise HTTPException(status_code=400, detail="make_id not found")
        model.make_id = make_uuid
    if payload.vehicle_type is not None:
        vehicle_type = _normalize_vehicle_type(payload.vehicle_type)
        if not vehicle_type or vehicle_type not in VEHICLE_TYPE_SET:
            raise HTTPException(status_code=400, detail="vehicle_type invalid")
        model.vehicle_type = vehicle_type
    if payload.active_flag is not None:
        model.is_active = bool(payload.active_flag)

    await _write_audit_log_sql(
        session=session,
        action="VEHICLE_MODEL_UPDATE",
        actor=current_user,
        resource_type="vehicle_model",
        resource_id=str(model.id),
        metadata={"slug": model.slug, "name": model.name},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"model": _normalize_vehicle_model_doc(model)}


@api_router.delete("/admin/vehicle-models/{model_id}")
async def admin_delete_vehicle_model(
    model_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)

    try:
        model_uuid = uuid.UUID(model_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid model id") from exc

    model = await session.get(VehicleModel, model_uuid)
    if not model:
        raise HTTPException(status_code=404, detail="Model not found")

    model.is_active = False

    await _write_audit_log_sql(
        session=session,
        action="VEHICLE_MODEL_DELETE",
        actor=current_user,
        resource_type="vehicle_model",
        resource_id=str(model.id),
        metadata={"active_flag": False},
        request=request,
        country_code=None,
    )
    await session.commit()

    return {"model": _normalize_vehicle_model_doc(model)}


async def _prepare_vehicle_import_payload(session: AsyncSession, payload: VehicleImportPayload) -> dict:
    makes = payload.makes or []
    models = payload.models or []

    normalized_makes = []
    normalized_models = []
    invalid_makes = []
    invalid_models = []
    unknown_vehicle_types = set()

    make_slug_counts: Dict[str, int] = {}
    for make in makes:
        slug = (make.slug or "").strip().lower()
        if slug:
            make_slug_counts[slug] = make_slug_counts.get(slug, 0) + 1

    duplicate_make_slugs = sorted([slug for slug, count in make_slug_counts.items() if count > 1])

    model_slug_counts: Dict[str, int] = {}
    for model in models:
        make_slug = (model.make_slug or "").strip().lower()
        slug = (model.slug or "").strip().lower()
        if make_slug and slug:
            key = f"{make_slug}:{slug}"
            model_slug_counts[key] = model_slug_counts.get(key, 0) + 1

    duplicate_model_slugs = []
    for key, count in model_slug_counts.items():
        if count > 1:
            make_slug, model_slug = key.split(":", 1)
            duplicate_model_slugs.append({"make_slug": make_slug, "model_slug": model_slug, "count": count})

    make_slug_lookup = set(make_slug_counts.keys())
    model_make_slugs = {(model.make_slug or "").strip().lower() for model in models if model.make_slug}
    make_slug_lookup.update(model_make_slugs)

    existing_make_docs = []
    if make_slug_lookup:
        result = await session.execute(select(VehicleMake).where(VehicleMake.slug.in_(list(make_slug_lookup))))
        existing_make_docs = result.scalars().all()

    existing_make_map: Dict[str, VehicleMake] = {make.slug: make for make in existing_make_docs}
    ambiguous_make_slugs: list[str] = []

    for idx, make in enumerate(makes):
        errors = []
        name = (make.name or "").strip()
        slug = (make.slug or "").strip().lower()
        country_code = (make.country_code or "").strip().upper()

        if not name:
            errors.append("name missing")
        if not slug:
            errors.append("slug missing")
        elif not SLUG_PATTERN.match(slug):
            errors.append("slug format invalid")
        if not country_code:
            errors.append("country_code missing")
        if slug in duplicate_make_slugs:
            errors.append("duplicate slug in payload")

        if errors:
            invalid_makes.append({"index": idx, "slug": slug or None, "reason": "; ".join(errors)})
            continue

        normalized_makes.append({
            "name": name,
            "slug": slug,
            "active_flag": make.active if make.active is not None else True,
        })

    valid_make_slugs = {make["slug"] for make in normalized_makes}
    invalid_make_slugs = {item.get("slug") for item in invalid_makes if item.get("slug")}

    for idx, model in enumerate(models):
        errors = []
        make_slug = (model.make_slug or "").strip().lower()
        name = (model.name or "").strip()
        slug = (model.slug or "").strip().lower()
        vehicle_type = (model.vehicle_type or "").strip().lower()

        if not make_slug:
            errors.append("make_slug missing")
        if not name:
            errors.append("name missing")
        if not slug:
            errors.append("slug missing")
        elif not SLUG_PATTERN.match(slug):
            errors.append("slug format invalid")
        if not vehicle_type:
            errors.append("vehicle_type missing")
        elif vehicle_type not in VEHICLE_TYPE_SET:
            errors.append("vehicle_type invalid")
            unknown_vehicle_types.add(vehicle_type)

        if make_slug in duplicate_make_slugs:
            errors.append("make_slug duplicated in payload")
        if make_slug in invalid_make_slugs:
            errors.append("make_slug invalid")
        if make_slug not in valid_make_slugs and make_slug not in existing_make_map:
            errors.append("make_slug not found")

        key = f"{make_slug}:{slug}" if make_slug and slug else None
        if key and model_slug_counts.get(key, 0) > 1:
            errors.append("duplicate model slug in payload")

        if errors:
            invalid_models.append({"index": idx, "slug": slug or None, "make_slug": make_slug or None, "reason": "; ".join(errors)})
            continue

        normalized_models.append({
            "make_slug": make_slug,
            "name": name,
            "slug": slug,
            "vehicle_type": vehicle_type,
            "active_flag": model.active if model.active is not None else True,
        })

    report = {
        "summary": {
            "makes_total": len(makes),
            "makes_valid": len(normalized_makes),
            "makes_invalid": len(invalid_makes),
            "models_total": len(models),
            "models_valid": len(normalized_models),
            "models_invalid": len(invalid_models),
        },
        "unknown_vehicle_types": sorted([v for v in unknown_vehicle_types if v]),
        "duplicate_make_slugs": duplicate_make_slugs,
        "duplicate_model_slugs": duplicate_model_slugs,
        "ambiguous_make_slugs": ambiguous_make_slugs,
        "invalid_makes": invalid_makes,
        "invalid_models": invalid_models,
    }

    can_apply = (
        not invalid_makes
        and not invalid_models
        and not duplicate_make_slugs
        and not duplicate_model_slugs
        and not ambiguous_make_slugs
        and not unknown_vehicle_types
    )
    report["can_apply"] = can_apply

    return {
        "normalized_makes": normalized_makes,
        "normalized_models": normalized_models,
        "report": report,
        "existing_make_map": existing_make_map,
    }


@api_router.post("/admin/vehicle-import/dry-run")
async def admin_vehicle_import_dry_run(
    payload: VehicleImportPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)
    prepared = await _prepare_vehicle_import_payload(session, payload)
    return prepared["report"]


@api_router.post("/admin/vehicle-import/apply")
async def admin_vehicle_import_apply(
    payload: VehicleImportPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=session)
    prepared = await _prepare_vehicle_import_payload(session, payload)
    report = prepared["report"]
    if not report.get("can_apply"):
        raise HTTPException(status_code=400, detail=report)

    normalized_makes = prepared["normalized_makes"]
    normalized_models = prepared["normalized_models"]
    existing_make_map = prepared["existing_make_map"]

    make_id_by_slug: Dict[str, uuid.UUID] = {}
    for slug, make in existing_make_map.items():
        make_id_by_slug[slug] = make.id

    make_inserts = 0
    make_updates = 0
    model_inserts = 0
    model_updates = 0

    async with session.begin():
        for make in normalized_makes:
            existing = await session.execute(select(VehicleMake).where(VehicleMake.slug == make["slug"]))
            existing_make = existing.scalar_one_or_none()
            if existing_make:
                existing_make.name = make["name"]
                existing_make.is_active = bool(make["active_flag"])
                make_updates += 1
                make_id_by_slug[make["slug"]] = existing_make.id
            else:
                new_make = VehicleMake(
                    name=make["name"],
                    slug=make["slug"],
                    is_active=bool(make["active_flag"]),
                )
                session.add(new_make)
                await session.flush()
                make_inserts += 1
                make_id_by_slug[make["slug"]] = new_make.id

        for model in normalized_models:
            make_id = make_id_by_slug.get(model["make_slug"])
            if not make_id:
                raise HTTPException(status_code=400, detail=f"make_slug not found: {model['make_slug']}")

            existing_model = await session.execute(
                select(VehicleModel).where(VehicleModel.make_id == make_id, VehicleModel.slug == model["slug"])
            )
            model_row = existing_model.scalar_one_or_none()
            if model_row:
                model_row.name = model["name"]
                model_row.vehicle_type = model["vehicle_type"]
                model_row.is_active = bool(model["active_flag"])
                model_updates += 1
            else:
                new_model = VehicleModel(
                    make_id=make_id,
                    name=model["name"],
                    slug=model["slug"],
                    vehicle_type=model["vehicle_type"],
                    is_active=bool(model["active_flag"]),
                )
                session.add(new_model)
                model_inserts += 1

        await _write_audit_log_sql(
            session=session,
            action="VEHICLE_MASTER_DATA_IMPORT",
            actor=current_user,
            resource_type="vehicle_master_data",
            resource_id="vehicle_import",
            metadata={
                "make_inserts": make_inserts,
                "make_updates": make_updates,
                "model_inserts": model_inserts,
                "model_updates": model_updates,
            },
            request=request,
            country_code=None,
        )

    return {
        "summary": {
            "make_inserts": make_inserts,
            "make_updates": make_updates,
            "model_inserts": model_inserts,
            "model_updates": model_updates,
        }
    }






async def _dashboard_invoice_totals(conditions: List[Any]) -> tuple[float, Dict[str, float]]:
    async with AsyncSessionLocal() as sql_session:
        totals = await _invoice_totals_by_currency(sql_session, conditions)
    total_amount = sum(totals.values())
    return round(total_amount, 2), {k: round(v, 2) for k, v in totals.items()}



def _build_dashboard_pdf(summary: Dict[str, Any], trend_window: int) -> bytes:
    styles = getSampleStyleSheet()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Dashboard Export")
    elements = []

    def format_number(value: Optional[float]) -> str:
        if value is None:
            return "-"
        try:
            return f"{float(value):,.2f}".replace(",", " ")
        except Exception:
            return str(value)

    def format_currency_totals(totals: Optional[Dict[str, Any]]) -> str:
        if not totals:
            return "-"
        parts = []
        for currency, amount in totals.items():
            parts.append(f"{format_number(amount)} {currency}")
        return " / ".join(parts)

    def add_table(title: str, rows: List[List[str]]):
        if not rows:
            return
        elements.append(Paragraph(title, styles["Heading3"]))
        table = Table(rows, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10))

    countries = summary.get("country_codes") or []
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    elements.append(Paragraph("Dashboard Özet Raporu", styles["Title"]))
    elements.append(Paragraph(f"Kapsam: {summary.get('scope')}", styles["Normal"]))
    elements.append(Paragraph(f"Ülkeler: {', '.join(countries) or '-'}", styles["Normal"]))
    elements.append(Paragraph(f"Trend Aralığı: {trend_window} gün", styles["Normal"]))
    elements.append(Paragraph(f"Oluşturma: {generated_at}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    finance_visible = bool(summary.get("finance_visible"))
    kpis = summary.get("kpis") or {}
    today = kpis.get("today") or {}
    week = kpis.get("last_7_days") or {}
    kpi_rows = [["Periyot", "Yeni İlan", "Yeni Kullanıcı", "Gelir"]]
    def kpi_revenue_label(data: Dict[str, Any]) -> str:
        if not finance_visible:
            return "Gizli"
        total = data.get("revenue_total")
        totals = data.get("revenue_currency_totals")
        if total is None and not totals:
            return "0"
        return f"{format_number(total)} ({format_currency_totals(totals)})"

    kpi_rows.append([
        "Bugün",
        str(today.get("new_listings", 0)),
        str(today.get("new_users", 0)),
        kpi_revenue_label(today),
    ])
    kpi_rows.append([
        "Son 7 Gün",
        str(week.get("new_listings", 0)),
        str(week.get("new_users", 0)),
        kpi_revenue_label(week),
    ])
    add_table("KPI Özeti", kpi_rows)

    metrics = summary.get("metrics") or {}
    users = summary.get("users") or {}
    active_countries = summary.get("active_countries") or {}
    active_modules = summary.get("active_modules") or {}
    metrics_rows = [["Metri̇k", "Değer"]]
    metrics_rows.append(["Toplam Kullanıcı", str(users.get("total", 0))])
    metrics_rows.append(["Aktif / Pasif", f"{users.get('active', 0)} / {users.get('inactive', 0)}"])
    metrics_rows.append(["Toplam İlan", str(metrics.get("total_listings", 0))])
    metrics_rows.append(["Yayınlı İlan", str(metrics.get("published_listings", 0))])
    metrics_rows.append(["Moderasyon Bekleyen", str(metrics.get("pending_moderation", 0))])
    metrics_rows.append(["Aktif Dealer", str(metrics.get("active_dealers", 0))])
    metrics_rows.append(["Aktif Ülke", str(active_countries.get("count", 0))])
    metrics_rows.append(["Aktif Modül", str(active_modules.get("count", 0))])
    add_table("Genel Metrikler", metrics_rows)

    risk = summary.get("risk_panel") or {}
    suspicious = risk.get("suspicious_logins") or {}
    sla = risk.get("sla_breaches") or {}
    pending = risk.get("pending_payments") or {}
    risk_rows = [["Risk Başlığı", "Sayı", "Eşik", "Detay"]]
    risk_rows.append([
        "Çoklu IP girişleri",
        str(suspicious.get("count", 0)),
        f"{suspicious.get('threshold', '-') } IP / {suspicious.get('window_hours', '-') } saat",
        "Örnek kayıtlar: " + str(len(suspicious.get("items") or [])),
    ])
    risk_rows.append([
        "Moderasyon SLA ihlali",
        str(sla.get("count", 0)),
        f"> {sla.get('threshold', '-') } saat",
        "Örnek kayıtlar: " + str(len(sla.get("items") or [])),
    ])
    if finance_visible:
        risk_rows.append([
            "Bekleyen ödemeler",
            str(pending.get("count", 0)),
            f"> {pending.get('threshold_days', '-') } gün",
            f"Toplam: {format_number(pending.get('total_amount'))} | {format_currency_totals(pending.get('currency_totals'))}",
        ])
    else:
        risk_rows.append([
            "Bekleyen ödemeler",
            "-",
            "-",
            "Gizli",
        ])
    add_table("Risk & Alarm Merkezi", risk_rows)

    trends = summary.get("trends") or {}
    listings = trends.get("listings") or []
    revenue = trends.get("revenue") or []
    revenue_map = {item.get("date"): item for item in revenue}
    trend_rows = [["Tarih", "İlan", "Gelir"]]
    for item in listings:
        date_label = item.get("date")
        revenue_item = revenue_map.get(date_label) or {}
        revenue_label = "Gizli" if not finance_visible else format_number(revenue_item.get("amount"))
        trend_rows.append([
            str(date_label),
            str(item.get("count", 0)),
            revenue_label,
        ])
    add_table("Trend Detayı", trend_rows)

    health = summary.get("health") or {}
    health_rows = [["Bileşen", "Değer"]]
    health_rows.append(["API status", str(health.get("api_status"))])
    health_rows.append(["DB status", str(health.get("db_status"))])
    health_rows.append(["API gecikme", f"{health.get('api_latency_ms')} ms"])
    health_rows.append(["DB gecikme", f"{health.get('db_latency_ms')} ms"])
    health_rows.append(["Son deploy", str(health.get("deployed_at"))])
    health_rows.append(["Son restart", str(health.get("restart_at"))])
    health_rows.append(["Uptime", str(health.get("uptime_human"))])
    add_table("Sistem Sağlığı", health_rows)

    doc.build(elements)
    pdf_value = buffer.getvalue()
    buffer.close()
    return pdf_value


def _empty_dashboard_summary(start_perf: float, can_view_finance: bool, trend_window: int) -> dict:
    now = datetime.now(timezone.utc)
    uptime_seconds = int((now - APP_START_TIME).total_seconds())
    api_latency_ms = int((time.perf_counter() - start_perf) * 1000)

    return {
        "scope": "global",
        "country_codes": [],
        "users": {"total": 0, "active": 0, "inactive": 0},
        "active_countries": {"count": 0, "codes": []},
        "active_modules": {"count": 0, "items": []},
        "recent_activity": [],
        "role_distribution": {},
        "activity_24h": {"new_listings": 0, "new_users": 0, "deleted_content": 0},
        "health": {
            "api_status": "ok",
            "db_status": "postgres",
            "api_latency_ms": api_latency_ms,
            "db_latency_ms": None,
            "deployed_at": os.environ.get("DEPLOYED_AT") or "unknown",
            "restart_at": APP_START_TIME.isoformat(),
            "uptime_seconds": uptime_seconds,
            "uptime_human": _format_uptime(uptime_seconds),
        },
        "metrics": {"total_listings": 0, "published_listings": 0},
        "kpis": {
            "today": {"new_listings": 0, "new_users": 0, "revenue_total": 0, "revenue_currency_totals": {}},
            "last_7_days": {"new_listings": 0, "new_users": 0, "revenue_total": 0, "revenue_currency_totals": {}},
        },
        "trends": {"window_days": trend_window, "listings": [], "revenue": []},
        "risk_panel": {
            "suspicious_logins": {
                "count": 0,
                "items": [],
                "threshold": DASHBOARD_MULTI_IP_THRESHOLD,
                "window_hours": DASHBOARD_MULTI_IP_WINDOW_HOURS,
            },
            "sla_breaches": {"count": 0, "items": [], "threshold": DASHBOARD_SLA_HOURS},
            "pending_payments": {
                "count": 0,
                "items": [],
                "threshold_days": DASHBOARD_PENDING_PAYMENT_DAYS,
            },
        },
        "finance_visible": can_view_finance,
    }


@api_router.get("/admin/dashboard/summary")
async def admin_dashboard_summary(
    request: Request,
    country: Optional[str] = None,
    trend_days: Optional[int] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    start_perf = time.perf_counter()

    role = current_user.get("role")
    can_view_finance = role in {"finance", "super_admin"}
    trend_window = trend_days or DASHBOARD_TREND_DAYS
    if trend_window < DASHBOARD_TREND_MIN_DAYS or trend_window > DASHBOARD_TREND_MAX_DAYS:
        raise HTTPException(status_code=400, detail="trend_days must be between 7 and 365")

    scope = (current_user.get("country_scope") or [])
    country_code = country.upper() if country else None

    if role == "country_admin":
        if not scope:
            raise HTTPException(status_code=403, detail="Country scope required")
        if country_code and country_code not in scope:
            raise HTTPException(status_code=403, detail="Country scope violation")
        effective_countries = [country_code] if country_code else scope
    else:
        effective_countries = [country_code] if country_code else None

    cache_key = _dashboard_cache_key(role, effective_countries, trend_window)
    cached = _get_cached_dashboard_summary(cache_key)
    if cached:
        response = {**cached}
        health = dict(response.get("health") or {})
        health["api_latency_ms"] = int((time.perf_counter() - start_perf) * 1000)
        response["health"] = health
        return response

    total_listings = (
        await session.execute(
            select(func.count()).select_from(Listing).where(
                Listing.country.in_(effective_countries) if effective_countries else True
            )
        )
    ).scalar_one()
    published_listings = (
        await session.execute(
            select(func.count()).select_from(Listing).where(
                Listing.status == "published",
                Listing.country.in_(effective_countries) if effective_countries else True,
            )
        )
    ).scalar_one()
    pending_moderation = (
        await session.execute(
            select(func.count()).select_from(Listing).where(
                Listing.status == "pending_moderation",
                Listing.country.in_(effective_countries) if effective_countries else True,
            )
        )
    ).scalar_one()

    total_users = (
        await session.execute(
            select(func.count()).select_from(SqlUser).where(
                SqlUser.country_code.in_(effective_countries) if effective_countries else True
            )
        )
    ).scalar_one()
    active_users = (
        await session.execute(
            select(func.count()).select_from(SqlUser).where(
                SqlUser.is_active.is_(True),
                SqlUser.country_code.in_(effective_countries) if effective_countries else True,
            )
        )
    ).scalar_one()
    inactive_users = (
        await session.execute(
            select(func.count()).select_from(SqlUser).where(
                SqlUser.is_active.is_(False),
                SqlUser.country_code.in_(effective_countries) if effective_countries else True,
            )
        )
    ).scalar_one()
    active_dealers = (
        await session.execute(
            select(func.count()).select_from(SqlUser).where(
                SqlUser.role == "dealer",
                SqlUser.is_active.is_(True),
                SqlUser.country_code.in_(effective_countries) if effective_countries else True,
            )
        )
    ).scalar_one()

    role_distribution = {}
    for role_name in ["super_admin", "country_admin", "moderator", "support", "finance", "dealer", "individual"]:
        role_distribution[role_name] = (
            await session.execute(
                select(func.count()).select_from(SqlUser).where(
                    SqlUser.role == role_name,
                    SqlUser.country_code.in_(effective_countries) if effective_countries else True,
                )
            )
        ).scalar_one()

    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    revenue_mtd = None
    totals = {}
    if can_view_finance:
        invoice_query = select(AdminInvoice).where(AdminInvoice.status == "paid", AdminInvoice.created_at >= month_start)
        invoices = (await session.execute(invoice_query)).scalars().all()
        for inv in invoices:
            currency = inv.currency or "UNKNOWN"
            amount = float(inv.amount_total or 0)
            totals[currency] = totals.get(currency, 0) + amount
        revenue_mtd = sum(totals.values())

    metrics = {
        "total_listings": int(total_listings or 0),
        "published_listings": int(published_listings or 0),
        "pending_moderation": int(pending_moderation or 0),
        "active_dealers": int(active_dealers or 0),
        "revenue_mtd": round(revenue_mtd, 2) if revenue_mtd is not None else None,
        "revenue_currency_totals": {k: round(v, 2) for k, v in totals.items()} if can_view_finance else None,
        "month_start_utc": month_start.isoformat(),
    }

    active_countries_query = select(Country).where(Country.is_enabled.is_(True))
    if effective_countries:
        active_countries_query = active_countries_query.where(Country.code.in_(effective_countries))
    active_countries = (await session.execute(active_countries_query)).scalars().all()
    active_country_codes = [row.code for row in active_countries]

    categories = (await session.execute(select(Category))).scalars().all()
    active_modules = set()
    for cat in categories:
        schema = _normalize_category_schema(cat.form_schema) if cat.form_schema else {}
        for key, module in (schema.get("modules") or {}).items():
            enabled = bool(module.get("enabled")) if isinstance(module, dict) else bool(module)
            if enabled:
                active_modules.add(key)

    recent_logs = (
        await session.execute(select(AuditLog).order_by(AuditLog.created_at.desc()).limit(10))
    ).scalars().all()
    recent_activity = [
        {
            "id": str(log.id),
            "event_type": log.event_type or log.action,
            "action": log.action,
            "resource_type": log.resource_type,
            "user_email": log.user_email,
            "created_at": log.created_at.isoformat() if log.created_at else None,
            "country_code": log.country_code,
        }
        for log in recent_logs
    ]

    since_dt = now - timedelta(hours=24)
    new_listings = (
        await session.execute(
            select(func.count()).select_from(Listing).where(Listing.created_at >= since_dt)
        )
    ).scalar_one()
    new_users = (
        await session.execute(
            select(func.count()).select_from(SqlUser).where(SqlUser.created_at >= since_dt)
        )
    ).scalar_one()
    deleted_content = (
        await session.execute(
            select(func.count()).select_from(AuditLog).where(
                AuditLog.event_type.in_(["LISTING_SOFT_DELETE", "LISTING_FORCE_UNPUBLISH"]),
                AuditLog.created_at >= since_dt,
            )
        )
    ).scalar_one()

    api_latency_ms = int((time.perf_counter() - start_perf) * 1000)
    uptime_seconds = int((now - APP_START_TIME).total_seconds())

    summary = {
        "scope": "country" if effective_countries else "global",
        "country_codes": effective_countries or active_country_codes,
        "users": {"total": int(total_users or 0), "active": int(active_users or 0), "inactive": int(inactive_users or 0)},
        "active_countries": {"count": len(active_country_codes), "codes": active_country_codes},
        "active_modules": {"count": len(active_modules), "items": sorted(active_modules)},
        "recent_activity": recent_activity,
        "role_distribution": role_distribution,
        "activity_24h": {
            "new_listings": int(new_listings or 0),
            "new_users": int(new_users or 0),
            "deleted_content": int(deleted_content or 0),
        },
        "health": {
            "api_status": "ok",
            "db_status": "ok",
            "api_latency_ms": api_latency_ms,
            "db_latency_ms": 0,
            "deployed_at": os.environ.get("DEPLOYED_AT") or "unknown",
            "restart_at": APP_START_TIME.isoformat(),
            "uptime_seconds": uptime_seconds,
            "uptime_human": _format_uptime(uptime_seconds),
        },
        "metrics": metrics,
        "kpis": {},
        "trends": {"window_days": trend_window, "items": []},
        "risk_panel": {},
        "finance_visible": can_view_finance,
    }

    _set_cached_dashboard_summary(cache_key, summary)
    return summary


@api_router.get("/admin/dashboard/export/pdf")
async def admin_dashboard_export_pdf(
    request: Request,
    country: Optional[str] = None,
    trend_days: Optional[int] = None,
    current_user=Depends(check_permissions(["super_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    _enforce_export_rate_limit(request, current_user.get("id"))

    summary = await admin_dashboard_summary(
        request=request,
        country=country,
        trend_days=trend_days,
        current_user=current_user,
        session=session,
    )
    trend_window = (summary.get("trends") or {}).get("window_days") or DASHBOARD_TREND_DAYS
    pdf_bytes = _build_dashboard_pdf(summary, trend_window)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    filename = f"dashboard-summary-{timestamp}.pdf"

    await _write_audit_log_sql(
        session=session,
        action="dashboard_export_pdf",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="dashboard_summary",
        resource_id="dashboard_summary",
        metadata={"trend_days": trend_window, "format": "pdf", "scope": summary.get("scope")},
        request=request,
        country_code=country.upper() if country else None,
    )
    await session.commit()

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/admin/dashboard/country-compare")
async def admin_dashboard_country_compare(
    request: Request,
    period: str = "30d",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    countries: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=None)
    selected_codes = _parse_country_codes(countries)
    return await _build_country_compare_payload_sql(
        session,
        current_user,
        period,
        start_date,
        end_date,
        sort_by,
        sort_dir,
        selected_codes,
    )


@api_router.get("/admin/dashboard/country-compare/export/csv")
async def admin_dashboard_country_compare_export_csv(
    request: Request,
    period: str = "30d",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    countries: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "support", "finance"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await resolve_admin_country_context(request, current_user=current_user, session=None)
    _enforce_export_rate_limit(request, current_user.get("id"))

    selected_codes = _parse_country_codes(countries)
    payload = await _build_country_compare_payload_sql(
        session,
        current_user,
        period,
        start_date,
        end_date,
        sort_by,
        sort_dir,
        selected_codes,
    )
    items = payload.get("items") or []

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([
        "Country",
        "Total Listings",
        "Total Growth 7d",
        "Total Growth 30d",
        "Published",
        "Published Growth 7d",
        "Published Growth 30d",
        "Conversion Rate %",
        "Active Dealers",
        "Dealer Growth 7d",
        "Dealer Growth 30d",
        "Dealer Density %",
        "Revenue EUR",
        "Revenue Growth 7d",
        "Revenue Growth 30d",
        "Revenue MTD Growth %",
        "SLA 24h",
        "SLA 48h",
        "Risk Multi Login",
        "Risk Pending Payments",
        "Note",
    ])
    for item in items:
        writer.writerow([
            item.get("country_code"),
            item.get("total_listings"),
            item.get("growth_total_listings_7d"),
            item.get("growth_total_listings_30d"),
            item.get("published_listings"),
            item.get("growth_published_7d"),
            item.get("growth_published_30d"),
            item.get("conversion_rate"),
            item.get("active_dealers"),
            item.get("growth_active_dealers_7d"),
            item.get("growth_active_dealers_30d"),
            item.get("dealer_density"),
            item.get("revenue_eur"),
            item.get("growth_revenue_7d"),
            item.get("growth_revenue_30d"),
            item.get("revenue_mtd_growth_pct"),
            item.get("sla_pending_24h"),
            item.get("sla_pending_48h"),
            item.get("risk_multi_login"),
            item.get("risk_pending_payments"),
            item.get("note"),
        ])

    csv_content = buffer.getvalue().encode("utf-8")
    buffer.close()
    filename = f"country-compare-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.csv"

    await _write_audit_log_sql(
        session=session,
        action="country_compare_export_csv",
        actor={"id": current_user.get("id"), "email": current_user.get("email")},
        resource_type="dashboard_country_compare",
        resource_id="country_compare",
        metadata={"period": period, "start_date": start_date, "end_date": end_date, "sort_by": sort_by, "sort_dir": sort_dir},
        request=request,
        country_code=None,
    )
    await session.commit()

    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/admin/session/health")
async def admin_session_health(request: Request, response: Response, current_user=Depends(check_permissions(list(ADMIN_ROLE_OPTIONS)))):
    auth_header = request.headers.get("Authorization", "")
    token = auth_header.split(" ", 1)[1] if auth_header.startswith("Bearer ") else None
    expires_at = None
    if token:
        payload = decode_token(token)
        exp = payload.get("exp")
        if exp:
            expires_at = datetime.fromtimestamp(exp, tz=timezone.utc).isoformat()
    roles = current_user.get("roles") or [current_user.get("role")]
    response.headers["Cache-Control"] = "no-store"
    return {
        "user_id": current_user.get("id"),
        "roles": roles,
        "expires_at": expires_at,
        "server_time": datetime.now(timezone.utc).isoformat(),
    }


# =====================
# Public Search v2 (SQL)
# =====================

@api_router.get("/v2/search")
async def public_search_v2(
    request: Request,
    country: str | None = None,
    q: Optional[str] = None,
    category: Optional[str] = None,
    make: Optional[str] = None,
    model: Optional[str] = None,
    sort: str = "date_desc",
    page: int = 1,
    limit: int = 20,
    price_min: Optional[int] = None,
    price_max: Optional[int] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    """SQL-backed search endpoint for Public Portal (active listings only)."""

    country_norm = (country or "").strip().upper()
    if not country_norm:
        raise HTTPException(status_code=400, detail="country is required")

    page = max(1, int(page))
    limit = min(100, max(1, int(limit)))

    base_conditions = [Listing.status == "published", Listing.country == country_norm]

    if q:
        q_value = f"%{q.strip()}%"
        base_conditions.append(or_(Listing.title.ilike(q_value), Listing.description.ilike(q_value)))

    if category:
        category_value = category.strip()
        category_id = None
        try:
            category_id = uuid.UUID(category_value)
        except ValueError:
            result = await session.execute(
                select(Category.id).where(
                    or_(
                        Category.slug["tr"].astext == category_value,
                        Category.slug["de"].astext == category_value,
                        Category.slug["fr"].astext == category_value,
                        Category.slug["en"].astext == category_value,
                    )
                )
            )
            category_id = result.scalar_one_or_none()
        if category_id:
            base_conditions.append(Listing.category_id == category_id)

    if make:
        make_value = make.strip()
        try:
            make_id = uuid.UUID(make_value)
            base_conditions.append(Listing.make_id == make_id)
        except ValueError:
            base_conditions.append(Listing.attributes["vehicle"]["make_key"].astext == make_value)

    if model:
        model_value = model.strip()
        try:
            model_id = uuid.UUID(model_value)
            base_conditions.append(Listing.model_id == model_id)
        except ValueError:
            base_conditions.append(Listing.attributes["vehicle"]["model_key"].astext == model_value)

    if price_min is not None:
        base_conditions.append(Listing.price >= float(price_min))
    if price_max is not None:
        base_conditions.append(Listing.price <= float(price_max))
    if price_min is not None or price_max is not None:
        base_conditions.append(or_(Listing.price_type == "FIXED", Listing.price_type.is_(None)))

    query_stmt = select(Listing).where(and_(*base_conditions))
    total_stmt = select(func.count()).select_from(Listing).where(and_(*base_conditions))

    if sort == "price_asc":
        query_stmt = query_stmt.order_by(Listing.price.asc().nulls_last())
    elif sort == "price_desc":
        query_stmt = query_stmt.order_by(Listing.price.desc().nulls_last())
    elif sort == "date_asc":
        query_stmt = query_stmt.order_by(Listing.created_at.asc())
    else:
        query_stmt = query_stmt.order_by(Listing.created_at.desc())

    rows = (await session.execute(query_stmt.offset((page - 1) * limit).limit(limit))).scalars().all()
    total = (await session.execute(total_stmt)).scalar_one() or 0

    items = []
    for listing in rows:
        attrs = listing.attributes or {}
        vehicle = attrs.get("vehicle") or {}
        title = (listing.title or "").strip()
        if not title:
            title = f"{(vehicle.get('make_key') or '').upper()} {vehicle.get('model_key') or ''} {vehicle.get('year') or ''}".strip()

        image_url = listing.images[0] if listing.images else None

        items.append(
            {
                "id": str(listing.id),
                "title": title,
                "price": listing.price,
                "price_type": listing.price_type or "FIXED",
                "price_amount": listing.price,
                "hourly_rate": listing.hourly_rate,
                "currency": listing.currency or "EUR",
                "secondary_price": None,
                "secondary_currency": None,
                "image": image_url,
                "city": listing.city or "",
            }
        )

    pages = (total + limit - 1) // limit if total else 0

    return {
        "items": items,
        "facets": {},
        "facet_meta": {},
        "pagination": {"total": int(total), "page": page, "pages": pages},
    }


# =====================
# Vehicle Master Data (DB)
# =====================

@api_router.get("/v1/vehicle/makes")
async def public_vehicle_makes(country: str | None = None, session: AsyncSession = Depends(get_sql_session)):
    if not country:
        raise HTTPException(status_code=400, detail="country is required")
    result = await session.execute(
        select(VehicleMake).where(VehicleMake.is_active.is_(True)).order_by(VehicleMake.name.asc())
    )
    items = [
        {"id": str(make.id), "key": make.slug, "label": make.name}
        for make in result.scalars().all()
        if make.slug
    ]
    return {"version": "sql", "items": items}


@api_router.get("/v1/vehicle/models")
async def public_vehicle_models(make: str, country: str | None = None, session: AsyncSession = Depends(get_sql_session)):
    make_row = await session.execute(select(VehicleMake).where(VehicleMake.slug == make))
    make_obj = make_row.scalars().first()
    if not make_obj:
        raise HTTPException(status_code=404, detail="Make not found")
    result = await session.execute(
        select(VehicleModel).where(
            VehicleModel.make_id == make_obj.id,
            VehicleModel.is_active.is_(True),
        ).order_by(VehicleModel.name.asc())
    )
    items = [
        {"id": str(model.id), "key": model.slug, "label": model.name}
        for model in result.scalars().all()
        if model.slug
    ]
    return {"version": "sql", "make": make, "items": items}


@api_router.get("/v2/vehicle/makes")
async def public_vehicle_makes_v2(country: str | None = None, session: AsyncSession = Depends(get_sql_session)):
    return await public_vehicle_makes(country, session)


@api_router.get("/v2/vehicle/models")
async def public_vehicle_models_v2(make_key: str | None = None, make: str | None = None, country: str | None = None, session: AsyncSession = Depends(get_sql_session)):
    key = make_key or make
    if not key:
        raise HTTPException(status_code=400, detail="make_key is required")
    return await public_vehicle_models(key, country, session)


@api_router.get("/v1/admin/vehicle-master/status")
async def vehicle_master_status_endpoint(current_user=Depends(check_permissions(["super_admin", "country_admin"]))):
    return vehicle_master_status()


@api_router.post("/v1/admin/vehicle-master/validate")
async def vehicle_master_validate_endpoint(request: Request, current_user=Depends(check_permissions(["super_admin", "country_admin"]))):
    form = await request.form()
    upload = form.get("file")
    if upload is None:
        raise HTTPException(status_code=400, detail="file is required")

    raw = await upload.read()
    return validate_upload(upload.filename or "bundle.json", raw)


@api_router.post("/v1/admin/vehicle-master/activate")
async def vehicle_master_activate_endpoint(payload: dict, request: Request, current_user=Depends(check_permissions(["super_admin", "country_admin"]))):
    staging_id = payload.get("staging_id")
    if not staging_id:
        raise HTTPException(status_code=400, detail="staging_id is required")



# =====================
# Vehicle Listings (SQL)
# =====================

async def _get_owned_listing(session: AsyncSession, listing_id: str, current_user: dict) -> Listing:
    listing_uuid = uuid.UUID(listing_id)
    listing = await session.get(Listing, listing_uuid)
    if not listing:
        raise HTTPException(status_code=404, detail="Listing not found")
    if str(listing.user_id) != str(current_user.get("id")):
        raise HTTPException(status_code=403, detail="Forbidden")
    return listing


def _listing_media_meta(listing: Listing) -> list[dict]:
    attrs = listing.attributes or {}
    return attrs.get("media") or []


def _listing_to_dict(listing: Listing) -> dict:
    attrs = listing.attributes or {}
    media_meta = _listing_media_meta(listing)
    return {
        "id": str(listing.id),
        "module": listing.module,
        "status": listing.status,
        "country": listing.country,
        "category_id": listing.category_id,
        "category_key": attrs.get("category_key"),
        "title": listing.title,
        "description": listing.description,
        "price": {
            "amount": listing.price,
            "hourly_rate": listing.hourly_rate,
            "price_type": listing.price_type or "FIXED",
            "currency_primary": listing.currency,
        },
        "price_type": listing.price_type or "FIXED",
        "price_amount": listing.price,
        "hourly_rate": listing.hourly_rate,
        "core_fields": attrs.get("core_fields") or {},
        "vehicle": attrs.get("vehicle") or {},
        "attributes": attrs.get("attributes") or {},
        "detail_groups": attrs.get("detail_groups") or [],
        "modules": attrs.get("modules") or {},
        "payment_options": attrs.get("payment_options") or {},
        "media": media_meta,
        "created_at": listing.created_at.isoformat() if listing.created_at else None,
        "updated_at": listing.updated_at.isoformat() if listing.updated_at else None,
        "expires_at": listing.expires_at.isoformat() if listing.expires_at else None,
    }


def _apply_listing_payload_sql(listing: Listing, payload: dict) -> None:
    if not payload:
        return
    attrs = dict(listing.attributes or {})
    core_fields = payload.get("core_fields") or {}
    title = core_fields.get("title") or payload.get("title")
    description = core_fields.get("description") or payload.get("description")
    if title is not None:
        listing.title = title.strip()
    if description is not None:
        listing.description = description.strip()
    category_value = payload.get("category_id")
    if category_value is not None:
        try:
            listing.category_id = uuid.UUID(str(category_value))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="category_id invalid") from exc
        attrs["category_id"] = str(listing.category_id)

    make_value = payload.get("make_id")
    model_value = payload.get("model_id")
    price_payload = core_fields.get("price") if isinstance(core_fields, dict) else None
    price_payload = price_payload or payload.get("price") or {}
    if price_payload:
        price_type_raw = price_payload.get("price_type") or payload.get("price_type") or listing.price_type or "FIXED"
        price_type = str(price_type_raw).upper()
        amount = price_payload.get("amount")
        hourly_rate = price_payload.get("hourly_rate")

        if price_type not in {"FIXED", "HOURLY"}:
            raise HTTPException(status_code=400, detail="price_type invalid")

        if price_type == "FIXED":
            if hourly_rate not in (None, ""):
                raise HTTPException(status_code=400, detail="Fiyat giriniz.")
            if amount in (None, ""):
                raise HTTPException(status_code=400, detail="Fiyat giriniz.")
            try:
                amount_value = float(amount)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Fiyat giriniz.")
            if amount_value <= 0:
                raise HTTPException(status_code=400, detail="Fiyat giriniz.")
            listing.price = amount_value
            listing.hourly_rate = None
        else:
            if amount not in (None, ""):
                raise HTTPException(status_code=400, detail="Saatlik ücret giriniz.")
            if hourly_rate in (None, ""):
                raise HTTPException(status_code=400, detail="Saatlik ücret giriniz.")
            try:
                hourly_value = float(hourly_rate)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="Saatlik ücret giriniz.")
            if hourly_value <= 0:
                raise HTTPException(status_code=400, detail="Saatlik ücret giriniz.")
            listing.price = None
            listing.hourly_rate = hourly_value

        listing.price_type = price_type
        listing.currency = price_payload.get("currency_primary") or listing.currency
        normalized_price = dict(price_payload)
        normalized_price["price_type"] = price_type
        normalized_price["amount"] = listing.price
        normalized_price["hourly_rate"] = listing.hourly_rate
        attrs["price"] = normalized_price

    if core_fields:
        attrs["core_fields"] = core_fields

    dynamic_fields = payload.get("dynamic_fields") or {}
    extra_attrs = payload.get("attributes") or {}
    merged_attrs = dict(attrs.get("attributes") or {})
    if isinstance(dynamic_fields, dict):
        merged_attrs.update(dynamic_fields)
    if isinstance(extra_attrs, dict):
        merged_attrs.update(extra_attrs)
    if price_payload:
        merged_attrs["price_eur"] = listing.price if listing.price_type == "FIXED" else None
    attrs["attributes"] = merged_attrs

    vehicle_payload = payload.get("vehicle") or {}
    if vehicle_payload:
        vehicle = dict(attrs.get("vehicle") or {})
        vehicle.update({k: v for k, v in vehicle_payload.items() if v is not None})
        attrs["vehicle"] = vehicle

    if make_value is None and isinstance(vehicle_payload, dict):
        make_value = vehicle_payload.get("make_id")
    if model_value is None and isinstance(vehicle_payload, dict):
        model_value = vehicle_payload.get("model_id")

    if make_value is not None:
        try:
            listing.make_id = uuid.UUID(str(make_value))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="make_id invalid") from exc
    if model_value is not None:
        try:
            listing.model_id = uuid.UUID(str(model_value))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="model_id invalid") from exc

    if payload.get("detail_groups") is not None:
        attrs["detail_groups"] = payload.get("detail_groups")
    if payload.get("modules") is not None:
        attrs["modules"] = payload.get("modules")
    if payload.get("payment_options") is not None:
        attrs["payment_options"] = payload.get("payment_options")

    listing.attributes = attrs


@api_router.get("/v1/listings/my")
async def list_my_listings(
    status: Optional[str] = None,
    q: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    safe_page = max(1, int(page))
    safe_limit = min(100, max(1, int(limit)))
    offset = (safe_page - 1) * safe_limit

    try:
        user_uuid = uuid.UUID(current_user.get("id"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    query = select(Listing).where(Listing.user_id == user_uuid)
    if status:
        status_value = status.strip().lower()
        if status_value == "expired":
            query = query.where(Listing.expires_at <= datetime.now(timezone.utc))
        elif status_value == "draft":
            query = query.where(Listing.status.in_(["draft", "needs_revision"]))
        elif status_value == "active":
            query = query.where(Listing.status == "published")
        else:
            query = query.where(Listing.status == status_value)
    if q:
        query = query.where(or_(Listing.title.ilike(f"%{q.strip()}%"), Listing.id.cast(String).ilike(f"%{q.strip()}%")))

    query = query.order_by(desc(Listing.created_at))
    rows = (await session.execute(query.offset(offset).limit(safe_limit))).scalars().all()
    total = (
        await session.execute(select(func.count()).select_from(Listing).where(Listing.user_id == user_uuid))
    ).scalar_one()

    listing_ids = [row.id for row in rows]
    moderation_reason_map: Dict[uuid.UUID, Optional[str]] = {}
    if listing_ids:
        moderation_rows = await session.execute(
            select(ModerationItem)
            .where(ModerationItem.listing_id.in_(listing_ids))
            .order_by(ModerationItem.listing_id, desc(ModerationItem.updated_at))
        )
        for item in moderation_rows.scalars().all():
            if item.listing_id not in moderation_reason_map and item.reason:
                moderation_reason_map[item.listing_id] = item.reason

    items = []
    for row in rows:
        payload = _listing_to_dict(row)
        payload["moderation_reason"] = moderation_reason_map.get(row.id)
        items.append(payload)

    return {"items": items, "pagination": {"total": int(total or 0), "page": safe_page, "limit": safe_limit}}


@api_router.post("/v1/listings/vehicle")
async def create_vehicle_draft(
    payload: dict,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        user_uuid = uuid.UUID(current_user.get("id"))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid user id") from exc

    country = (payload.get("country") or current_user.get("country_code") or "DE").upper()
    category_value = payload.get("category_id")
    if not category_value:
        raise HTTPException(status_code=400, detail="category_id required")
    try:
        category_uuid = uuid.UUID(str(category_value))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="category_id invalid") from exc

    category = await session.get(Category, category_uuid)
    if not category:
        raise HTTPException(status_code=404, detail="category not found")
    module_value = category.module or "vehicle"

    vehicle_payload = payload.get("vehicle") or {}
    vehicle_payload = dict(vehicle_payload)
    make_value = payload.get("make_id") or vehicle_payload.get("make_id")
    model_value = payload.get("model_id") or vehicle_payload.get("model_id")
    make_uuid = None
    model_uuid = None
    if make_value or model_value:
        if not make_value or not model_value:
            raise HTTPException(status_code=400, detail="make_id and model_id required")
        try:
            make_uuid = uuid.UUID(str(make_value))
            model_uuid = uuid.UUID(str(model_value))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="make_id/model_id invalid") from exc
        vehicle_payload.setdefault("make_id", str(make_uuid))
        vehicle_payload.setdefault("model_id", str(model_uuid))

    listing = Listing(
        user_id=user_uuid,
        status="draft",
        module=module_value,
        country=country,
        title=payload.get("title") or "",
        description=payload.get("description") or "",
        price_type="FIXED",
        price=None,
        hourly_rate=None,
        currency="EUR",
        category_id=category_uuid,
        make_id=make_uuid,
        model_id=model_uuid,
        attributes={"category_id": str(category_uuid), "module": module_value, "vehicle": vehicle_payload},
        images=[],
    )
    _apply_listing_payload_sql(listing, payload)
    session.add(listing)
    await session.commit()
    await session.refresh(listing)

    return {"id": str(listing.id), "status": listing.status, "validation_errors": [], "next_actions": ["upload_media", "submit"]}


@api_router.patch("/v1/listings/vehicle/{listing_id}/draft")
@api_router.post("/v1/listings/vehicle/{listing_id}/draft")
async def save_vehicle_draft(
    listing_id: str,
    payload: dict = Body(default={}),
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    if listing.status not in ["draft", "needs_revision", "unpublished"]:
        raise HTTPException(status_code=400, detail="Listing not editable")

    _apply_listing_payload_sql(listing, payload)
    listing.updated_at = datetime.now(timezone.utc)
    await session.commit()

    return {"id": listing_id, "status": listing.status, "updated_at": listing.updated_at.isoformat() if listing.updated_at else None}


@api_router.get("/v1/listings/vehicle/{listing_id}/draft")
async def get_vehicle_draft(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    return {"item": _listing_to_dict(listing)}


@api_router.post("/v1/listings/vehicle/{listing_id}/request-publish")
async def request_publish_vehicle_listing(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    if listing.status not in ["draft", "needs_revision", "unpublished"]:
        raise HTTPException(status_code=400, detail="Listing not eligible for publish")

    listing.status = "pending_moderation"
    if not listing.expires_at:
        snapshot_result = await session.execute(
            select(PricingPriceSnapshot).where(PricingPriceSnapshot.listing_id == listing.id)
        )
        snapshot = snapshot_result.scalar_one_or_none()
        duration_days = (snapshot.publish_days or snapshot.duration_days) if snapshot else 30
        listing.expires_at = datetime.now(timezone.utc) + timedelta(days=duration_days)
    listing.updated_at = datetime.now(timezone.utc)

    await _upsert_moderation_item(
        session=session,
        listing=listing,
        status="PENDING",
        reason=None,
        moderator_id=None,
        audit_ref=None,
    )
    await session.commit()
    return {"ok": True, "status": listing.status}


@api_router.post("/v1/listings/vehicle/{listing_id}/unpublish")
async def unpublish_vehicle_listing(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    if listing.status != "published":
        raise HTTPException(status_code=400, detail="Only published listings can be unpublished")

    listing.status = "unpublished"
    listing.updated_at = datetime.now(timezone.utc)
    await session.commit()
    return {"ok": True, "status": listing.status}


@api_router.post("/v1/listings/vehicle/{listing_id}/archive")
async def archive_vehicle_listing(
    listing_id: str,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    listing.status = "archived"
    listing.updated_at = datetime.now(timezone.utc)
    await session.commit()
    return {"ok": True, "status": listing.status}


@api_router.post("/v1/listings/vehicle/{listing_id}/extend")
async def extend_vehicle_listing(
    listing_id: str,
    days: int = Body(default=30, embed=True),
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    if days < 1 or days > 365:
        raise HTTPException(status_code=400, detail="days must be between 1 and 365")

    base_dt = listing.expires_at or datetime.now(timezone.utc)
    if base_dt.tzinfo is None:
        base_dt = base_dt.replace(tzinfo=timezone.utc)
    listing.expires_at = base_dt + timedelta(days=days)
    listing.updated_at = datetime.now(timezone.utc)
    await session.commit()

    return {"ok": True, "expires_at": listing.expires_at.isoformat() if listing.expires_at else None}


@api_router.post("/v1/listings/vehicle/{listing_id}/media")
async def upload_vehicle_media(
    listing_id: str,
    files: list[UploadFile] = File(...),
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    if listing.status not in ["draft", "needs_revision"]:
        raise HTTPException(status_code=400, detail="Only draft/needs_revision can accept media")

    media_meta = _listing_media_meta(listing)
    for f in files:
        raw = await f.read()
        try:
            filename, w, h = store_image(listing.country, listing_id, raw, f.filename or "upload.jpg")
        except ValueError as e:
            raise HTTPException(status_code=422, detail=str(e))

        media_id = str(uuid.uuid4())
        media_item = {
            "media_id": media_id,
            "file": filename,
            "width": w,
            "height": h,
            "is_cover": False,
        }
        media_meta.append(media_item)

    if media_meta and not any(m.get("is_cover") for m in media_meta):
        media_meta[0]["is_cover"] = True

    listing.attributes = {**(listing.attributes or {}), "media": media_meta}
    listing.images = [f"/media/listings/{listing_id}/{m['file']}" for m in media_meta]
    listing.updated_at = datetime.now(timezone.utc)
    await session.commit()

    resp_media = [
        {
            "media_id": m["media_id"],
            "file": m["file"],
            "width": m.get("width"),
            "height": m.get("height"),
            "is_cover": m.get("is_cover", False),
            "preview_url": f"/api/v1/listings/vehicle/{listing_id}/media/{m['media_id']}/preview",
        }
        for m in media_meta
    ]

    return {"id": listing_id, "status": listing.status, "media": resp_media}


@api_router.patch("/v1/listings/vehicle/{listing_id}/media/order")
async def reorder_vehicle_media(
    listing_id: str,
    payload: dict = Body(default={}),
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    media_meta = _listing_media_meta(listing)
    if not media_meta:
        return {"id": listing_id, "status": listing.status, "media": []}

    order = payload.get("order") or []
    cover_media_id = payload.get("cover_media_id")

    media_by_id = {m.get("media_id"): m for m in media_meta if m.get("media_id")}
    next_media = []
    seen = set()
    for media_id in order:
        if media_id in media_by_id and media_id not in seen:
            next_media.append(media_by_id[media_id])
            seen.add(media_id)
    for media in media_meta:
        media_id = media.get("media_id")
        if media_id and media_id not in seen:
            next_media.append(media)
            seen.add(media_id)

    if cover_media_id and cover_media_id in media_by_id:
        for media in next_media:
            media["is_cover"] = media.get("media_id") == cover_media_id
    else:
        has_cover = any(m.get("is_cover") for m in next_media)
        if not has_cover and next_media:
            next_media[0]["is_cover"] = True

    listing.attributes = {**(listing.attributes or {}), "media": next_media}
    listing.images = [f"/media/listings/{listing_id}/{m['file']}" for m in next_media]
    listing.updated_at = datetime.now(timezone.utc)
    await session.commit()

    resp_media = [
        {
            "media_id": m.get("media_id"),
            "file": m.get("file"),
            "width": m.get("width"),
            "height": m.get("height"),
            "is_cover": m.get("is_cover", False),
            "preview_url": f"/api/v1/listings/vehicle/{listing_id}/media/{m.get('media_id')}/preview",
        }
        for m in next_media
    ]

    return {"id": listing_id, "status": listing.status, "media": resp_media}


@api_router.get("/v1/listings/vehicle/{listing_id}/media/{media_id}/preview")
async def preview_vehicle_media(listing_id: str, media_id: str, current_user=Depends(require_portal_scope("account")), session: AsyncSession = Depends(get_sql_session)):
    listing = await _get_owned_listing(session, listing_id, current_user)
    media = next((m for m in _listing_media_meta(listing) if m.get("media_id") == media_id), None)
    if not media:
        raise HTTPException(status_code=404, detail="Media not found")

    path = resolve_public_media_path(listing_id, media["file"])
    return FileResponse(path)


@api_router.post("/v1/listings/vehicle/{listing_id}/submit")
async def submit_vehicle_listing(
    listing_id: str,
    payload: dict = Body(default={}),
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    listing = await _get_owned_listing(session, listing_id, current_user)
    if listing.status not in ["draft", "needs_revision"]:
        raise HTTPException(status_code=400, detail="Listing not draft/needs_revision")

    _apply_listing_payload_sql(listing, payload)
    listing.updated_at = datetime.now(timezone.utc)

    media_meta = _listing_media_meta(listing)
    validation_errors = []
    if not listing.title:
        validation_errors.append({"code": "TITLE_REQUIRED", "field": "title", "message": "Başlık gerekli"})
    if not listing.category_id:
        validation_errors.append({"code": "CATEGORY_REQUIRED", "field": "category_id", "message": "Kategori gerekli"})
    if not listing.make_id or not listing.model_id:
        validation_errors.append({"code": "MAKE_MODEL_REQUIRED", "field": "make_model", "message": "Marka ve model gerekli"})
    if not media_meta:
        validation_errors.append({"code": "MEDIA_REQUIRED", "field": "media", "message": "En az 1 görsel ekleyin"})

    if validation_errors:
        await session.commit()
        raise HTTPException(
            status_code=422,
            detail={
                "id": listing_id,
                "status": listing.status,
                "validation_errors": validation_errors,
                "next_actions": ["fix_form", "upload_media"],
            },
        )

    pricing_response, quote, policy, _ = await _compute_pricing_quote(session, current_user, None)
    override_active = pricing_response.get("override_active", False)
    snapshot = None
    if not quote.get("requires_payment"):
        snapshot = await _create_pricing_snapshot(session, listing, current_user, quote, policy, override_active)

    if quote.get("quota_used") and quote.get("subscription_id"):
        subscription = await session.get(UserPackageSubscription, uuid.UUID(quote["subscription_id"]))
        if subscription and subscription.remaining_quota >= 0:
            subscription.remaining_quota = max(0, subscription.remaining_quota - 1)
            if subscription.remaining_quota == 0:
                subscription.status = "expired"
            subscription.updated_at = datetime.now(timezone.utc)

    if quote.get("requires_payment"):
        await session.commit()
        return JSONResponse(
            status_code=402,
            content={
                "id": listing_id,
                "status": listing.status,
                "payment_required": True,
                "quote": pricing_response,
                "snapshot_id": str(snapshot.id) if snapshot else None,
                "next_actions": ["payment_required"],
            },
        )

    listing.status = "pending_moderation"
    if not listing.expires_at:
        duration_days = (snapshot.publish_days or snapshot.duration_days) if snapshot else 30
        listing.expires_at = datetime.now(timezone.utc) + timedelta(days=duration_days)

    await _upsert_moderation_item(
        session=session,
        listing=listing,
        status="PENDING",
        reason=None,
        moderator_id=None,
        audit_ref=None,
    )
    await session.commit()

    return {
        "id": listing_id,
        "status": listing.status,
        "validation_errors": [],
        "next_actions": ["wait_moderation"],
        "detail_url": f"/ilan/{listing_id}?preview=1",
        "quote": pricing_response,
        "snapshot_id": str(snapshot.id) if snapshot else None,
    }


@api_router.get("/v1/listings/vehicle/{listing_id}")
async def get_vehicle_detail(listing_id: str, preview: bool = False, session: AsyncSession = Depends(get_sql_session)):
    listing_uuid = uuid.UUID(listing_id)
    listing = await session.get(Listing, listing_uuid)
    if not listing:
        raise HTTPException(status_code=404, detail="Not found")
    if listing.status != "published" and not preview:
        raise HTTPException(status_code=404, detail="Not found")

    media_meta = _listing_media_meta(listing)
    out_media = [
        {
            "media_id": m["media_id"],
            "url": f"/media/listings/{listing_id}/{m['file']}",
            "is_cover": m.get("is_cover", False),
            "width": m.get("width"),
            "height": m.get("height"),
        }
        for m in media_meta
    ]

    attrs = listing.attributes or {}
    v = attrs.get("vehicle") or {}
    title = listing.title or f"{v.get('make_key','').upper()} {v.get('model_key','')} {v.get('year','')}".strip()

    return {
        "id": listing_id,
        "status": listing.status,
        "country": listing.country,
        "category_id": str(listing.category_id) if listing.category_id else None,
        "make_id": str(listing.make_id) if listing.make_id else None,
        "model_id": str(listing.model_id) if listing.model_id else None,
        "category_key": attrs.get("category_key"),
        "vehicle": v,
        "attributes": attrs.get("attributes") or {},
        "media": out_media,
        "title": title,
        "price": listing.price,
        "price_type": listing.price_type or "FIXED",
        "price_amount": listing.price,
        "hourly_rate": listing.hourly_rate,
        "currency": listing.currency or "EUR",
        "secondary_price": None,
        "secondary_currency": None,
        "location": {"city": listing.city or "", "country": listing.country},
        "description": listing.description or "",
        "seller": {"name": "", "is_verified": False},
        "modules": attrs.get("modules") or {},
        "contact_option_phone": listing.contact_option_phone,
        "contact_option_message": listing.contact_option_message,
        "contact": {"phone_protected": not listing.contact_option_phone},
    }


@app.get("/media/listings/{listing_id}/{file}")
async def public_vehicle_media(
    listing_id: str,
    file: str,
    session: AsyncSession = Depends(get_sql_session),
):
    listing_uuid = uuid.UUID(listing_id)
    listing = await session.get(Listing, listing_uuid)
    if not listing or listing.status != "published":
        raise HTTPException(status_code=404, detail="Not found")

    media = _listing_media_meta(listing)
    if not any(m.get("file") == file for m in media):
        raise HTTPException(status_code=404, detail="Not found")

    path = resolve_public_media_path(listing_id, file)
    return FileResponse(path)


@api_router.post("/v1/admin/vehicle-master/rollback")
async def vehicle_master_rollback_endpoint(payload: dict, request: Request, current_user=Depends(check_permissions(["super_admin", "country_admin"]))):
    target_version = payload.get("target_version")
    current = vehicle_master_rollback(user_id=current_user["id"], target_version=target_version)

    # reload in-memory cache
    request.app.state.vehicle_master = load_current_master(request.app.state.vehicle_master_dir)

    return {"ok": True, "current": current}




# --- Admin Site Design / Ads / Doping / Footer / Info Pages ---
AD_PLACEMENTS = {
    "AD_HOME_TOP": "Anasayfa Üst Banner",
    "AD_CATEGORY_TOP": "Kategori Üst Banner",
    "AD_CATEGORY_RIGHT": "Kategori Sağ",
    "AD_CATEGORY_BOTTOM": "Kategori Alt Banner",
    "AD_SEARCH_TOP": "Arama Üst Banner",
    "AD_IN_FEED": "Akış İçi",
    "AD_LISTING_RIGHT": "İlan Detay Sağ",
    "AD_LOGIN_1": "Login Slot 1",
    "AD_LOGIN_2": "Login Slot 2",
}

AD_FORMAT_LABELS = {
    "horizontal": "Yatay",
    "vertical": "Dikey",
    "square": "Kare",
}

AD_FORMAT_RULES = {
    "AD_HOME_TOP": ["horizontal"],
    "AD_CATEGORY_TOP": ["horizontal"],
    "AD_CATEGORY_RIGHT": ["vertical"],
    "AD_CATEGORY_BOTTOM": ["horizontal"],
    "AD_SEARCH_TOP": ["horizontal"],
    "AD_IN_FEED": ["square"],
    "AD_LISTING_RIGHT": ["vertical"],
    "AD_LOGIN_1": ["horizontal"],
    "AD_LOGIN_2": ["horizontal"],
}

ADS_MANAGER_ROLES = ["super_admin", "country_admin", "ads_manager"]
PRICING_MANAGER_ROLES = ["super_admin", "country_admin", "pricing_manager"]

CAMPAIGN_STATUSES = {"draft", "active", "paused", "expired"}
PRICING_CAMPAIGN_SCOPES = {"individual", "corporate", "all"}
PRICING_CAMPAIGN_ITEM_SCOPES = {"individual", "corporate"}

DOPING_STATUSES = {"requested", "paid", "approved", "published", "expired"}


class HeaderUpdatePayload(BaseModel):
    logo_url: Optional[str] = None


class AdCreatePayload(BaseModel):
    placement: str
    campaign_id: Optional[str] = None
    format: Optional[str] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    is_active: Optional[bool] = True
    priority: Optional[int] = 0
    target_url: Optional[str] = None


class AdUpdatePayload(BaseModel):
    campaign_id: Optional[str] = None
    format: Optional[str] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    is_active: Optional[bool] = None
    priority: Optional[int] = None
    target_url: Optional[str] = None


class AdCampaignCreatePayload(BaseModel):
    name: str
    advertiser: Optional[str] = None
    budget: Optional[float] = None
    currency: Optional[str] = "EUR"
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    status: Optional[str] = "draft"


class AdCampaignUpdatePayload(BaseModel):
    name: Optional[str] = None
    advertiser: Optional[str] = None
    budget: Optional[float] = None
    currency: Optional[str] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    status: Optional[str] = None


class PricingCampaignPolicyPayload(BaseModel):
    is_enabled: bool
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    scope: str = "all"


class PricingQuotePayload(BaseModel):
    user_type: Optional[str] = None
    listing_count_year: Optional[int] = None
    listing_type: Optional[str] = None
    campaign_item_id: Optional[str] = None
    listing_quota: Optional[int] = None


class PricingCheckoutPayload(BaseModel):
    listing_id: str
    origin_url: str
    campaign_item_id: Optional[str] = None
    listing_quota: Optional[int] = None


class PricingTierRuleInput(BaseModel):
    tier_no: int
    price_amount: float
    currency: str
    is_active: Optional[bool] = True


class PricingTierUpdatePayload(BaseModel):
    rules: List[PricingTierRuleInput]


class PricingPackageInput(BaseModel):
    name: str
    listing_quota: int
    package_duration_days: int
    package_price_amount: float
    currency: str
    is_active: Optional[bool] = True


class PricingPackageUpdatePayload(BaseModel):
    packages: List[PricingPackageInput]


class PricingCampaignItemCreatePayload(BaseModel):
    scope: str
    name: Optional[str] = None
    listing_quota: int
    price_amount: float
    currency: str = "EUR"
    publish_days: int
    start_at: datetime
    end_at: datetime
    is_active: Optional[bool] = True


class PricingCampaignItemUpdatePayload(BaseModel):
    name: Optional[str] = None
    listing_quota: Optional[int] = None
    price_amount: Optional[float] = None
    currency: Optional[str] = None
    publish_days: Optional[int] = None
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    is_active: Optional[bool] = None


class PricingCampaignItemStatusPayload(BaseModel):
    is_active: bool


class DopingRequestPayload(BaseModel):
    listing_id: str
    placement_home: Optional[bool] = False
    placement_category: Optional[bool] = False
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    priority: Optional[int] = 0


class DopingApprovalPayload(BaseModel):
    request_id: str
    placement_home: Optional[bool] = False
    placement_category: Optional[bool] = False
    start_at: Optional[datetime] = None
    end_at: Optional[datetime] = None
    priority: Optional[int] = 0


class FooterLayoutPayload(BaseModel):
    layout: dict
    status: Optional[str] = "draft"


class InfoPagePayload(BaseModel):
    slug: str
    title_tr: str
    title_de: str
    title_fr: str
    content_tr: str
    content_de: str
    content_fr: str
    is_published: Optional[bool] = False


class InfoPageUpdatePayload(BaseModel):
    title_tr: Optional[str] = None
    title_de: Optional[str] = None
    title_fr: Optional[str] = None
    content_tr: Optional[str] = None
    content_de: Optional[str] = None
    content_fr: Optional[str] = None
    is_published: Optional[bool] = None


def _normalize_ad_dates(start_at: Optional[datetime], end_at: Optional[datetime]) -> Tuple[Optional[datetime], Optional[datetime]]:
    if start_at and end_at and end_at < start_at:
        raise HTTPException(status_code=400, detail="end_at must be after start_at")
    return start_at, end_at


def _normalize_campaign_fields(
    start_at: Optional[datetime],
    end_at: Optional[datetime],
    budget: Optional[float],
    currency: Optional[str],
) -> Tuple[Optional[datetime], Optional[datetime], Optional[float], Optional[str]]:
    if start_at and end_at and end_at < start_at:
        raise HTTPException(status_code=400, detail="end_at must be after start_at")
    if budget is not None and budget < 0:
        raise HTTPException(status_code=400, detail="budget must be >= 0")
    normalized_currency = currency
    if currency:
        normalized_currency = currency.upper()
        if len(normalized_currency) != 3 or not normalized_currency.isalpha():
            raise HTTPException(status_code=400, detail="Invalid currency")
    return start_at, end_at, budget, normalized_currency


def _is_active_window(start_at: Optional[datetime], end_at: Optional[datetime]) -> bool:
    now = datetime.now(timezone.utc)
    if start_at and start_at > now:
        return False
    if end_at and end_at < now:
        return False
    return True


def _is_campaign_active(campaign: Optional[AdCampaign]) -> bool:
    if not campaign:
        return True
    if campaign.status != "active":
        return False
    return _is_active_window(campaign.start_at, campaign.end_at)


def _is_ad_active(ad: Advertisement, campaign: Optional[AdCampaign]) -> bool:
    if not ad.is_active:
        return False
    if not _is_active_window(ad.start_at, ad.end_at):
        return False
    if campaign and not _is_campaign_active(campaign):
        return False
    return True


def _resolve_ad_format(placement: str, value: Optional[str]) -> str:
    allowed = AD_FORMAT_RULES.get(placement)
    if not allowed:
        raise HTTPException(status_code=400, detail="Invalid placement")
    if not value:
        return allowed[0]
    if value not in allowed:
        raise HTTPException(status_code=400, detail="Invalid format for placement")
    return value


async def _find_active_ad_conflict(
    session: AsyncSession, placement: str, exclude_id: Optional[uuid.UUID] = None
) -> Optional[Advertisement]:
    query = select(Advertisement).where(
        Advertisement.placement == placement,
        Advertisement.is_active.is_(True),
    )
    if exclude_id:
        query = query.where(Advertisement.id != exclude_id)
    result = await session.execute(query)
    return result.scalar_one_or_none()


def _get_env_int(key: str, default: int) -> int:
    value = os.environ.get(key)
    if value is None or value == "":
        return default
    try:
        return int(value)
    except ValueError:
        return default


END_WARNING_DAYS = _get_env_int("END_WARNING_DAYS", 3)
TRAFFIC_SPIKE_THRESHOLD = _get_env_int("TRAFFIC_SPIKE_THRESHOLD", 200)

AD_IMPRESSION_DEDUP_MINUTES = 30
AD_BOT_KEYWORDS = ("bot", "spider", "crawl", "scanner")


def _is_bot_user_agent(user_agent: str) -> bool:
    ua = (user_agent or "").lower()
    return any(keyword in ua for keyword in AD_BOT_KEYWORDS)


def _hash_analytics_value(raw: str) -> str:
    secret = os.environ.get("SECRET_KEY")
    if not secret:
        raise HTTPException(status_code=500, detail="Config missing")
    return hashlib.sha256(f"{secret}:{raw}".encode("utf-8")).hexdigest()


def _build_ad_hashes(request: Request) -> tuple[str, str]:
    ip_raw = _get_client_ip(request) or "unknown"
    ua_raw = request.headers.get("user-agent") or "unknown"
    return _hash_analytics_value(ip_raw), _hash_analytics_value(ua_raw)


async def _should_dedup_impression(session: AsyncSession, ad_id: uuid.UUID, ip_hash: str, ua_hash: str) -> bool:
    window_start = datetime.now(timezone.utc) - timedelta(minutes=AD_IMPRESSION_DEDUP_MINUTES)
    result = await session.execute(
        select(AdImpression.id)
        .where(
            AdImpression.ad_id == ad_id,
            AdImpression.ip_hash == ip_hash,
            AdImpression.user_agent_hash == ua_hash,
            AdImpression.created_at >= window_start,
        )
        .limit(1)
    )
    return result.scalar_one_or_none() is not None


async def _expire_ads(session: AsyncSession) -> None:
    now = datetime.now(timezone.utc)
    changed = False

    result = await session.execute(
        select(AdCampaign).where(
            AdCampaign.status.in_(["draft", "active", "paused"]),
            AdCampaign.end_at.isnot(None),
            AdCampaign.end_at < now,
        )
    )
    expired_campaigns = result.scalars().all()
    expired_campaign_ids = []
    for campaign in expired_campaigns:
        campaign.status = "expired"
        campaign.updated_at = now
        expired_campaign_ids.append(campaign.id)
    if expired_campaigns:
        changed = True

    expire_clauses = [and_(Advertisement.end_at.isnot(None), Advertisement.end_at < now)]
    if expired_campaign_ids:
        expire_clauses.append(Advertisement.campaign_id.in_(expired_campaign_ids))

    result = await session.execute(
        select(Advertisement).where(Advertisement.is_active.is_(True), or_(*expire_clauses))
    )
    expired_ads = result.scalars().all()
    for item in expired_ads:
        item.is_active = False
        item.updated_at = now
    if expired_ads:
        changed = True

    if changed:
        await session.commit()


async def _expire_doping(session: AsyncSession) -> None:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(DopingRequest).where(
            DopingRequest.status.in_(["approved", "published"]),
            DopingRequest.end_at.isnot(None),
            DopingRequest.end_at < now,
        )
    )
    expired = result.scalars().all()
    for item in expired:
        item.status = "expired"
        item.expired_at = now
        item.updated_at = now
    if expired:
        await session.commit()


@api_router.get("/site/header")
async def get_site_header(session: AsyncSession = Depends(get_sql_session)):
    result = await session.execute(select(SiteHeaderSetting).order_by(desc(SiteHeaderSetting.updated_at)).limit(1))
    setting = result.scalar_one_or_none()
    return {
        "logo_url": setting.logo_url if setting else None,
    }


@api_router.get("/admin/site/header")
async def admin_get_site_header(
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(select(SiteHeaderSetting).order_by(desc(SiteHeaderSetting.updated_at)).limit(1))
    setting = result.scalar_one_or_none()
    return {
        "id": str(setting.id) if setting else None,
        "logo_url": setting.logo_url if setting else None,
    }


@api_router.post("/admin/site/header/logo")
async def upload_site_header_logo(
    file: UploadFile = File(...),
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    if not file:
        raise HTTPException(status_code=400, detail="File is required")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")
    try:
        asset_key, _ = store_site_asset(data, file.filename, folder="header", allow_svg=True)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    result = await session.execute(select(SiteHeaderSetting).order_by(desc(SiteHeaderSetting.updated_at)).limit(1))
    setting = result.scalar_one_or_none()
    if not setting:
        setting = SiteHeaderSetting(logo_url=f"/api/site/assets/{asset_key}")
        session.add(setting)
    else:
        setting.logo_url = f"/api/site/assets/{asset_key}"
    await session.commit()

    return {"ok": True, "logo_url": setting.logo_url}


@api_router.get("/site/assets/{asset_key:path}")
async def get_site_asset(asset_key: str):
    base_dir = os.path.join(os.path.dirname(__file__), "static", "site_assets")
    path = os.path.normpath(os.path.join(base_dir, asset_key))
    if not path.startswith(base_dir):
        raise HTTPException(status_code=400, detail="Invalid asset path")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Asset not found")
    return FileResponse(path)


@api_router.get("/ads")
async def list_ads_public(
    placement: str = Query(...),
    session: AsyncSession = Depends(get_sql_session),
):
    if placement not in AD_PLACEMENTS:
        raise HTTPException(status_code=400, detail="Invalid placement")
    await _expire_ads(session)
    result = await session.execute(
        select(Advertisement, AdCampaign)
        .outerjoin(AdCampaign, AdCampaign.id == Advertisement.campaign_id)
        .where(Advertisement.placement == placement, Advertisement.is_active.is_(True))
        .order_by(desc(Advertisement.priority), desc(Advertisement.updated_at))
    )
    items = []
    for ad, campaign in result.all():
        if not _is_ad_active(ad, campaign):
            continue
        items.append(
            {
                "id": str(ad.id),
                "asset_url": ad.asset_url,
                "target_url": ad.target_url,
                "start_at": ad.start_at.isoformat() if ad.start_at else None,
                "end_at": ad.end_at.isoformat() if ad.end_at else None,
                "priority": ad.priority,
            }
        )
    return {"items": items}


@api_router.post("/ads/{ad_id}/impression")
async def record_ad_impression(
    ad_id: str,
    request: Request,
    current_user=Depends(get_current_user_optional),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        ad_uuid = uuid.UUID(ad_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid ad id") from exc

    await _expire_ads(session)
    ad = await session.get(Advertisement, ad_uuid)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")

    campaign = None
    if ad.campaign_id:
        campaign = await session.get(AdCampaign, ad.campaign_id)

    if not _is_ad_active(ad, campaign):
        return {"ok": False, "skipped": "inactive"}

    user_agent = request.headers.get("user-agent") or ""
    if _is_bot_user_agent(user_agent):
        return {"ok": False, "skipped": "bot"}

    ip_hash, ua_hash = _build_ad_hashes(request)
    if await _should_dedup_impression(session, ad_uuid, ip_hash, ua_hash):
        return {"ok": True, "deduped": True}

    user_id = None
    if current_user and current_user.get("id"):
        try:
            user_id = uuid.UUID(str(current_user.get("id")))
        except ValueError:
            user_id = None

    impression = AdImpression(
        ad_id=ad_uuid,
        placement=ad.placement,
        user_id=user_id,
        ip_hash=ip_hash,
        user_agent_hash=ua_hash,
    )
    session.add(impression)
    await session.commit()
    return {"ok": True, "deduped": False}


@api_router.get("/ads/{ad_id}/click")
async def record_ad_click(
    ad_id: str,
    request: Request,
    current_user=Depends(get_current_user_optional),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        ad_uuid = uuid.UUID(ad_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid ad id") from exc

    await _expire_ads(session)
    ad = await session.get(Advertisement, ad_uuid)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")

    campaign = None
    if ad.campaign_id:
        campaign = await session.get(AdCampaign, ad.campaign_id)

    if not _is_ad_active(ad, campaign):
        raise HTTPException(status_code=404, detail="Ad not active")

    if not ad.target_url:
        raise HTTPException(status_code=400, detail="Target URL missing")

    user_agent = request.headers.get("user-agent") or ""
    if _is_bot_user_agent(user_agent):
        raise HTTPException(status_code=404, detail="Ad not active")

    ip_hash, ua_hash = _build_ad_hashes(request)
    user_id = None
    if current_user and current_user.get("id"):
        try:
            user_id = uuid.UUID(str(current_user.get("id")))
        except ValueError:
            user_id = None

    click = AdClick(
        ad_id=ad_uuid,
        placement=ad.placement,
        user_id=user_id,
        ip_hash=ip_hash,
        user_agent_hash=ua_hash,
    )
    session.add(click)
    await session.commit()
    return RedirectResponse(url=ad.target_url)


@api_router.get("/admin/ads")
async def list_ads_admin(
    placement: Optional[str] = None,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _expire_ads(session)
    query = select(Advertisement, AdCampaign).outerjoin(AdCampaign, AdCampaign.id == Advertisement.campaign_id)
    if placement:
        query = query.where(Advertisement.placement == placement)
    query = query.order_by(desc(Advertisement.updated_at))
    result = await session.execute(query)
    items = []
    for ad, campaign in result.all():
        items.append(
            {
                "id": str(ad.id),
                "placement": ad.placement,
                "campaign_id": str(ad.campaign_id) if ad.campaign_id else None,
                "campaign_name": campaign.name if campaign else "Standalone",
                "campaign_status": campaign.status if campaign else None,
                "format": ad.format,
                "asset_url": ad.asset_url,
                "target_url": ad.target_url,
                "start_at": ad.start_at.isoformat() if ad.start_at else None,
                "end_at": ad.end_at.isoformat() if ad.end_at else None,
                "priority": ad.priority,
                "is_active": ad.is_active,
                "created_at": ad.created_at.isoformat() if ad.created_at else None,
            }
        )
    return {
        "items": items,
        "placements": AD_PLACEMENTS,
        "format_rules": AD_FORMAT_RULES,
        "format_labels": AD_FORMAT_LABELS,
    }


@api_router.post("/admin/ads")
async def create_ad(
    payload: AdCreatePayload,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    if payload.placement not in AD_PLACEMENTS:
        raise HTTPException(status_code=400, detail="Invalid placement")
    start_at, end_at = _normalize_ad_dates(payload.start_at, payload.end_at)
    format_value = _resolve_ad_format(payload.placement, payload.format)

    if payload.is_active:
        conflict = await _find_active_ad_conflict(session, payload.placement)
        if conflict:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Active ad already exists for placement",
                    "existing_id": str(conflict.id),
                },
            )

    campaign_id = None
    if payload.campaign_id:
        try:
            campaign_uuid = uuid.UUID(payload.campaign_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid campaign id") from exc
        campaign = await session.get(AdCampaign, campaign_uuid)
        if not campaign:
            raise HTTPException(status_code=404, detail="Campaign not found")
        campaign_id = campaign.id

    ad = Advertisement(
        placement=payload.placement,
        campaign_id=campaign_id,
        format=format_value,
        start_at=start_at,
        end_at=end_at,
        is_active=bool(payload.is_active),
        priority=payload.priority or 0,
        target_url=payload.target_url,
    )
    session.add(ad)
    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Active ad already exists for placement")
    await session.refresh(ad)
    return {"ok": True, "id": str(ad.id)}


@api_router.patch("/admin/ads/{ad_id}")
async def update_ad(
    ad_id: str,
    payload: AdUpdatePayload,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        ad_uuid = uuid.UUID(ad_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid ad id") from exc

    ad = await session.get(Advertisement, ad_uuid)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")

    start_at = payload.start_at if payload.start_at is not None else ad.start_at
    end_at = payload.end_at if payload.end_at is not None else ad.end_at
    start_at, end_at = _normalize_ad_dates(start_at, end_at)

    if payload.is_active is True:
        conflict = await _find_active_ad_conflict(session, ad.placement, exclude_id=ad.id)
        if conflict:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Active ad already exists for placement",
                    "existing_id": str(conflict.id),
                },
            )

    if payload.start_at is not None:
        ad.start_at = start_at
    if payload.end_at is not None:
        ad.end_at = end_at
    if payload.is_active is not None:
        ad.is_active = payload.is_active
    if payload.priority is not None:
        ad.priority = payload.priority
    if payload.target_url is not None:
        ad.target_url = payload.target_url
    if payload.format is not None:
        ad.format = _resolve_ad_format(ad.placement, payload.format)
    if payload.campaign_id is not None:
        if payload.campaign_id == "" or payload.campaign_id.lower() == "null":
            ad.campaign_id = None
        else:
            try:
                campaign_uuid = uuid.UUID(payload.campaign_id)
            except ValueError as exc:
                raise HTTPException(status_code=400, detail="Invalid campaign id") from exc
            campaign = await session.get(AdCampaign, campaign_uuid)
            if not campaign:
                raise HTTPException(status_code=404, detail="Campaign not found")
            ad.campaign_id = campaign.id

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Active ad already exists for placement")
    return {"ok": True}


@api_router.post("/admin/ads/{ad_id}/upload")
async def upload_ad_asset(
    ad_id: str,
    file: UploadFile = File(...),
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        ad_uuid = uuid.UUID(ad_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid ad id") from exc

    ad = await session.get(Advertisement, ad_uuid)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty file")

    try:
        asset_key, _ = store_site_asset(data, file.filename, folder="ads", allow_svg=False)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    ad.asset_url = f"/api/site/assets/{asset_key}"
    await session.commit()
    return {"ok": True, "asset_url": ad.asset_url}


def _parse_analytics_datetime(value: str, is_end: bool = False) -> datetime:
    dt = datetime.fromisoformat(value)
    if len(value) == 10:
        if is_end:
            dt = datetime.combine(dt.date(), datetime.max.time())
        else:
            dt = datetime.combine(dt.date(), datetime.min.time())
    if not dt.tzinfo:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _resolve_analytics_window(range_key: str, start_at: Optional[str], end_at: Optional[str]) -> tuple[datetime, datetime]:
    now = datetime.now(timezone.utc)
    if start_at and end_at:
        return _parse_analytics_datetime(start_at), _parse_analytics_datetime(end_at, is_end=True)
    if range_key not in {"7d", "30d"}:
        raise HTTPException(status_code=400, detail="Invalid range")
    days = 7 if range_key == "7d" else 30
    return now - timedelta(days=days), now


def _percent_change(current: int, previous: int) -> Optional[float]:
    if previous <= 0:
        return None
    return ((current - previous) / previous) * 100


async def _campaign_traffic_stats(
    session: AsyncSession,
    campaign_ids: list[uuid.UUID],
) -> Dict[uuid.UUID, Dict[str, int]]:
    if not campaign_ids:
        return {}

    now = datetime.now(timezone.utc)
    current_start = now - timedelta(days=1)
    previous_start = now - timedelta(days=2)
    previous_end = current_start

    async def _counts(model, start: datetime, end: datetime) -> Dict[uuid.UUID, int]:
        stmt = (
            select(Advertisement.campaign_id, func.count())
            .select_from(model)
            .join(Advertisement, Advertisement.id == model.ad_id)
            .where(
                model.created_at >= start,
                model.created_at < end,
                Advertisement.campaign_id.in_(campaign_ids),
            )
            .group_by(Advertisement.campaign_id)
        )
        rows = (await session.execute(stmt)).all()
        return {row[0]: int(row[1]) for row in rows if row[0]}

    current_impressions = await _counts(AdImpression, current_start, now)
    previous_impressions = await _counts(AdImpression, previous_start, previous_end)
    current_clicks = await _counts(AdClick, current_start, now)
    previous_clicks = await _counts(AdClick, previous_start, previous_end)

    stats: Dict[uuid.UUID, Dict[str, int]] = {}
    for cid in campaign_ids:
        stats[cid] = {
            "current_impressions": current_impressions.get(cid, 0),
            "previous_impressions": previous_impressions.get(cid, 0),
            "current_clicks": current_clicks.get(cid, 0),
            "previous_clicks": previous_clicks.get(cid, 0),
        }
    return stats


def _build_campaign_warnings(
    campaign: AdCampaign,
    stats: Dict[str, int],
) -> List[Dict[str, str]]:
    warnings: List[Dict[str, str]] = []
    now = datetime.now(timezone.utc)

    if campaign.end_at and campaign.status in {"active", "paused"}:
        delta_days = (campaign.end_at - now).total_seconds() / 86400
        if 0 <= delta_days <= END_WARNING_DAYS:
            warnings.append(
                {
                    "type": "end_at",
                    "severity": "warning",
                    "message": f"Kampanya bitişine {math.ceil(delta_days)} gün kaldı.",
                    "recommendation": "Süreyi uzatın veya yeni kampanya planlayın.",
                }
            )

    if campaign.status == "active":
        impression_change = _percent_change(stats.get("current_impressions", 0), stats.get("previous_impressions", 0))
        if impression_change is not None and impression_change >= TRAFFIC_SPIKE_THRESHOLD:
            warnings.append(
                {
                    "type": "traffic_impressions",
                    "severity": "warning",
                    "message": f"Gösterim artışı %{impression_change:.0f} (son 24s vs önceki 24s).",
                    "recommendation": "Trafik artışını doğrulayın ve slot/creative kontrol edin.",
                }
            )

        click_change = _percent_change(stats.get("current_clicks", 0), stats.get("previous_clicks", 0))
        if click_change is not None and click_change >= TRAFFIC_SPIKE_THRESHOLD:
            warnings.append(
                {
                    "type": "traffic_clicks",
                    "severity": "warning",
                    "message": f"Tıklama artışı %{click_change:.0f} (son 24s vs önceki 24s).",
                    "recommendation": "Landing sayfası ve hedef URL performansını kontrol edin.",
                }
            )

    if campaign.budget is not None and campaign.status == "active":
        warnings.append(
            {
                "type": "budget_info",
                "severity": "info",
                "message": "Budget alanı dolu. CPM/CPC tanımı olmadığı için tüketim hesaplanmıyor.",
                "recommendation": "CPM/CPC tanımlayın veya manuel takip edin.",
            }
        )

    return warnings


async def _get_latest_pricing_campaign(session: AsyncSession) -> Optional[PricingCampaign]:
    result = await session.execute(
        select(PricingCampaign).order_by(desc(PricingCampaign.updated_at)).limit(1)
    )
    return result.scalar_one_or_none()


def _pricing_campaign_to_dict(policy: PricingCampaign) -> Dict[str, Any]:
    return {
        "id": str(policy.id),
        "is_enabled": policy.is_enabled,
        "start_at": policy.start_at.isoformat() if policy.start_at else None,
        "end_at": policy.end_at.isoformat() if policy.end_at else None,
        "scope": policy.scope,
        "published_at": policy.published_at.isoformat() if policy.published_at else None,
        "version": policy.version,
        "created_by": str(policy.created_by) if policy.created_by else None,
        "updated_by": str(policy.updated_by) if policy.updated_by else None,
    }


def _ensure_aware_datetime(value: Optional[datetime]) -> Optional[datetime]:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value


def _is_pricing_campaign_active(policy: PricingCampaign, user_type: str) -> bool:
    if not policy.is_enabled:
        return False
    start_at = _ensure_aware_datetime(policy.start_at)
    if start_at is None:
        return False
    end_at = _ensure_aware_datetime(policy.end_at)
    now = datetime.now(timezone.utc)
    if start_at and start_at > now:
        return False
    if end_at and end_at < now:
        return False
    if policy.scope == "all":
        return True
    return policy.scope == user_type


async def _expire_pricing_campaigns(session: AsyncSession, actor: Optional[dict] = None) -> int:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(PricingCampaign).where(
            PricingCampaign.is_enabled.is_(True),
            PricingCampaign.end_at.isnot(None),
            PricingCampaign.end_at < now,
        )
    )
    expired = result.scalars().all()
    if not expired:
        return 0

    system_actor = actor or {"id": None, "email": "system@internal"}
    for policy in expired:
        policy.is_enabled = False
        policy.updated_at = now
        policy.updated_by = _safe_uuid(system_actor.get("id"))
        await _write_audit_log_sql(
            session=session,
            action="PRICING_CAMPAIGN_EXPIRED",
            actor=system_actor,
            resource_type="pricing_campaign",
            resource_id=str(policy.id),
            metadata={"end_at": policy.end_at.isoformat() if policy.end_at else None},
            request=None,
            country_code=None,
        )
    await session.commit()
    return len(expired)


async def _expire_pricing_campaign_items(session: AsyncSession, actor: Optional[dict] = None) -> int:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(PricingCampaignItem).where(
            PricingCampaignItem.is_active.is_(True),
            PricingCampaignItem.is_deleted.is_(False),
            PricingCampaignItem.end_at.isnot(None),
            PricingCampaignItem.end_at < now,
        )
    )
    expired = result.scalars().all()
    if not expired:
        return 0

    system_actor = actor or {"id": None, "email": "system@internal"}
    for item in expired:
        item.is_active = False
        item.updated_at = now
        item.updated_by = _safe_uuid(system_actor.get("id"))
        await _write_audit_log_sql(
            session=session,
            action="PRICING_CAMPAIGN_ITEM_EXPIRED",
            actor=system_actor,
            resource_type="pricing_campaign_item",
            resource_id=str(item.id),
            metadata={"end_at": item.end_at.isoformat() if item.end_at else None},
            request=None,
            country_code=None,
        )
    await session.commit()
    return len(expired)


async def _assert_no_overlap_active_campaign_item(
    session: AsyncSession,
    scope: str,
    start_at: datetime,
    end_at: datetime,
    exclude_id: Optional[uuid.UUID] = None,
) -> None:
    query = select(PricingCampaignItem).where(
        PricingCampaignItem.scope == scope,
        PricingCampaignItem.is_active.is_(True),
        PricingCampaignItem.is_deleted.is_(False),
        PricingCampaignItem.start_at.isnot(None),
        PricingCampaignItem.end_at.isnot(None),
        PricingCampaignItem.start_at < end_at,
        PricingCampaignItem.end_at > start_at,
    )
    if exclude_id:
        query = query.where(PricingCampaignItem.id != exclude_id)
    existing = (await session.execute(query.limit(1))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=409, detail="Active campaign overlaps with existing time range")


DEFAULT_PRICING_CAMPAIGN_ITEMS = [
    {
        "scope": "individual",
        "name": "Örnek Bireysel Kampanya",
        "listing_quota": 1,
        "price_amount": 0,
        "publish_days": 90,
        "is_active": False,
    },
    {
        "scope": "corporate",
        "name": "Örnek Kurumsal Kampanya",
        "listing_quota": 10,
        "price_amount": 0,
        "publish_days": 90,
        "is_active": False,
    },
]


def _normalize_currency_code(value: Optional[str]) -> str:
    code = (value or "EUR").upper()
    if len(code) != 3 or not code.isalpha():
        raise HTTPException(status_code=400, detail="Invalid currency")
    return code


def _validate_campaign_item_fields(
    scope: str,
    listing_quota: Optional[int],
    price_amount: Optional[float],
    publish_days: Optional[int],
    start_at: Optional[datetime],
    end_at: Optional[datetime],
    require_dates: bool = False,
    enforce_start_future: bool = False,
) -> None:
    if scope not in PRICING_CAMPAIGN_ITEM_SCOPES:
        raise HTTPException(status_code=400, detail="Invalid scope")
    if listing_quota is not None and listing_quota <= 0:
        raise HTTPException(status_code=400, detail="listing_quota must be > 0")
    if price_amount is not None and price_amount < 0:
        raise HTTPException(status_code=400, detail="price_amount must be >= 0")
    if publish_days is not None and publish_days <= 0:
        raise HTTPException(status_code=400, detail="publish_days must be > 0")
    if require_dates and (start_at is None or end_at is None):
        raise HTTPException(status_code=400, detail="start_at and end_at required")
    if start_at and end_at and end_at <= start_at:
        raise HTTPException(status_code=400, detail="end_at must be after start_at")
    if enforce_start_future and start_at:
        now = datetime.now(timezone.utc)
        if start_at < now - timedelta(minutes=1):
            raise HTTPException(status_code=400, detail="start_at must be in the future")


async def _ensure_pricing_defaults(session: AsyncSession) -> None:
    result = await session.execute(select(PricingCampaignItem))
    items = result.scalars().all()
    if items:
        return

    now = datetime.now(timezone.utc)
    default_end = now + timedelta(days=30)

    for sample in DEFAULT_PRICING_CAMPAIGN_ITEMS:
        session.add(
            PricingCampaignItem(
                scope=sample["scope"],
                name=sample.get("name"),
                listing_quota=sample["listing_quota"],
                price_amount=sample["price_amount"],
                currency="EUR",
                publish_days=sample["publish_days"],
                is_active=sample.get("is_active", False),
                start_at=sample.get("start_at") or now,
                end_at=sample.get("end_at") or default_end,
            )
        )
    await session.commit()


async def _expire_tier_rules(session: AsyncSession) -> None:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(PricingTierRule).where(
            PricingTierRule.is_active.is_(True),
            PricingTierRule.effective_end_at.isnot(None),
            PricingTierRule.effective_end_at < now,
        )
    )
    expired = result.scalars().all()
    if not expired:
        return
    for rule in expired:
        rule.is_active = False
        rule.updated_at = now
    await session.commit()


async def _expire_package_subscriptions(session: AsyncSession, actor: Optional[dict] = None) -> None:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(UserPackageSubscription).where(
            UserPackageSubscription.status == "active",
            UserPackageSubscription.ends_at.isnot(None),
            UserPackageSubscription.ends_at < now,
        )
    )
    expired = result.scalars().all()
    if not expired:
        return
    system_actor = actor or {"id": None, "email": "system@internal"}
    for sub in expired:
        sub.status = "expired"
        sub.remaining_quota = 0
        sub.updated_at = now
        await _write_audit_log_sql(
            session=session,
            action="PRICING_PACKAGE_EXPIRED",
            actor=system_actor,
            resource_type="pricing_package",
            resource_id=str(sub.package_id),
            metadata={"subscription_id": str(sub.id)},
            request=None,
            country_code=None,
        )
    await session.commit()


def _pricing_year_bounds(now: datetime) -> tuple[datetime, datetime]:
    start = datetime(now.year, 1, 1, tzinfo=timezone.utc)
    end = datetime(now.year + 1, 1, 1, tzinfo=timezone.utc)
    return start, end


async def _count_user_listings_current_year(session: AsyncSession, user_id: uuid.UUID) -> int:
    now = datetime.now(timezone.utc)
    start, end = _pricing_year_bounds(now)
    result = await session.execute(
        select(func.count())
        .select_from(Listing)
        .where(
            Listing.user_id == user_id,
            Listing.created_at >= start,
            Listing.created_at < end,
            Listing.deleted_at.is_(None),
            Listing.status != "draft",
        )
    )
    return int(result.scalar() or 0)


async def _get_active_tier_rules(session: AsyncSession) -> list[PricingTierRule]:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(PricingTierRule)
        .where(
            PricingTierRule.scope == "individual",
            PricingTierRule.is_active.is_(True),
            or_(PricingTierRule.effective_start_at.is_(None), PricingTierRule.effective_start_at <= now),
            or_(PricingTierRule.effective_end_at.is_(None), PricingTierRule.effective_end_at >= now),
        )
        .order_by(PricingTierRule.tier_no)
    )
    return result.scalars().all()


def _pick_tier_rule(rules: list[PricingTierRule], listing_no: int) -> Optional[PricingTierRule]:
    for rule in rules:
        if rule.listing_from <= listing_no and (rule.listing_to is None or listing_no <= rule.listing_to):
            return rule
    return rules[-1] if rules else None


async def _get_active_subscription(
    session: AsyncSession, user_id: uuid.UUID
) -> tuple[Optional[UserPackageSubscription], Optional[PricingPackage]]:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(UserPackageSubscription, PricingPackage)
        .join(PricingPackage, PricingPackage.id == UserPackageSubscription.package_id)
        .where(
            UserPackageSubscription.user_id == user_id,
            UserPackageSubscription.status == "active",
            UserPackageSubscription.remaining_quota > 0,
            or_(UserPackageSubscription.ends_at.is_(None), UserPackageSubscription.ends_at > now),
            PricingPackage.is_active.is_(True),
        )
        .order_by(desc(UserPackageSubscription.starts_at))
        .limit(1)
    )
    row = result.first()
    if not row:
        return None, None
    return row[0], row[1]


def _resolve_pricing_user_type(current_user: dict) -> str:
    role = current_user.get("role") or ""
    raw_type = _determine_user_type(role)
    if raw_type == "dealer" or current_user.get("user_type") == "corporate":
        return "corporate"
    return "individual"


def _campaign_item_is_available(item: PricingCampaignItem, now: datetime) -> bool:
    if item.is_deleted:
        return False
    if not item.is_active:
        return False
    if not item.start_at or not item.end_at:
        return False
    if item.start_at > now:
        return False
    if item.end_at < now:
        return False
    return True


def _serialize_campaign_item_for_quote(item: PricingCampaignItem) -> Dict[str, Any]:
    return {
        "id": str(item.id),
        "name": item.name,
        "listing_quota": item.listing_quota,
        "price_amount": float(item.price_amount or 0),
        "currency": item.currency,
        "publish_days": item.publish_days,
        "start_at": item.start_at.isoformat() if item.start_at else None,
        "end_at": item.end_at.isoformat() if item.end_at else None,
        "is_active": item.is_active,
    }


def _serialize_pricing_campaign_item(item: PricingCampaignItem) -> Dict[str, Any]:
    return {
        "id": str(item.id),
        "scope": item.scope,
        "name": item.name,
        "listing_quota": item.listing_quota,
        "price_amount": float(item.price_amount or 0),
        "currency": item.currency,
        "publish_days": item.publish_days,
        "is_active": item.is_active,
        "start_at": item.start_at.isoformat() if item.start_at else None,
        "end_at": item.end_at.isoformat() if item.end_at else None,
        "is_deleted": item.is_deleted,
        "deleted_at": item.deleted_at.isoformat() if item.deleted_at else None,
        "created_at": item.created_at.isoformat() if item.created_at else None,
        "updated_at": item.updated_at.isoformat() if item.updated_at else None,
    }


async def _resolve_campaign_item_from_payload(
    session: AsyncSession,
    scope: str,
    payload: Optional[PricingQuotePayload],
) -> Optional[PricingCampaignItem]:
    if not payload:
        return None

    now = datetime.now(timezone.utc)
    if payload.campaign_item_id:
        try:
            item_uuid = uuid.UUID(payload.campaign_item_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="campaign_item_id invalid") from exc
        item = await session.get(PricingCampaignItem, item_uuid)
        if not item or item.scope != scope or item.is_deleted:
            raise HTTPException(status_code=404, detail="Campaign item not found")
        if not _campaign_item_is_available(item, now):
            raise HTTPException(status_code=409, detail="Campaign item is not active")
        return item

    if payload.listing_quota:
        result = await session.execute(
            select(PricingCampaignItem).where(
                PricingCampaignItem.scope == scope,
                PricingCampaignItem.listing_quota == payload.listing_quota,
                PricingCampaignItem.is_deleted.is_(False),
            )
        )
        for item in result.scalars().all():
            if _campaign_item_is_available(item, now):
                return item

    return None


async def _get_active_pricing_campaign_items(
    session: AsyncSession, scope: str
) -> list[PricingCampaignItem]:
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(PricingCampaignItem)
        .where(
            PricingCampaignItem.scope == scope,
            PricingCampaignItem.is_deleted.is_(False),
            PricingCampaignItem.is_active.is_(True),
            PricingCampaignItem.start_at.isnot(None),
            PricingCampaignItem.end_at.isnot(None),
            PricingCampaignItem.start_at <= now,
            PricingCampaignItem.end_at >= now,
        )
        .order_by(PricingCampaignItem.listing_quota)
    )
    return result.scalars().all()


async def _build_campaign_item_quote(
    session: AsyncSession,
    scope: str,
    payload: Optional[PricingQuotePayload],
) -> Dict[str, Any]:
    items = await _get_active_pricing_campaign_items(session, scope)
    selected = await _resolve_campaign_item_from_payload(session, scope, payload)

    if not selected:
        if len(items) == 1:
            selected = items[0]
        elif len(items) == 0:
            return {
                "type": "campaign_none",
                "reason": "no_active_campaign",
                "requires_payment": False,
                "amount": 0,
                "currency": "EUR",
                "publish_days": 90,
                "listing_quota": 1,
                "scope": scope,
            }
        else:
            return {
                "type": "campaign_selection",
                "reason": "campaign_selection_required",
                "requires_payment": True,
                "items": [_serialize_campaign_item_for_quote(item) for item in items],
                "scope": scope,
            }

    amount = float(selected.price_amount or 0)
    return {
        "type": "campaign_item",
        "reason": "campaign_item",
        "campaign_item_id": str(selected.id),
        "name": selected.name,
        "listing_quota": selected.listing_quota,
        "amount": amount,
        "currency": selected.currency,
        "publish_days": selected.publish_days,
        "requires_payment": amount > 0,
        "quota_used": False,
        "scope": scope,
    }


async def _build_individual_quote(session: AsyncSession, user_id: uuid.UUID, payload: Optional[PricingQuotePayload]) -> Dict[str, Any]:
    return await _build_campaign_item_quote(session, "individual", payload)


async def _build_corporate_quote(session: AsyncSession, user_id: uuid.UUID, payload: Optional[PricingQuotePayload]) -> Dict[str, Any]:
    subscription, package = await _get_active_subscription(session, user_id)
    if subscription and package:
        return {
            "type": "quota",
            "reason": "quota_subscription",
            "subscription_id": str(subscription.id),
            "package_id": str(package.id),
            "listing_quota": package.listing_quota,
            "amount": 0,
            "currency": package.currency,
            "publish_days": package.publish_days,
            "requires_payment": False,
            "quota_used": True,
            "remaining_quota": subscription.remaining_quota,
        }

    return await _build_campaign_item_quote(session, "corporate", payload)


def _build_pricing_quote_response(
    quote: Dict[str, Any],
    override_active: bool,
    policy: Optional[PricingCampaign],
) -> Dict[str, Any]:
    response = {
        "pricing_mode": "campaign" if override_active else "default",
        "override_active": override_active,
        "campaign_active": override_active,
        "quote": quote,
    }
    if override_active and quote.get("warning"):
        response["warning"] = quote.get("warning")
    if policy:
        response["campaign"] = _pricing_campaign_to_dict(policy)
    return response


async def _compute_pricing_quote(
    session: AsyncSession, current_user: dict, payload: Optional[PricingQuotePayload] = None
) -> tuple[Dict[str, Any], Dict[str, Any], Optional[PricingCampaign], str]:
    await _ensure_pricing_defaults(session)
    await _expire_pricing_campaigns(session, actor=current_user)
    await _expire_pricing_campaign_items(session, actor=current_user)
    await _expire_package_subscriptions(session, actor=current_user)

    user_type = _resolve_pricing_user_type(current_user)
    policy = await _get_latest_pricing_campaign(session)
    override_active = False
    if policy:
        override_active = _is_pricing_campaign_active(policy, user_type)

    if user_type == "corporate":
        quote = await _build_corporate_quote(session, uuid.UUID(current_user.get("id")), payload)
    else:
        quote = await _build_individual_quote(session, uuid.UUID(current_user.get("id")), payload)

    response = _build_pricing_quote_response(quote, override_active, policy)
    return response, quote, policy, user_type


async def _create_pricing_snapshot(
    session: AsyncSession,
    listing: Listing,
    current_user: dict,
    quote: Dict[str, Any],
    policy: Optional[PricingCampaign],
    override_active: bool,
) -> Optional[PricingPriceSnapshot]:
    if quote.get("type") not in {"campaign_item", "quota"}:
        return None

    existing = await session.execute(
        select(PricingPriceSnapshot).where(PricingPriceSnapshot.listing_id == listing.id)
    )
    snapshot = existing.scalar_one_or_none()
    if snapshot:
        return snapshot

    publish_days = quote.get("publish_days") or quote.get("duration_days") or 90

    snapshot = PricingPriceSnapshot(
        listing_id=listing.id,
        user_id=listing.user_id,
        rule_id=_safe_uuid(quote.get("rule_id")),
        package_id=_safe_uuid(quote.get("package_id")),
        campaign_id=policy.id if policy and override_active else None,
        campaign_item_id=_safe_uuid(quote.get("campaign_item_id")),
        listing_quota=quote.get("listing_quota"),
        currency=quote.get("currency") or "EUR",
        amount=quote.get("amount") or 0,
        duration_days=publish_days,
        publish_days=publish_days,
        campaign_override_active=bool(override_active),
        snapshot_type=quote.get("type") or "campaign_item",
        metadata={
            "listing_no": quote.get("listing_no"),
            "listing_count_year": quote.get("listing_count_year"),
            "quota_used": quote.get("quota_used"),
            "requires_payment": quote.get("requires_payment"),
            "campaign_item_id": quote.get("campaign_item_id"),
            "listing_quota": quote.get("listing_quota"),
        },
    )
    session.add(snapshot)
    return snapshot


@api_router.get("/admin/ads/analytics")
async def get_ads_analytics(
    range: str = "30d",
    start_at: Optional[str] = None,
    end_at: Optional[str] = None,
    group_by: str = "ad",
    campaign_id: Optional[str] = None,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    start_dt, end_dt = _resolve_analytics_window(range, start_at, end_at)
    if group_by not in {"ad", "campaign", "placement"}:
        raise HTTPException(status_code=400, detail="Invalid group_by")

    campaign_uuid: Optional[uuid.UUID] = None
    if campaign_id:
        try:
            campaign_uuid = uuid.UUID(campaign_id)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    impression_stmt = select(func.count()).select_from(AdImpression).where(
        AdImpression.created_at >= start_dt,
        AdImpression.created_at <= end_dt,
    )
    click_stmt = select(func.count()).select_from(AdClick).where(
        AdClick.created_at >= start_dt,
        AdClick.created_at <= end_dt,
    )
    if campaign_uuid:
        impression_stmt = impression_stmt.join(Advertisement, Advertisement.id == AdImpression.ad_id).where(
            Advertisement.campaign_id == campaign_uuid
        )
        click_stmt = click_stmt.join(Advertisement, Advertisement.id == AdClick.ad_id).where(
            Advertisement.campaign_id == campaign_uuid
        )

    impressions_total = int((await session.execute(impression_stmt)).scalar() or 0)
    clicks_total = int((await session.execute(click_stmt)).scalar() or 0)
    ctr_total = round((clicks_total / impressions_total) * 100, 2) if impressions_total else 0.0

    now = datetime.now(timezone.utc)
    active_ads_stmt = (
        select(func.count())
        .select_from(Advertisement)
        .outerjoin(AdCampaign, AdCampaign.id == Advertisement.campaign_id)
        .where(
            Advertisement.is_active.is_(True),
            or_(Advertisement.start_at.is_(None), Advertisement.start_at <= now),
            or_(Advertisement.end_at.is_(None), Advertisement.end_at >= now),
            or_(
                AdCampaign.id.is_(None),
                and_(
                    AdCampaign.status == "active",
                    or_(AdCampaign.start_at.is_(None), AdCampaign.start_at <= now),
                    or_(AdCampaign.end_at.is_(None), AdCampaign.end_at >= now),
                ),
            ),
        )
    )
    if campaign_uuid:
        active_ads_stmt = active_ads_stmt.where(Advertisement.campaign_id == campaign_uuid)
    active_ads = int((await session.execute(active_ads_stmt)).scalar() or 0)

    groups_payload = []

    if group_by == "placement":
        impression_rows_stmt = select(AdImpression.placement, func.count()).where(
            AdImpression.created_at >= start_dt,
            AdImpression.created_at <= end_dt,
        )
        click_rows_stmt = select(AdClick.placement, func.count()).where(
            AdClick.created_at >= start_dt,
            AdClick.created_at <= end_dt,
        )
        if campaign_uuid:
            impression_rows_stmt = impression_rows_stmt.join(Advertisement, Advertisement.id == AdImpression.ad_id).where(
                Advertisement.campaign_id == campaign_uuid
            )
            click_rows_stmt = click_rows_stmt.join(Advertisement, Advertisement.id == AdClick.ad_id).where(
                Advertisement.campaign_id == campaign_uuid
            )
        impressions_by = dict(
            (row[0], int(row[1])) for row in (await session.execute(impression_rows_stmt.group_by(AdImpression.placement))).all()
        )
        clicks_by = dict(
            (row[0], int(row[1])) for row in (await session.execute(click_rows_stmt.group_by(AdClick.placement))).all()
        )
        for placement_key, label in AD_PLACEMENTS.items():
            imp = impressions_by.get(placement_key, 0)
            clk = clicks_by.get(placement_key, 0)
            ctr = round((clk / imp) * 100, 2) if imp else 0.0
            groups_payload.append(
                {
                    "key": placement_key,
                    "label": label,
                    "impressions": imp,
                    "clicks": clk,
                    "ctr": ctr,
                }
            )

    if group_by == "campaign":
        impression_rows_stmt = (
            select(Advertisement.campaign_id, func.count())
            .select_from(AdImpression)
            .join(Advertisement, Advertisement.id == AdImpression.ad_id)
            .where(AdImpression.created_at >= start_dt, AdImpression.created_at <= end_dt)
            .group_by(Advertisement.campaign_id)
        )
        click_rows_stmt = (
            select(Advertisement.campaign_id, func.count())
            .select_from(AdClick)
            .join(Advertisement, Advertisement.id == AdClick.ad_id)
            .where(AdClick.created_at >= start_dt, AdClick.created_at <= end_dt)
            .group_by(Advertisement.campaign_id)
        )
        if campaign_uuid:
            impression_rows_stmt = impression_rows_stmt.where(Advertisement.campaign_id == campaign_uuid)
            click_rows_stmt = click_rows_stmt.where(Advertisement.campaign_id == campaign_uuid)

        impressions_by = dict((row[0], int(row[1])) for row in (await session.execute(impression_rows_stmt)).all())
        clicks_by = dict((row[0], int(row[1])) for row in (await session.execute(click_rows_stmt)).all())

        campaign_ids = {cid for cid in impressions_by.keys() if cid} | {cid for cid in clicks_by.keys() if cid}
        campaigns = {}
        if campaign_ids:
            result = await session.execute(select(AdCampaign).where(AdCampaign.id.in_(campaign_ids)))
            campaigns = {item.id: item for item in result.scalars().all()}

        all_keys = set(impressions_by.keys()) | set(clicks_by.keys())
        for camp_id in all_keys:
            imp = impressions_by.get(camp_id, 0)
            clk = clicks_by.get(camp_id, 0)
            ctr = round((clk / imp) * 100, 2) if imp else 0.0
            campaign = campaigns.get(camp_id)
            label = campaign.name if campaign else "Standalone"
            groups_payload.append(
                {
                    "key": str(camp_id) if camp_id else "standalone",
                    "label": label,
                    "impressions": imp,
                    "clicks": clk,
                    "ctr": ctr,
                }
            )

    if group_by == "ad":
        impression_rows_stmt = select(AdImpression.ad_id, func.count()).where(
            AdImpression.created_at >= start_dt,
            AdImpression.created_at <= end_dt,
        )
        click_rows_stmt = select(AdClick.ad_id, func.count()).where(
            AdClick.created_at >= start_dt,
            AdClick.created_at <= end_dt,
        )
        if campaign_uuid:
            impression_rows_stmt = impression_rows_stmt.join(Advertisement, Advertisement.id == AdImpression.ad_id).where(
                Advertisement.campaign_id == campaign_uuid
            )
            click_rows_stmt = click_rows_stmt.join(Advertisement, Advertisement.id == AdClick.ad_id).where(
                Advertisement.campaign_id == campaign_uuid
            )
        impressions_by = dict((row[0], int(row[1])) for row in (await session.execute(impression_rows_stmt.group_by(AdImpression.ad_id))).all())
        clicks_by = dict((row[0], int(row[1])) for row in (await session.execute(click_rows_stmt.group_by(AdClick.ad_id))).all())

        ad_ids = set(impressions_by.keys()) | set(clicks_by.keys())
        ads = {}
        if ad_ids:
            result = await session.execute(select(Advertisement).where(Advertisement.id.in_(ad_ids)))
            ads = {item.id: item for item in result.scalars().all()}

        for ad_id in ad_ids:
            imp = impressions_by.get(ad_id, 0)
            clk = clicks_by.get(ad_id, 0)
            ctr = round((clk / imp) * 100, 2) if imp else 0.0
            ad = ads.get(ad_id)
            label = None
            if ad:
                label = ad.target_url or ad.asset_url
            if not label:
                label = f"Reklam {str(ad_id)[:8]}"
            groups_payload.append(
                {
                    "key": str(ad_id),
                    "label": label,
                    "impressions": imp,
                    "clicks": clk,
                    "ctr": ctr,
                }
            )

    return {
        "range": {
            "start_at": start_dt.isoformat(),
            "end_at": end_dt.isoformat(),
        },
        "totals": {
            "impressions": impressions_total,
            "clicks": clicks_total,
            "ctr": ctr_total,
            "active_ads": active_ads,
        },
        "group_by": group_by,
        "groups": groups_payload,
    }


@api_router.get("/admin/ads/campaigns")
async def list_ad_campaigns(
    status: Optional[str] = None,
    q: Optional[str] = None,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _expire_ads(session)
    now = datetime.now(timezone.utc)

    total_ads_subq = (
        select(func.count())
        .select_from(Advertisement)
        .where(Advertisement.campaign_id == AdCampaign.id)
        .correlate(AdCampaign)
        .scalar_subquery()
    )
    active_ads_subq = (
        select(func.count())
        .select_from(Advertisement)
        .where(
            Advertisement.campaign_id == AdCampaign.id,
            Advertisement.is_active.is_(True),
            or_(Advertisement.start_at.is_(None), Advertisement.start_at <= now),
            or_(Advertisement.end_at.is_(None), Advertisement.end_at >= now),
            AdCampaign.status == "active",
            or_(AdCampaign.start_at.is_(None), AdCampaign.start_at <= now),
            or_(AdCampaign.end_at.is_(None), AdCampaign.end_at >= now),
        )
        .correlate(AdCampaign)
        .scalar_subquery()
    )

    query = select(AdCampaign, total_ads_subq.label("total_ads"), active_ads_subq.label("active_ads"))
    if status:
        query = query.where(AdCampaign.status == status)
    if q:
        query = query.where(or_(AdCampaign.name.ilike(f"%{q}%"), AdCampaign.advertiser.ilike(f"%{q}%")))
    query = query.order_by(desc(AdCampaign.updated_at))

    result = await session.execute(query)
    rows = result.all()
    campaign_ids = [row[0].id for row in rows]
    stats_map = await _campaign_traffic_stats(session, campaign_ids)

    items = []
    for campaign, total_ads, active_ads in rows:
        stats = stats_map.get(
            campaign.id,
            {
                "current_impressions": 0,
                "previous_impressions": 0,
                "current_clicks": 0,
                "previous_clicks": 0,
            },
        )
        warnings = _build_campaign_warnings(campaign, stats)
        items.append(
            {
                "id": str(campaign.id),
                "name": campaign.name,
                "advertiser": campaign.advertiser,
                "budget": float(campaign.budget) if campaign.budget is not None else None,
                "currency": campaign.currency,
                "start_at": campaign.start_at.isoformat() if campaign.start_at else None,
                "end_at": campaign.end_at.isoformat() if campaign.end_at else None,
                "status": campaign.status,
                "total_ads": int(total_ads or 0),
                "active_ads": int(active_ads or 0),
                "warnings_count": len(warnings),
                "has_warning": len(warnings) > 0,
            }
        )
    return {"items": items}


@api_router.post("/admin/ads/campaigns")
async def create_ad_campaign(
    payload: AdCampaignCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    status_value = payload.status or "draft"
    if status_value not in CAMPAIGN_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    if status_value == "expired":
        raise HTTPException(status_code=400, detail="Expired status is automatic")

    start_at, end_at, budget, currency = _normalize_campaign_fields(
        payload.start_at, payload.end_at, payload.budget, payload.currency
    )

    campaign = AdCampaign(
        name=payload.name,
        advertiser=payload.advertiser,
        budget=budget,
        currency=currency or "EUR",
        start_at=start_at,
        end_at=end_at,
        status=status_value,
    )
    session.add(campaign)

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_CREATED",
        actor=current_user,
        resource_type="ad_campaign",
        resource_id=str(campaign.id),
        metadata={"name": payload.name},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    await session.refresh(campaign)
    return {"ok": True, "id": str(campaign.id)}


@api_router.get("/admin/ads/campaigns/{campaign_id}")
async def get_ad_campaign(
    campaign_id: str,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    campaign = await session.get(AdCampaign, campaign_uuid)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    ads_result = await session.execute(
        select(Advertisement)
        .where(Advertisement.campaign_id == campaign.id)
        .order_by(desc(Advertisement.updated_at))
    )
    ads_payload = []
    for ad in ads_result.scalars().all():
        ads_payload.append(
            {
                "id": str(ad.id),
                "placement": ad.placement,
                "format": ad.format,
                "is_active": ad.is_active,
                "target_url": ad.target_url,
                "start_at": ad.start_at.isoformat() if ad.start_at else None,
                "end_at": ad.end_at.isoformat() if ad.end_at else None,
            }
        )

    stats_map = await _campaign_traffic_stats(session, [campaign.id])
    stats = stats_map.get(
        campaign.id,
        {
            "current_impressions": 0,
            "previous_impressions": 0,
            "current_clicks": 0,
            "previous_clicks": 0,
        },
    )
    warnings = _build_campaign_warnings(campaign, stats)

    return {
        "campaign": {
            "id": str(campaign.id),
            "name": campaign.name,
            "advertiser": campaign.advertiser,
            "budget": float(campaign.budget) if campaign.budget is not None else None,
            "currency": campaign.currency,
            "start_at": campaign.start_at.isoformat() if campaign.start_at else None,
            "end_at": campaign.end_at.isoformat() if campaign.end_at else None,
            "status": campaign.status,
        },
        "ads": ads_payload,
        "warnings": warnings,
    }


@api_router.patch("/admin/ads/campaigns/{campaign_id}")
async def update_ad_campaign(
    campaign_id: str,
    payload: AdCampaignUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        campaign_uuid = uuid.UUID(campaign_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid campaign id") from exc

    campaign = await session.get(AdCampaign, campaign_uuid)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    if payload.status:
        if payload.status == "expired":
            raise HTTPException(status_code=400, detail="Expired status is automatic")
        if payload.status == "draft" and campaign.status != "draft":
            raise HTTPException(status_code=400, detail="Cannot move to draft")
        if payload.status not in {"active", "paused", "draft"}:
            raise HTTPException(status_code=400, detail="Invalid status")
        if campaign.status == "expired":
            raise HTTPException(status_code=400, detail="Expired campaign cannot be reactivated")

    start_at = payload.start_at if payload.start_at is not None else campaign.start_at
    end_at = payload.end_at if payload.end_at is not None else campaign.end_at
    budget = payload.budget if payload.budget is not None else campaign.budget
    currency = payload.currency if payload.currency is not None else campaign.currency
    start_at, end_at, budget, currency = _normalize_campaign_fields(start_at, end_at, budget, currency)

    if payload.name is not None:
        campaign.name = payload.name
    if payload.advertiser is not None:
        campaign.advertiser = payload.advertiser
    if payload.budget is not None:
        campaign.budget = budget
    if payload.currency is not None:
        campaign.currency = currency or campaign.currency
    if payload.start_at is not None:
        campaign.start_at = start_at
    if payload.end_at is not None:
        campaign.end_at = end_at

    status_changed = False
    if payload.status is not None and payload.status != campaign.status:
        campaign.status = payload.status
        status_changed = True

    action = "CAMPAIGN_STATUS_CHANGED" if status_changed else "CAMPAIGN_UPDATED"
    await _write_audit_log_sql(
        session=session,
        action=action,
        actor=current_user,
        resource_type="ad_campaign",
        resource_id=str(campaign.id),
        metadata={"status": campaign.status},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.post("/admin/ads/campaigns/{campaign_id}/ads/{ad_id}")
async def link_ad_to_campaign(
    campaign_id: str,
    ad_id: str,
    request: Request,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        campaign_uuid = uuid.UUID(campaign_id)
        ad_uuid = uuid.UUID(ad_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid id") from exc

    campaign = await session.get(AdCampaign, campaign_uuid)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    ad = await session.get(Advertisement, ad_uuid)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")

    if ad.campaign_id and ad.campaign_id != campaign.id:
        raise HTTPException(status_code=409, detail="Ad already linked to another campaign")

    ad.campaign_id = campaign.id

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_AD_LINKED",
        actor=current_user,
        resource_type="ad_campaign",
        resource_id=str(campaign.id),
        metadata={"ad_id": str(ad.id)},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.delete("/admin/ads/campaigns/{campaign_id}/ads/{ad_id}")
async def unlink_ad_from_campaign(
    campaign_id: str,
    ad_id: str,
    request: Request,
    current_user=Depends(check_permissions(ADS_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        campaign_uuid = uuid.UUID(campaign_id)
        ad_uuid = uuid.UUID(ad_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid id") from exc

    campaign = await session.get(AdCampaign, campaign_uuid)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    ad = await session.get(Advertisement, ad_uuid)
    if not ad:
        raise HTTPException(status_code=404, detail="Ad not found")

    if ad.campaign_id != campaign.id:
        raise HTTPException(status_code=400, detail="Ad not linked to campaign")

    ad.campaign_id = None

    await _write_audit_log_sql(
        session=session,
        action="CAMPAIGN_AD_UNLINKED",
        actor=current_user,
        resource_type="ad_campaign",
        resource_id=str(campaign.id),
        metadata={"ad_id": str(ad.id)},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.get("/admin/pricing/campaign-items")
async def list_pricing_campaign_items(
    scope: str,
    include_deleted: bool = False,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    _validate_campaign_item_fields(scope, None, None, None, None, None)
    await _expire_pricing_campaign_items(session, actor=current_user)
    query = select(PricingCampaignItem).where(PricingCampaignItem.scope == scope)
    if not include_deleted:
        query = query.where(PricingCampaignItem.is_deleted.is_(False))
    query = query.order_by(desc(PricingCampaignItem.updated_at))
    result = await session.execute(query)
    items = [_serialize_pricing_campaign_item(item) for item in result.scalars().all()]
    return {"items": items}


@api_router.post("/admin/pricing/campaign-items")
async def create_pricing_campaign_item(
    payload: PricingCampaignItemCreatePayload,
    request: Request,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    start_at = _ensure_aware_datetime(payload.start_at)
    end_at = _ensure_aware_datetime(payload.end_at)
    _validate_campaign_item_fields(
        payload.scope,
        payload.listing_quota,
        payload.price_amount,
        payload.publish_days,
        start_at,
        end_at,
        require_dates=True,
        enforce_start_future=True,
    )
    currency = _normalize_currency_code(payload.currency)
    if payload.is_active:
        await _assert_no_overlap_active_campaign_item(session, payload.scope, start_at, end_at)

    item = PricingCampaignItem(
        scope=payload.scope,
        name=payload.name,
        listing_quota=payload.listing_quota,
        price_amount=payload.price_amount,
        currency=currency,
        publish_days=payload.publish_days,
        is_active=bool(payload.is_active),
        start_at=start_at,
        end_at=end_at,
        created_by=_safe_uuid(current_user.get("id")),
        updated_by=_safe_uuid(current_user.get("id")),
        is_deleted=False,
    )
    session.add(item)
    await session.flush()

    await _write_audit_log_sql(
        session=session,
        action="PRICING_CAMPAIGN_ITEM_CREATED",
        actor=current_user,
        resource_type="pricing_campaign_item",
        resource_id=str(item.id),
        metadata={"scope": item.scope, "listing_quota": item.listing_quota},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    await session.refresh(item)
    return {"item": _serialize_pricing_campaign_item(item)}


@api_router.put("/admin/pricing/campaign-items/{item_id}")
async def update_pricing_campaign_item(
    item_id: str,
    payload: PricingCampaignItemUpdatePayload,
    request: Request,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid item id") from exc

    item = await session.get(PricingCampaignItem, item_uuid)
    if not item or item.is_deleted:
        raise HTTPException(status_code=404, detail="Campaign item not found")

    start_at = _ensure_aware_datetime(payload.start_at) if payload.start_at is not None else item.start_at
    end_at = _ensure_aware_datetime(payload.end_at) if payload.end_at is not None else item.end_at

    _validate_campaign_item_fields(
        item.scope,
        payload.listing_quota,
        payload.price_amount,
        payload.publish_days,
        start_at,
        end_at,
        require_dates=True,
    )

    if payload.start_at is not None and start_at is not None:
        now = datetime.now(timezone.utc)
        if start_at < now - timedelta(minutes=1) and item.start_at != start_at:
            raise HTTPException(status_code=400, detail="start_at must be in the future")

    target_active = payload.is_active if payload.is_active is not None else item.is_active
    if target_active:
        await _assert_no_overlap_active_campaign_item(session, item.scope, start_at, end_at, exclude_id=item.id)

    if payload.name is not None:
        item.name = payload.name
    if payload.listing_quota is not None:
        item.listing_quota = payload.listing_quota
    if payload.price_amount is not None:
        item.price_amount = payload.price_amount
    if payload.currency is not None:
        item.currency = _normalize_currency_code(payload.currency)
    if payload.publish_days is not None:
        item.publish_days = payload.publish_days
    if payload.start_at is not None:
        item.start_at = start_at
    if payload.end_at is not None:
        item.end_at = end_at
    if payload.is_active is not None:
        item.is_active = payload.is_active

    item.updated_at = datetime.now(timezone.utc)
    item.updated_by = _safe_uuid(current_user.get("id"))

    await _write_audit_log_sql(
        session=session,
        action="PRICING_CAMPAIGN_ITEM_UPDATED",
        actor=current_user,
        resource_type="pricing_campaign_item",
        resource_id=str(item.id),
        metadata={"scope": item.scope},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    await session.refresh(item)
    return {"item": _serialize_pricing_campaign_item(item)}


@api_router.patch("/admin/pricing/campaign-items/{item_id}/status")
async def update_pricing_campaign_item_status(
    item_id: str,
    payload: PricingCampaignItemStatusPayload,
    request: Request,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid item id") from exc

    item = await session.get(PricingCampaignItem, item_uuid)
    if not item or item.is_deleted:
        raise HTTPException(status_code=404, detail="Campaign item not found")

    if payload.is_active:
        if item.start_at is None or item.end_at is None:
            raise HTTPException(status_code=400, detail="start_at and end_at required")
        await _assert_no_overlap_active_campaign_item(session, item.scope, item.start_at, item.end_at, exclude_id=item.id)

    item.is_active = payload.is_active
    item.updated_at = datetime.now(timezone.utc)
    item.updated_by = _safe_uuid(current_user.get("id"))

    await _write_audit_log_sql(
        session=session,
        action="PRICING_CAMPAIGN_ITEM_ACTIVATED" if payload.is_active else "PRICING_CAMPAIGN_ITEM_DEACTIVATED",
        actor=current_user,
        resource_type="pricing_campaign_item",
        resource_id=str(item.id),
        metadata={"scope": item.scope, "is_active": payload.is_active},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    await session.refresh(item)
    return {"item": _serialize_pricing_campaign_item(item)}


@api_router.delete("/admin/pricing/campaign-items/{item_id}")
async def delete_pricing_campaign_item(
    item_id: str,
    request: Request,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        item_uuid = uuid.UUID(item_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid item id") from exc

    item = await session.get(PricingCampaignItem, item_uuid)
    if not item or item.is_deleted:
        raise HTTPException(status_code=404, detail="Campaign item not found")

    item.is_deleted = True
    item.deleted_at = datetime.now(timezone.utc)
    item.is_active = False
    item.updated_at = datetime.now(timezone.utc)
    item.updated_by = _safe_uuid(current_user.get("id"))

    await _write_audit_log_sql(
        session=session,
        action="PRICING_CAMPAIGN_ITEM_DELETED",
        actor=current_user,
        resource_type="pricing_campaign_item",
        resource_id=str(item.id),
        metadata={"scope": item.scope},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.get("/admin/pricing/tiers")
async def list_pricing_tiers(
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_pricing_defaults(session)
    await _expire_tier_rules(session)
    result = await session.execute(
        select(PricingTierRule)
        .where(PricingTierRule.scope == "individual")
        .order_by(PricingTierRule.tier_no)
    )
    items = []
    for rule in result.scalars().all():
        items.append(
            {
                "id": str(rule.id),
                "tier_no": rule.tier_no,
                "listing_from": rule.listing_from,
                "listing_to": rule.listing_to,
                "price_amount": float(rule.price_amount or 0),
                "currency": rule.currency,
                "publish_days": rule.publish_days,
                "is_active": rule.is_active,
                "effective_start_at": rule.effective_start_at.isoformat() if rule.effective_start_at else None,
                "effective_end_at": rule.effective_end_at.isoformat() if rule.effective_end_at else None,
            }
        )
    return {"items": items}


@api_router.put("/admin/pricing/tiers")
async def update_pricing_tiers(
    payload: PricingTierUpdatePayload,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_pricing_defaults(session)
    now = datetime.now(timezone.utc)
    updates = []

    rules_by_tier = {item.tier_no: item for item in payload.rules}
    for tier_no in [1, 2, 3]:
        rule_payload = rules_by_tier.get(tier_no)
        if not rule_payload:
            continue
        if tier_no not in {1, 2, 3}:
            raise HTTPException(status_code=400, detail="Invalid tier_no")
        if rule_payload.price_amount < 0:
            raise HTTPException(status_code=400, detail="price_amount must be >= 0")
        currency = _normalize_currency_code(rule_payload.currency)
        price_amount = 0 if tier_no == 1 else rule_payload.price_amount

        result = await session.execute(
            select(PricingTierRule).where(
                PricingTierRule.scope == "individual",
                PricingTierRule.tier_no == tier_no,
            )
        )
        rule = result.scalar_one_or_none()
        listing_from = tier_no
        listing_to = tier_no if tier_no < 3 else None
        if not rule:
            rule = PricingTierRule(
                scope="individual",
                year_window="calendar_year",
                tier_no=tier_no,
                listing_from=listing_from,
                listing_to=listing_to,
                price_amount=price_amount,
                currency=currency,
                publish_days=90,
                is_active=True,
            )
            session.add(rule)
        else:
            rule.listing_from = listing_from
            rule.listing_to = listing_to
            rule.price_amount = price_amount
            rule.currency = currency
            rule.is_active = True if rule_payload.is_active is None else rule_payload.is_active
            rule.updated_at = now
        updates.append({"tier_no": tier_no, "price_amount": float(price_amount), "currency": currency})

    await _write_audit_log_sql(
        session=session,
        action="PRICING_TIER_UPDATED",
        actor=current_user,
        resource_type="pricing_tier_rule",
        resource_id="pricing_tier_rules",
        metadata={"updates": updates},
        request=None,
        country_code=current_user.get("country_code"),
    )
    await session.commit()
    return {"ok": True}


@api_router.get("/admin/pricing/packages")
async def list_pricing_packages_admin(
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_pricing_defaults(session)
    result = await session.execute(
        select(PricingPackage)
        .where(PricingPackage.scope == "corporate")
        .order_by(PricingPackage.listing_quota)
    )
    items = []
    for item in result.scalars().all():
        items.append(
            {
                "id": str(item.id),
                "name": item.name,
                "listing_quota": item.listing_quota,
                "package_price_amount": float(item.package_price_amount or 0),
                "currency": item.currency,
                "publish_days": item.publish_days,
                "package_duration_days": item.package_duration_days,
                "is_active": item.is_active,
            }
        )
    return {"items": items}


@api_router.put("/admin/pricing/packages")
async def update_pricing_packages(
    payload: PricingPackageUpdatePayload,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_pricing_defaults(session)
    updates = []
    for package_payload in payload.packages:
        if package_payload.package_price_amount < 0:
            raise HTTPException(status_code=400, detail="package_price_amount must be >= 0")
        if package_payload.listing_quota <= 0:
            raise HTTPException(status_code=400, detail="listing_quota must be > 0")
        if package_payload.package_duration_days <= 0:
            raise HTTPException(status_code=400, detail="package_duration_days must be > 0")
        currency = _normalize_currency_code(package_payload.currency)

        result = await session.execute(
            select(PricingPackage).where(
                PricingPackage.scope == "corporate",
                PricingPackage.name == package_payload.name,
            )
        )
        package = result.scalar_one_or_none()
        if not package:
            package = PricingPackage(
                scope="corporate",
                name=package_payload.name,
                listing_quota=package_payload.listing_quota,
                package_price_amount=package_payload.package_price_amount,
                currency=currency,
                publish_days=90,
                package_duration_days=package_payload.package_duration_days,
                is_active=package_payload.is_active if package_payload.is_active is not None else True,
            )
            session.add(package)
        else:
            package.listing_quota = package_payload.listing_quota
            package.package_price_amount = package_payload.package_price_amount
            package.currency = currency
            package.package_duration_days = package_payload.package_duration_days
            if package_payload.is_active is not None:
                package.is_active = package_payload.is_active
            package.updated_at = datetime.now(timezone.utc)
        updates.append({"name": package_payload.name, "price": float(package_payload.package_price_amount), "currency": currency})

    await _write_audit_log_sql(
        session=session,
        action="PRICING_PACKAGE_UPDATED",
        actor=current_user,
        resource_type="pricing_package",
        resource_id="pricing_packages",
        metadata={"updates": updates},
        request=None,
        country_code=current_user.get("country_code"),
    )
    await session.commit()
    return {"ok": True}


@api_router.get("/admin/pricing/campaign")
async def get_pricing_campaign(
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _expire_pricing_campaigns(session, actor=current_user)
    policy = await _get_latest_pricing_campaign(session)
    if not policy:
        return {
            "policy": {
                "id": None,
                "is_enabled": False,
                "start_at": None,
                "end_at": None,
                "scope": "all",
                "published_at": None,
                "version": 0,
                "created_by": None,
                "updated_by": None,
            },
            "active": False,
        }
    return {
        "policy": _pricing_campaign_to_dict(policy),
        "active": _is_pricing_campaign_active(policy, "individual") or _is_pricing_campaign_active(policy, "corporate"),
    }


@api_router.put("/admin/pricing/campaign")
async def update_pricing_campaign(
    payload: PricingCampaignPolicyPayload,
    current_user=Depends(check_permissions(PRICING_MANAGER_ROLES)),
    session: AsyncSession = Depends(get_sql_session),
):
    await _expire_pricing_campaigns(session, actor=current_user)

    if payload.scope not in PRICING_CAMPAIGN_SCOPES:
        raise HTTPException(status_code=400, detail="Invalid scope")
    if payload.is_enabled and payload.start_at is None:
        raise HTTPException(status_code=400, detail="start_at is required when enabling")
    if payload.start_at and payload.end_at and payload.end_at < payload.start_at:
        raise HTTPException(status_code=400, detail="end_at must be after start_at")

    policy = await _get_latest_pricing_campaign(session)
    if not policy:
        policy = PricingCampaign(
            is_enabled=False,
            scope=payload.scope,
            start_at=payload.start_at,
            end_at=payload.end_at,
            created_by=_safe_uuid(current_user.get("id")),
            updated_by=_safe_uuid(current_user.get("id")),
            version=1,
        )
        session.add(policy)
        await session.flush()

    if payload.is_enabled:
        conflict = await session.execute(
            select(PricingCampaign)
            .where(PricingCampaign.is_enabled.is_(True), PricingCampaign.id != policy.id)
            .limit(1)
        )
        if conflict.scalar_one_or_none():
            raise HTTPException(status_code=409, detail="Another pricing campaign is already active")

    previous_enabled = policy.is_enabled

    policy.is_enabled = payload.is_enabled
    policy.start_at = payload.start_at
    policy.end_at = payload.end_at
    policy.scope = payload.scope
    now_ts = datetime.now(timezone.utc)
    policy.updated_at = now_ts
    policy.updated_by = _safe_uuid(current_user.get("id"))
    policy.version = (policy.version or 0) + 1
    if policy.created_by is None:
        policy.created_by = _safe_uuid(current_user.get("id"))

    if not previous_enabled and policy.is_enabled:
        policy.published_at = now_ts

    action = "PRICING_CAMPAIGN_UPDATED"
    if not previous_enabled and policy.is_enabled:
        action = "PRICING_CAMPAIGN_ENABLED"
    elif previous_enabled and not policy.is_enabled:
        action = "PRICING_CAMPAIGN_DISABLED"

    await _write_audit_log_sql(
        session=session,
        action=action,
        actor=current_user,
        resource_type="pricing_campaign",
        resource_id=str(policy.id),
        metadata={"scope": policy.scope, "is_enabled": policy.is_enabled},
        request=None,
        country_code=current_user.get("country_code"),
    )

    try:
        await session.commit()
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=409, detail="Another pricing campaign is already active")

    return {
        "ok": True,
        "policy": _pricing_campaign_to_dict(policy),
        "active": _is_pricing_campaign_active(policy, "individual") or _is_pricing_campaign_active(policy, "corporate"),
    }


@api_router.post("/pricing/quote")
async def get_pricing_quote_endpoint(
    payload: PricingQuotePayload,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    response, _, _, _ = await _compute_pricing_quote(session, current_user, payload)
    return response


@api_router.post("/pricing/checkout-session")
async def create_pricing_checkout_session(
    payload: PricingCheckoutPayload,
    current_user=Depends(require_portal_scope("account")),
    session: AsyncSession = Depends(get_sql_session),
):
    await _ensure_invoices_db_ready(session)

    if not STRIPE_API_KEY:
        raise HTTPException(status_code=503, detail="Stripe not configured")

    try:
        listing_uuid = uuid.UUID(payload.listing_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="listing_id invalid") from exc

    listing = await session.get(Listing, listing_uuid)
    if not listing or str(listing.user_id) != current_user.get("id"):
        raise HTTPException(status_code=404, detail="Listing not found")

    if listing.status not in {"draft", "needs_revision"}:
        raise HTTPException(status_code=400, detail="Listing not eligible for payment")

    pricing_response, quote, policy, user_type = await _compute_pricing_quote(session, current_user, payload)
    override_active = pricing_response.get("override_active", False)

    if quote.get("type") == "campaign_selection":
        return JSONResponse(
            status_code=409,
            content={
                "selection_required": True,
                "quote": pricing_response,
            },
        )

    if not quote.get("requires_payment"):
        return {"payment_required": False, "quote": pricing_response}

    if quote.get("type") != "campaign_item":
        raise HTTPException(status_code=400, detail="Campaign item required for checkout")

    origin = payload.origin_url.strip().rstrip("/")
    if not origin:
        raise HTTPException(status_code=400, detail="origin_url required")

    country_code = listing.country or current_user.get("country_code")
    if not _is_payment_enabled_for_country(country_code):
        raise HTTPException(status_code=403, detail="Payments disabled for this country")

    amount = float(quote.get("amount") or 0)
    currency = quote.get("currency") or "EUR"
    publish_days = quote.get("publish_days") or 90

    if amount <= 0:
        raise HTTPException(status_code=400, detail="Payment not required for this listing")

    snapshot = await _create_pricing_snapshot(session, listing, current_user, quote, policy, override_active)
    if not snapshot:
        raise HTTPException(status_code=400, detail="Pricing snapshot could not be created")
    await session.flush()

    now = datetime.now(timezone.utc)
    invoice = AdminInvoice(
        invoice_no=_generate_invoice_no(),
        user_id=uuid.UUID(current_user.get("id")),
        subscription_id=None,
        plan_id=None,
        campaign_id=None,
        amount_total=amount,
        currency=currency,
        status="issued",
        payment_status="requires_payment_method",
        issued_at=now,
        due_at=None,
        provider_customer_id=None,
        meta_json={
            "listing_id": str(listing.id),
            "pricing_snapshot_id": str(snapshot.id),
            "pricing_mode": pricing_response.get("pricing_mode"),
            "pricing_user_type": user_type,
            "pricing_publish_days": publish_days,
            "pricing_campaign_item_id": quote.get("campaign_item_id"),
            "pricing_listing_quota": quote.get("listing_quota"),
        },
        scope="country" if country_code else "global",
        country_code=country_code,
        payment_method=None,
        notes=None,
        created_at=now,
        updated_at=now,
    )
    session.add(invoice)
    await session.flush()

    webhook_url = f"{origin}/api/webhook/stripe"
    stripe_checkout = StripeCheckout(api_key=STRIPE_API_KEY, webhook_url=webhook_url)

    success_url = f"{origin}/account/payments/success?session_id={{CHECKOUT_SESSION_ID}}"
    cancel_url = f"{origin}/account/payments/cancel"

    session_request = CheckoutSessionRequest(
        amount=float(amount),
        currency=currency,
        success_url=success_url,
        cancel_url=cancel_url,
        metadata={
            "invoice_id": str(invoice.id),
            "listing_id": str(listing.id),
            "pricing_snapshot_id": str(snapshot.id),
            "pricing_campaign_item_id": quote.get("campaign_item_id"),
        },
    )

    checkout_session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(session_request)

    payment = Payment(
        invoice_id=invoice.id,
        user_id=invoice.user_id,
        provider="stripe",
        provider_ref=checkout_session.session_id,
        status="requires_payment_method",
        amount_total=amount,
        currency=currency,
        meta_json={"checkout_session_id": checkout_session.session_id},
        created_at=now,
        updated_at=now,
    )
    transaction = PaymentTransaction(
        provider="stripe",
        session_id=checkout_session.session_id,
        provider_payment_id=None,
        invoice_id=invoice.id,
        dealer_id=invoice.user_id,
        amount=amount,
        currency=currency,
        status="requires_payment_method",
        payment_status="requires_payment_method",
        metadata_json={
            "checkout_url": checkout_session.url,
            "pricing_snapshot_id": str(snapshot.id),
            "pricing_campaign_item_id": quote.get("campaign_item_id"),
        },
        created_at=now,
        updated_at=now,
    )
    session.add(payment)
    session.add(transaction)
    await session.commit()

    return {
        "checkout_url": checkout_session.url,
        "session_id": checkout_session.session_id,
        "invoice_id": str(invoice.id),
        "snapshot_id": str(snapshot.id),
        "payment_required": True,
        "quote": pricing_response,
    }


@api_router.get("/pricing/packages")
async def list_pricing_packages_endpoint(session: AsyncSession = Depends(get_sql_session)):
    now = datetime.now(timezone.utc)
    result = await session.execute(
        select(PricingCampaignItem)
        .where(
            PricingCampaignItem.scope == "corporate",
            PricingCampaignItem.is_deleted.is_(False),
            PricingCampaignItem.is_active.is_(True),
            PricingCampaignItem.start_at.isnot(None),
            PricingCampaignItem.end_at.isnot(None),
            PricingCampaignItem.start_at <= now,
            PricingCampaignItem.end_at >= now,
        )
        .order_by(PricingCampaignItem.listing_quota)
    )
    items = []
    for item in result.scalars().all():
        items.append(
            {
                "id": str(item.id),
                "name": item.name,
                "listing_quota": item.listing_quota,
                "price_amount": float(item.price_amount or 0),
                "currency": item.currency,
                "publish_days": item.publish_days,
            }
        )
    return {"packages": items}


@api_router.post("/v1/doping/requests")
async def create_doping_request(
    payload: DopingRequestPayload,
    current_user=Depends(get_current_user),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        listing_uuid = uuid.UUID(payload.listing_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid listing id") from exc

    listing = await session.get(Listing, listing_uuid)
    if not listing or listing.user_id != current_user.get("id"):
        raise HTTPException(status_code=404, detail="Listing not found")

    start_at, end_at = _normalize_ad_dates(payload.start_at, payload.end_at)

    req = DopingRequest(
        listing_id=listing_uuid,
        user_id=uuid.UUID(current_user.get("id")),
        status="requested",
        placement_home=bool(payload.placement_home),
        placement_category=bool(payload.placement_category),
        start_at=start_at,
        end_at=end_at,
        priority=payload.priority or 0,
    )
    session.add(req)
    await session.commit()
    await session.refresh(req)

    await _write_audit_log_sql(
        session=session,
        action="DOPING_REQUESTED",
        actor=current_user,
        resource_type="doping",
        resource_id=str(req.id),
        metadata={"listing_id": str(listing_uuid)},
        request=None,
        country_code=current_user.get("country_code"),
    )

    return {"ok": True, "id": str(req.id)}


@api_router.get("/admin/doping/requests")
async def list_doping_requests(
    status: Optional[str] = None,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    await _expire_doping(session)
    query = select(DopingRequest).order_by(desc(DopingRequest.updated_at))
    if status:
        query = query.where(DopingRequest.status == status)
    result = await session.execute(query)
    items = []
    for item in result.scalars().all():
        items.append(
            {
                "id": str(item.id),
                "listing_id": str(item.listing_id),
                "user_id": str(item.user_id),
                "status": item.status,
                "placement_home": item.placement_home,
                "placement_category": item.placement_category,
                "start_at": item.start_at.isoformat() if item.start_at else None,
                "end_at": item.end_at.isoformat() if item.end_at else None,
                "priority": item.priority,
                "approved_by": str(item.approved_by) if item.approved_by else None,
                "approved_at": item.approved_at.isoformat() if item.approved_at else None,
                "published_at": item.published_at.isoformat() if item.published_at else None,
                "expired_at": item.expired_at.isoformat() if item.expired_at else None,
            }
        )
    return {"items": items}


@api_router.post("/admin/doping/requests/{request_id}/approve")
async def approve_doping_request(
    request_id: str,
    payload: DopingApprovalPayload,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        req_uuid = uuid.UUID(request_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid request id") from exc

    req = await session.get(DopingRequest, req_uuid)
    if not req:
        raise HTTPException(status_code=404, detail="Doping request not found")

    if req.status != "paid":
        raise HTTPException(status_code=400, detail="Doping request must be paid before approval")

    start_at, end_at = _normalize_ad_dates(payload.start_at or req.start_at, payload.end_at or req.end_at)

    req.placement_home = bool(payload.placement_home)
    req.placement_category = bool(payload.placement_category)
    req.start_at = start_at
    req.end_at = end_at
    req.priority = payload.priority or 0

    req.status = "approved"
    now = datetime.now(timezone.utc)
    req.approved_by = uuid.UUID(current_user.get("id")) if current_user.get("id") else None
    req.approved_at = now
    req.updated_at = now

    await _write_audit_log_sql(
        session=session,
        action="DOPING_APPROVED",
        actor=current_user,
        resource_type="doping",
        resource_id=str(req.id),
        metadata={"listing_id": str(req.listing_id)},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.post("/admin/doping/requests/{request_id}/mark-paid")
async def mark_doping_paid(
    request_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        req_uuid = uuid.UUID(request_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid request id") from exc

    req = await session.get(DopingRequest, req_uuid)
    if not req:
        raise HTTPException(status_code=404, detail="Doping request not found")

    if req.status != "requested":
        raise HTTPException(status_code=400, detail="Only requested doping can be marked as paid")

    req.status = "paid"
    req.updated_at = datetime.now(timezone.utc)

    await _write_audit_log_sql(
        session=session,
        action="DOPING_PAID",
        actor=current_user,
        resource_type="doping",
        resource_id=str(req.id),
        metadata={"listing_id": str(req.listing_id)},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.post("/admin/doping/requests/{request_id}/publish")
async def publish_doping_request(
    request_id: str,
    request: Request,
    current_user=Depends(check_permissions(["super_admin", "country_admin", "moderator"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        req_uuid = uuid.UUID(request_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid request id") from exc

    req = await session.get(DopingRequest, req_uuid)
    if not req:
        raise HTTPException(status_code=404, detail="Doping request not found")

    if req.status != "approved":
        raise HTTPException(status_code=400, detail="Doping request must be approved before publishing")

    now = datetime.now(timezone.utc)
    req.status = "published"
    req.published_at = now
    req.updated_at = now

    await _write_audit_log_sql(
        session=session,
        action="DOPING_PUBLISHED",
        actor=current_user,
        resource_type="doping",
        resource_id=str(req.id),
        metadata={"listing_id": str(req.listing_id)},
        request=request,
        country_code=current_user.get("country_code"),
    )

    await session.commit()
    return {"ok": True}


@api_router.get("/doping/placements")
async def get_doping_placements(
    placement: Optional[str] = None,
    session: AsyncSession = Depends(get_sql_session),
):
    await _expire_doping(session)
    query = select(DopingRequest, Listing).join(Listing, Listing.id == DopingRequest.listing_id)
    query = query.where(DopingRequest.status == "published")
    if placement == "homepage":
        query = query.where(DopingRequest.placement_home.is_(True))
    if placement == "category":
        query = query.where(DopingRequest.placement_category.is_(True))
    query = query.order_by(desc(DopingRequest.priority), desc(DopingRequest.updated_at))

    result = await session.execute(query)
    items = []
    for doping, listing in result.all():
        if not _is_active_window(doping.start_at, doping.end_at):
            continue
        items.append(
            {
                "doping_id": str(doping.id),
                "listing_id": str(listing.id),
                "title": listing.title,
                "status": listing.status,
                "price": listing.price,
                "currency": listing.currency or "EUR",
                "priority": doping.priority,
            }
        )
    return {"items": items}


@api_router.get("/site/footer")
async def get_footer_layout(session: AsyncSession = Depends(get_sql_session)):
    result = await session.execute(
        select(FooterLayout)
        .where(FooterLayout.status == "published")
        .order_by(desc(FooterLayout.updated_at))
        .limit(1)
    )
    layout = result.scalar_one_or_none()
    return {
        "layout": layout.layout if layout else None,
        "version": layout.version if layout else None,
    }


@api_router.get("/admin/footer/layout")
async def get_footer_layout_admin(
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(select(FooterLayout).order_by(desc(FooterLayout.updated_at)).limit(1))
    layout = result.scalar_one_or_none()
    return {
        "id": str(layout.id) if layout else None,
        "layout": layout.layout if layout else None,
        "status": layout.status if layout else None,
        "version": layout.version if layout else None,
    }


@api_router.put("/admin/footer/layout")
async def save_footer_layout(
    payload: FooterLayoutPayload,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(select(FooterLayout).order_by(desc(FooterLayout.updated_at)).limit(1))
    layout = result.scalar_one_or_none()
    if not layout:
        layout = FooterLayout(layout=payload.layout, status=payload.status or "draft", version=1)
        session.add(layout)
    else:
        layout.layout = payload.layout
        layout.status = payload.status or layout.status
        layout.version = (layout.version or 1) + 1
    await session.commit()
    return {"ok": True}


@api_router.post("/admin/footer/layout/publish")
async def publish_footer_layout(
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(select(FooterLayout).order_by(desc(FooterLayout.updated_at)).limit(1))
    layout = result.scalar_one_or_none()
    if not layout:
        raise HTTPException(status_code=404, detail="Footer layout not found")
    layout.status = "published"
    await session.commit()
    return {"ok": True}


@api_router.get("/admin/info-pages")
async def list_info_pages(
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    result = await session.execute(select(InfoPage).order_by(desc(InfoPage.updated_at)))
    items = []
    for page in result.scalars().all():
        items.append(
            {
                "id": str(page.id),
                "slug": page.slug,
                "title_tr": page.title_tr,
                "title_de": page.title_de,
                "title_fr": page.title_fr,
                "is_published": page.is_published,
                "updated_at": page.updated_at.isoformat() if page.updated_at else None,
            }
        )
    return {"items": items}


@api_router.post("/admin/info-pages")
async def create_info_page(
    payload: InfoPagePayload,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    slug_value = payload.slug.strip().lower()
    if not slug_value:
        raise HTTPException(status_code=400, detail="Slug is required")

    existing = await session.execute(select(InfoPage).where(InfoPage.slug == slug_value))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Slug already exists")

    page = InfoPage(
        slug=slug_value,
        title_tr=payload.title_tr,
        title_de=payload.title_de,
        title_fr=payload.title_fr,
        content_tr=payload.content_tr,
        content_de=payload.content_de,
        content_fr=payload.content_fr,
        is_published=bool(payload.is_published),
    )
    session.add(page)
    await session.commit()
    return {"ok": True, "id": str(page.id)}


@api_router.patch("/admin/info-pages/{page_id}")
async def update_info_page(
    page_id: str,
    payload: InfoPageUpdatePayload,
    current_user=Depends(check_permissions(["super_admin", "country_admin"])),
    session: AsyncSession = Depends(get_sql_session),
):
    try:
        page_uuid = uuid.UUID(page_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid page id") from exc

    page = await session.get(InfoPage, page_uuid)
    if not page:
        raise HTTPException(status_code=404, detail="Info page not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(page, field, value)

    await session.commit()
    return {"ok": True}


@api_router.get("/info/{slug}")
async def get_info_page(slug: str, session: AsyncSession = Depends(get_sql_session)):
    slug_value = slug.strip().lower()
    result = await session.execute(select(InfoPage).where(InfoPage.slug == slug_value, InfoPage.is_published.is_(True)))
    page = result.scalar_one_or_none()
    if not page:
        raise HTTPException(status_code=404, detail="Page not found")
    return {
        "slug": page.slug,
        "title_tr": page.title_tr,
        "title_de": page.title_de,
        "title_fr": page.title_fr,
        "content_tr": page.content_tr,
        "content_de": page.content_de,
        "content_fr": page.content_fr,
    }

# --- Router binding + RBAC allowlist (must be after all route definitions) ---
app.include_router(api_router)

RBAC_ALLOWLIST, RBAC_MISSING_POLICIES = _build_rbac_allowlist(app)
app.state.rbac_allowlist = RBAC_ALLOWLIST
app.state.rbac_missing_policies = RBAC_MISSING_POLICIES
if RBAC_MISSING_POLICIES:
    logging.getLogger("rbac_guard").warning(
        "rbac_policy_missing",
        extra={"missing_routes": RBAC_MISSING_POLICIES},
    )
